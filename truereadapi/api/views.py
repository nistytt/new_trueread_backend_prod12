from datetime import date
from django.db.models import Q, Exists, OuterRef
from calendar import monthrange
from django.db import DatabaseError, IntegrityError
from api.models import SupervsiorLocation
import uuid
from api.models import SupervisorLogin
from rest_framework.decorators import api_view
from .services.uptime_service import get_lambda_uptime, get_rds_uptime,get_lambda_uptime_by_range
from datetime import date, timedelta
from django.shortcuts import render
from rest_framework.response import Response
import requests
from rest_framework.decorators import api_view, permission_classes
from .models import Consumers, MeterReaderRegistration, Office, SupervisorLogin, UserManagement
from .serializers import (
    MeterReaderRegistrationSerializer,
    ConsumerDataSerializer,
    ConsumerWiseDetailsSerializer,
    MridSerializer,
    Serail,
    ConsumersMeterRegistration,
    SupervisorLoginSerializer,
    UserManagementSerializer,
)
from django.db.models import Q
from rest_framework import status
import datetime
from .serializers import ConsumerSerializer
from rest_framework.decorators import parser_classes

from rest_framework.parsers import MultiPartParser, FormParser
from django.db import connection
import json
import jwt
import requests
from django.contrib.auth import authenticate
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.pagination import LimitOffsetPagination

from .serializers import FailedImageSerializer
from django.db import transaction
from django.utils.dateparse import parse_date

from django.http import JsonResponse
from django.db.models import Count, Case, When, IntegerField, F

# from decouple import config
from django.db.models import Q
from rest_framework.pagination import PageNumberPagination
from django.core.paginator import Paginator
import base64
from datetime import datetime, timedelta, date
from copy import deepcopy
import math
from rest_framework.views import APIView
from rest_framework.response import Response
from openpyxl import Workbook, load_workbook
from django.http import HttpResponse
from geojson import Point, Feature, FeatureCollection
from pyproj import CRS

from django.forms.models import model_to_dict


# import xlswriter
from django_filters import FilterSet

# Create your views here.

SECRETKEY = "6AZJYQ2T317WGPXC0UHVLDOR49FIBS8N5ME"


# ------------------------------v6.2.19--------------------------------------
# @parser_classes([MultiPartParser, FormParser])
# @api_view(["POST"])
# def consumers(request):
#     data = request.data.copy()
#     # _mutable = data._mutable
#     # data._mutable = True
#     rdng_date = data["rdng_date"]
#     cons_name = data["cons_name"]
#     cons_ac_no = data["cons_ac_no"]
#     if cons_name == "Test":
#         return Response({"status": True, "message": "Test data not inserted"})

#     # GET OFFICE DATA

#     reading_date_db = rdng_date[:10]
#     adddate = "-01"
#     bill_month_add = reading_date_db[:7] + adddate
#     print(bill_month_add)

#     reading_year = reading_date_db[:4]
#     reading_month = reading_date_db[5:7]
#     print(reading_date_db)
#     print(reading_month)
#     status = ""
#     # comment this line
#     data["reading_date_db"] = reading_date_db
#     data["bill_month_dt"] = bill_month_add
#     char = 0
#     ba_bl_id = data["ba_bl_id"]
#     try:
#         if data["prsnt_mtr_status"] == "Ok":
#             if data["prsnt_ocr_rdng"] != "Not Found":
#                 prsnt_ocr_rdng_temp = str(int(data["prsnt_ocr_rdng"]))
#                 prsnt_rdng_temp = str(int(data["prsnt_rdng"]))
#                 temp = max(int(prsnt_ocr_rdng_temp), int(prsnt_rdng_temp))
#                 temp1 = min(int(prsnt_ocr_rdng_temp), int(prsnt_rdng_temp))
#                 print("prsnt_ocr_rdng_temp--------->", prsnt_ocr_rdng_temp)
#                 print("prsnt_rdng_temp--------->", prsnt_rdng_temp)
#                 print("data['prsnt_ocr_rdng']--------->", data["prsnt_ocr_rdng"])
#                 print("data['prsnt_rdng']--------->", data["prsnt_rdng"])
#                 flag = False
#                 if (prsnt_ocr_rdng_temp) == (prsnt_rdng_temp):
#                     status = "Exact"

#                 elif (len(prsnt_ocr_rdng_temp) - len(prsnt_rdng_temp)) in (-1, 1):
#                     # temp =max(int(prsnt_ocr_rdng_temp),int(prsnt_rdng_temp))
#                     # temp1 =min(int(prsnt_ocr_rdng_temp),int(prsnt_rdng_temp))

#                     for x in range(len(str(temp))):
#                         temp_t = str(temp)
#                         l = list(temp_t)
#                         l.pop(x)
#                         if l == list(str(temp1)):
#                             flag = True
#                     if flag == True:
#                         status = "1_val_miss"
#                     else:
#                         status = "diff"
#                 elif len(prsnt_ocr_rdng_temp) == len(prsnt_rdng_temp):
#                     for x in range(len(prsnt_ocr_rdng_temp)):
#                         u, v = prsnt_ocr_rdng_temp, prsnt_rdng_temp
#                         if u[x] != v[x]:
#                             char += 1

#                     if char == 1:
#                         status = "1_val_diff"
#                     else:
#                         status = "diff"

#                 else:
#                     print("OOKOOKOKOKOKOK")
#                     if (str(temp1) in str(temp)) and (len(str(temp1)) > 1):
#                         status = "subs"
#                     else:
#                         status = "diff"

#                 if status == "1_val_diff" or status == "1_val_miss" or status == "subs":
#                     print("INSIDE UPDATESSSSS")
#                     data["prsnt_rdng_ocr_odv"] = data["prsnt_ocr_rdng"]
#                     data["rdng_ocr_status_odv"] = data["rdng_ocr_status"]
#                     data["prsnt_ocr_excep_old_values"] = data["prsnt_rdng_ocr_excep"]
#                     data["prsnt_ocr_rdng"] = data["prsnt_rdng"]
#                     data["rdng_ocr_status"] = "Passed"
#                     data["manual_update_flag"] = "true"
#                     data["prsnt_rdng_ocr_excep"] = ""
#                     data["rdng_ocr_status_changed_by"] = "Backend_CC"
#     except:
#         pass
#     try:
#         if data["prsnt_mtr_status"] == "Ok" and data["rdng_ocr_status"] == "":
#             data["rdng_ocr_status"] = "Failed"
#             data["manual_update_flag"] = "true"
#             data["rdng_ocr_status_changed_by"] = "Backend_RERUN"
#     except:
#         pass

#     # data._mutable = _mutable
#     cursor = connection.cursor()
#     # if bill id is present check the consumer ac no and month

#     newid = (
#         Consumers.objects.filter(
#             Q(bill_month_dt=bill_month_add) & Q(cons_ac_no=cons_ac_no)
#         )
#         .order_by("-id")
#         .first()
#     )

#     print("newid", newid)
#     if newid is not None:
#         print("updated")
#         serializer = ConsumerSerializer(newid, data=data, partial=True)
#         if serializer.is_valid():
#             serializer.save()
#             return Response({"status": True, "message": "Data Updated successfully"})
#         return Response(serializer.errors)

#     # if bill id is not present then insert
#     print("inserted")
#     serializer = ConsumerSerializer(data=data)
#     if serializer.is_valid():
#         serializer.save()
#         return Response(
#             {"status": True, "message": "Data added successfully", "version": "28"}
#         )
#     return Response(serializer.errors)
@parser_classes([MultiPartParser, FormParser])
@api_view(["POST"])
def consumers(request):
    data = request.data.copy()
    # _mutable = data._mutable
    # data._mutable = True
    rdng_date = data["rdng_date"]
    cons_name = data["cons_name"]
    cons_ac_no = data["cons_ac_no"]
    ofc_section = data['ofc_section']

    if cons_name == "Test":
        return Response({"status": True, "message": "Test data not inserted"})

    reading_date_db = rdng_date[:10]
    adddate = "-01"
    bill_month_add = reading_date_db[:7] + adddate
    print(bill_month_add)

    reading_year = reading_date_db[:4]
    reading_month = reading_date_db[5:7]
    print(reading_date_db)
    print(reading_month)
    status = ""
    # comment this line
    data["reading_date_db"] = reading_date_db
    data["bill_month_dt"] = bill_month_add

    # NEW LOGIC : IF OCR READING IS NOT FOUND SET IMAGE BLUR
    if data.get("prsnt_ocr_rdng") == "Not Found":
        data["prsnt_rdng_ocr_excep"] = "Image blur"

    char = 0
    ba_bl_id = data["ba_bl_id"]
    try:
        if data["prsnt_mtr_status"] == "Ok":
            if data["prsnt_ocr_rdng"] != "Not Found":
                prsnt_ocr_rdng_temp = str(int(data["prsnt_ocr_rdng"]))
                prsnt_rdng_temp = str(int(data["prsnt_rdng"]))
                temp = max(int(prsnt_ocr_rdng_temp), int(prsnt_rdng_temp))
                temp1 = min(int(prsnt_ocr_rdng_temp), int(prsnt_rdng_temp))
                print("prsnt_ocr_rdng_temp--------->", prsnt_ocr_rdng_temp)
                print("prsnt_rdng_temp--------->", prsnt_rdng_temp)
                print("data['prsnt_ocr_rdng']--------->",
                      data["prsnt_ocr_rdng"])
                print("data['prsnt_rdng']--------->", data["prsnt_rdng"])
                flag = False
                if (prsnt_ocr_rdng_temp) == (prsnt_rdng_temp):
                    status = "Exact"

                elif (len(prsnt_ocr_rdng_temp) - len(prsnt_rdng_temp)) in (-1, 1):
                    # temp =max(int(prsnt_ocr_rdng_temp),int(prsnt_rdng_temp))
                    # temp1 =min(int(prsnt_ocr_rdng_temp),int(prsnt_rdng_temp))

                    for x in range(len(str(temp))):
                        temp_t = str(temp)
                        l = list(temp_t)
                        l.pop(x)
                        if l == list(str(temp1)):
                            flag = True
                    if flag == True:
                        status = "1_val_miss"
                    else:
                        status = "diff"
                elif len(prsnt_ocr_rdng_temp) == len(prsnt_rdng_temp):
                    for x in range(len(prsnt_ocr_rdng_temp)):
                        u, v = prsnt_ocr_rdng_temp, prsnt_rdng_temp
                        if u[x] != v[x]:
                            char += 1

                    if char == 1:
                        status = "1_val_diff"
                    else:
                        status = "diff"

                else:
                    print("OOKOOKOKOKOKOK")
                    if (str(temp1) in str(temp)) and (len(str(temp1)) > 1):
                        status = "subs"
                    else:
                        status = "diff"

                if status == "1_val_diff" or status == "1_val_miss" or status == "subs":
                    print("INSIDE UPDATESSSSS")
                    data["prsnt_rdng_ocr_odv"] = data["prsnt_ocr_rdng"]
                    data["rdng_ocr_status_odv"] = data["rdng_ocr_status"]
                    data["prsnt_ocr_excep_old_values"] = data["prsnt_rdng_ocr_excep"]
                    data["prsnt_ocr_rdng"] = data["prsnt_rdng"]
                    data["rdng_ocr_status"] = "Passed"
                    data["manual_update_flag"] = "true"
                    data["prsnt_rdng_ocr_excep"] = ""
                    data["rdng_ocr_status_changed_by"] = "Backend_CC"
    except:
        pass
    try:
        if data["prsnt_mtr_status"] == "Ok" and data["rdng_ocr_status"] == "":
            data["rdng_ocr_status"] = "Failed"
            data["manual_update_flag"] = "true"
            data["rdng_ocr_status_changed_by"] = "Backend_RERUN"
    except:
        pass

    # if bill id is present check the consumer ac no and month

    newid = (
        Consumers.objects.filter(
            Q(bill_month_dt=bill_month_add) & Q(
                cons_ac_no=cons_ac_no) & Q(ofc_section=ofc_section)
        )
        .order_by("-id")
        .first()
    )

    print("newid", newid)
    if newid is not None:
        # Check all the columns only then update
        if (
            newid.ofc_discom == data['ofc_discom'] and
            newid.ofc_zone == data['ofc_zone'] and
            newid.ofc_circle == data['ofc_circle'] and
            newid.ofc_division == data['ofc_division'] and
            newid.ofc_sub_div_code == data['ofc_sub_div_code'] and
            newid.ofc_subdivision == data['ofc_subdivision'] and
            newid.ofc_section == data['ofc_section'] and
            newid.mr_unit == data['mr_unit'] and
            newid.bl_area_code == data['bl_area_code'] and
            newid.bl_agnc_type == data['bl_agnc_type'] and
            newid.bl_agnc_name == data['bl_agnc_name'] and
            newid.mr_id == data['mr_id'] and
            newid.mr_ph_no == data['mr_ph_no'] and
            newid.cons_ac_no == data['cons_ac_no'] and
            newid.cons_name == data['cons_name'] and
            newid.con_trf_cat == data['con_trf_cat'] and
            newid.con_mtr_sl_no == data['con_mtr_sl_no'] and
            newid.con_mtr_phs == data['con_mtr_phs'] and
            newid.rdng_req_val == data['rdng_req_val'] and
            newid.prev_rdng == data['prev_rdng'] and
            newid.prev_md == data['prev_md'] and
            newid.prev_pf_rdng == data['prev_pf_rdng'] and
            newid.prev_rdng_date == data['prev_rdng_date'] and
            newid.prev_rdng_status == data['prev_rdng_status'] and
            newid.bl_mnth == data['bl_mnth'] and
            newid.rdng_date == data['rdng_date'] and
            newid.geo_lat == data['geo_lat'] and
            newid.geo_long == data['geo_long'] and
            newid.prsnt_mtr_status == data['prsnt_mtr_status'] and
            newid.abnormality == data['abnormality'] and
            newid.mr_rmrk == data['mr_rmrk'] and
            newid.rdng_ocr_status == data['rdng_ocr_status'] and
            newid.prsnt_ocr_rdng == data['prsnt_ocr_rdng'] and
            newid.prsnt_rdng == data['prsnt_rdng'] and
            newid.prsnt_rdng_ocr_excep == data['prsnt_rdng_ocr_excep'] and
            newid.rdng_img == data['rdng_img'] and
            newid.ocr_md_status == data['ocr_md_status'] and
            newid.prsnt_md_rdng_ocr == data['prsnt_md_rdng_ocr'] and
            newid.prsnt_md_rdng == data['prsnt_md_rdng'] and
            newid.md_ocr_excep == data['md_ocr_excep'] and
            newid.md_img == data['md_img'] and
            newid.ocr_pf_status == data['ocr_pf_status'] and
            newid.ocr_pf_reading == data['ocr_pf_reading'] and
            newid.pf_manual_reading == data['pf_manual_reading'] and
            newid.pf_ocr_exception == data['pf_ocr_exception'] and
            newid.pf_image == data['pf_image'] and
            newid.ai_mdl_ver == data['ai_mdl_ver'] and
            newid.ph_name == data['ph_name'] and
            newid.cmra_res == data['cmra_res'] and
            newid.andr_ver == data['andr_ver'] and
            newid.qc_req == data['qc_req'] and
            newid.ba_cons_id == data['ba_cons_id'] and
            newid.ba_ac_id == data['ba_ac_id'] and
            newid.ba_prsnt_rdng_status == data['ba_prsnt_rdng_status'] and
            newid.ba_mrc == data['ba_mrc'] and
            newid.ba_mru == data['ba_mru'] and
            newid.ba_subdiv == data['ba_subdiv'] and
            newid.ba_div == data['ba_div'] and
            newid.ba_agnc_id == data['ba_agnc_id'] and
            newid.ba_bl_id == data['ba_bl_id'] and
            newid.ba_bl_date == data['ba_bl_date'] and
            newid.ba_prev_rdng_status == data['ba_prev_rdng_status'] and
            newid.qc_done == data['qc_done'] and
            newid.qc_done_user_id == data['qc_done_user_id'] and
            newid.qc_date == data['qc_date'] and
            newid.qc_flag == data['qc_flag'] and
            newid.qc_rmrk == data['qc_rmrk'] and
            newid.ai_retrain == data['ai_retrain'] and
            newid.is_object_meter == data['is_object_meter'] and
            newid.mr_success_feedback == data['mr_success_feedback'] and
            newid.reading_parameter_type == data['reading_parameter_type'] and
            newid.md_reading_parameter_type == data['md_reading_parameter_type'] and
            newid.pf_reading_parameter_type == data['pf_reading_parameter_type'] and
            newid.rdng_ocr_status_changed_by == data['rdng_ocr_status_changed_by'] and
            newid.prsnt_rdng_ocr_odv == data['prsnt_rdng_ocr_odv'] and
            newid.rdng_ocr_status_odv == data['rdng_ocr_status_odv'] and
            newid.prsnt_ocr_excep_old_values == data['prsnt_ocr_excep_old_values']
        ):
            print("updatation does not takes place")
            return Response({"status": True, "message": "No change in Data"})
        else:
            print("updating as some of the values are changed")

            serializer = ConsumerSerializer(newid, data=data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({"status": True, "message": "Data Updated successfully"})
            return Response(serializer.errors)

    # if bill id is not present then insert
    print("inserted")
    serializer = ConsumerSerializer(data=data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"status": True, "message": "Data added successfully", "version": "28"}
        )
    return Response(serializer.errors)


@api_view(["POST"])
def consumers_bulk(request):
    data_list = request.data.copy()
    print("datalist--->", data_list)
    count_insert = 0
    count_update = 0
    failed_consumers = []
    for data in data_list:
        # ensure defaults for new keys
        data['kvah_manual'] = data.get('kvah_manual', None)
        data['kvah_Status'] = data.get('kvah_Status', None)
        data['mtr_sr_no'] = data.get('mtr_sr_no', None)

        rdng_date = data["rdng_date"]
        cons_name = data["cons_name"]
        cons_ac_no = data["cons_ac_no"]
        ofc_section = data['ofc_section']

        if cons_name == "Test":
            return Response({"status": True, "message": "Test data not inserted"})

        reading_date_db = rdng_date[:10]
        adddate = "-01"
        bill_month_add = reading_date_db[:7] + adddate
        print(bill_month_add)
 
        reading_year = reading_date_db[:4]
        reading_month = reading_date_db[5:7]
        print(reading_date_db)
        print(reading_month)
        status = ""
        # comment this line
        data["reading_date_db"] = reading_date_db
        data["bill_month_dt"] = bill_month_add


        char = 0
        ba_bl_id = data["ba_bl_id"]

        rdngImg = data.get("rdng_img")

        # Ensure rdngImg is a clean string (not list or nested)
        if isinstance(rdngImg, list):
            rdngImg = rdngImg[0]
        elif isinstance(rdngImg, dict) and "url" in rdngImg:
            rdngImg = rdngImg["url"]
        elif not isinstance(rdngImg, str):
            rdngImg = str(rdngImg)

        # strip extra quotes or spaces
        rdngImg = rdngImg.strip('"').strip()

        try:
            if data["prsnt_mtr_status"] == "Ok":
                if data["prsnt_ocr_rdng"] != "Not Found":
                    prsnt_ocr_rdng_temp = str(int(data["prsnt_ocr_rdng"]))
                    prsnt_rdng_temp = str(int(data["prsnt_rdng"]))
                    temp = max(int(prsnt_ocr_rdng_temp), int(prsnt_rdng_temp))
                    temp1 = min(int(prsnt_ocr_rdng_temp), int(prsnt_rdng_temp))
                    print("prsnt_ocr_rdng_temp--------->", prsnt_ocr_rdng_temp)
                    print("prsnt_rdng_temp--------->", prsnt_rdng_temp)
                    print("data['prsnt_ocr_rdng']--------->",
                          data["prsnt_ocr_rdng"])
                    print("data['prsnt_rdng']--------->", data["prsnt_rdng"])
                    flag = False
                    if (prsnt_ocr_rdng_temp) == (prsnt_rdng_temp):
                        status = "Exact"
 
                    elif (len(prsnt_ocr_rdng_temp) - len(prsnt_rdng_temp)) in (-1, 1):
                        for x in range(len(str(temp))):
                            temp_t = str(temp)
                            l = list(temp_t)
                            l.pop(x)
                            if l == list(str(temp1)):
                                flag = True
                        if flag == True:
                            status = "1_val_miss"
                        else:
                            status = "diff"
                    elif len(prsnt_ocr_rdng_temp) == len(prsnt_rdng_temp):
                        for x in range(len(prsnt_ocr_rdng_temp)):
                            u, v = prsnt_ocr_rdng_temp, prsnt_rdng_temp
                            if u[x] != v[x]:
                                char += 1
 
                        if char == 1:
                            status = "1_val_diff"
                        else:
                            status = "diff"
 
                    else:
                        print("OOKOOKOKOKOKOK")
                        if (str(temp1) in str(temp)) and (len(str(temp1)) > 1):
                            status = "subs"
                        else:
                            status = "diff"
 
                    if status == "1_val_diff" or status == "1_val_miss" or status == "subs":
                        print("INSIDE UPDATESSSSS")
                        data["prsnt_rdng_ocr_odv"] = data["prsnt_ocr_rdng"]
                        data["rdng_ocr_status_odv"] = data["rdng_ocr_status"]
                        data["prsnt_ocr_excep_old_values"] = data["prsnt_rdng_ocr_excep"]
                        data["prsnt_ocr_rdng"] = data["prsnt_rdng"]
                        data["rdng_ocr_status"] = "Passed"
                        data["manual_update_flag"] = "true"
                        data["prsnt_rdng_ocr_excep"] = ""
                        data["rdng_ocr_status_changed_by"] = "Backend_CC"
        except:
            pass
        try:
            if data["prsnt_mtr_status"] == "Ok" and data["rdng_ocr_status"] == "" :
                data["rdng_ocr_status"] = "Failed"
                data["manual_update_flag"] = "true"
                data["rdng_ocr_status_changed_by"] = "Backend_RERUN"
        except:
            pass
        
                # ---------------- METER COMPATIBILITY LOGIC ----------------
# new logic fro exception addition in the new column

        # prsnt_ocr_status = data.get("prsnt_ocr_status")
        # ocr_md_status = data.get("ocr_md_status")
        # kvah_status = data.get("kvah_Status")
        # ocr_pf_status = data.get("ocr_pf_status")

        # meter_status_result = "Incompatible meter"  # default

        # # -------- PASSED --------
        # if (
        #     prsnt_ocr_status == "Passed" and
        #     ocr_md_status == "Passed" and
        #     (
        #         (kvah_status is None and ocr_pf_status is None) or
        #         (kvah_status == "Passed" and ocr_pf_status == "Passed")
        #     )
        # ):
        #     meter_status_result = "Passed"

        # # -------- WITH EXCEPTION --------
        # elif (
        #     prsnt_ocr_status == "Failed" and
        #     ocr_md_status == "Failed" and
        #     (
        #         (kvah_status is None or ocr_pf_status is None) or
        #         (kvah_status == "Failed" and ocr_pf_status == "Failed")
        #     )
        # ):
        #     meter_status_result = "with exception"

        # data["meter_status_result"] = meter_status_result

        # if bill id is present check the consumer ac no and month
 
        try:
            newid = (
                Consumers.objects.filter(
                    Q(bill_month_dt=bill_month_add) & Q(
                        cons_ac_no=cons_ac_no) & Q(ofc_section=ofc_section)
                )
                .order_by("-id")
                .first()
            )
            print("newid", newid)
            if newid is not None:
                # Check all the columns only then update
                if (
                    newid.ofc_discom == data['ofc_discom'] and
                    newid.ofc_zone == data['ofc_zone'] and
                    newid.ofc_circle == data['ofc_circle'] and
                    newid.ofc_division == data['ofc_division'] and
                    newid.ofc_sub_div_code == data['ofc_sub_div_code'] and
                    newid.ofc_subdivision == data['ofc_subdivision'] and
                    newid.ofc_section == data['ofc_section'] and
                    newid.mr_unit == data['mr_unit'] and
                    newid.bl_area_code == data['bl_area_code'] and
                    newid.bl_agnc_type == data['bl_agnc_type'] and
                    newid.bl_agnc_name == data['bl_agnc_name'] and
                    newid.mr_id == data['mr_id'] and
                    newid.mr_ph_no == data['mr_ph_no'] and
                    newid.cons_ac_no == data['cons_ac_no'] and
                    newid.cons_name == data['cons_name'] and
                    newid.con_trf_cat == data['con_trf_cat'] and
                    newid.con_mtr_sl_no == data['con_mtr_sl_no'] and
                    newid.con_mtr_phs == data['con_mtr_phs'] and
                    newid.rdng_req_val == data['rdng_req_val'] and
                    newid.prev_rdng == data['prev_rdng'] and
                    newid.prev_md == data['prev_md'] and
                    newid.prev_pf_rdng == data['prev_pf_rdng'] and
                    newid.prev_rdng_date == data['prev_rdng_date'] and
                    newid.prev_rdng_status == data['prev_rdng_status'] and
                    newid.bl_mnth == data['bl_mnth'] and
                    newid.rdng_date == data['rdng_date'] and
                    newid.geo_lat == data['geo_lat'] and
                    newid.geo_long == data['geo_long'] and
                    newid.prsnt_mtr_status == data['prsnt_mtr_status'] and
                    newid.abnormality == data['abnormality'] and
                    newid.mr_rmrk == data['mr_rmrk'] and
                    newid.rdng_ocr_status == data['rdng_ocr_status'] and
                    newid.prsnt_ocr_rdng == data['prsnt_ocr_rdng'] and
                    newid.prsnt_rdng == data['prsnt_rdng'] and
                    newid.prsnt_rdng_ocr_excep == data['prsnt_rdng_ocr_excep'] and
                    newid.rdng_img == data['rdng_img'] and
                    newid.ocr_md_status == data['ocr_md_status'] and
                    newid.prsnt_md_rdng_ocr == data['prsnt_md_rdng_ocr'] and
                    newid.prsnt_md_rdng == data['prsnt_md_rdng'] and
                    newid.md_ocr_excep == data['md_ocr_excep'] and
                    newid.md_img == data['md_img'] and
                    newid.ocr_pf_status == data['ocr_pf_status'] and
                    newid.ocr_pf_reading == data['ocr_pf_reading'] and
                    newid.pf_manual_reading == data['pf_manual_reading'] and
                    newid.pf_ocr_exception == data['pf_ocr_exception'] and
                    newid.pf_image == data['pf_image'] and
                    newid.ai_mdl_ver == data['ai_mdl_ver'] and
                    newid.ph_name == data['ph_name'] and
                    newid.cmra_res == data['cmra_res'] and
                    newid.andr_ver == data['andr_ver'] and
                    newid.data_sync_date == data['data_sync_date'] and
                    newid.qc_req == data['qc_req'] and
                    newid.ba_cons_id == data['ba_cons_id'] and
                    newid.ba_ac_id == data['ba_ac_id'] and
                    newid.ba_prsnt_rdng_status == data['ba_prsnt_rdng_status'] and
                    newid.ba_mrc == data['ba_mrc'] and
                    newid.ba_mru == data['ba_mru'] and
                    newid.ba_subdiv == data['ba_subdiv'] and
                    newid.ba_div == data['ba_div'] and
                    newid.ba_agnc_id == data['ba_agnc_id'] and
                    newid.ba_bl_id == data['ba_bl_id'] and
                    newid.ba_bl_date == data['ba_bl_date'] and
                    newid.ba_prev_rdng_status == data['ba_prev_rdng_status'] and
                    newid.qc_done == data['qc_done'] and
                    newid.qc_done_user_id == data['qc_done_user_id'] and
                    newid.qc_date == data['qc_date'] and
                    newid.qc_flag == data['qc_flag'] and
                    newid.qc_rmrk == data['qc_rmrk'] and
                    newid.ai_retrain == data['ai_retrain'] and
                    newid.is_object_meter == data['is_object_meter'] and
                    newid.mr_success_feedback == data['mr_success_feedback'] and
                    newid.reading_parameter_type == data['reading_parameter_type'] and
                    newid.md_reading_parameter_type == data['md_reading_parameter_type'] and
                    newid.pf_reading_parameter_type == data['pf_reading_parameter_type'] and
                    newid.rdng_ocr_status_changed_by == data['rdng_ocr_status_changed_by'] and
                    newid.prsnt_rdng_ocr_odv == data['prsnt_rdng_ocr_odv'] and
                    newid.rdng_ocr_status_odv == data['rdng_ocr_status_odv'] and
                    newid.prsnt_ocr_excep_old_values == data['prsnt_ocr_excep_old_values'] and
                    newid.kvah_rdng == data['kvah_rdng'] and
                    newid.kvah_img == data['kvah_img'] and
                    newid.kvah_manual == data['kvah_manual'] and
                    newid.kvah_Status == data['kvah_Status'] and
                    newid.mtr_sr_no == data['mtr_sr_no']
                ):
                    print("updatation does not takes place")
                    # return Response({"status": True, "message": "No change in Data"})
                else:
                    print("updating as some of the values are changed")

                    serializer = ConsumerSerializer(
                        newid, data=data, partial=True)
                    if serializer.is_valid():
                        serializer.save()
                        count_update += 1

            # if bill id is not present then insert
 
            else:
                print("inserted")
                serializer = ConsumerSerializer(data=data)
                if serializer.is_valid():
                    serializer.save()
                    count_insert += 1

        except Exception as e:
            failedCons = {"message": str(e), "cons_ac_no": data['cons_ac_no']}
            failed_consumers.append(failedCons)
    print({"status": True, "message": f"Data inserted {count_insert} and Data updated {count_update}", "version": "28"})
    if len(failed_consumers) > 0:
        return Response(
            {"status": False, "message": f"Data inserted {count_insert} and Data updated {count_update}",
                "failed_consumers": failed_consumers, "version": "28"}
        )
    return Response(
        {"status": True, "message": f"Data inserted {count_insert} and Data updated {count_update}",
            "failed_consumers": failed_consumers, "version": "28"}
    )


@api_view(["GET"])
def getconsumers(request):
    data = Consumers.objects.all().order_by("-id")
    serializer = ConsumerSerializer(data, many=True)
    return Response(serializer.data)


@api_view(["POST"])
def deleteconsumers(request):
    try:
        mrId = request.data.get("mrId")
        user = MeterReaderRegistration.objects.get(mrId=mrId)
        user.delete()
        return Response(
            {"status": True, "message": "MR deleted successfully(from api)"}
        )
    except MeterReaderRegistration.DoesNotExist:
        return Response(
            {"status": False, "message": "MR Does not Exist(from api)"},
            status=status.HTTP_200_OK,
        )


@api_view(["GET"])
def getofficedatalist(request):
    data = Office.objects.all()
    cursor = connection.cursor()
    query = f"""
    select * from office

    """
    cursor.execute(query)
    result = dictfetchall(cursor)
    return Response(result)
    pass


@parser_classes([MultiPartParser, FormParser])
@api_view(["POST"])
def meterReaderRegistrationfun(request):
    data = request.data
    serializer = MeterReaderRegistrationSerializer(data=data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"status": True, "message": "Registration Successfull (from api)"},
            status=status.HTTP_201_CREATED,
        )
    return Response(
        {"status": False,
            "message": "Meter Reader Id Already exists (from api)"},
        status=status.HTTP_200_OK,
    )


@parser_classes([MultiPartParser, FormParser])
@api_view(["POST"])
def meterReaderRegistrationUpdateOffice(request):
    data = request.data
    mrId = data["mrId"]
    id = MeterReaderRegistration.objects.get(mrId=mrId)
    if id is None:
        return Response(
            {"status": False,
                "message": "Meter Reader Id Does Not exists (from api)"},
            status=status.HTTP_200_OK,
        )
    serializer = MeterReaderRegistrationSerializer(id, data=data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {
                "status": True,
                "message": "Registration Data Updated Successfull (from api)",
            },
            status=status.HTTP_201_CREATED,
        )
    return Response(
        {"status": False, "message": "Error (from api)"}, status=status.HTTP_200_OK
    )


@parser_classes([MultiPartParser, FormParser])
@api_view(["POST"])
def metereReaderlogin(request):
    newdata = request.data
    mrId = newdata["mrId"]
    token = newdata["androidToken"]
    try:
        id = MeterReaderRegistration.objects.get(mrId=mrId)
        serializer = MeterReaderRegistrationSerializer(
            id, data=newdata, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(
                {
                    "status": True,
                    "message": "Login Successfull(from api)",
                    "data": serializer.data,
                },
                status=status.HTTP_201_CREATED,
            )
        return Response(
            {"status": False, "message": "Something wrong(from api)"},
            status=status.HTTP_200_OK,
        )
    except MeterReaderRegistration.DoesNotExist:
        return Response(
            {"status": False, "message": "MR Does not Exist(from api)"},
            status=status.HTTP_200_OK,
        )


@api_view(["GET"])
def getregdata(request):
    role_to_fetch = request.query_params.get('role', 'meterreader').lower()
    discom = request.query_params.get('discom', 'all').upper()

    if role_to_fetch == 'supervisor':
        today = date.today()
        location_exists = SupervsiorLocation.objects.filter(
            supervisor_number=OuterRef('supervisor_number'),
            date=today
        )
        filters = {}
        if discom != 'ALL':
            filters['discom'] = discom

        # Step 1: Valid Postgres DISTINCT ON query
        qs = (
            SupervisorLogin.objects
            .filter(**filters)
            .filter(
                id__in=SupervisorLogin.objects.values('supervisor_number')
                .distinct()
                .values_list('id', flat=True)
            )
            .annotate(location=Exists(location_exists))
            # must match DISTINCT ON rule
            .order_by('supervisor_number', '-location')
            .distinct('supervisor_number')
        )

        # Step 2: force evaluation
        result = list(qs)

        # Step 3: Python sort → location=True first
        result.sort(key=lambda x: not x.location)

        serializer = SupervisorLoginSerializer(result, many=True)
        return Response(serializer.data)
    else:
        data = MeterReaderRegistration.objects.all()
        serializer = MeterReaderRegistrationSerializer(data, many=True)
        return Response(serializer.data)


@api_view(["GET"])
def mvcards(request):
    offset = request.query_params.get("offset", None)
    new = []

    def listfun(dict):
        new.append(dict.copy())
        return new

    cursor = connection.cursor()
    new_dict = {}
    if offset is not None:
        clause = "OFFSET %s"
        params = [offset]
    else:
        clause = ""
        params = ""
    try:
        query = f"""select m.mr_id,m.rdng_date,m.prsnt_mtr_status,m.prsnt_ocr_rdng,m.prsnt_rdng,m.ocr_pf_reading,m.cons_name,m.prsnt_md_rdng_ocr,m.rdng_ocr_status,m.rdng_img,m.prsnt_md_rdng,m.id

                    from readingmaster m ORDER BY m.rdng_date DESC LIMIT 10 {clause}

    """

        cursor.execute(query, params)
        result = cursor.fetchall()
        for row in result:
            new_dict["id"] = row[11]
            new_dict["mrid"] = row[0]
            new_dict["rdngDate"] = row[1]
            new_dict["prsntmtrstatus"] = row[2]
            new_dict["prsntOcrRdng"] = row[3]
            new_dict["prsntRdng"] = row[4]
            new_dict["prsntPf"] = row[5]
            new_dict["consName"] = row[6]
            new_dict["prsntMdRdngOcr"] = row[7]
            new_dict["rdngocrstatus"] = row[8]
            new_dict["rdngImg"] = row[9]
            new_dict["prsntMdRdng"] = row[10]
            newdata = listfun(new_dict)
        return Response(newdata)
    except:
        return Response([])


@api_view(["GET"])
def get_officedata(request):
    with open("D:/SBPDCL.json") as f:
        data = json.load(f)
        for i in data:
            data = Office.objects.create(
                discom=i["discom"],
                zone=i["zone"],
                circlename=i["circlename"],
                divisionname=i["divisionname"],
                divisioncode=i["divisioncode"],
                subdivision=i["subdivision"],
                subdivisioncode=i["subdivisioncode"],
                sectionname=i["sectionname"],
                sectioncode=i["sectioncode"],
            )
            data.save()
    return Response("data added Successfully")


# Create your views here.
@api_view(["POST"])
def loginuser(request):
    data = request.data
    email = data["email"]
    password = data["password"]
    print("email", email)
    print("password", password)
    try:
        if email == "payfinix@gmail.com" and password == "payfinix#123":
            print("hello")
            token = jwt.encode({"agentid": email},
                               SECRETKEY, algorithm="HS256")
            # return Response({"user":"admin","token":token})
            return Response({"user": "admin", "accessToken": token})

        return Response(
            {"status": "False", "Msg": "Email or Password did not match"},
            status=status.HTTP_404_NOT_FOUND,
        )
    except:
        return Response(
            {"status": "False", "Msg": "Email or Password did not match"},
            status=status.HTTP_404_NOT_FOUND,
        )


@api_view(["GET"])
def get_discom(request):
    newlist = []
    discom = Office.objects.values_list("discom").distinct()
    for row in discom:
        print(row[0])
        newlist.append(row[0])
        print(newlist)
    return Response(newlist)


@api_view(["POST"])
def get_zone(request):
    newlist = []
    zone = (
        Office.objects.filter(discom=request.data["discom"])
        .values_list("zone")
        .distinct()
    )
    for row in zone:
        print(row[0])
        newlist.append(row[0])
        print(newlist)
    return Response(newlist)


@api_view(["POST"])
def get_circle(request):
    newlist = []
    circle = (
        Office.objects.filter(zone=request.data["zone"])
        .values_list("circlename")
        .distinct()
    )
    for row in circle:
        print(row[0])
        newlist.append(row[0])
        print(newlist)
    return Response(newlist)


@api_view(["POST"])
def get_division(request):
    newlist = []
    division = (
        Office.objects.filter(circlename=request.data["circle"])
        .values_list("divisionname")
        .distinct()
    )
    for row in division:
        print(row[0])
        newlist.append(row[0])
        print(newlist)
    return Response(newlist)


@api_view(["POST"])
def get_subdivision(request):
    newlist = []
    subdivision = (
        Office.objects.filter(divisionname=request.data["division"])
        .values_list("subdivision")
        .distinct()
    )
    for row in subdivision:
        print(row[0])
        newlist.append(row[0])
        print(newlist)
    return Response(newlist)


@api_view(["POST"])
def get_section(request):
    new = []

    def listfun(dict):
        print(dict)
        new.append(dict.copy())
        return new

    newdict = {}
    section = (
        Office.objects.filter(subdivision=request.data["subdivision"])
        .values_list("sectionname", "sectioncode")
        .distinct()
    )

    print("section", section)
    for row in section:
        newdict["sectionName"] = row[0]
        newdict["sectionCode"] = row[1]

        newdata = listfun(newdict)
    return Response(newdata)


@api_view(["POST"])
def get_sectionforuser(request):
    newlist = []
    section = (
        Office.objects.filter(subdivision=request.data["subdivision"])
        .values_list("sectionname")
        .distinct()
    )
    print("section", section)
    for row in section:
        newlist.append(row[0])
    return Response(newlist)


@api_view(["GET"])
def get_meter_summary(request):
    month = request.query_params.get("month", None)
    startdate = request.query_params.get("startdate", None)
    enddate = request.query_params.get("enddate", None)
    section = request.query_params.get("section", None)
    new = []

    def listfun(dict):
        print(dict)
        new.append(dict.copy())
        return new

    newdict = {}
    clause = ""
    cursor = connection.cursor()
    if month is not None:
        month = month.split("-")[1]
        clause = "WHERE EXTRACT (MONTH FROM readingmaster.reading_date_db)=%s"
        params = [month]
    elif startdate and enddate:
        clause = "WHERE readingmaster.reading_date_db BETWEEN %s AND %s"
        print(clause)
        params = [startdate, enddate]
    elif section:
        clause = " WHERE readingmaster.ofc_section=%s"
        params = [section]
        pass
    else:
        clause = ""
        params = ""

    query = f"""select readingmaster.mr_id,count(readingmaster.id),count(readingmaster.prsnt_mtr_status='Ok' or NULL),count(readingmaster.rdng_ocr_status='Passed' or NULL),count(readingmaster.rdng_ocr_status='Failed' or NULL),count(readingmaster.prsnt_mtr_status='Meter Defective' or NULL),count(readingmaster.prsnt_mtr_status='Door Locked' or NULL)
            from readingmaster {clause}  group by readingmaster.mr_id
        """
    cursor.execute(query, params)
    results = cursor.fetchall()
    try:
        for row in results:
            total = row[1]
            okreadings = row[2]
            ocrreadings = row[3]
            ocrwithexcep = row[4]
            meterdefective = row[5]
            doorlocked = row[6]

            # Percentage
            okreadpercent = math.floor(((okreadings / total) * 100))
            ocrreadingpercent = math.floor(
                (((ocrreadings / okreadings) if okreadings else 0) * 100)
            )
            ocrwithexceppercent = math.floor(
                (((ocrwithexcep / okreadings) if okreadings else 0) * 100)
            )
            meterdefectivepercent = math.floor(
                ((meterdefective / total) * 100))
            doorlockedpercent = math.floor(((doorlocked / total) * 100))

            # add to dictionary
            newdict["mrid"] = row[0]
            newdict["totalReadings"] = row[1]
            newdict["OKreadings"] = okreadings
            newdict["OKreadingspercent"] = okreadpercent
            newdict["OCRReadings"] = ocrreadings
            newdict["OCRReadingspercent"] = ocrreadingpercent
            newdict["OCRwithException"] = ocrwithexcep
            newdict["OCRwithExceptionpercent"] = ocrwithexceppercent
            newdict["MeterDefective"] = meterdefective
            newdict["MeterDefectivepercent"] = meterdefectivepercent
            newdict["DoorLocked"] = doorlocked
            newdict["DoorLockedpercent"] = doorlockedpercent
            # add to list
            newdata = listfun(newdict)
        return Response(newdata)
    except:
        return Response([])


@api_view(["POST"])
def locationdiscom(request):
    data = request.data.get("discom")
    print("discom", data)
    newlist = []
    if data == "all":
        discom = Office.objects.values_list("zone").distinct()
        for row in discom:
            print(row[0])
            newlist.append(row[0])
            print(newlist)
        return Response(newlist)
    zone = (
        Office.objects.filter(discom=request.data["discom"])
        .values_list("zone")
        .distinct()
    )
    for row in zone:
        newlist.append(row[0])
        print(newlist)
    return Response(newlist)


# @api_view(['POST'])
# def locationzone(request):

#     data = request.data.get('zone')
#     print('discom', data)
#     newlist = []
#     if data == 'all':

#         discom = (Office.objects.values_list('circlename').distinct())
#         for row in discom:
#             print(row[0])
#             newlist.append(row[0])
#             print(newlist)
#         return Response(newlist)
#     circle = Office.objects.filter(
#         zone=request.data['zone']).values_list('circlename').distinct()
#     for row in circle:
#         newlist.append(row[0])
#         print(newlist)
#     return Response(newlist)


@api_view(["POST"])
def locationcircle(request):
    data = request.data.get("circle")
    print("discom", data)
    newlist = []
    if data == "all":
        discom = Office.objects.values_list("divisionname").distinct()
        for row in discom:
            print(row[0])
            newlist.append(row[0])
            print(newlist)
        return Response(newlist)
    circle = (
        Office.objects.filter(circlename=request.data["circle"])
        .values_list("divisionname")
        .distinct()
    )
    for row in circle:
        newlist.append(row[0])
        print(newlist)
    return Response(newlist)


@api_view(["POST"])
def locationdivision(request):
    data = request.data.get("division")
    print("discom", data)
    newlist = []
    if data == "all":
        division = Office.objects.values_list("subdivision").distinct()
        for row in division:
            print(row[0])
            newlist.append(row[0])
            print(newlist)
        return Response(newlist)
    division = (
        Office.objects.filter(divisionname=request.data["division"])
        .values_list("subdivision")
        .distinct()
    )
    for row in division:
        newlist.append(row[0])
        print(newlist)
    return Response(newlist)


@api_view(["POST"])
def locationsubdivision(request):
    data = request.data.get("subdivision")
    print("discom", data)
    newlist = []
    if data == "all":
        division = Office.objects.values_list("sectionname").distinct()
        for row in division:
            print(row[0])
            newlist.append(row[0])
            print(newlist)
        return Response(newlist)
    division = (
        Office.objects.filter(subdivision=request.data["subdivision"])
        .values_list("sectionname")
        .distinct()
    )
    for row in division:
        newlist.append(row[0])
        print(newlist)
    return Response(newlist)


@api_view(["POST"])
def locationwisediscom(request):
    month = date.today().month

    newdict = {}
    new = []

    def listfun(dict):
        print(dict)
        new.append(dict.copy())
        return new

    cursor = connection.cursor()
    locationwise = request.query_params.get("locationwise", None)
    locationname = request.query_params.get("locationname", None)
    if (locationwise is not None) and (locationname == "all"):
        location = "ofc" + locationwise
        cursor.execute(
            f"""
   select ofc_zone as location, count(distinct mr_id),count(prsnt_mtr_status='Ok' or NULL),count(rdng_ocr_status='Passed' or NULL),count(rdng_ocr_status='Failed' or NULL),count(prsnt_mtr_status='Meter Defective' or NULL),count(prsnt_mtr_status='Door Locked' or NULL),count(mr_id)
    from readingmaster where extract(month from reading_date_db)='{month}' and ofc_zone!='' GROUP BY ofc_zone
    """
        )

    if (locationwise is not None) and (locationname != "all"):
        location = "ofc_" + locationwise
        clause = "WHERE ofc_" + locationwise + "='" + locationname + "' "
        cursor.execute(
            f"""
   select ofc_zone as location, count(distinct mr_id),count(prsnt_mtr_status='Ok' or NULL),count(rdng_ocr_status='Passed' or NULL),count(rdng_ocr_status='Failed' or NULL),count(prsnt_mtr_status='Meter Defective' or NULL),count(prsnt_mtr_status='Door Locked' or NULL),count(mr_id)
    from readingmaster {clause} and extract(month from reading_date_db)='{month}' and ofc_zone!='' GROUP BY ofc_zone
    """
        )
    result = cursor.fetchall()
    print("result", result)
    try:
        for row in result:
            locationname = row[0]
            total = row[7]
            mrid = row[1]
            okreadings = row[2]
            OcrReadings = row[3]
            Ocrwithexception = row[4]
            meterDefective = row[5]
            doorLocked = row[6]
            okreadpercent = math.floor((okreadings / total) * 100)
            ocrreadingpercent = math.floor(
                ((OcrReadings / okreadings) if okreadings else 0) * 100
            )
            ocrwithexceppercent = math.floor(
                ((Ocrwithexception / okreadings) if okreadings else 0) * 100
            )
            meterdefectivepercent = math.floor((meterDefective / total) * 100)
            doorlockedpercent = math.floor((doorLocked / total) * 100)
            newdict["locationname"] = row[0]
            newdict["mrid"] = row[1]
            newdict["okreadings"] = row[2]
            newdict["okreadingspercent"] = okreadpercent
            newdict["OcrReadings"] = row[3]
            newdict["OcrReadingspercent"] = ocrreadingpercent
            newdict["Ocrwithexception"] = row[4]
            newdict["Ocrwithexceptionpercent"] = ocrwithexceppercent
            newdict["meterDefective"] = row[5]
            newdict["meterDefectivepercent"] = meterdefectivepercent
            newdict["doorLocked"] = row[6]
            newdict["doorLockedpercent"] = doorlockedpercent
            newdict["total"] = total
            data = listfun(newdict)
        return Response(data)
    except:
        return Response([])


# @api_view(['POST'])
# def locationwisezone(request):
#     newdict = {}
#     new = []

#     def listfun(dict):
#         print(dict)
#         new.append(dict.copy())
#         return new
#     cursor = connection.cursor()
#     locationwise = request.query_params.get('locationwise', None)
#     locationname = request.query_params.get('locationname', None)

#     if ((locationwise is not None) and (locationname == 'all')):
#         location = 'ofc_'+locationwise
#         cursor.execute(f'''
#    select ofc_circle as location, count(distinct mr_id),count(prsnt_mtr_status='Ok' or NULL),count(rdng_ocr_status='Passed' or NULL),count(rdng_ocr_status='Failed' or NULL),count(prsnt_mtr_status='Meter Defective' or NULL),count(prsnt_mtr_status='Door Locked' or NULL),count(mr_id)
#     from readingmaster  GROUP BY ofc_circle
#     ''')
#     if ((locationwise is not None) and (locationname != 'all')):
#         location = 'ofc_'+locationwise
#         clause = "WHERE ofc_"+locationwise+"='"+locationname+"' "

#         cursor.execute(f'''
#    select ofc_circle as location, count(distinct mr_id),count(prsnt_mtr_status='Ok' or NULL),count(rdng_ocr_status='Passed' or NULL),count(rdng_ocr_status='Failed' or NULL),count(prsnt_mtr_status='Meter Defective' or NULL),count(prsnt_mtr_status='Door Locked' or NULL),count(mr_id)
#     from readingmaster {clause} GROUP BY ofc_circle
#     ''')

#     result = cursor.fetchall()
#     print("result", result)
#     try:
#         for row in result:
#             locationname = row[0]
#             total = row[7]
#             mrid = row[1]
#             okreadings = row[2]
#             OcrReadings = row[3]
#             Ocrwithexception = row[4]
#             meterDefective = row[5]
#             doorLocked = row[6]
#             okreadpercent = math.floor((okreadings/total)*100)
#             ocrreadingpercent = math.floor(
#                 ((OcrReadings/okreadings) if okreadings else 0)*100)
#             ocrwithexceppercent = math.floor(
#                 ((Ocrwithexception/okreadings) if okreadings else 0)*100)
#             meterdefectivepercent = math.floor((meterDefective/total)*100)
#             doorlockedpercent = math.floor((doorLocked/total)*100)
#             newdict['locationname'] = row[0]
#             newdict['mrid'] = row[1]
#             newdict['okreadings'] = row[2]
#             newdict['okreadingspercent'] = okreadpercent
#             newdict['OcrReadings'] = row[3]
#             newdict['OcrReadingspercent'] = ocrreadingpercent
#             newdict['Ocrwithexception'] = row[4]
#             newdict['Ocrwithexceptionpercent'] = ocrwithexceppercent
#             newdict['meterDefective'] = row[5]
#             newdict['meterDefectivepercent'] = meterdefectivepercent
#             newdict['doorLocked'] = row[6]
#             newdict['doorLockedpercent'] = doorlockedpercent
#             data = listfun(newdict)
#         return Response(data)
#     except:
#         return Response([])


@api_view(["POST"])
def locationwisecircle(request):
    newdict = {}
    new = []

    def listfun(dict):
        print(dict)
        new.append(dict.copy())
        return new

    cursor = connection.cursor()
    locationwise = request.query_params.get("locationwise", None)
    locationname = request.query_params.get("locationname", None)

    if (locationwise is not None) and (locationname == "all"):
        location = "ofc_" + locationwise
        cursor.execute(
            f"""
   select ofc_division as location, count(distinct mr_id),count(prsnt_mtr_status='Ok' or NULL),count(rdng_ocr_status='Passed' or NULL),count(rdng_ocr_status='Failed' or NULL),count(prsnt_mtr_status='Meter Defective' or NULL),count(prsnt_mtr_status='Door Locked' or NULL),count(mr_id)
    from readingmaster  GROUP BY ofc_division
    """
        )
    if (locationwise is not None) and (locationname != "all"):
        location = "ofc_" + locationwise
        clause = "WHERE ofc_" + locationwise + "='" + locationname + "' "

        cursor.execute(
            f"""
   select ofc_division as location, count(distinct mr_id),count(prsnt_mtr_status='Ok' or NULL),count(rdng_ocr_status='Passed' or NULL),count(rdng_ocr_status='Failed' or NULL),count(prsnt_mtr_status='Meter Defective' or NULL),count(prsnt_mtr_status='Door Locked' or NULL),count(mr_id)
    from readingmaster {clause} GROUP BY ofc_division
    """
        )

    result = cursor.fetchall()
    print("result", result)
    try:
        for row in result:
            locationname = row[0]
            total = row[7]
            mrid = row[1]
            okreadings = row[2]
            OcrReadings = row[3]
            Ocrwithexception = row[4]
            meterDefective = row[5]
            doorLocked = row[6]
            okreadpercent = math.floor((okreadings / total) * 100)
            ocrreadingpercent = math.floor(
                ((OcrReadings / okreadings) if okreadings else 0) * 100
            )
            ocrwithexceppercent = math.floor(
                ((Ocrwithexception / okreadings) if okreadings else 0) * 100
            )
            meterdefectivepercent = math.floor((meterDefective / total) * 100)
            doorlockedpercent = math.floor((doorLocked / total) * 100)
            newdict["locationname"] = row[0]
            newdict["mrid"] = row[1]
            newdict["okreadings"] = row[2]
            newdict["okreadingspercent"] = okreadpercent
            newdict["OcrReadings"] = row[3]
            newdict["OcrReadingspercent"] = ocrreadingpercent
            newdict["Ocrwithexception"] = row[4]
            newdict["Ocrwithexceptionpercent"] = ocrwithexceppercent
            newdict["meterDefective"] = row[5]
            newdict["meterDefectivepercent"] = meterdefectivepercent
            newdict["doorLocked"] = row[6]
            newdict["doorLockedpercent"] = doorlockedpercent
            data = listfun(newdict)
        return Response(data)
    except:
        return Response([])


@api_view(["POST"])
def locationwisedivision(request):
    month = date.today().month
    newdict = {}
    new = []

    def listfun(dict):
        print(dict)
        new.append(dict.copy())
        return new

    cursor = connection.cursor()
    locationwise = request.query_params.get("locationwise", None)
    locationname = request.query_params.get("locationname", None)
    if (locationwise is not None) and (locationname == "all"):
        location = "ofc_" + locationwise
        cursor.execute(
            f"""
   select ofc_subdivision as location, count(distinct mr_id),count(prsnt_mtr_status='Ok' or NULL),count(rdng_ocr_status='Passed' or NULL),count(rdng_ocr_status='Failed' or NULL),count(prsnt_mtr_status='Meter Defective' or NULL),count(prsnt_mtr_status='Door Locked' or NULL),count(mr_id)
    from readingmaster  GROUP BY ofc_subdivision
    """
        )
    if (locationwise is not None) and (locationname != "all"):
        location = "ofc_" + locationwise
        clause = "WHERE ofc_" + locationwise + "='" + locationname + "' "

        cursor.execute(
            f"""
   select ofc_subdivision as location, count(distinct mr_id),count(prsnt_mtr_status='Ok' or NULL),count(rdng_ocr_status='Passed' or NULL),count(rdng_ocr_status='Failed' or NULL),count(prsnt_mtr_status='Meter Defective' or NULL),count(prsnt_mtr_status='Door Locked' or NULL),count(mr_id)
    from readingmaster {clause} GROUP BY ofc_subdivision
    """
        )

    result = cursor.fetchall()
    print("result", result)
    try:
        for row in result:
            locationname = row[0]
            total = row[7]
            mrid = row[1]
            okreadings = row[2]
            OcrReadings = row[3]
            Ocrwithexception = row[4]
            meterDefective = row[5]
            doorLocked = row[6]
            okreadpercent = math.floor((okreadings / total) * 100)
            ocrreadingpercent = math.floor(
                ((OcrReadings / okreadings) if okreadings else 0) * 100
            )
            ocrwithexceppercent = math.floor(
                ((Ocrwithexception / okreadings) if okreadings else 0) * 100
            )
            meterdefectivepercent = math.floor((meterDefective / total) * 100)
            doorlockedpercent = math.floor((doorLocked / total) * 100)
            newdict["locationname"] = row[0]
            newdict["mrid"] = row[1]
            newdict["okreadings"] = row[2]
            newdict["okreadingspercent"] = okreadpercent
            newdict["OcrReadings"] = row[3]
            newdict["OcrReadingspercent"] = ocrreadingpercent
            newdict["Ocrwithexception"] = row[4]
            newdict["Ocrwithexceptionpercent"] = ocrwithexceppercent
            newdict["meterDefective"] = row[5]
            newdict["meterDefectivepercent"] = meterdefectivepercent
            newdict["doorLocked"] = row[6]
            newdict["doorLockedpercent"] = doorlockedpercent
            data = listfun(newdict)
        return Response(data)
    except:
        return Response([])


@api_view(["POST"])
def locationwisesubdivision(request):
    month = date.today().month
    newdict = {}
    new = []

    def listfun(dict):
        print(dict)
        new.append(dict.copy())
        return new

    cursor = connection.cursor()
    locationwise = request.query_params.get("locationwise", None)
    locationname = request.query_params.get("locationname", None)

    if (locationwise is not None) and (locationname == "all"):
        location = "ofc_" + locationwise
        cursor.execute(
            f"""
   select ofc_section as location, count(distinct mr_id),count(prsnt_mtr_status='Ok' or NULL),count(rdng_ocr_status='Passed' or NULL),count(rdng_ocr_status='Failed' or NULL),count(prsnt_mtr_status='Meter Defective' or NULL),count(prsnt_mtr_status='Door Locked' or NULL),count(mr_id)
    from readingmaster where extract(month from reading_date_db)='{month}' and ofc_section!=''  GROUP BY ofc_section
    """
        )
    if (locationwise is not None) and (locationname != "all"):
        location = "ofc_" + locationwise
        clause = "where ofc_" + locationwise + "='" + locationname + "' "

        cursor.execute(
            f"""
    select o.sectionname as location, count(distinct mr_id),count(prsnt_mtr_status='Ok' or NULL),count(rdng_ocr_status='Passed' or NULL),count(rdng_ocr_status='Failed' or NULL),count(prsnt_mtr_status='Meter Defective' or NULL),count(prsnt_mtr_status='Door Locked' or NULL),count(mr_id) as total
    from readingmaster r full join office o on r.ofc_section=o.sectioncode {clause} and extract(month from reading_date_db)='{month}' and o.sectionname!=''  GROUP BY o.sectionname
    """
        )

    result = cursor.fetchall()
    print("result", result)
    try:
        for row in result:
            locationname = row[0]
            total = row[7]
            mrid = row[1]
            okreadings = row[2]
            OcrReadings = row[3]
            Ocrwithexception = row[4]
            meterDefective = row[5]
            doorLocked = row[6]
            okreadpercent = math.floor((okreadings / total) * 100)
            ocrreadingpercent = math.floor(
                ((OcrReadings / okreadings) if okreadings else 0) * 100
            )
            ocrwithexceppercent = math.floor(
                ((Ocrwithexception / okreadings) if okreadings else 0) * 100
            )
            meterdefectivepercent = math.floor((meterDefective / total) * 100)
            doorlockedpercent = math.floor((doorLocked / total) * 100)
            newdict["locationname"] = row[0]
            newdict["mrid"] = row[1]
            newdict["okreadings"] = row[2]
            newdict["okreadingspercent"] = okreadpercent
            newdict["OcrReadings"] = row[3]
            newdict["OcrReadingspercent"] = ocrreadingpercent
            newdict["Ocrwithexception"] = row[4]
            newdict["Ocrwithexceptionpercent"] = ocrwithexceppercent
            newdict["meterDefective"] = row[5]
            newdict["meterDefectivepercent"] = meterdefectivepercent
            newdict["doorLocked"] = row[6]
            newdict["doorLockedpercent"] = doorlockedpercent
            newdict["total"] = total
            data = listfun(newdict)
        return Response(data)
    except:
        return Response([])


@api_view(["GET"])
def ocrsummary(request):
    month = request.query_params.get("month", None)
    startdate = request.query_params.get("startdate", None)
    enddate = request.query_params.get("enddate", None)
    clause = ""
    if month is not None:
        month = month.split("-")[1]
        clause = "WHERE EXTRACT (MONTH FROM readingmaster.reading_date_db)=%s"
        params = [month]
    elif startdate and enddate:
        clause = "WHERE readingmaster.reading_date_db BETWEEN %s AND %s"
        print(clause)
        params = [startdate, enddate]
    else:
        clause = ""
        params = ""
    new = []

    def listfun(dict):
        print(dict)
        new.append(dict.copy())
        return new

    cursor = connection.cursor()
    newdict = {}
    query = f"""
   select mr_id,count(mr_id),count(prsnt_mtr_status='Ok' or NULL),count(rdng_ocr_status='Passed' or NULL),count(readingmaster.rdng_ocr_status='Failed' or NULL),count(prsnt_rdng_ocr_excep='Parameters Incorrect' or NULL),count(prsnt_rdng_ocr_excep='Parameters Unavailable' or NULL),count(prsnt_rdng_ocr_excep='Parameters Unclear' or NULL),count(prsnt_rdng_ocr_excep='Image Invalid' or NULL),count(prsnt_rdng_ocr_excep='Image Unclear' or NULL),count(prsnt_rdng_ocr_excep='Image Spoofed' or NULL),count(prsnt_rdng_ocr_excep='Image Stain on Decimal' or NULL),count(prsnt_rdng_ocr_excep='Meter Mismatched' or NULL),count(prsnt_rdng_ocr_excep='Meter On Height' or NULL),count(prsnt_rdng_ocr_excep='Meter Dirty' or NULL),count(prsnt_rdng_ocr_excep='Meter Display Broken' or NULL),count(prsnt_rdng_ocr_excep='Reflection Daylight Reflection On Meter' or NULL),count(prsnt_rdng_ocr_excep='Reflection Backlight Reflection' or NULL),count(prsnt_mtr_status='Meter Defective' or NULL),count(prsnt_mtr_status='Door Locked' or NULL)
    from readingmaster {clause} GROUP BY mr_id
    """

    cursor.execute(query, params)
    results = cursor.fetchall()
    # Convert tuple
    for row in results:
        total = row[1]
        okreadings = row[2]
        ocrreadings = row[3]
        ocrwithexcep = row[4]
        parametersincorrect = row[5]
        parametersunavailable = row[6]
        parametersunclear = row[7]
        imageinvalid = row[8]
        imageunclear = row[9]
        imagespoofed = row[10]
        imagestainondecimal = row[11]
        metermismatched = row[12]
        meteronheight = row[13]
        meterdirty = row[14]
        meterdisplaybroken = row[15]
        reflectiondaylight = row[16]
        reflectionbacklight = row[17]
        meterdefective = row[18]
        doorlocked = row[19]
        # Percentage
        okreadpercent = math.floor((okreadings / total) * 100)
        ocrreadingpercent = math.floor(
            ((ocrreadings / okreadings) if okreadings else 0) * 100
        )
        ocrwithexceppercent = math.floor(
            ((ocrwithexcep / okreadings) if okreadings else 0) * 100
        )
        parametersincorrectpercent = math.floor(
            ((parametersincorrect / okreadings) if okreadings else 0) * 100
        )
        parametersunavailablepercent = math.floor(
            ((parametersunavailable / okreadings) if okreadings else 0) * 100
        )

        parametersunclearpercent = math.floor(
            ((parametersunclear / okreadings) if okreadings else 0) * 100
        )
        imageinvalidpercent = math.floor(
            ((imageinvalid / okreadings) if okreadings else 0) * 100
        )
        imageunclearpercent = math.floor(
            ((imageunclear / okreadings) if okreadings else 0) * 100
        )
        imagespoofedpercent = math.floor(
            ((imagespoofed / okreadings) if okreadings else 0) * 100
        )
        imagestainondecimalpercent = math.floor(
            ((imagestainondecimal / okreadings) if okreadings else 0) * 100
        )
        metermismatchedpercent = math.floor(
            ((metermismatched / okreadings) if okreadings else 0) * 100
        )
        meteronheightpercent = math.floor(
            ((meteronheight / okreadings) if okreadings else 0) * 100
        )
        meterdirtypercent = math.floor(
            ((meterdirty / okreadings) if okreadings else 0) * 100
        )
        meterdisplaybrokenpercent = math.floor(
            ((meterdisplaybroken / okreadings) if okreadings else 0) * 100
        )
        reflectiondaylightpercent = math.floor(
            ((reflectiondaylight / okreadings) if okreadings else 0) * 100
        )
        reflectionbacklightpercent = math.floor(
            ((reflectionbacklight / okreadings) if okreadings else 0) * 100
        )
        meterdefectivepercent = math.floor((meterdefective / total) * 100)
        doorlockedpercent = math.floor((doorlocked / total) * 100)
        # add to dictionary
        newdict["mrid"] = row[0]
        newdict["totalReadings"] = row[1]
        newdict["OKreadings"] = okreadings
        newdict["OKreadingspercent"] = okreadpercent
        newdict["OCRReadings"] = ocrreadings
        newdict["OCRReadingspercent"] = ocrreadingpercent
        newdict["OCRwithException"] = ocrwithexcep
        newdict["OCRwithExceptionpercent"] = ocrwithexceppercent
        newdict["parametersincorrect"] = parametersincorrect
        newdict["parametersincorrectpercent"] = parametersincorrectpercent
        newdict["parametersunavailable"] = parametersunavailable
        newdict["parametersunavailablepercent"] = parametersunavailablepercent
        newdict["parametersunclear"] = parametersunclear
        newdict["parametersunclearpercent"] = parametersunclearpercent
        newdict["imageinvalid"] = imageinvalid
        newdict["imageinvalidpercent"] = imageinvalidpercent
        newdict["imageunclear"] = imageunclear
        newdict["imageunclearpercent"] = imageunclearpercent
        newdict["imageunclearpercent"] = imageunclearpercent
        newdict["imagespoofed"] = imagespoofed
        newdict["imagespoofedpercent"] = imagespoofedpercent
        newdict["imagestainondecimal"] = imagestainondecimal
        newdict["imagestainondecimalpercent"] = imagestainondecimalpercent
        newdict["metermismatched"] = metermismatched
        newdict["metermismatchedpercent"] = metermismatchedpercent
        newdict["meteronheight"] = meteronheight
        newdict["meteronheightpercent"] = meteronheightpercent
        newdict["meterdirty"] = meterdirty
        newdict["meterdirtypercent"] = meterdirtypercent
        newdict["meterdisplaybroken"] = meterdisplaybroken
        newdict["meterdisplaybrokenpercent"] = meterdisplaybrokenpercent
        newdict["reflectiondaylight"] = reflectiondaylight
        newdict["reflectiondaylightpercent"] = reflectiondaylightpercent
        newdict["reflectionbacklight"] = reflectionbacklight
        newdict["reflectionbacklightpercent"] = reflectionbacklightpercent
        newdict["MeterDefective"] = meterdefective
        newdict["MeterDefectivepercent"] = meterdefectivepercent
        newdict["DoorLocked"] = doorlocked
        newdict["DoorLockedpercent"] = doorlockedpercent
        # add to list
        newdata = listfun(newdict)
    return Response(newdata)


@api_view(["POST"])
def agencywisesummary(request):
    data = request.query_params.get("location", None)
    month = request.query_params.get("month", None)
    startdate = request.query_params.get("startdate", None)
    enddate = request.query_params.get("enddate", None)
    new = []

    def listfun(dict):
        print(dict)
        new.append(dict.copy())
        return new

    if month and data:
        month = month.split("-")[1]
        clause = "WHERE EXTRACT (MONTH FROM readingmaster.reading_date_db)=%s and readingmaster.ofc_discom=%s"
        params = [month, data]

    else:
        clause = ""
        params = ""
    cursor = connection.cursor()
    newdict = {}

    query = f"""select readingmaster.bl_agnc_name,count(readingmaster.id),count(readingmaster.prsnt_mtr_status='Ok' or NULL),count(readingmaster.rdng_ocr_status='Passed' or NULL),count(readingmaster.rdng_ocr_status='Failed' or NULL),count(readingmaster.prsnt_mtr_status='Meter Defective' or NULL),count(readingmaster.prsnt_mtr_status='Door Locked' or NULL)
                    from readingmaster {clause} group by readingmaster.bl_agnc_name
    """
    # if data=="NBPDCL":
    #     cursor.execute("""select readingmaster.bl_agnc_type,count(readingmaster.id),count(readingmaster.prsnt_mtr_status='Ok' or NULL),count(readingmaster.rdng_ocr_status='Passed' or NULL),count(readingmaster.prsnt_rdng_ocr_excep='Parameters Incorrect' or readingmaster.prsnt_rdng_ocr_excep='Meter On Height' or readingmaster.prsnt_rdng_ocr_excep='Image Invalid' or readingmaster.prsnt_rdng_ocr_excep='Image Spoofed' or readingmaster.prsnt_rdng_ocr_excep='Meter Dirty' or readingmaster.prsnt_rdng_ocr_excep='Meter Display Broken' or readingmaster.prsnt_rdng_ocr_excep='Meter Display Broken' or NULL),count(readingmaster.prsnt_mtr_status='Meter Defective' or NULL),count(readingmaster.prsnt_mtr_status='Door Locked' or NULL)
    #                 from readingmaster where ofc_discom='NBPDCL' group by readingmaster.bl_agnc_type
    # """)
    # if data=="SBPDCL":
    #     cursor.execute("""select readingmaster.bl_agnc_type,count(readingmaster.id),count(readingmaster.prsnt_mtr_status='Ok' or NULL),count(readingmaster.rdng_ocr_status='Passed' or NULL),count(readingmaster.prsnt_rdng_ocr_excep='Parameters Incorrect' or readingmaster.prsnt_rdng_ocr_excep='Meter On Height' or readingmaster.prsnt_rdng_ocr_excep='Image Invalid' or readingmaster.prsnt_rdng_ocr_excep='Image Spoofed' or readingmaster.prsnt_rdng_ocr_excep='Meter Dirty' or readingmaster.prsnt_rdng_ocr_excep='Meter Display Broken' or readingmaster.prsnt_rdng_ocr_excep='Meter Display Broken' or NULL),count(readingmaster.prsnt_mtr_status='Meter Defective' or NULL),count(readingmaster.prsnt_mtr_status='Door Locked' or NULL)
    #                 from readingmaster where ofc_discom='SBPDCL' group by  readingmaster.bl_agnc_type
    # """)

    try:
        cursor.execute(query, params)
        result = cursor.fetchall()
        for row in result:
            total = row[1]
            okreadings = row[2]
            ocrreadings = row[3]
            ocrwithexcep = row[4]
            meterdefective = row[5]
            doorlocked = row[6]
            # Percentage
            okreadpercent = math.floor((okreadings / total) * 100)
            ocrreadingpercent = math.floor(
                ((ocrreadings / okreadings) if okreadings else 0) * 100
            )
            ocrwithexceppercent = math.floor(
                ((ocrwithexcep / okreadings) if okreadings else 0) * 100
            )
            meterdefectivepercent = math.floor((meterdefective / total) * 100)
            doorlockedpercent = math.floor((doorlocked / total) * 100)

            # add to dictionary
            newdict["agency"] = row[0]
            newdict["totalReadings"] = row[1]
            newdict["OKreadings"] = okreadings
            newdict["OKreadingspercent"] = okreadpercent
            newdict["OCRReadings"] = ocrreadings
            newdict["OCRReadingspercent"] = ocrreadingpercent
            newdict["OCRwithException"] = ocrwithexcep
            newdict["OCRwithExceptionpercent"] = ocrwithexceppercent
            newdict["MeterDefective"] = meterdefective
            newdict["MeterDefectivepercent"] = meterdefectivepercent
            newdict["DoorLocked"] = doorlocked
            newdict["DoorLockedpercent"] = doorlockedpercent
            # add to list
            newdata = listfun(newdict)
        return Response(newdata)
    except:
        return Response([])


# @api_view(['GET'])
# def get_meter_reader_detail(request):
#     monthByData = request.query_params.get("getMonth", None)
#     mrid = request.query_params.get('mrid', None)
#     startDate = request.query_params.get('startdate', None)
#     endDate = request.query_params.get('enddate', None)
#     data = Consumers.objects.all()
#     if monthByData:
#         monthByData = monthByData.split('-')[1]
#         print("month", monthByData)
#         data = Consumers.objects.filter(reading_date_db__month=monthByData)
#         serializer = ConsumerDataSerializer(data, many=True)
#         return Response(serializer.data)
#     if startDate and endDate:
#         datedata = data.filter(reading_date_db__range=[startDate, endDate])
#         serializer = ConsumerDataSerializer(datedata, many=True)
#         return Response(serializer.data)
#     if mrid:
#         mrdata = Consumers.objects.filter(mr_id=mrid)
#         serializer = ConsumerDataSerializer(mrdata, many=True)
#         return Response(serializer.data)
# else:
#     month=1
#     mrdata = Consumers.objects.filter(reading_date_db__month=month)
#     serializer=ConsumerDataSerializer(mrdata,many=True)
#     return Response(serializer.data)


@api_view(["GET"])
def get_meter_reader_detail(request):
    pagesize = request.query_params.get("pagesize", 10)
    paginator = PageNumberPagination()
    paginator.page_size = pagesize
    monthByData = request.query_params.get("getMonth", None)
    mrid = request.query_params.get("mrid", None)
    startDate = request.query_params.get("startdate", None)
    endDate = request.query_params.get("enddate", None)
    person_objects = []
    if monthByData:
        monthByData = monthByData.split("-")[1]
        person_objects = Consumers.objects.filter(
            reading_date_db__month=monthByData
        ).order_by("-id")
    elif startDate and endDate:
        person_objects = Consumers.objects.filter(
            reading_date_db__range=[startDate, endDate]
        ).order_by("-id")
    elif mrid:
        person_objects = Consumers.objects.filter(mr_id=mrid).order_by("-id")
    else:
        person_objects = []
    result_page = paginator.paginate_queryset(person_objects, request)
    serializer = ConsumerDataSerializer(result_page, many=True)
    return paginator.get_paginated_response(serializer.data)


@api_view(["GET"])
def minidashboard(request):
    cursor = connection.cursor()
    cursor.execute(
        f"""select count(readingmaster.rdng_ocr_status='Passed' or NULL),count(readingmaster.rdng_ocr_status='Failed' or NULL),count(readingmaster.prsnt_mtr_status='Meter Defective' or NULL),count(readingmaster.prsnt_mtr_status='Door Locked' or NULL),count(readingmaster.prsnt_mtr_status='Ok' or NULL),count(rdng_img)
                    from readingmaster
    """
    )
    consumers = Consumers.objects.values_list("cons_ac_no").distinct()
    mr = Consumers.objects.values_list("mr_id").distinct()
    consumerslength = len(consumers)
    mrid = len(mr)
    newdict = {}
    result = cursor.fetchall()
    for row in result:
        metervisionaireadings = row[0]
        metervisionaireadingswithocrexception = row[1]
        doorlockedreadings = row[3]
        meterdefectivereadings = row[2]
        okreadings = row[4]

    newdict["mrid"] = mrid
    newdict["okreadings"] = okreadings
    newdict["consumerslength"] = consumerslength
    newdict["ocrreadings"] = metervisionaireadings
    newdict["ocrwithexception"] = metervisionaireadingswithocrexception
    newdict["doorlocked"] = doorlockedreadings
    newdict["meterdefective"] = meterdefectivereadings
    newdict["totalreadings"] = row[5]
    print(newdict)
    return Response(newdict)


# @api_view(['GET'])
# def exceptionlist(request):
#     month = date.today().month
#     agency = request.query_params.get("agency")
#     if agency != 'null':
#         total = Consumers.objects.filter(prsnt_mtr_status='Ok', bl_agnc_name=agency,
#                                          reading_date_db__month=month).aggregate(Count('id'))['id__count']
#     else:
#         total = Consumers.objects.filter(
#             prsnt_mtr_status='Ok', reading_date_db__month=month).aggregate(Count('id'))['id__count']

#     # total = len(ok)

#     cursor = connection.cursor()
#     clause = ''
#     if agency != 'null':
#         clause = f"and bl_agnc_name='{agency}'"

#     query = (f'''
#    select count(prsnt_mtr_status='Ok' or NULL),count(readingmaster.prsnt_rdng_ocr_excep='Parameters Incorrect' or NULL),count(prsnt_rdng_ocr_excep='Parameters Unavailable' or NULL),count(prsnt_rdng_ocr_excep='Parameters Unclear' or NULL),count(prsnt_rdng_ocr_excep='Image Invalid' or NULL),count(prsnt_rdng_ocr_excep='Image Unclear' or NULL),count(prsnt_rdng_ocr_excep='Image Spoofed' or NULL),count(prsnt_rdng_ocr_excep='Image Stain on Decimal' or NULL),count(prsnt_rdng_ocr_excep='Meter Mismatched' or NULL),count(prsnt_rdng_ocr_excep='Meter On Height' or NULL),count(prsnt_rdng_ocr_excep='Meter Dirty' or NULL),count(prsnt_rdng_ocr_excep='Meter Display Broken' or NULL),count(prsnt_rdng_ocr_excep='Daylight Reflection On Meter' or NULL),count(prsnt_rdng_ocr_excep='Backlight Reflection' or NULL),
#    count(prsnt_rdng_ocr_excep='Incorrect Reading' or NULL)
#     from readingmaster where extract(month from reading_date_db)='{month}' {clause}
#     ''')

#     cursor.execute(query)
#     results = cursor.fetchall()
#     print(results)
#     newdict = {}
#     for i in results:
#         newdict['Parameters Incorrect'] = i[1]
#         newdict['Parameters Unavailable'] = i[2]
#         newdict['Parameters Unclear'] = i[3]
#         newdict['Image Invalid'] = i[4]
#         newdict['Image Unclear'] = i[5]
#         newdict['Image Spoofed'] = i[6]
#         newdict['Image Stain on Decimal'] = i[7]
#         newdict['Meter Mismatched'] = i[8]
#         newdict['Meter On Height'] = i[9]
#         newdict['Meter Dirty'] = i[10]
#         newdict['Meter Display Broken'] = i[11]
#         newdict['Reflection Daylight Reflection On Meter'] = i[12]
#         newdict['Reflection Backlight Reflection'] = i[13]
#         newdict['Incorrect Reading'] = i[14]

#     return Response({"total": total, "data": newdict})


@api_view(["GET"])
def exceptionlist(request):
    month = date.today().month
    agency = request.query_params.get("agency")
    # ofc_discom=request.query_params.get("ofc_discom")

    cursor = connection.cursor()
    clause = ""
    if agency != "null":
        clause = f"and bl_agnc_name='{agency}'"
    # if ofc_discom !='null':
    #     clause=f"and ofc_discom='{ofc_discom}'"

    query1 = f"""
    select count(prsnt_mtr_status='Ok' or NULL),
    count(prsnt_rdng_ocr_excep='Spoofed Image' or NULL),
    count(prsnt_rdng_ocr_excep='Image blur' or NULL),
    count(prsnt_rdng_ocr_excep='Meter Dirty' or NULL),
    count(prsnt_rdng_ocr_excep='Incorrect Reading' or NULL),
    count(reading_parameter_type='Parameters Mismatch' or null),
    count(reading_parameter_type='Parameters Unavailable' or null),
    count(reading_parameter_type='' or null),
    count(prsnt_rdng_ocr_excep='No Exception Found' or prsnt_rdng_ocr_excep=''  or null)

    from readingmaster where extract(month from reading_date_db)='{month}' and rdng_ocr_status='Failed' {clause}
    """
    query2 = f"""
           select count(prsnt_mtr_status='Ok' or null) as OK from readingmaster where extract(month from reading_date_db)='{month}' {clause}
           """

    cursor.execute(query1)
    results = cursor.fetchall()
    print(results)
    newdict = {}
    for i in results:
        newdict["OK"] = i[0]
        newdict["Image Spoofed"] = i[1]
        newdict["Image blur"] = i[2]
        newdict["Meter Dirty"] = i[3]
        newdict["Incorrect Reading"] = i[4]
        newdict["Parameters Mismatch"] = i[5]
        newdict["Parameters Unavailable"] = i[6]
        newdict["Others"] = i[7]
        newdict["Others."] = i[8]
    cursor.execute(query2)
    resulttotal = dictfetchall(cursor)

    return Response(
        {"total": resulttotal[0]["ok"] if resulttotal else 0, "data": newdict}
    )


@api_view(["GET"])
def consumer_wise_details(request):
    acno = request.query_params.get("acno", None)
    print(acno)
    if acno is not None:
        data = Consumers.objects.filter(cons_ac_no=acno)
        serializer = ConsumerWiseDetailsSerializer(data, many=True)
        return Response(serializer.data)
    return Response({"status": False, "consumer": "Not Available"})


@api_view(["GET"])
def getconsumerscount(request):
    uniqueConsAcc = Consumers.objects.values_list("cons_ac_no").distinct()
    consumerscount = len(uniqueConsAcc)
    return Response({"Consumers": consumerscount})


@api_view(["GET"])
def totalcounts(request):
    cursor = connection.cursor()
    # cursor.execute(f'''select count(consumers.rdng_ocr_status='Passed' or NULL),count(consumers.prsnt_rdng_ocr_excep='Parameters Incorrect' or consumers.prsnt_rdng_ocr_excep='Meter On Height' or consumers.prsnt_rdng_ocr_excep='Image Invalid' or consumers.prsnt_rdng_ocr_excep='Image Spoofed' or consumers.prsnt_rdng_ocr_excep='Meter Dirty' or consumers.prsnt_rdng_ocr_excep='Meter Display Broken' or consumers.prsnt_rdng_ocr_excep='Meter Display Broken' or prsnt_rdng_ocr_excep='Parameters Incorrect' or NULL),count(consumers.prsnt_mtr_status='Meter Defective' or NULL),count(consumers.prsnt_mtr_status='Door Locked' or NULL)
    #                 from consumers
    # ''')
    cursor.execute(
        f"""select count(readingmaster.rdng_ocr_status='Passed' or NULL),count(readingmaster.rdng_ocr_status='Failed' or NULL),count(readingmaster.prsnt_mtr_status='Meter Defective' or NULL),count(readingmaster.prsnt_mtr_status='Door Locked' or NULL)
                    from readingmaster
    """
    )

    newdict = {}
    result = cursor.fetchall()
    for row in result:
        metervisionaireadings = row[0]
        metervisionaireadingswithocrexception = row[1]
        doorlockedreadings = row[2]
        meterdefectivereadings = row[3]
    newdict["metervisionaireadings"] = metervisionaireadings
    newdict[
        "metervisionaireadingswithocrexception"
    ] = metervisionaireadingswithocrexception
    newdict["doorlockedreadings"] = doorlockedreadings
    newdict["meterdefectivereadings"] = meterdefectivereadings

    print(newdict)
    return Response(newdict)


@api_view(["POST"])
def getmridforSection(request):
    data = request.data.get("sectioncode", None)
    newlist = []
    mridData = (
        Consumers.objects.filter(
            ofc_section=data).values_list("mr_id").distinct()
    )
    # serializer = ConsumerDataSerializer(mridData,many=True)
    for row in mridData:
        print(row[0])
        newlist.append(row[0])
        print(newlist)
    return Response(newlist)


@api_view(["GET"])
def test(request):
    offset = request.query_params.get("offset", None)
    new = []

    def listfun(dict):
        new.append(dict.copy())
        return new

    cursor = connection.cursor()
    new_dict = {}
    if offset is not None:
        clause = "OFFSET %s"
        params = [offset]
    else:
        clause = ""
        params = ""

    query = f"""select m.mr_id,m.rdng_date,m.prsnt_mtr_status,m.prsnt_ocr_rdng,m.prsnt_rdng,m.ocr_pf_reading,m.cons_name,m.prsnt_md_rdng_ocr,m.rdng_ocr_status,m.rdng_img,m.prsnt_md_rdng,m.id,r."mrPhoto"

                    from readingmaster m,meterreaderregistration r where m.mr_id=r."mrId" ORDER BY m.rdng_date DESC LIMIT 10 {clause}
    """

    cursor.execute(query, params)
    result = cursor.fetchall()
    print("RESULTS____________", result)
    try:
        for row in result:
            new_dict["id"] = row[11]
            new_dict["mrid"] = row[0]
            new_dict["rdngDate"] = row[1]
            new_dict["prsntmtrstatus"] = row[2]
            new_dict["prsntOcrRdng"] = row[3]
            new_dict["prsntRdng"] = row[4]
            new_dict["prsntPf"] = row[5]
            new_dict["consName"] = row[6]
            new_dict["prsntMdRdngOcr"] = row[7]
            new_dict["rdngocrstatus"] = row[8]
            new_dict["rdngImg"] = row[9]
            new_dict["prsntMdRdng"] = row[10]
            new_dict["avatar"] = row[12]
            newdata = listfun(new_dict)
        return Response(newdata)
    except:
        return Response([])


# @api_view(['GET'])
# def topmeterreaders(request):
#     new = []

#     def listfun(dict):
#         print("dict------->", dict)
#         new.append(dict.copy())
#         return new
#     new_dict = {}
#     cursor = connection.cursor()
#     # query=(f'''select m.mr_id,r."mrName",r."mrPhone",r."mrPhoto",count(m.prsnt_mtr_status='Ok' or NULL)

#     #                 from readingmaster m,meterreaderregistration r where m.mr_id=r."mrId" GROUP BY m.id  LIMIT 10
#     # ''')
#     # query=(f'''select m.mr_id,SUM(CASE WHEN m.rdng_ocr_status='Passed' THEN 1 ELSE 0 END),SUM(CASE WHEN m.rdng_ocr_status='Failed' THEN 1 ELSE 0 END)

#     #                 from readingmaster m GROUP BY m.mr_id  LIMIT 10
#     # ''')

#     # query=(f'''select m.mr_id,SUM(CASE WHEN m.rdng_ocr_status='Passed' THEN 1 WHEN m.rdng_ocr_status='Failed' THEN-1 ELSE 0 END) as pass_fail,SUM(CASE WHEN m.rdng_ocr_status='Passed' THEN 1 ELSE 0 END),SUM(CASE WHEN m.rdng_ocr_status='Failed' THEN 1 ELSE 0 END)

#     #                 from readingmaster m GROUP BY m.mr_id order by pass_fail DESC  LIMIT 10
#     # ''')

#     # query=(f'''select r."mrName",SUM(CASE WHEN m.rdng_ocr_status='Passed' THEN 1 WHEN m.rdng_ocr_status='Failed' THEN-1 ELSE 0 END) as pass_fail,SUM(CASE WHEN m.rdng_ocr_status='Passed' THEN 1 ELSE 0 END),SUM(CASE WHEN m.rdng_ocr_status='Failed' THEN 1 ELSE 0 END),count(m.mr_id),count(m.prsnt_mtr_status='Ok' or NULL),count(m.prsnt_mtr_status='Meter Defective' or NULL),count(m.prsnt_mtr_status='Door Locked' or NULL),count(r."mrName")

#     #                 from readingmaster m,meterreaderregistration r where m.mr_id=r."mrId" GROUP BY r."mrName" order by pass_fail DESC
#     # ''')
#     query = (f'''select r."mrName",
# count(rdng_ocr_status='Passed' or NULL) as passed,
# count(rdng_ocr_status='Failed' or NULL) as failed,
# count(m.prsnt_mtr_status='Ok' or NULL),
# count(m.prsnt_mtr_status='Meter Defective' or NULL),
# count(m.prsnt_mtr_status='Door Locked' or NULL),
# count(m.mr_id) as total,r."mrPhoto",r."mrPhone"
# from readingmaster m,meterreaderregistration r
# where m.mr_id=r."mrId"
# AND EXTRACT(MONTH FROM m.reading_date_db)='1'
# GROUP BY r."mrName",r."mrPhoto",r."mrPhone"
# order by passed DESC  limit 10
#     ''')

#     cursor.execute(query)
#     result = cursor.fetchall()
#     # print("RESULTS____________", result)
#     try:
#         for row in result:
#             meterreader = row[0]
#             ocrreadings = row[1]
#             ocrwithexcep = row[2]
#             okreadings = row[3]
#             meterdefective = row[4]
#             doorlocked = row[5]
#             total = row[6]

#             okreadpercent = round(((okreadings/total)*100), 2)
#             ocrreadingpercent = round((
#                 ((ocrreadings/okreadings) if okreadings else 0)*100), 2)
#             ocrwithexceppercent = round((
#                 ((ocrwithexcep/okreadings) if okreadings else 0)*100), 2)
#             meterdefectivepercent = round(((meterdefective/total)*100), 2)
#             doorlockedpercent = round(((doorlocked/total)*100), 2)

#             new_dict['meterreader'] = row[0]
#             new_dict['mrphoto'] = row[7]
#             new_dict['mrphone'] = row[8]

#             new_dict['OK'] = row[3]
#             new_dict['okreadpercent'] = okreadpercent
#             new_dict['ocrpassed'] = row[1]
#             new_dict['ocrreadingpercent'] = ocrreadingpercent
#             new_dict['ocrexception'] = row[2]
#             new_dict['ocrwithexceppercent'] = ocrwithexceppercent
#             new_dict['DoorLocked'] = row[5]
#             new_dict['doorlockedpercent'] = doorlockedpercent
#             new_dict['MeterDefective'] = row[4]
#             new_dict['meterdefectivepercent'] = meterdefectivepercent
#             new_dict['total'] = total

#             newdata = listfun(new_dict)
#         return Response(newdata)
#     except:
#         return Response([])


@api_view(["GET"])
def mrwisedailydata(request):
    new = []

    def listfun(dict):
        new.append(dict.copy())
        return new

    new_dict = {}
    today = date.today()
    print("Today's date:", today)

    cursor = connection.cursor()
    # query=f''''
    # from readingmaster r where r.reading_date_db={today}
    # '''
    result = dict(cursor.execute(
        "SELECT r.mr_id as mrid FROM readingmaster r"))
    print("result---->", result)

    return Response("result")


def dictfetchall(cursor):
    columns = [col[0] for col in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


@api_view(["GET"])
def consdetail(request):
    print("hi")
    acno = request.query_params.get("acno", None)
    cursor = connection.cursor()

    cursor.execute(
        f"""
    SELECT  m.ofc_discom,m.ofc_zone,m.ofc_circle,m.ofc_division,m.ofc_subdivision,m.ofc_section,
    m.id,m.cons_name,m.cons_ac_no,cons_address,m.cons_ph_no,m.con_trf_cat,m.mr_unit,
    r."mrId",r."mrName",r."mrPhone",r."mrPhoto" as avatar,m.con_mtr_sl_no,
    m.rdng_date,m.prsnt_mtr_status,m.prsnt_md_rdng,m.ocr_pf_reading,m.abnormality,m.prsnt_rdng_ocr_excep,m.md_ocr_excep,m.mr_rmrk,m.qc_req,m.ai_mdl_ver,m.ph_name,m.cmra_res,m.andr_ver,m.reading_date_db,
    m.rdng_img,m.md_img,m.pf_image,m.prsnt_ocr_rdng,m.prsnt_rdng,m.prsnt_md_rdng_ocr,m.rdng_ocr_status,m.ocr_pf_status,m.pf_manual_reading,m.prsnt_md_rdng,m.ocr_md_status,m.md_img
      FROM
    readingmaster m,meterreaderregistration r where m.mr_id=r."mrId" AND m.cons_ac_no='{acno}'

    """
    )
    results = dictfetchall(cursor)

    return Response(results)


@api_view(["GET"])
def qccheck(request):
    new = []

    def listfun(dict):
        new.append(dict.copy())
        return new

    cursor = connection.cursor()

    cursor.execute(
        f"""select r.id,r.cons_name,r.cons_ac_no,r.con_mtr_sl_no,r.prsnt_rdng,r.prsnt_ocr_rdng,r.rdng_ocr_status,r.rdng_img,

                   r.prsnt_mtr_status,r.abnormality from readingmaster r where r.qc_req='Yes' order by id DESC LIMIT 1

    """
    )
    newdict = {}
    result = cursor.fetchall()
    for row in result:
        id = row[0]
        consumername = row[1]
        consumeraccno = row[2]
        consumermeterslno = row[3]
        prsntrdng = row[4]
        prsntocrrdng = row[5]
        rdngocrstatus = row[6]
        rdngimg = row[7]
        prsntmtrstatus = row[8]
        abnormality = row[9]

    newdict["id"] = id
    newdict["consumername"] = consumername
    newdict["consumeraccno"] = consumeraccno
    newdict["consumermeterslno"] = consumermeterslno
    newdict["prsntrdng"] = prsntrdng
    newdict["prsntocrrdng"] = prsntocrrdng
    newdict["rdngocrstatus"] = rdngocrstatus
    newdict["rdngimg"] = rdngimg
    newdict["prsntmtrstatus"] = prsntmtrstatus
    newdict["abnormality"] = abnormality

    newdata = listfun(newdict)
    print(newdict)
    return Response(newdict)


@api_view(["POST"])
def qccheckupdate(request):  # sourcery skip: avoid-builtin-shadow
    newdata = request.data
    print("---------------------------------->", newdata)
    try:
        id = Consumers.objects.get(id=newdata["id"])
        serializer = ConsumerDataSerializer(id, data=newdata, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response({"status": True, "message": "QC DONE SUCCESSFULLY"})
        return Response(
            {"status": False, "message": "Something wrong"}, status=status.HTTP_200_OK
        )
    except Consumers.DoesNotExist:
        return Response(
            {"status": False, "message": "id not present"}, status=status.HTTP_200_OK
        )


# @api_view(['GET'])
# def qcreport(request):
#     new=[]
#     def listfun(dict):
#         new.append(dict.copy())
#         return new
#     cursor=connection.cursor()

#     cursor.execute(f'''select r.mr_id,r.cons_ac_no,r.rdng_ocr_status,r.prsnt_mtr_status,r.abnormality from readingmaster r                  where r.qc_req='no' order by id

#     ''')
#     newdict={}
#     result=cursor.fetchall()
#     for row in result:
#         id=row[0]
#         consumername=row[1]
#         consumeraccno=row[2]
#         consumermeterslno=row[3]
#         prsntrdng=row[4]
#         prsntocrrdng=row[5]
#         rdngocrstatus=row[6]
#         rdngimg=row[7]
#         prsntmtrstatus=row[8]
#         abnormality=row[9]


#     newdict['id']=id
#     newdict['consumername']=consumername
#     newdict['consumeraccno']=consumeraccno
#     newdict['consumermeterslno']=consumermeterslno
#     newdict['prsntrdng']=prsntrdng
#     newdict['prsntocrrdng']=prsntocrrdng
#     newdict['rdngocrstatus']=rdngocrstatus
#     newdict['rdngimg']=rdngimg
#     newdict['prsntmtrstatus']=prsntmtrstatus
#     newdict['abnormality']=abnormality

#     newdata=listfun(newdict)
#     print(newdict)
#     return Response(newdict)

# @api_view(['GET',])
# def mvcheck(request):
#     todaydate=date.today()
#     pagesize = request.query_params.get("pagesize",)
#     paginator = PageNumberPagination()
#     paginator.page_size = pagesize
#     monthByData = request.query_params.get("getMonth", None)
#     orderby = request.query_params.get("orderby", None)
#     mrid = request.query_params.get('mrid', None)
#     startDate = request.query_params.get('startdate', None)
#     endDate = request.query_params.get('enddate', None)
#     endDate = request.query_params.get('enddate', None)
#     searchdata = request.query_params.get('searchdata', None)

#     # all=request.query_params.get('all',None)
#     # person_objects=Consumers.objects.all().order_by('-id')
#     clause = ''
#     if monthByData:
#         monthByData = monthByData.split('-')[1]
#         print("month--after---------", monthByData)
#         clause = 'where EXTRACT(MONTH FROM m.reading_date_db)=%s'
#         params = ([monthByData])
#         # person_objects=person_objects.filter(data_sync_date__month=12)
#     elif startDate and endDate:
#         clause = 'where m.reading_date_db BETWEEN %s AND %s'
#         params = ([startDate, endDate])
#     elif mrid:
#         clause = 'where m.mr_id=%s'
#         params = ([mrid])
#     elif searchdata:
#         clause = 'where m.mr_id=%s'
#         print(clause)

#         params = [searchdata]

#         pass
#     else:
#         clause = f"where m.reading_date_db='{todaydate}'"
#         params = ''
#     cursor = connection.cursor()
#     query = (f'''select m.mr_id as "mrId",m.rdng_date,m.prsnt_mtr_status,m.prsnt_ocr_rdng,m.prsnt_rdng,m.ocr_pf_reading,m.cons_name,m.prsnt_md_rdng_ocr,m.rdng_ocr_status,m.rdng_img,m.prsnt_md_rdng,m.id,r."mrPhoto"
#                     from readingmaster m left outer join meterreaderregistration r on m.mr_id=r."mrId" {clause} order by m.rdng_date {orderby}

#         ''')
#     print(query)
#     cursor.execute(query, params)
#     person_objects = dictfetchall(cursor)
#     result_page = paginator.paginate_queryset(person_objects, request)
#     # serializer = ConsumerDataSerializer1(result_page, many=True)
#     return paginator.get_paginated_response(result_page)


@api_view(
    [
        "GET",
    ]
)
def mvcheck(request):
    pagesize = request.query_params.get(
        "pagesize",
    )
    page = request.query_params.get(
        "page",
    )
    offset = (int(pagesize) * int(page)) - int(pagesize)
    monthByData = request.query_params.get("getMonth", None)
    orderby = request.query_params.get("orderby", None)
    mrid = request.query_params.get("mrid", None)
    startDate = request.query_params.get("startdate", None)
    endDate = request.query_params.get("enddate", None)
    searchdata = request.query_params.get("searchdata", None)
    clause = ""
    if monthByData:
        monthByData = monthByData.split("-")[1]
        print("month--after---------", monthByData)
        clause = f"where EXTRACT(MONTH FROM m.reading_date_db)='{monthByData}'"

    elif startDate and endDate:
        clause = f"where m.reading_date_db BETWEEN {startDate} AND {endDate}"

    elif mrid:
        clause = f"where m.mr_id='{mrid}'"

    elif searchdata:
        clause = f"where m.mr_id='{searchdata}' or m.cons_ac_no='{searchdata}' or m.cons_name='{searchdata}'"
        print(clause)
    else:
        clause = ""
    query = Consumers.objects.raw(
        f"""select m.mr_id as "mrId",m.rdng_date,m.prsnt_mtr_status,m.prsnt_ocr_rdng,m.prsnt_rdng,m.ocr_pf_reading,m.cons_name,m.cons_ac_no,m.prsnt_md_rdng_ocr,m.rdng_ocr_status,m.rdng_img,m.prsnt_md_rdng,m.id,r."mrPhoto",count(*) over () as total_count
                    from readingmaster m left outer join meterreaderregistration r on m.mr_id=r."mrId" {clause} order by m.rdng_date {orderby} limit {pagesize} offset {offset}

    """
    )
    print(query)
    serializer = ConsumersMeterRegistration(query, many=True)
    count = serializer.data[0]["total_count"] if serializer.data else 0
    return Response({"count": count, "results": serializer.data})


@api_view(["GET"])
def minidashboardmonth(request):
    month = date.today().month
    today = date.today()
    print("month", month)
    agency = request.query_params.get("agency")
    clause = ""
    if agency != "null":
        clause = f"and bl_agnc_name='{agency}'"
        print("clause---------->", clause)

    cursor = connection.cursor()
    query = f"""select
             count(readingmaster.rdng_ocr_status='Passed' or NULL),
             count(readingmaster.rdng_ocr_status='Failed' or NULL),
             count(readingmaster.prsnt_mtr_status='Meter Defective' or NULL),
             count(readingmaster.prsnt_mtr_status='Door Locked' or NULL),
             count(readingmaster.prsnt_mtr_status='Ok' or NULL),
             count(*),
             count(distinct mr_id)
             from readingmaster WHERE EXTRACT (MONTH FROM reading_date_db)='{month}' {clause}
    """
    cursor.execute(query)

    result = cursor.fetchall()
    for row in result:
        metervisionaireadings = row[0]
        metervisionaireadingswithocrexception = row[1]
        doorlockedreadings = row[3]
        meterdefectivereadings = row[2]
        okreadings = row[4]
        mrid = row[6]

    newdict = {
        "mrid": mrid,
        "okreadings": okreadings,
        "ocrreadings": metervisionaireadings,
        "ocrwithexception": metervisionaireadingswithocrexception,
        "doorlocked": doorlockedreadings,
        "meterdefective": meterdefectivereadings,
        "totalreadings": row[5],
    }
    print(newdict)

    return Response(newdict)


# from django.db.models import Count
# from django.db.models.functions import ExtractMonth
# from .models import Consumers

# @api_view(['GET'])
# def minidashboardmonth(request):
#     month = date.today().month
#     agency = request.query_params.get("agency")
#     clause = {}

#     if agency != 'null':
#         clause['bl_agnc_name'] = agency

#     readings = Consumers.objects.filter(reading_date_db__month=month, **clause).aggregate(
#         ocr_passed_count=Count('rdng_ocr_status', filter=Q(rdng_ocr_status='Passed')),
#         ocr_failed_count=Count('rdng_ocr_status', filter=Q(rdng_ocr_status='Failed')),
#         meter_defective_count=Count('prsnt_mtr_status', filter=Q(prsnt_mtr_status='Meter Defective')),
#         door_locked_count=Count('prsnt_mtr_status', filter=Q(prsnt_mtr_status='Door Locked')),
#         ok_count=Count('prsnt_mtr_status', filter=Q(prsnt_mtr_status='Ok')),
#         rdng_img_count=Count('rdng_img'),
#         mrid_count=Count('mr_id', distinct=True)
#     )

#     newdict = {
#         'mrid': readings['mrid_count'],
#         'okreadings': readings['ok_count'],
#         'ocrreadings': readings['ocr_passed_count'],
#         'ocrwithexception': readings['ocr_failed_count'],
#         'doorlocked': readings['door_locked_count'],
#         'meterdefective': readings['meter_defective_count'],
#         'totalreadings': readings['rdng_img_count'],
#     }

#     return Response(newdict)


@parser_classes([MultiPartParser, FormParser])
@api_view(["POST"])
def consumerstest(request):
    data = request.data
    _mutable = data._mutable
    data._mutable = True
    rdng_date = data["rdng_date"]
    print("rdng_date", rdng_date)
    reading_date_db = rdng_date[:10]
    print("reading_date_db", reading_date_db)
    data["reading_date_db"] = reading_date_db
    data._mutable = _mutable

    return Response("ok")


@api_view(["GET"])
def clusters(request):
    # paginator = PageNumberPagination()
    # paginator.page_size = 100
    today = date.today()

    # data = Consumers.objects.all()

    # serializer = ConsumerSerializer(data, many=True)
    # result_page = paginator.paginate_queryset(serializer.data, request)
    # # serializer = ConsumerDataSerializer1(result_page, many=True)
    # clause = ''
    cursor = connection.cursor()
    query = f"""
    select mr_id,rdng_date,cons_name,geo_lat,geo_long,prsnt_mtr_status,rdng_ocr_status,rdng_img from readingmaster where reading_date_db='{today}'
    """
    cursor.execute(query)
    result = dictfetchall(cursor)
    return Response(result)


@api_view(["GET"])
def monthdata(request):
    month = date.today().month
    new = []

    def listfun(dict):
        print(dict)
        new.append(dict.copy())
        return new

    newdict = {}
    cursor = connection.cursor()

    #     query = f"""SELECT r.mr_id,r.ofc_discom,r.ofc_zone,r.ofc_circle,r.ofc_division,r.ofc_subdivision,count(r.cons_ac_no) as billed_consumers,
    # count(r.prsnt_mtr_status='Ok' OR NULL) as ok_readings,count(r.rdng_ocr_status='Passed' or NULL) as OCRwithoutException,
    # count(r.rdng_ocr_status='Failed' OR NULL) as OCRwithException,count(r.prsnt_mtr_status='Meter Defective' OR NULL) as MeterDefective,
    # count(r.prsnt_mtr_status='Door Locked' OR NULL) as DoorLocked
    # FROM readingmaster r
    # WHERE EXTRACT (MONTH FROM reading_date_db)='{month}'
    # GROUP BY r.mr_id,r.ofc_discom,r.ofc_zone,r.ofc_circle,r.ofc_division,r.ofc_subdivision;
    query = f"""SELECT r.mr_id,r.ofc_discom,r.ofc_zone,r.ofc_circle,r.ofc_division,r.ofc_subdivision,r.bl_agnc_name as Agency,
    count(r.cons_ac_no) as billed_consumers,
count(r.prsnt_mtr_status='Ok' OR NULL) as ok_readings,
count(r.rdng_ocr_status='Passed' or NULL) as OCRwithoutException,
count(r.rdng_ocr_status='Failed' OR NULL) as OCRwithException,
count(r.prsnt_mtr_status='Meter Defective' OR NULL) as MeterDefective,
count(r.prsnt_mtr_status='Door Locked' OR NULL) as DoorLocked
FROM readingmaster r
WHERE EXTRACT (MONTH FROM reading_date_db)='{month}'
GROUP BY r.mr_id,r.ofc_discom,r.ofc_zone,r.ofc_circle,r.ofc_division,r.ofc_subdivision,r.bl_agnc_name ;

        """
    cursor.execute(query)
    results = cursor.fetchall()

    try:
        for row in results:
            total = row[7]
            okreadings = row[8]
            ocrreadings = row[9]
            ocrwithexcep = row[10]
            meterdefective = row[11]
            doorlocked = row[12]

            # Percentage
            okreadpercent = round(((okreadings / total) * 100), 2)
            ocrreadingpercent = round(
                (((ocrreadings / okreadings) if okreadings else 0) * 100), 2
            )
            ocrwithexceppercent = round(
                (((ocrwithexcep / okreadings) if okreadings else 0) * 100), 2
            )
            meterdefectivepercent = round(
                (((meterdefective / total) if total else 0) * 100), 2
            )
            doorlockedpercent = round(
                (((doorlocked / total) if total else 0) * 100), 2)
            # add to dictionary
            newdict["mrid"] = row[0]
            # newdict['mrPhone'] = row[10]
            newdict["ofc_discom"] = row[1]
            newdict["ofc_zone"] = row[2]
            newdict["ofc_circle"] = row[3]
            newdict["ofc_division"] = row[4]
            newdict["ofc_subdivision"] = row[5]
            newdict["agency"] = row[6]
            newdict["billed_consumers"] = row[7]
            newdict["meterdefective"] = row[11]
            newdict["doorlocked"] = row[12]

            newdict["OKreadings"] = okreadings
            newdict["OCRReadings"] = ocrreadings
            newdict["OCRwithException"] = ocrwithexcep
            newdict["OKreadingspercent"] = okreadpercent
            newdict["OCRReadingspercent"] = ocrreadingpercent
            newdict["OCRwithExceptionpercent"] = ocrwithexceppercent
            newdict["MeterDefectivePercent"] = meterdefectivepercent
            newdict["DoorLockedOercent"] = doorlockedpercent
            # add to list
            newdata = listfun(newdict)
        wb = Workbook()
        ws = wb.active
        ws.title = "JSON Data"
        headers = list(newdata[0].keys())
        for j, header in enumerate(headers):
            ws.cell(row=1, column=j + 1, value=header)
        for i, row in enumerate(newdata, start=2):
            for j, key in enumerate(headers):
                ws.cell(row=i, column=j + 1, value=row[key])
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=data.xlsx"
        wb.save(response)
        return response
    except:
        return Response([])


@api_view(["GET"])
def dailydata(request):
    todaydate = date.today()
    new = []

    def listfun(dict):
        new.append(dict.copy())
        return new

    newdict = {}
    cursor = connection.cursor()
    query = f"""SELECT r.mr_id,r.ofc_discom,r.ofc_zone,r.ofc_circle,r.ofc_division,r.ofc_subdivision,r.bl_agnc_name as Agency,
    count(r.cons_ac_no) as billed_consumers,
count(r.prsnt_mtr_status='Ok' OR NULL) as ok_readings,
count(r.rdng_ocr_status='Passed' or NULL) as OCRwithoutException,
count(r.rdng_ocr_status='Failed' OR NULL) as OCRwithException
FROM readingmaster r
WHERE reading_date_db ='{todaydate}'
GROUP BY r.mr_id,r.ofc_discom,r.ofc_zone,r.ofc_circle,r.ofc_division,r.ofc_subdivision,r.bl_agnc_name ;

        """
    cursor.execute(query)
    results = cursor.fetchall()
    try:
        for row in results:
            total = row[7]
            okreadings = row[8]
            ocrreadings = row[9]
            ocrwithexcep = row[10]
            # Percentage
            okreadpercent = round(((okreadings / total) * 100), 2)
            ocrreadingpercent = round(
                (((ocrreadings / okreadings) if okreadings else 0) * 100), 2
            )
            ocrwithexceppercent = round(
                (((ocrwithexcep / okreadings) if okreadings else 0) * 100), 2
            )
            # add to dictionary
            newdict["mrid"] = row[0]
            # newdict['mrPhone'] = row[10]
            newdict["ofc_discom"] = row[1]
            newdict["ofc_zone"] = row[2]
            newdict["ofc_circle"] = row[3]
            newdict["ofc_division"] = row[4]
            newdict["ofc_subdivision"] = row[5]
            newdict["agency"] = row[6]
            newdict["billed_consumers"] = row[7]
            newdict["OKreadings"] = okreadings
            newdict["OCRReadings"] = ocrreadings
            newdict["OCRwithException"] = ocrwithexcep
            newdict["OKreadingspercent"] = okreadpercent
            newdict["OCRReadingspercent"] = ocrreadingpercent
            newdict["OCRwithExceptionpercent"] = ocrwithexceppercent
            # add to list
            newdata = listfun(newdict)
        wb = Workbook()
        ws = wb.active
        ws.title = "JSON Data"
        headers = list(newdata[0].keys())
        for j, header in enumerate(headers):
            ws.cell(row=1, column=j + 1, value=header)
        for i, row in enumerate(newdata, start=2):
            for j, key in enumerate(headers):
                ws.cell(row=i, column=j + 1, value=row[key])
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=data.xlsx"
        wb.save(response)
        return response
    except:
        return Response([])


@api_view(["GET"])
def testdata(request):
    todaydate = date.today()
    new = []

    def listfun(dict):
        print(dict)
        new.append(dict.copy())
        return new

    newdict = {}
    cursor = connection.cursor()
    query = f"""SELECT r.mr_id,r.ofc_discom,r.ofc_zone,r.ofc_circle,r.ofc_division,r.ofc_subdivision,count(r.cons_ac_no) as billed_consumers,
count(r.prsnt_mtr_status='Ok' OR NULL) as ok_readings,count(r.rdng_ocr_status='Passed' or NULL) as OCRwithoutException,
count(r.rdng_ocr_status='Failed' OR NULL) as OCRwithException, m."mrPhone"
FROM readingmaster r
JOIN meterreaderregistration m ON m."mrId" = r.mr_id
AND reading_date_db ='{todaydate}'
GROUP BY r.mr_id,r.ofc_discom,r.ofc_zone,r.ofc_circle,r.ofc_division,r.ofc_subdivision, m."mrPhone";
        """
    cursor.execute(query)
    results = cursor.fetchall()
    try:
        for row in results:
            total = row[6]
            okreadings = row[7]
            ocrreadings = row[8]
            ocrwithexcep = row[9]

            # Percentage
            okreadpercent = round(((okreadings / total) * 100), 2)
            ocrreadingpercent = round(
                (((ocrreadings / okreadings) if okreadings else 0) * 100), 2
            )
            ocrwithexceppercent = round(
                (((ocrwithexcep / okreadings) if okreadings else 0) * 100), 2
            )

            # add to dictionary
            newdict["mrid"] = row[0]
            newdict["mrPhone"] = row[10]
            newdict["ofc_discom"] = row[1]
            newdict["ofc_zone"] = row[2]
            newdict["ofc_circle"] = row[3]
            newdict["ofc_division"] = row[4]
            newdict["ofc_subdivision"] = row[5]
            newdict["billed_consumers"] = row[6]

            newdict["OKreadings"] = okreadings
            newdict["OCRReadings"] = ocrreadings
            newdict["OCRwithException"] = ocrwithexcep

            newdict["OKreadingspercent"] = okreadpercent
            newdict["OCRReadingspercent"] = ocrreadingpercent
            newdict["OCRwithExceptionpercent"] = ocrwithexceppercent

            # add to list
            newdata = listfun(newdict)
        print("newdata--", newdata)
        wb = Workbook()
        ws = wb.active
        ws.title = "JSON Data"
        headers = list(newdata[0].keys())
        for j, header in enumerate(headers):
            ws.cell(row=1, column=j + 1, value=header)

        for i, row in enumerate(newdata, start=2):
            for j, key in enumerate(headers):
                ws.cell(row=i, column=j + 1, value=row[key])
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=data.xlsx"
        wb.save(response)
        return response

    except:
        return Response([])


@api_view(["GET"])
def geocluster(request):
    today = date.today()
    print(today)
    cursor = connection.cursor()
    query = f"""
    select m.mr_id as id,m.geo_lat,m.geo_long from readingmaster m where reading_date_db='{today}'

    """

    cursor.execute(query)

    result = dictfetchall(cursor)
    geo = to_geojson(result)

    return Response(geo)

    pass


def convertdicttolist(lsts, key):
    return [x.get(key) for x in lsts]


# @api_view(['POST'])
# def locationdiscom(request):
#     locationwise=request.data.get('locationwise')
#     locationname=request.data.get('locationname')
#     groupby=request.data.get('groupby')
#     print("groupby",groupby)
#     cursor = connection.cursor()
#     if ((locationwise is not None) and (locationname == 'all')):
#         location = 'ofc_'+locationwise
#         cursor.execute(f'''
#    select {groupby} as location
#     from office  GROUP BY {groupby}
#     ''')
#     if ((locationwise is not None) and (locationname != 'all')):
#         location = 'ofc_'+locationwise
#         clause = "WHERE "+locationwise+"='"+locationname+"' "
#         cursor.execute(f'''
#    select {groupby} as location
#     from office {clause}  GROUP BY {groupby}
#     ''')

#     result = dictfetchall(cursor)
#     res=(convertdicttolist(result, 'location'))
#     return Response(res)


@api_view(["POST"])
def locationzone(request):
    locationwise = request.data.get("locationwise")
    locationname = request.data.get("locationname")
    groupby = request.data.get("groupby")
    where = request.data.get("where")
    previouslocation = request.data.get("previouslocation")
    clause = ""
    print("groupby", groupby)
    cursor = connection.cursor()
    if (locationwise is not None) and (locationname == "all"):
        location = "ofc_" + locationwise
        clause = "WHERE " + previouslocation + "='" + where + "' "
        cursor.execute(
            f"""
   select {groupby} as location
    from office {clause}  GROUP BY {groupby}
    """
        )
    if (locationwise is not None) and (locationname != "all"):
        clause = "WHERE " + locationwise + "='" + locationname + "' "
        cursor.execute(
            f"""
   select {groupby} as location
    from office {clause}  GROUP BY {groupby}
    """
        )

    result = dictfetchall(cursor)
    res = convertdicttolist(result, "location")
    return Response(res)


@api_view(["POST"])
def locationwisezone(request):
    month = date.today().month

    newdict = {}
    new = []

    def listfun(dict):
        print(dict)
        new.append(dict.copy())
        return new

    cursor = connection.cursor()
    locationwise = request.data.get("locationwise", None)
    locationname = request.data.get("locationname", None)
    groupby = request.data.get("groupby")
    where = request.data.get("where")
    previouslocation = request.data.get("previouslocation")
    clause = ""
    if (locationwise is not None) and (locationname == "all"):
        location = "ofc_" + locationwise
        clause = "WHERE " + previouslocation + "='" + where + "' "
        cursor.execute(
            f"""
   select {groupby} as location, count(distinct mr_id),count(prsnt_mtr_status='Ok' or NULL),count(rdng_ocr_status='Passed' or NULL),count(rdng_ocr_status='Failed' or NULL),count(prsnt_mtr_status='Meter Defective' or NULL),count(prsnt_mtr_status='Door Locked' or NULL),count(mr_id)
    from readingmaster {clause} and extract(month from reading_date_db)='{month}' and {groupby}!=''  GROUP BY {groupby}
    """
        )
    if (locationwise is not None) and (locationname != "all"):
        location = "ofc_" + locationwise
        clause = "WHERE " + locationwise + "='" + locationname + "' "

        cursor.execute(
            f"""
   select {groupby} as location, count(distinct mr_id),count(prsnt_mtr_status='Ok' or NULL),count(rdng_ocr_status='Passed' or NULL),count(rdng_ocr_status='Failed' or NULL),count(prsnt_mtr_status='Meter Defective' or NULL),count(prsnt_mtr_status='Door Locked' or NULL),count(mr_id)
    from readingmaster {clause} and extract(month from reading_date_db)='{month}' and {groupby}!='' GROUP BY {groupby}
    """
        )

    result = cursor.fetchall()
    print("result", result)
    try:
        for row in result:
            locationname = row[0]
            total = row[7]
            mrid = row[1]
            okreadings = row[2]
            OcrReadings = row[3]
            Ocrwithexception = row[4]
            meterDefective = row[5]
            doorLocked = row[6]
            okreadpercent = math.floor((okreadings / total) * 100)
            ocrreadingpercent = math.floor(
                ((OcrReadings / okreadings) if okreadings else 0) * 100
            )
            ocrwithexceppercent = math.floor(
                ((Ocrwithexception / okreadings) if okreadings else 0) * 100
            )
            meterdefectivepercent = math.floor((meterDefective / total) * 100)
            doorlockedpercent = math.floor((doorLocked / total) * 100)
            newdict["locationname"] = row[0]
            newdict["mrid"] = row[1]
            newdict["okreadings"] = row[2]
            newdict["okreadingspercent"] = okreadpercent
            newdict["OcrReadings"] = row[3]
            newdict["OcrReadingspercent"] = ocrreadingpercent
            newdict["Ocrwithexception"] = row[4]
            newdict["Ocrwithexceptionpercent"] = ocrwithexceppercent
            newdict["meterDefective"] = row[5]
            newdict["meterDefectivepercent"] = meterdefectivepercent
            newdict["doorLocked"] = row[6]
            newdict["doorLockedpercent"] = doorlockedpercent
            newdict["total"] = total
            data = listfun(newdict)
        return Response(data)
    except:
        return Response([])


def to_geojson(entries):
    features = []
    for entry in entries:
        if entry["geo_lat"] != "" and entry["geo_long"] != "":
            geolat = entry["geo_lat"] = float(entry["geo_lat"])
            geolong = entry["geo_long"] = float(entry["geo_long"])
            point = Point([entry["geo_long"], entry["geo_lat"]])

            del entry["geo_lat"]
            del entry["geo_long"]
            feature = Feature(geometry=point, properties=entry)
            features.append(feature)
    crs = {"type": "name", "properties": {
        "name": "urn:ogc:def:crs:OGC:1.3:CRS84"}}
    feature_collection = FeatureCollection(crs=crs, features=features)
    return feature_collection


def filtermethod():
    pass


@api_view(["GET"])
def tester(request):
    pass


@api_view(["GET"])
def topmeterreaders1(request):
    month = date.today().month
    agency = request.query_params.get("agency")
    new = []
    cursor = connection.cursor()
    clause = ""
    if agency != "null":
        clause = f"and m.bl_agnc_name='{agency}'"
    query = f"""
SELECT r."mrName",

       SUM(CASE WHEN m.rdng_ocr_status='Passed' THEN 1 ELSE 0 END) as passed,

       SUM(CASE WHEN m.rdng_ocr_status='Failed' THEN 1 ELSE 0 END) as failed,

       count(m.prsnt_mtr_status='Ok' or NULL) AS ok,



       ROUND((cast(count(prsnt_mtr_status='Meter Defective' or null) as float)

/ COALESCE(cast(count(mr_id) as float),1) * 100)::numeric, 2) as Meter_Defective,

       ROUND((cast(count(prsnt_mtr_status='Door Locked' or null) as float)

/ COALESCE(cast(count(mr_id) as float),1) * 100)::numeric, 2) as Door_locked,



        count(m.mr_id) as total,



       CASE

           WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0

           ELSE ROUND((cast(count(rdng_ocr_status='Passed' or null) as float)

           / cast(count(prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)

       END as passed_percent,



       CASE

           WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0

           ELSE ROUND((cast(count(rdng_ocr_status='Failed' or null) as float)

           / cast(count(prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)

       END as failed_percent,



       ROUND((cast(count(prsnt_mtr_status='Ok' or null) as float)

/ COALESCE(cast(count(mr_id) as float),1) * 100)::numeric, 2) as OK_percent,

       r."mrPhoto",

       r."mrPhone",

       r."mrId",



       (DATE_PART('day', CURRENT_DATE) * 30) as "current_day_product"

FROM readingmaster m , meterreaderregistration r

     where m.mr_id = r."mrId"

AND EXTRACT(MONTH FROM m.reading_date_db)='{month}' {clause}

GROUP BY r."mrName",

         r."mrPhoto",

         r."mrPhone",

         r."mrId"

HAVING count(m.mr_id)> (DATE_PART('day', CURRENT_DATE) * 30) and ( ROUND((cast(count(prsnt_mtr_status='Door Locked' or null) as float)

/ COALESCE(cast(count(mr_id) as float),1) * 100)::numeric, 2) <=2) and (ROUND((cast(count(prsnt_mtr_status='Meter Defective' or null) as float)

/ COALESCE(cast(count(mr_id) as float),1) * 100)::numeric, 2) <=20)

ORDER BY passed_percent  DESC

limit 10

"""
    print("query")

    cursor.execute(query)
    result = cursor.fetchall()

    try:
        for row in result:
            meterreader = row[0]
            passed = row[1]
            failed = row[2]
            okreadings = row[3]
            meterdefective = row[4]
            doorlocked = row[5]
            total = row[6]
            passed_percent = row[7]
            failed_percent = row[8]
            ok_percent = row[9]
            mr_photo = row[10]
            mr_phone = row[11]
            mr_id = row[12]

            new_dict = {
                "meterreader": meterreader,
                "passed": passed,
                "failed": failed,
                "OK": okreadings,
                "meterdefectivepercent": meterdefective,
                "doorlockedpercent": doorlocked,
                "total": total,
                "ocrreadingpercent": passed_percent,
                "ocrwithexceppercent": failed_percent,
                "okreadpercent": ok_percent,
                "mrphoto": mr_photo,
                "mrphone": mr_phone,
                "mrId": mr_id,
            }
            new.append(new_dict)

        return Response(new)
    except:
        return Response(status=status.HTTP_404_NOT_FOUND)


@api_view(["GET"])
def minidashboardsbpdclmonth(request):
    month = date.today().month
    today = date.today()
    print("month", month)
    cursor = connection.cursor()
    agency = request.query_params.get("agency")
    clause = ""
    if agency != "null":
        clause = f"and bl_agnc_name='{agency}'"
        print("clause---------->", clause)

    cursor.execute(
        f"""select count(readingmaster.rdng_ocr_status='Passed' or NULL),count(readingmaster.rdng_ocr_status='Failed' or NULL),count(readingmaster.prsnt_mtr_status='Meter Defective' or NULL),count(readingmaster.prsnt_mtr_status='Door Locked' or NULL),count(readingmaster.prsnt_mtr_status='Ok' or NULL),count(*),count(distinct mr_id)
                    from readingmaster WHERE EXTRACT (MONTH FROM reading_date_db)='{month}' and readingmaster.ofc_discom='SBPDCL' {clause}
    """
    )

    newdict = {}
    result = cursor.fetchall()
    for row in result:
        metervisionaireadings = row[0]
        metervisionaireadingswithocrexception = row[1]
        doorlockedreadings = row[3]
        meterdefectivereadings = row[2]
        okreadings = row[4]

    newdict["mrid"] = row[6]
    newdict["okreadings"] = okreadings

    newdict["ocrreadings"] = metervisionaireadings
    newdict["ocrwithexception"] = metervisionaireadingswithocrexception
    newdict["doorlocked"] = doorlockedreadings
    newdict["meterdefective"] = meterdefectivereadings
    newdict["totalreadings"] = row[5]

    print(newdict)
    return Response(newdict)


@api_view(["GET"])
def minidashboardnbpdclmonth(request):
    month = date.today().month
    today = date.today()
    agency = request.query_params.get("agency")
    clause = ""
    print("month", month)
    cursor = connection.cursor()
    if agency != "null":
        clause = f"and bl_agnc_name='{agency}'"

    cursor.execute(
        f"""select count(readingmaster.rdng_ocr_status='Passed' or NULL),count(readingmaster.rdng_ocr_status='Failed' or NULL),count(readingmaster.prsnt_mtr_status='Meter Defective' or NULL),count(readingmaster.prsnt_mtr_status='Door Locked' or NULL),count(readingmaster.prsnt_mtr_status='Ok' or NULL),count(*),count(distinct mr_id)
                    from readingmaster WHERE EXTRACT (MONTH FROM reading_date_db)='{month}' and readingmaster.ofc_discom='NBPDCL' {clause}
    """
    )

    newdict = {}
    result = cursor.fetchall()
    for row in result:
        metervisionaireadings = row[0]
        metervisionaireadingswithocrexception = row[1]
        doorlockedreadings = row[3]
        meterdefectivereadings = row[2]
        okreadings = row[4]

    newdict["mrid"] = row[6]
    newdict["okreadings"] = okreadings
    newdict["ocrreadings"] = metervisionaireadings
    newdict["ocrwithexception"] = metervisionaireadingswithocrexception
    newdict["doorlocked"] = doorlockedreadings
    newdict["meterdefective"] = meterdefectivereadings
    newdict["totalreadings"] = row[5]

    print(newdict)
    return Response(newdict)


@api_view(["GET"])
def exceptionlistsbpdclmonth(request):
    month = date.today().month

    agency = request.query_params.get("agency")
    cursor = connection.cursor()
    clause = ""
    if agency != "null":
        clause = f"and bl_agnc_name='{agency}'"

    query1 = f"""
   select count(prsnt_mtr_status='Ok' or NULL),
   count(prsnt_rdng_ocr_excep='Spoofed Image' or NULL),
   count(prsnt_rdng_ocr_excep='Image blur' or NULL),
   count(prsnt_rdng_ocr_excep='Meter Dirty' or NULL),
   count(prsnt_rdng_ocr_excep='Incorrect Reading' or NULL),
   count(reading_parameter_type='Parameters Mismatch' or null),
   count(reading_parameter_type='Parameters Unavailable' or null),
   count(reading_parameter_type='' or null),
   count(prsnt_rdng_ocr_excep='No Exception Found' or prsnt_rdng_ocr_excep=''  or null)

    from readingmaster where extract(month from reading_date_db)='{month}' and ofc_discom='SBPDCL' and rdng_ocr_status='Failed' {clause}
    """
    query2 = f"""
           select count(prsnt_mtr_status='Ok' or null) as OK from readingmaster where extract(month from reading_date_db)='{month}' and ofc_discom='SBPDCL'{clause}
           """

    cursor.execute(query1)
    results = cursor.fetchall()
    print(results)
    newdict = {}
    for i in results:
        newdict["OK"] = i[0]
        newdict["Image Spoofed"] = i[1]
        newdict["Image blur"] = i[2]
        newdict["Meter Dirty"] = i[3]
        newdict["Incorrect Reading"] = i[4]
        newdict["Parameters Mismatch"] = i[5]
        newdict["Parameters Unavailable"] = i[6]
        newdict["Others"] = i[7]
        newdict["Others."] = i[8]
    cursor.execute(query2)
    resulttotal = dictfetchall(cursor)

    return Response(
        {"total": resulttotal[0]["ok"] if resulttotal else 0, "data": newdict}
    )


@api_view(["GET"])
def exceptionlistnbpdclmonth(request):
    month = date.today().month

    agency = request.query_params.get("agency")

    cursor = connection.cursor()
    clause = ""
    if agency != "null":
        clause = f"and bl_agnc_name='{agency}'"

    query1 = f"""
   select count(prsnt_mtr_status='Ok' or NULL),
   count(prsnt_rdng_ocr_excep='Spoofed Image' or NULL),
   count(prsnt_rdng_ocr_excep='Image blur' or NULL),
   count(prsnt_rdng_ocr_excep='Meter Dirty' or NULL),
   count(prsnt_rdng_ocr_excep='Incorrect Reading' or NULL),
   count(reading_parameter_type='Parameters Mismatch' or null),
   count(reading_parameter_type='Parameters Unavailable' or null),
   count(reading_parameter_type='' or null),
   count(prsnt_rdng_ocr_excep='No Exception Found' or prsnt_rdng_ocr_excep=''  or null)

    from readingmaster where extract(month from reading_date_db)='{month}' and ofc_discom='NBPDCL' and rdng_ocr_status='Failed' {clause}
    """
    query2 = f"""
           select count(prsnt_mtr_status='Ok' or null) as OK from readingmaster where extract(month from reading_date_db)='{month}' and ofc_discom='NBPDCL'{clause}
           """

    cursor.execute(query1)
    results = cursor.fetchall()
    print(results)
    newdict = {}
    for i in results:
        newdict["OK"] = i[0]
        newdict["Image Spoofed"] = i[1]
        newdict["Image blur"] = i[2]
        newdict["Meter Dirty"] = i[3]
        newdict["Incorrect Reading"] = i[4]
        newdict["Parameters Mismatch"] = i[5]
        newdict["Parameters Unavailable"] = i[6]
        newdict["Others"] = i[7]
        newdict["Others."] = i[8]
    cursor.execute(query2)
    resulttotal = dictfetchall(cursor)

    return Response(
        {"total": resulttotal[0]["ok"] if resulttotal else 0, "data": newdict}
    )


@api_view(["GET"])
def topmeterreaderssbpdcl(request):
    new = []
    month = date.today().month
    cursor = connection.cursor()
    agency = request.query_params.get("agency")
    clause = ""
    if agency != "null":
        clause = f"and bl_agnc_name='{agency}'"
    query = f"""
SELECT r."mrName",
       SUM(CASE WHEN m.rdng_ocr_status='Passed' THEN 1 ELSE 0 END) as passed,
       SUM(CASE WHEN m.rdng_ocr_status='Failed' THEN 1 ELSE 0 END) as failed,
       count(m.prsnt_mtr_status='Ok' or NULL) AS ok,

       ROUND((cast(count(prsnt_mtr_status='Meter Defective' or null) as float)
/ COALESCE(cast(count(mr_id) as float),1) * 100)::numeric, 2) as Meter_Defective,
       ROUND((cast(count(prsnt_mtr_status='Door Locked' or null) as float)
/ COALESCE(cast(count(mr_id) as float),1) * 100)::numeric, 2) as Door_locked,

        count(m.mr_id) as total,

       CASE
           WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
           ELSE ROUND((cast(count(rdng_ocr_status='Passed' or null) as float)
           / cast(count(prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
       END as passed_percent,

       CASE
           WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
           ELSE ROUND((cast(count(rdng_ocr_status='Failed' or null) as float)
           / cast(count(prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
       END as failed_percent,

       ROUND((cast(count(prsnt_mtr_status='Ok' or null) as float)
/ COALESCE(cast(count(mr_id) as float),1) * 100)::numeric, 2) as OK_percent,

       r."mrPhoto",
       r."mrPhone",
       r."mrId",

       (DATE_PART('day', CURRENT_DATE) * 30) as "current_day_product"
FROM readingmaster m ,
     meterreaderregistration r
     where m.mr_id = r."mrId" AND ofc_discom='SBPDCL' {clause}
AND EXTRACT(MONTH FROM m.reading_date_db)='{month}'
GROUP BY r."mrName",
         r."mrPhoto",
         r."mrPhone",
         r."mrId"
HAVING count(m.mr_id)> (DATE_PART('day', CURRENT_DATE) * 30) and ( ROUND((cast(count(prsnt_mtr_status='Door Locked' or null) as float)
/ COALESCE(cast(count(mr_id) as float),1) * 100)::numeric, 2) <=2) and (ROUND((cast(count(prsnt_mtr_status='Meter Defective' or null) as float)
/ COALESCE(cast(count(mr_id) as float),1) * 100)::numeric, 2) <=20)
ORDER BY passed_percent  DESC
limit 10
"""
    print("query")

    cursor.execute(query)
    result = cursor.fetchall()

    try:
        for row in result:
            meterreader = row[0]
            passed = row[1]
            failed = row[2]
            okreadings = row[3]
            meterdefective = row[4]
            doorlocked = row[5]
            total = row[6]
            passed_percent = row[7]
            failed_percent = row[8]
            ok_percent = row[9]
            mr_photo = row[10]
            mr_phone = row[11]
            mr_id = row[12]

            new_dict = {
                "meterreader": meterreader,
                "passed": passed,
                "failed": failed,
                "OK": okreadings,
                "meterdefectivepercent": meterdefective,
                "doorlockedpercent": doorlocked,
                "total": total,
                "ocrreadingpercent": passed_percent,
                "ocrwithexceppercent": failed_percent,
                "okreadpercent": ok_percent,
                "mrphoto": mr_photo,
                "mrphone": mr_phone,
                "mrId": mr_id,
            }
            new.append(new_dict)

        return Response(new)
    except:
        return Response(status=status.HTTP_404_NOT_FOUND)


@api_view(["GET"])
def topmeterreadersnbpdcl(request):
    new = []
    cursor = connection.cursor()
    month = date.today().month
    agency = request.query_params.get("agency")
    clause = ""
    if agency != "null":
        clause = f"and bl_agnc_name='{agency}'"
    query = f"""
SELECT r."mrName",
       SUM(CASE WHEN m.rdng_ocr_status='Passed' THEN 1 ELSE 0 END) as passed,
       SUM(CASE WHEN m.rdng_ocr_status='Failed' THEN 1 ELSE 0 END) as failed,
       count(m.prsnt_mtr_status='Ok' or NULL) AS ok,

       ROUND((cast(count(prsnt_mtr_status='Meter Defective' or null) as float)
/ COALESCE(cast(count(mr_id) as float),1) * 100)::numeric, 2) as Meter_Defective,
       ROUND((cast(count(prsnt_mtr_status='Door Locked' or null) as float)
/ COALESCE(cast(count(mr_id) as float),1) * 100)::numeric, 2) as Door_locked,

        count(m.mr_id) as total,

       CASE
           WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
           ELSE ROUND((cast(count(rdng_ocr_status='Passed' or null) as float)
           / cast(count(prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
       END as passed_percent,

       CASE
           WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
           ELSE ROUND((cast(count(rdng_ocr_status='Failed' or null) as float)
           / cast(count(prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
       END as failed_percent,

       ROUND((cast(count(prsnt_mtr_status='Ok' or null) as float)
/ COALESCE(cast(count(mr_id) as float),1) * 100)::numeric, 2) as OK_percent,

       r."mrPhoto",
       r."mrPhone",
       r."mrId",

       (DATE_PART('day', CURRENT_DATE) * 30) as "current_day_product"
FROM readingmaster m ,
     meterreaderregistration r
     where m.mr_id = r."mrId" AND ofc_discom='NBPDCL'
AND EXTRACT(MONTH FROM m.reading_date_db)='{month}' {clause}
GROUP BY r."mrName",
         r."mrPhoto",
         r."mrPhone",
         r."mrId"
HAVING count(m.mr_id)> (DATE_PART('day', CURRENT_DATE) * 30) and ( ROUND((cast(count(prsnt_mtr_status='Door Locked' or null) as float)
/ COALESCE(cast(count(mr_id) as float),1) * 100)::numeric, 2) <=2) and (ROUND((cast(count(prsnt_mtr_status='Meter Defective' or null) as float)
/ COALESCE(cast(count(mr_id) as float),1) * 100)::numeric, 2) <=20)
ORDER BY passed_percent  DESC
limit 10

"""
    print("query")

    cursor.execute(query)
    result = cursor.fetchall()

    try:
        for row in result:
            meterreader = row[0]
            passed = row[1]
            failed = row[2]
            okreadings = row[3]
            meterdefective = row[4]
            doorlocked = row[5]
            total = row[6]
            passed_percent = row[7]
            failed_percent = row[8]
            ok_percent = row[9]
            mr_photo = row[10]
            mr_phone = row[11]
            mr_id = row[12]

            new_dict = {
                "meterreader": meterreader,
                "passed": passed,
                "failed": failed,
                "OK": okreadings,
                "meterdefectivepercent": meterdefective,
                "doorlockedpercent": doorlocked,
                "total": total,
                "ocrreadingpercent": passed_percent,
                "ocrwithexceppercent": failed_percent,
                "okreadpercent": ok_percent,
                "mrphoto": mr_photo,
                "mrphone": mr_phone,
                "mrId": mr_id,
            }
            new.append(new_dict)

        return Response(new)
    except:
        return Response(status=status.HTTP_404_NOT_FOUND)


@api_view(["GET"])
def clusterstest(request):
    mrid = request.query_params.get("mrid")
    today = date.today()
    clause = ""
    if mrid is not None:
        clause = "and mr_id" "='" + mrid + "' "
        print(clause)

    cursor = connection.cursor()
    query = f"""
    select mr_id,rdng_date,cons_name,geo_lat,geo_long,prsnt_mtr_status,rdng_ocr_status,rdng_img from readingmaster where reading_date_db='{today}' {clause}
    """
    cursor.execute(query)
    result = dictfetchall(cursor)
    return Response(result)


# @api_view(['GET'])
# def clusterstestnew(request):
#     mrid = request.query_params.get("mrid")
#     today = date.today()
#     clause = ""
#     if mrid is not None:
#         clause = "and mr_id""='"+mrid+"' "
#         print(clause)

#         cursor = connection.cursor()
#         query = f'''
#         select mr_id,rdng_date,cons_name,geo_lat,geo_long,prsnt_mtr_status,rdng_ocr_status,rdng_img from readingmaster where reading_date_db='{today}' {clause}
#         '''
#         cursor.execute(query)
#         result = dictfetchall(cursor)
#         return Response(result)
#     return Response([])


@api_view(["GET"])
def get_meter_summarytest(request):
    filters = request.query_params.dict()
    print("filters", filters)
    new = []

    def listfun(dict):
        # print(dict)
        new.append(dict.copy())
        return new

    newdict = {}
    cursor = connection.cursor()
    clause = f""""""

    if filters:
        clause += "WHERE "
        for i, (key, value) in enumerate(filters.items()):
            if i > 0:
                clause += "AND "
            if key == "month":
                print("key['month']", key)
                print("value['month']", value)
                month = value.split("-")[1]
                key = "extract(month from reading_date_db)"
                value = month

            clause += f"{key}='{value}'"
        print("clause", clause)
    # print("query",query)
    query = f"""select readingmaster.mr_id,count(readingmaster.id),count(readingmaster.prsnt_mtr_status='Ok' or NULL),count(readingmaster.rdng_ocr_status='Passed' or NULL),count(readingmaster.rdng_ocr_status='Failed' or NULL),count(readingmaster.prsnt_mtr_status='Meter Defective' or NULL),count(readingmaster.prsnt_mtr_status='Door Locked' or NULL)
            from readingmaster {clause}  group by readingmaster.mr_id
        """
    print("query", query)
    cursor.execute(query)
    results = cursor.fetchall()
    try:
        for row in results:
            total = row[1]
            okreadings = row[2]
            ocrreadings = row[3]
            ocrwithexcep = row[4]
            meterdefective = row[5]
            doorlocked = row[6]

            # Percentage
            okreadpercent = math.floor(((okreadings / total) * 100))
            ocrreadingpercent = math.floor(
                (((ocrreadings / okreadings) if okreadings else 0) * 100)
            )
            ocrwithexceppercent = math.floor(
                (((ocrwithexcep / okreadings) if okreadings else 0) * 100)
            )
            meterdefectivepercent = math.floor(
                ((meterdefective / total) * 100))
            doorlockedpercent = math.floor(((doorlocked / total) * 100))

            # add to dictionary
            newdict["mrid"] = row[0]
            newdict["totalReadings"] = row[1]
            newdict["OKreadings"] = okreadings
            newdict["OKreadingspercent"] = okreadpercent
            newdict["OCRReadings"] = ocrreadings
            newdict["OCRReadingspercent"] = ocrreadingpercent
            newdict["OCRwithException"] = ocrwithexcep
            newdict["OCRwithExceptionpercent"] = ocrwithexceppercent
            newdict["MeterDefective"] = meterdefective
            newdict["MeterDefectivepercent"] = meterdefectivepercent
            newdict["DoorLocked"] = doorlocked
            newdict["DoorLockedpercent"] = doorlockedpercent
            # add to list
            newdata = listfun(newdict)
        return Response(newdata)
    except:
        return Response([])


@api_view(["GET"])
def consumerwisemap(request):
    consacno = request.query_params.get("consacno")
    clause = ""

    if consacno is not None:
        clause = "where cons_ac_no" "='" + consacno + "' "
        print(clause)

    cursor = connection.cursor()
    query = f"""
    select mr_id,rdng_date,cons_name,geo_lat,geo_long,prsnt_mtr_status,rdng_ocr_status,rdng_img from readingmaster  {clause}
    """
    cursor.execute(query)
    result = dictfetchall(cursor)
    return Response(result)


# @api_view(['GET'])
# def geoclusternew(request):
#     statusby = request.query_params.get('statusby', None)
#     today = date.today()
#     print(today)
#     cursor = connection.cursor()
#     if statusby:
#         query = f'''
#         select m.mr_id as id,m.geo_lat,m.geo_long from readingmaster m where reading_date_db='{today}' and (m.prsnt_mtr_status='{statusby}' or NULL)
#     '''
#     else:
#         query = f'''
#         select m.mr_id as id,m.geo_lat,m.geo_long from readingmaster m where reading_date_db='{today}'
#         '''
#     cursor.execute(query)
#     result = dictfetchall(cursor)
#     geo = to_geojson(result)
#     return Response(geo)
#     pass


@api_view(["POST"])
def geoclusternew(request):
    print("qwerty")
    data = request.data.get("filters", None)
    today = date.today()
    clause = ""
    # try:
    if data:
        print("rtyui")
        clause += "WHERE"
        for i, (key, value) in enumerate(data.items()):
            if i > 0:
                clause += " AND "
            if key == "prsnt_mtr_status":
                clause += f" {key}='{value}'"
            if key == "bl_agnc_name":
                clause += f" {key}='{value}'"
    cursor = connection.cursor()
    if data:
        query = f"""
        select m.mr_id as id,m.geo_lat,m.geo_long from readingmaster m {clause} and reading_date_db='{today}'
    """
    else:
        query = f"""
        select m.mr_id as id,m.geo_lat,m.geo_long from readingmaster m where reading_date_db='{today}'
        """
    print(query)
    cursor.execute(query)
    result = dictfetchall(cursor)
    geo = to_geojson(result)
    return Response(geo)


@api_view(["GET"])
def dailybilling(request):
    todaydate = date.today()
    month = date.today().month
    new = []

    def listfun(dict):
        print(dict)
        new.append(dict.copy())
        return new

    newdict = {}
    cursor = connection.cursor()
    # query = f"""select o.zone as Zone,o.divisionname as Division,(count(distinct m.id)) as total_billed_today from office o left join readingmaster m on o.zone=m.ofc_zone and o.divisionname=m.ofc_division AND m.reading_date_db='{todaydate}' group by o.zone,o.divisionname order by total_billed_today desc;

    #     """
    # query = f"""select o.zone as Zone,o.divisionname as Division,o.subdivision,o.no_of_consumers,(count(distinct m.id)) as total_billed_today from office_consumers o left join readingmaster m on o.zone=m.ofc_zone and o.divisionname=m.ofc_division and o.subdivision=m.ofc_subdivision AND m.reading_date_db='{todaydate}' group by o.zone,o.divisionname,o.subdivision,o.no_of_consumers order by total_billed_today desc;"""
    query = f"""select o.zone as Zone, o.divisionname as Division, o.subdivision, o.no_of_consumers,
(count(distinct case when m.reading_date_db='{todaydate}' then m.id end)) as total_billed_today,
(count(distinct case when extract(month from m.reading_date_db)='{month}'then m.cons_ac_no end)) as total_billed_this_month
from office_consumers o
left join readingmaster m on o.zone=m.ofc_zone and o.divisionname=m.ofc_division
and o.subdivision=m.ofc_subdivision
group by o.zone,o.divisionname,o.subdivision,o.no_of_consumers
order by total_billed_today desc
"""
    cursor.execute(query)
    results = dictfetchall(cursor)
    wb = Workbook()
    ws = wb.active
    ws.title = "JSON Data"
    headers = list(results[0].keys())
    for j, header in enumerate(headers):
        ws.cell(row=1, column=j + 1, value=header)
    for i, row in enumerate(results, start=2):
        for j, key in enumerate(headers):
            ws.cell(row=i, column=j + 1, value=row[key])
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=data.xlsx"
    wb.save(response)
    return response


@api_view(["GET", "POST"])
def mvchecktest(request):
    queryfilters = request.query_params.dict()
    print("queryfilters", queryfilters)
    pagesize = request.query_params.get("pagesize")
    orderby = request.query_params.get("orderby")
    filter = request.query_params.get("filters")
    if filter:
        filters = json.loads(filter)

    paginator = PageNumberPagination()
    paginator.page_size = pagesize
    clause = ""
    try:
        if filters:
            clause += "WHERE "
            for i, (key, value) in enumerate(filters.items()):
                if i > 0:
                    clause += "AND "
                if key == "month":
                    print("key['month']", key)
                    print("value['month']", value)
                    month = value.split("-")[1]
                    key = "extract(month from reading_date_db)"
                    value = month
                clause += f"{key}='{value}'"
    except:
        pass

    cursor = connection.cursor()
    query = f"""select m.mr_id as "mrId",m.rdng_date,m.prsnt_mtr_status,m.prsnt_ocr_rdng,m.prsnt_rdng,m.ocr_pf_reading,m.cons_name,m.prsnt_md_rdng_ocr,m.rdng_ocr_status,m.rdng_img,m.prsnt_md_rdng,m.id,r."mrPhoto"
                    from readingmaster m left outer join meterreaderregistration r on m.mr_id=r."mrId" {clause} order by m.rdng_date {orderby}

        """
    print("query", query)

    cursor.execute(query)
    person_objects = dictfetchall(cursor)
    result_page = paginator.paginate_queryset(person_objects, request)
    return paginator.get_paginated_response(result_page)


@api_view(["GET"])
def dashboarddailydata(request):
    today = date.today()
    agency = request.query_params.get("agency")
    newdict = {}
    clause = ""
    if agency != "null":
        clause = f"and bl_agnc_name='{agency}'"
    cursor = connection.cursor()
    query = f""" select count(distinct mr_id) as activemridtoday,count(rdng_date) as totalbilledtoday from readingmaster where reading_date_db='{today}' {clause}

    """
    cursor.execute(query)
    result = cursor.fetchall()
    for row in result:
        activemridtoday = row[0]
        totalbilledtoday = row[1]

    newdict["activemridtoday"] = activemridtoday
    newdict["totalbilledtoday"] = totalbilledtoday

    return Response(newdict)


@api_view(["POST"])
def dashboarddailydata1(request):
    today = date.today()
    agency = request.query_params.get("agency")
    newdict = {}
    clause = ""
    cursor = connection.cursor()
    filters = request.data.get("filters", None)
    result = {}

    try:
        if filters:
            clause += " AND "
            for i, (key, value) in enumerate(filters.items()):
                if i > 0:
                    clause += "AND "

                if key == "agency":
                    clause += f"bl_agnc_name='{value}'"
                    pass
                if key == "ofc_discom":
                    clause += f"ofc_discom='{value}'"
                    pass
    except:
        pass
    query = f""" select count(distinct mr_id) as activemridtoday,count(rdng_date) as totalbilledtoday from readingmaster where reading_date_db='{today}' {clause}

    """
    cursor.execute(query)
    result = cursor.fetchall()
    for row in result:
        activemridtoday = row[0]
        totalbilledtoday = row[1]

    newdict["activemridtoday"] = activemridtoday
    newdict["totalbilledtoday"] = totalbilledtoday

    return Response(newdict)


@api_view(["GET"])
def dashboarddailydatasbpdcl(request):
    today = date.today()
    newdict = {}
    cursor = connection.cursor()
    agency = request.query_params.get("agency")
    clause = ""
    if agency != "null":
        clause = f"and bl_agnc_name='{agency}'"
    query = f""" select count(distinct mr_id) as activemridtoday,count(rdng_date) as totalbilledtoday from readingmaster where reading_date_db='{today}' and ofc_discom='SBPDCL' {clause}

    """
    cursor.execute(query)
    result = cursor.fetchall()
    for row in result:
        activemridtoday = row[0]
        totalbilledtoday = row[1]

    newdict["activemridtoday"] = activemridtoday
    newdict["totalbilledtoday"] = totalbilledtoday

    return Response(newdict)


@api_view(["GET"])
def dashboarddailydatanbpdcl(request):
    today = date.today()
    agency = request.query_params.get("agency")
    clause = ""
    if agency != "null":
        clause = f"and bl_agnc_name='{agency}'"
    newdict = {}
    cursor = connection.cursor()

    query = f""" select count(distinct mr_id) as activemridtoday,count(rdng_date) as totalbilledtoday from readingmaster where reading_date_db='{today}' and ofc_discom='NBPDCL' {clause}
    """
    cursor.execute(query)
    result = cursor.fetchall()
    for row in result:
        activemridtoday = row[0]
        totalbilledtoday = row[1]

    newdict["activemridtoday"] = activemridtoday
    newdict["totalbilledtoday"] = totalbilledtoday

    return Response(newdict)


@api_view(["GET"])
def performancewisemrs(request):
    new = []
    month = date.today().month

    cursor = connection.cursor()
    query = f"""
SELECT r."mrName",

       SUM(CASE WHEN m.rdng_ocr_status='Passed' THEN 1 ELSE 0 END) as passed,

       SUM(CASE WHEN m.rdng_ocr_status='Failed' THEN 1 ELSE 0 END) as failed,

       count(m.prsnt_mtr_status='Ok' or NULL) AS ok,

       ROUND((cast(count(prsnt_mtr_status='Meter Defective' or null) as float)

/ COALESCE(cast(count(mr_id) as float),1) * 100)::numeric, 2) as Meter_Defective,

       ROUND((cast(count(prsnt_mtr_status='Door Locked' or null) as float)

/ COALESCE(cast(count(mr_id) as float),1) * 100)::numeric, 2) as Door_locked,



        count(m.mr_id) as total,



       CASE

           WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0

           ELSE ROUND((cast(count(rdng_ocr_status='Passed' or null) as float)

           / cast(count(prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)

       END as passed_percent,



       CASE

           WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0

           ELSE ROUND((cast(count(rdng_ocr_status='Failed' or null) as float)

           / cast(count(prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)

       END as failed_percent,



       ROUND((cast(count(prsnt_mtr_status='Ok' or null) as float)

/ COALESCE(cast(count(mr_id) as float),1) * 100)::numeric, 2) as OK_percent,

       r."mrPhoto",

       r."mrPhone",

       r."mrId",



       (DATE_PART('day', CURRENT_DATE) * 30) as "current_day_product"

FROM readingmaster m ,

     meterreaderregistration r

     where m.mr_id = r."mrId"

AND EXTRACT(MONTH FROM m.reading_date_db)='{month}'

GROUP BY r."mrName",

         r."mrPhoto",

         r."mrPhone",

         r."mrId"

HAVING count(m.mr_id)> (DATE_PART('day', CURRENT_DATE) * 30) and ( ROUND((cast(count(prsnt_mtr_status='Door Locked' or null) as float)

/ COALESCE(cast(count(mr_id) as float),1) * 100)::numeric, 2) <=2) and (ROUND((cast(count(prsnt_mtr_status='Meter Defective' or null) as float)

/ COALESCE(cast(count(mr_id) as float),1) * 100)::numeric, 2) <=20)

ORDER BY passed_percent  DESC

"""
    print("query")

    cursor.execute(query)
    result = cursor.fetchall()

    try:
        for row in result:
            meterreader = row[0]
            passed = row[1]
            failed = row[2]
            okreadings = row[3]
            meterdefective = row[4]
            doorlocked = row[5]
            total = row[6]
            passed_percent = row[7]
            failed_percent = row[8]
            ok_percent = row[9]
            mr_photo = row[10]
            mr_phone = row[11]
            mr_id = row[12]

            new_dict = {
                "meterreader": meterreader,
                "passed": passed,
                "failed": failed,
                "OK": okreadings,
                "meterdefectivepercent": meterdefective,
                "doorlockedpercent": doorlocked,
                "total": total,
                "ocrreadingpercent": passed_percent,
                "ocrwithexceppercent": failed_percent,
                "okreadpercent": ok_percent,
                "mrphoto": mr_photo,
                "mrphone": mr_phone,
                "mrId": mr_id,
            }
            new.append(new_dict)

        return Response(new)
    except:
        return Response(status=status.HTTP_404_NOT_FOUND)


@api_view(["GET"])
def performancewisemrssbpdcl(request):
    new = []
    month = date.today().month

    cursor = connection.cursor()
    query = f"""
SELECT r."mrName",
       SUM(CASE WHEN m.rdng_ocr_status='Passed' THEN 1 ELSE 0 END) as passed,
       SUM(CASE WHEN m.rdng_ocr_status='Failed' THEN 1 ELSE 0 END) as failed,
       count(m.prsnt_mtr_status='Ok' or NULL) AS ok,

       ROUND((cast(count(prsnt_mtr_status='Meter Defective' or null) as float)
/ COALESCE(cast(count(mr_id) as float),1) * 100)::numeric, 2) as Meter_Defective,
       ROUND((cast(count(prsnt_mtr_status='Door Locked' or null) as float)
/ COALESCE(cast(count(mr_id) as float),1) * 100)::numeric, 2) as Door_locked,

        count(m.mr_id) as total,

       CASE
           WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
           ELSE ROUND((cast(count(rdng_ocr_status='Passed' or null) as float)
           / cast(count(prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
       END as passed_percent,

       CASE
           WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
           ELSE ROUND((cast(count(rdng_ocr_status='Failed' or null) as float)
           / cast(count(prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
       END as failed_percent,

       ROUND((cast(count(prsnt_mtr_status='Ok' or null) as float)
/ COALESCE(cast(count(mr_id) as float),1) * 100)::numeric, 2) as OK_percent,

       r."mrPhoto",
       r."mrPhone",
       r."mrId",

       (DATE_PART('day', CURRENT_DATE) * 30) as "current_day_product"
FROM readingmaster m ,
     meterreaderregistration r
     where m.mr_id = r."mrId" AND ofc_discom='SBPDCL'
AND EXTRACT(MONTH FROM m.reading_date_db)='{month}'
GROUP BY r."mrName",
         r."mrPhoto",
         r."mrPhone",
         r."mrId"
HAVING count(m.mr_id)> (DATE_PART('day', CURRENT_DATE) * 30) and ( ROUND((cast(count(prsnt_mtr_status='Door Locked' or null) as float)
/ COALESCE(cast(count(mr_id) as float),1) * 100)::numeric, 2) <=2) and (ROUND((cast(count(prsnt_mtr_status='Meter Defective' or null) as float)
/ COALESCE(cast(count(mr_id) as float),1) * 100)::numeric, 2) <=20)
ORDER BY passed_percent  DESC

"""
    print("query")

    cursor.execute(query)
    result = cursor.fetchall()

    try:
        for row in result:
            meterreader = row[0]
            passed = row[1]
            failed = row[2]
            okreadings = row[3]
            meterdefective = row[4]
            doorlocked = row[5]
            total = row[6]
            passed_percent = row[7]
            failed_percent = row[8]
            ok_percent = row[9]
            mr_photo = row[10]
            mr_phone = row[11]
            mr_id = row[12]

            new_dict = {
                "meterreader": meterreader,
                "passed": passed,
                "failed": failed,
                "OK": okreadings,
                "meterdefectivepercent": meterdefective,
                "doorlockedpercent": doorlocked,
                "total": total,
                "ocrreadingpercent": passed_percent,
                "ocrwithexceppercent": failed_percent,
                "okreadpercent": ok_percent,
                "mrphoto": mr_photo,
                "mrphone": mr_phone,
                "mrId": mr_id,
            }
            new.append(new_dict)

        return Response(new)
    except:
        return Response(status=status.HTTP_404_NOT_FOUND)


@api_view(["GET"])
def performancewisemrsnbpdcl(request):
    new = []
    month = date.today().month

    cursor = connection.cursor()
    query = f"""
SELECT r."mrName",
       SUM(CASE WHEN m.rdng_ocr_status='Passed' THEN 1 ELSE 0 END) as passed,
       SUM(CASE WHEN m.rdng_ocr_status='Failed' THEN 1 ELSE 0 END) as failed,
       count(m.prsnt_mtr_status='Ok' or NULL) AS ok,

       ROUND((cast(count(prsnt_mtr_status='Meter Defective' or null) as float)
/ COALESCE(cast(count(mr_id) as float),1) * 100)::numeric, 2) as Meter_Defective,
       ROUND((cast(count(prsnt_mtr_status='Door Locked' or null) as float)
/ COALESCE(cast(count(mr_id) as float),1) * 100)::numeric, 2) as Door_locked,

        count(m.mr_id) as total,

       CASE
           WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
           ELSE ROUND((cast(count(rdng_ocr_status='Passed' or null) as float)
           / cast(count(prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
       END as passed_percent,

       CASE
           WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
           ELSE ROUND((cast(count(rdng_ocr_status='Failed' or null) as float)
           / cast(count(prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
       END as failed_percent,

       ROUND((cast(count(prsnt_mtr_status='Ok' or null) as float)
/ COALESCE(cast(count(mr_id) as float),1) * 100)::numeric, 2) as OK_percent,

       r."mrPhoto",
       r."mrPhone",
       r."mrId",

       (DATE_PART('day', CURRENT_DATE) * 30) as "current_day_product"
FROM readingmaster m ,
     meterreaderregistration r
     where m.mr_id = r."mrId" AND ofc_discom='NBPDCL'
AND EXTRACT(MONTH FROM m.reading_date_db)='{month}'
GROUP BY r."mrName",
         r."mrPhoto",
         r."mrPhone",
         r."mrId"
HAVING count(m.mr_id)> (DATE_PART('day', CURRENT_DATE) * 30) and ( ROUND((cast(count(prsnt_mtr_status='Door Locked' or null) as float)
/ COALESCE(cast(count(mr_id) as float),1) * 100)::numeric, 2) <=2) and (ROUND((cast(count(prsnt_mtr_status='Meter Defective' or null) as float)
/ COALESCE(cast(count(mr_id) as float),1) * 100)::numeric, 2) <=20)
ORDER BY passed_percent  DESC


"""
    print("query")

    cursor.execute(query)
    result = cursor.fetchall()

    try:
        for row in result:
            meterreader = row[0]
            passed = row[1]
            failed = row[2]
            okreadings = row[3]
            meterdefective = row[4]
            doorlocked = row[5]
            total = row[6]
            passed_percent = row[7]
            failed_percent = row[8]
            ok_percent = row[9]
            mr_photo = row[10]
            mr_phone = row[11]
            mr_id = row[12]

            new_dict = {
                "meterreader": meterreader,
                "passed": passed,
                "failed": failed,
                "OK": okreadings,
                "meterdefectivepercent": meterdefective,
                "doorlockedpercent": doorlocked,
                "total": total,
                "ocrreadingpercent": passed_percent,
                "ocrwithexceppercent": failed_percent,
                "okreadpercent": ok_percent,
                "mrphoto": mr_photo,
                "mrphone": mr_phone,
                "mrId": mr_id,
            }
            new.append(new_dict)

        return Response(new)
    except:
        return Response(status=status.HTTP_404_NOT_FOUND)


@api_view(["GET", "POST"])
def mvsummarycards(request):
    queryfilters = request.query_params.dict()
    print("queryfilters", queryfilters)
    pagesize = request.data.get("pagesize")
    page = request.data.get(
        "page",
    )
    offset = (int(pagesize) * int(page)) - int(pagesize)
    orderby = request.data.get("orderby")
    filters = request.data.get("filters")
    print("filters", filters)
    clause = ""
    try:
        if filters:
            clause += "WHERE "
            for i, (key, value) in enumerate(filters.items()):
                if i > 0:
                    clause += "AND "
                if key == "month":
                    print("key['month']", key)
                    print("value['month']", value)
                    month = value.split("-")[1]
                    key = "extract(month from reading_date_db)"
                    value = month
                clause += f"{key}='{value}'"

    except:
        pass
    cursor = connection.cursor()
    # query = (f'''select m.mr_id as "mrId",m.rdng_date,m.prsnt_mtr_status,m.prsnt_ocr_rdng,m.prsnt_rdng,m.ocr_pf_reading,m.cons_name,m.prsnt_md_rdng_ocr,m.rdng_ocr_status,m.rdng_img,m.prsnt_md_rdng,m.id,r."mrPhoto"
    #         from readingmaster m left outer join meterreaderregistration r on m.mr_id=r."mrId" {clause} order by m.rdng_date

    # ''')
    query = f"""select m.mr_id as "mrId",m.rdng_date,m.prsnt_mtr_status,m.prsnt_ocr_rdng,m.prsnt_rdng,m.ocr_pf_reading,m.cons_name,m.cons_ac_no,m.prsnt_md_rdng_ocr,m.rdng_ocr_status,m.rdng_img,m.prsnt_md_rdng,m.id,r."mrPhoto",m.prsnt_rdng_ocr_excep,m.reading_parameter_type, count(*)  over () as total_count
                    from readingmaster m left outer join meterreaderregistration r on m.mr_id=r."mrId" {clause} order by m.rdng_date DESC limit {pagesize} offset {offset}

        """
    print("query", query)
    cursor.execute(query)
    person_objects = dictfetchall(cursor)
    return Response(
        {"count": person_objects[0]["total_count"], "results": person_objects}
    )


# @api_view(['GET'])
# def locationwisehyperlink(request):
#     month = date.today().month
#     newdict = {}
#     new = []

#     def listfun(dict):
#         print(dict)
#         new.append(dict.copy())
#         return new
#     cursor = connection.cursor()
#     cursor.execute(f'''
#    select ofc_zone as location, count(distinct mr_id),count(prsnt_mtr_status='Ok' or NULL),count(rdng_ocr_status='Passed' or NULL),count(rdng_ocr_status='Failed' or NULL),count(prsnt_mtr_status='Meter Defective' or NULL),count(prsnt_mtr_status='Door Locked' or NULL),count(mr_id)
#     from readingmaster where extract(month from reading_date_db)='{month}'  GROUP BY ofc_zone
#     ''')
#     result = cursor.fetchall()
#     print("result", result)
#     try:
#         for row in result:
#             locationname = row[0]
#             total = row[7]
#             mrid = row[1]
#             okreadings = row[2]
#             OcrReadings = row[3]
#             Ocrwithexception = row[4]
#             meterDefective = row[5]
#             doorLocked = row[6]
#             okreadpercent = math.floor((okreadings/total)*100)
#             ocrreadingpercent = math.floor(
#                 ((OcrReadings/okreadings) if okreadings else 0)*100)
#             ocrwithexceppercent = math.floor(
#                 ((Ocrwithexception/okreadings) if okreadings else 0)*100)
#             meterdefectivepercent = math.floor((meterDefective/total)*100)
#             doorlockedpercent = math.floor((doorLocked/total)*100)
#             newdict['locationname'] = row[0]
#             newdict['mrid'] = row[1]
#             newdict['okreadings'] = row[2]
#             newdict['okreadingspercent'] = okreadpercent
#             newdict['OcrReadings'] = row[3]
#             newdict['OcrReadingspercent'] = ocrreadingpercent
#             newdict['Ocrwithexception'] = row[4]
#             newdict['Ocrwithexceptionpercent'] = ocrwithexceppercent
#             newdict['meterDefective'] = row[5]
#             newdict['meterDefectivepercent'] = meterdefectivepercent
#             newdict['doorLocked'] = row[6]
#             newdict['doorLockedpercent'] = doorlockedpercent
#             newdict['total'] = total
#             data = listfun(newdict)
#         return Response(data)
#     except:
#         return Response([])


# @api_view(['POST'])
# def locationwisehyperlinkclick(request):
#     month = date.today().month
#     newdict = {}
#     new = []

#     def listfun(dict):
#         print(dict)
#         new.append(dict.copy())
#         return new
#     cursor = connection.cursor()
#     locationwise = request.data.get('locationwise', None)
#     locationname = request.data.get('locationname', None)
#     groupby = request.data.get('groupby')
#     clause = ''
#     clause = "WHERE "+locationwise+"='"+locationname+"' "
#     if groupby == "ofc_section":

#         cursor.execute(f'''
#     select o.sectionname as location, count(distinct mr_id),count(prsnt_mtr_status='Ok' or NULL),count(rdng_ocr_status='Passed' or NULL),count(rdng_ocr_status='Failed' or NULL),count(prsnt_mtr_status='Meter Defective' or NULL),count(prsnt_mtr_status='Door Locked' or NULL),count(mr_id) as total,r.ofc_section
#     from readingmaster r left outer join office o on r.ofc_section=o.sectioncode {clause} and extract(month from reading_date_db)='{month}'  GROUP BY o.sectionname,r.ofc_section
#     ''')

#     else:
#         cursor.execute(f'''
#    select {groupby} as location, count(distinct mr_id),count(prsnt_mtr_status='Ok' or NULL),count(rdng_ocr_status='Passed' or NULL),count(rdng_ocr_status='Failed' or NULL),count(prsnt_mtr_status='Meter Defective' or NULL),count(prsnt_mtr_status='Door Locked' or NULL),count(mr_id)
#     from readingmaster {clause} and extract(month from reading_date_db)='{month}' GROUP BY {groupby}
#     ''')

#     result = cursor.fetchall()
#     print("result", result)
#     try:
#         for row in result:
#             locationname = row[0]
#             total = row[7]

#             mrid = row[1]
#             okreadings = row[2]
#             OcrReadings = row[3]
#             Ocrwithexception = row[4]
#             meterDefective = row[5]
#             doorLocked = row[6]
#             okreadpercent = math.floor((okreadings/total)*100)
#             ocrreadingpercent = math.floor(
#                 ((OcrReadings/okreadings) if okreadings else 0)*100)
#             ocrwithexceppercent = math.floor(
#                 ((Ocrwithexception/okreadings) if okreadings else 0)*100)
#             meterdefectivepercent = math.floor((meterDefective/total)*100)
#             doorlockedpercent = math.floor((doorLocked/total)*100)
#             newdict['locationname'] = row[0]
#             if groupby == "ofc_section":
#                 newdict['locationnumber'] = row[8]
#             newdict['mrid'] = row[1]
#             newdict['okreadings'] = row[2]
#             newdict['okreadingspercent'] = okreadpercent
#             newdict['OcrReadings'] = row[3]
#             newdict['OcrReadingspercent'] = ocrreadingpercent
#             newdict['Ocrwithexception'] = row[4]
#             newdict['Ocrwithexceptionpercent'] = ocrwithexceppercent
#             newdict['meterDefective'] = row[5]
#             newdict['meterDefectivepercent'] = meterdefectivepercent
#             newdict['doorLocked'] = row[6]
#             newdict['doorLockedpercent'] = doorlockedpercent
#             newdict['total'] = total
#             data = listfun(newdict)
#         return Response(data)
#     except:
#         return Response([])
# @api_view(['GET'])
# def locationwisehyperlink(request):

#     today = date.today()
#     thismonth = today.strftime('%Y-%m')
#     year = thismonth.split('-')[0]
#     month = thismonth.split('-')[1]

#     agency = request.query_params.get("agency")
#     clause = f"where extract(month from reading_date_db) = '{month}' AND extract(year from reading_date_db) = '{year}' and ofc_zone <>''"
#     if agency != 'null':
#         clause = f"where bl_agnc_name='{agency}'and  extract(month from reading_date_db) = '{month}' AND extract(year from reading_date_db) = '{year}' and ofc_zone <>''"
#         print("clause---------->", clause)
#     newdict = {}
#     new = []

#     def listfun(dict):
#         print(dict)
#         new.append(dict.copy())
#         return new
#     cursor = connection.cursor()
#     cursor.execute(f'''
#    select ofc_zone as location, count(distinct mr_id),count(prsnt_mtr_status='Ok' or NULL),count(rdng_ocr_status='Passed' or NULL),count(rdng_ocr_status='Failed' or NULL),count(prsnt_mtr_status='Meter Defective' or NULL),count(prsnt_mtr_status='Door Locked' or NULL),count(mr_id)
#     from readingmaster {clause}   GROUP BY ofc_zone
#     ''')
#     result = cursor.fetchall()
#     print("result", result)
#     try:
#         for row in result:
#             locationname = row[0]
#             total = row[7]
#             mrid = row[1]
#             okreadings = row[2]
#             OcrReadings = row[3]
#             Ocrwithexception = row[4]
#             meterDefective = row[5]
#             doorLocked = row[6]
#             okreadpercent = math.floor((okreadings/total)*100)
#             ocrreadingpercent = math.floor(
#                 ((OcrReadings/okreadings) if okreadings else 0)*100)
#             ocrwithexceppercent = math.floor(
#                 ((Ocrwithexception/okreadings) if okreadings else 0)*100)
#             meterdefectivepercent = math.floor((meterDefective/total)*100)
#             doorlockedpercent = math.floor((doorLocked/total)*100)
#             newdict['locationname'] = row[0]
#             newdict['mrid'] = row[1]
#             newdict['okreadings'] = row[2]
#             newdict['okreadingspercent'] = okreadpercent
#             newdict['OcrReadings'] = row[3]
#             newdict['OcrReadingspercent'] = ocrreadingpercent
#             newdict['Ocrwithexception'] = row[4]
#             newdict['Ocrwithexceptionpercent'] = ocrwithexceppercent
#             newdict['meterDefective'] = row[5]
#             newdict['meterDefectivepercent'] = meterdefectivepercent
#             newdict['doorLocked'] = row[6]
#             newdict['doorLockedpercent'] = doorlockedpercent
#             newdict['total'] = total
#             data = listfun(newdict)
#         return Response(data)
#     except:
#         return Response([])


# @api_view(['POST'])
# def locationwisehyperlinkclick(request):
#     today = date.today()
#     thismonth = today.strftime('%Y-%m')
#     year = thismonth.split('-')[0]
#     month = thismonth.split('-')[1]
#     newdict = {}
#     new = []

#     def listfun(dict):
#         new.append(dict.copy())
#         return new
#     cursor = connection.cursor()
#     locationwise = request.data.get('locationwise', None)
#     locationname = request.data.get('locationname', None)
#     groupby = request.data.get('groupby')
#     agency = request.data.get('agency')
#     print("agency", agency)
#     clause = ''
#     clause += "WHERE "+locationwise+"='"+locationname+"' "
#     if agency != None:
#         clause += f"AND bl_agnc_name='{agency}'and  extract(month from reading_date_db) = '{month}' AND extract(year from reading_date_db) = '{year}'"
#         pass
#     if groupby == "ofc_section":

#         query = (f'''
#     select o.sectionname as location, count(distinct mr_id),count(prsnt_mtr_status='Ok' or NULL),count(rdng_ocr_status='Passed' or NULL),count(rdng_ocr_status='Failed' or NULL),count(prsnt_mtr_status='Meter Defective' or NULL),count(prsnt_mtr_status='Door Locked' or NULL),count(mr_id) as total,r.ofc_section
#     from readingmaster r left outer join office o on r.ofc_section=o.sectioncode {clause} and  extract(month from reading_date_db) = '{month}' AND extract(year from reading_date_db) = '{year}'  GROUP BY o.sectionname,r.ofc_section
#     ''')

#     else:
#         query = (f'''
#    select {groupby} as location, count(distinct mr_id),count(prsnt_mtr_status='Ok' or NULL),count(rdng_ocr_status='Passed' or NULL),count(rdng_ocr_status='Failed' or NULL),count(prsnt_mtr_status='Meter Defective' or NULL),count(prsnt_mtr_status='Door Locked' or NULL),count(mr_id)
#     from readingmaster {clause} and  extract(month from reading_date_db) = '{month}' AND extract(year from reading_date_db) = '{year}' GROUP BY {groupby}
#     ''')
#     print("QUERY", query)

#     cursor.execute(query)
#     result = cursor.fetchall()
#     # print("result", result)
#     try:
#         for row in result:
#             locationname = row[0]
#             total = row[7]

#             mrid = row[1]
#             okreadings = row[2]
#             OcrReadings = row[3]
#             Ocrwithexception = row[4]
#             meterDefective = row[5]
#             doorLocked = row[6]
#             okreadpercent = math.floor((okreadings/total)*100)
#             ocrreadingpercent = math.floor(
#                 ((OcrReadings/okreadings) if okreadings else 0)*100)
#             ocrwithexceppercent = math.floor(
#                 ((Ocrwithexception/okreadings) if okreadings else 0)*100)
#             meterdefectivepercent = math.floor((meterDefective/total)*100)
#             doorlockedpercent = math.floor((doorLocked/total)*100)
#             newdict['locationname'] = row[0]
#             if groupby == "ofc_section":
#                 newdict['locationnumber'] = row[8]
#             newdict['mrid'] = row[1]
#             newdict['okreadings'] = row[2]
#             newdict['okreadingspercent'] = okreadpercent
#             newdict['OcrReadings'] = row[3]
#             newdict['OcrReadingspercent'] = ocrreadingpercent
#             newdict['Ocrwithexception'] = row[4]
#             newdict['Ocrwithexceptionpercent'] = ocrwithexceppercent
#             newdict['meterDefective'] = row[5]
#             newdict['meterDefectivepercent'] = meterdefectivepercent
#             newdict['doorLocked'] = row[6]
#             newdict['doorLockedpercent'] = doorlockedpercent
#             newdict['total'] = total
#             data = listfun(newdict)
#         return Response(data)
#     except:
#         return Response([])


@api_view(["GET"])
def locationwisehyperlink(request):
    today = date.today()
    thismonth = today.strftime("%Y-%m")
    year = thismonth.split("-")[0]
    month = thismonth.split("-")[1]

    agency = request.query_params.get("agency")
    ofc_discom = request.query_params.get("ofc_discom")
    user = request.query_params.get("user")

    clause = f"where extract(month from reading_date_db) = '{month}' AND extract(year from reading_date_db) = '{year}'"
    if user == "false":
        if agency != "null" and ofc_discom != "null":
            clause = f"where bl_agnc_name='{agency}'and ofc_discom='{ofc_discom}' and extract(month from reading_date_db) = '{month}' AND extract(year from reading_date_db) = '{year}'"

        if ofc_discom != "null" and agency == "null":
            clause = f"where ofc_discom='{ofc_discom}'and  extract(month from reading_date_db) = '{month}' AND extract(year from reading_date_db) = '{year}'"
    print("clause---------->", clause)
    newdict = {}
    new = []

    def listfun(dict):
        print(dict)
        new.append(dict.copy())
        return new

    cursor = connection.cursor()
    cursor.execute(
        f"""
   select ofc_zone as location, count(distinct mr_id),count(prsnt_mtr_status='Ok' or NULL),count(rdng_ocr_status='Passed' or NULL),count(rdng_ocr_status='Failed' or NULL),count(prsnt_mtr_status='Meter Defective' or NULL),count(prsnt_mtr_status='Door Locked' or NULL),count(mr_id)
    from readingmaster {clause} and ofc_zone!='' GROUP BY ofc_zone
    """
    )
    result = cursor.fetchall()
    print("result", result)
    try:
        for row in result:
            locationname = row[0]
            total = row[7]
            mrid = row[1]
            okreadings = row[2]
            OcrReadings = row[3]
            Ocrwithexception = row[4]
            meterDefective = row[5]
            doorLocked = row[6]
            okreadpercent = math.floor((okreadings / total) * 100)
            ocrreadingpercent = math.floor(
                ((OcrReadings / okreadings) if okreadings else 0) * 100
            )
            ocrwithexceppercent = math.floor(
                ((Ocrwithexception / okreadings) if okreadings else 0) * 100
            )
            meterdefectivepercent = math.floor((meterDefective / total) * 100)
            doorlockedpercent = math.floor((doorLocked / total) * 100)
            newdict["locationname"] = row[0]
            newdict["mrid"] = row[1]
            newdict["okreadings"] = row[2]
            newdict["okreadingspercent"] = okreadpercent
            newdict["OcrReadings"] = row[3]
            newdict["OcrReadingspercent"] = ocrreadingpercent
            newdict["Ocrwithexception"] = row[4]
            newdict["Ocrwithexceptionpercent"] = ocrwithexceppercent
            newdict["meterDefective"] = row[5]
            newdict["meterDefectivepercent"] = meterdefectivepercent
            newdict["doorLocked"] = row[6]
            newdict["doorLockedpercent"] = doorlockedpercent
            newdict["total"] = total
            data = listfun(newdict)
        return Response(data)
    except:
        return Response([])


@api_view(["POST"])
def locationwisehyperlinkclick(request):
    today = date.today()
    thismonth = today.strftime("%Y-%m")
    year = thismonth.split("-")[0]
    month = thismonth.split("-")[1]
    newdict = {}
    new = []

    def listfun(dict):
        new.append(dict.copy())
        return new

    cursor = connection.cursor()
    locationwise = request.data.get("locationwise", None)
    locationname = request.data.get("locationname", None)
    groupby = request.data.get("groupby")
    agency = request.data.get("agency")
    ofc_discom = request.data.get("ofc_discom")
    user = request.data.get("user")
    print("agency", agency)
    clause = ""
    clause += f"WHERE {locationwise}='{locationname}' and  extract(month from reading_date_db) = '{month}' AND extract(year from reading_date_db) = '{year}'"

    # if agency != None:
    # clause += f"AND bl_agnc_name='{agency}'and  extract(month from reading_date_db) = '{month}' AND extract(year from reading_date_db) = '{year}'"
    #     pass
    if user == False:
        if agency != None and ofc_discom != None:
            clause = f"where {locationwise}='{locationname}' and bl_agnc_name='{agency}'and ofc_discom='{ofc_discom}' and extract(month from reading_date_db) = '{month}' AND extract(year from reading_date_db) = '{year}'"

        if ofc_discom != None and agency == None:
            clause = f"where {locationwise}='{locationname}'and ofc_discom='{ofc_discom}'and  extract(month from reading_date_db) = '{month}' AND extract(year from reading_date_db) = '{year}'"

    print("caluse----------->", clause)
    if groupby == "ofc_section":
        query = f"""
    select o.sectionname as location, count(distinct mr_id),count(prsnt_mtr_status='Ok' or NULL),count(rdng_ocr_status='Passed' or NULL),count(rdng_ocr_status='Failed' or NULL),count(prsnt_mtr_status='Meter Defective' or NULL),count(prsnt_mtr_status='Door Locked' or NULL),count(mr_id) as total,r.ofc_section
    from readingmaster r left outer join office o on r.ofc_section=o.sectioncode {clause} and o.sectionname!=''   GROUP BY o.sectionname,r.ofc_section
    """

    else:
        query = f"""
   select {groupby} as location, count(distinct mr_id),count(prsnt_mtr_status='Ok' or NULL),count(rdng_ocr_status='Passed' or NULL),count(rdng_ocr_status='Failed' or NULL),count(prsnt_mtr_status='Meter Defective' or NULL),count(prsnt_mtr_status='Door Locked' or NULL),count(mr_id)
    from readingmaster {clause} and {groupby}!=''  GROUP BY {groupby}
    """
    print("QUERY", query)

    cursor.execute(query)
    result = cursor.fetchall()
    # print("result", result)
    try:
        for row in result:
            locationname = row[0]
            total = row[7]

            mrid = row[1]
            okreadings = row[2]
            OcrReadings = row[3]
            Ocrwithexception = row[4]
            meterDefective = row[5]
            doorLocked = row[6]
            okreadpercent = math.floor((okreadings / total) * 100)
            ocrreadingpercent = math.floor(
                ((OcrReadings / okreadings) if okreadings else 0) * 100
            )
            ocrwithexceppercent = math.floor(
                ((Ocrwithexception / okreadings) if okreadings else 0) * 100
            )
            meterdefectivepercent = math.floor((meterDefective / total) * 100)
            doorlockedpercent = math.floor((doorLocked / total) * 100)
            newdict["locationname"] = row[0]
            if groupby == "ofc_section":
                newdict["locationnumber"] = row[8]
            newdict["mrid"] = row[1]
            newdict["okreadings"] = row[2]
            newdict["okreadingspercent"] = okreadpercent
            newdict["OcrReadings"] = row[3]
            newdict["OcrReadingspercent"] = ocrreadingpercent
            newdict["Ocrwithexception"] = row[4]
            newdict["Ocrwithexceptionpercent"] = ocrwithexceppercent
            newdict["meterDefective"] = row[5]
            newdict["meterDefectivepercent"] = meterdefectivepercent
            newdict["doorLocked"] = row[6]
            newdict["doorLockedpercent"] = doorlockedpercent
            newdict["total"] = total
            data = listfun(newdict)
        return Response(data)
    except:
        return Response([])


# @api_view(['POST'])
# def meterWiseReport(request):

#     discom = request.data.get("ofc_discom", None)
#     orderby = request.data.get("orderby", None)
#     month = request.data.get("month", None)
#     agency = request.data.get("bl_agnc_name", None)
#     print("hj", orderby, discom)
#     data = request.data.get("filters", None)

#     try:
#         if discom and orderby:
#             clause = 'WHERE '
#             clause += f"extract(month from reading_date_db) = '{month.split('-')[1]}' AND extract(year from reading_date_db) = '{month.split('-')[0]} ' " if (
#                 month) else f"extract(month from reading_date_db) = '{date.today().month}' AND extract(year from reading_date_db) = '{date.today().year}' "
#             clause += f" AND bl_agnc_name='{agency}'" if (agency) else ''
#             clause += f" AND ofc_discom ='{discom}'" if (discom) else ''
#             cursor = connection.cursor()
#             query = (f'''
#             SELECT m.{orderby} as location, count(m.id) as id,
# count(m.prsnt_mtr_status='Ok' or null) as Ok,
# CASE WHEN count(m.prsnt_mtr_status='Ok' or null) = 0 THEN 0
#  ELSE ROUND((cast(count(m.prsnt_mtr_status='Ok' or null) as float)
# 			 / cast(count(m.id) as float) * 100)::numeric, 2) END as ok_persent,
# count(m.rdng_ocr_status='Passed' or null) as ocr_Passed,
# CASE WHEN count(m.rdng_ocr_status='Passed' or null) = 0 THEN 0
#  ELSE ROUND((cast(count(m.rdng_ocr_status='Passed' or null) as float)
# 			 / cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2) END as ocr_Passed_persent,
# count(m.prsnt_rdng_ocr_excep='Parameters Incorrect' or null) as Parameters_Incorrect,
# CASE WHEN count(m.prsnt_rdng_ocr_excep='Parameters Incorrect' or null) = 0 THEN 0
#  ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Parameters Incorrect' or null) as float)
# 			 / cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2) END as Parameters_Incorrect_Persent,
# count(m.prsnt_rdng_ocr_excep='Parameters Unclear' or null) as Parameters_Unclear,
# CASE WHEN count(m.prsnt_rdng_ocr_excep='Parameters Unclear' or null) = 0 THEN 0
#  ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Parameters Unclear' or null) as float)
# 			 / cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2) END as Parameters_Unclear_Persent,
# count(m.prsnt_rdng_ocr_excep='Parameters Unavailable' or null) as Parameters_Unavailable,
# CASE WHEN count(m.prsnt_rdng_ocr_excep='Parameters Unavailable' or null) = 0 THEN 0
#  ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Parameters Unavailable' or null) as float)
# 			 / cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2) END as Parameters_Unavailable_Persent,
# count(m.prsnt_rdng_ocr_excep='Image Invalid' or null) as Image_Invalid,
# CASE WHEN count(m.prsnt_rdng_ocr_excep='Image Invalid' or null) = 0 THEN 0
#  ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Image Invalid' or null) as float)
# 			 / cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2) END as Image_Invalid_Persent,
# count(m.prsnt_rdng_ocr_excep='Image Unclear' or null) as Image_Unclear,
# CASE WHEN count(m.prsnt_rdng_ocr_excep='Image Unclear' or null) = 0 THEN 0
#  ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Image Unclear' or null) as float)
# 			 / cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2) END as Image_Unclear_Persent,
# count(m.prsnt_rdng_ocr_excep='Image Spoofed' or null) as Image_Spoofed,
# CASE WHEN count(m.prsnt_rdng_ocr_excep='Image Spoofed' or null) = 0 THEN 0
#  ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Image Spoofed' or null) as float)
# 			 / cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2) END as Image_Spoofed_Persent,
# count(m.prsnt_rdng_ocr_excep='Image Stain on Decimal' or null) as Image_Stain_OnDecimal,
# CASE WHEN count(m.prsnt_rdng_ocr_excep='Image Stain on Decimal' or null) = 0 THEN 0
#  ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Image Stain on Decimal' or null) as float)
# 			 / cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2) END as Image_Stain_On_Decimal_persent,
# count(m.prsnt_rdng_ocr_excep='Meter Mismatched' or null) as Meter_Mismatched,
# CASE WHEN count(m.prsnt_rdng_ocr_excep='Meter Mismatched' or null) = 0 THEN 0
#  ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Meter Mismatched' or null) as float)
# 			 / cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2) END as Meter_Mismatched_persent,
# count(m.prsnt_rdng_ocr_excep='Meter On Height' or null) as Meter_On_Height,
# CASE WHEN count(m.prsnt_rdng_ocr_excep='Meter On Height' or null) = 0 THEN 0
#  ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Meter On Height' or null) as float)
# 			 / cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2) END as Meter_On_Height_Persent,
# count(m.prsnt_rdng_ocr_excep='Meter Dirty' or null) as Meter_Dirty,
# CASE WHEN count(m.prsnt_rdng_ocr_excep='Meter Dirty' or null) = 0 THEN 0
#  ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Meter Dirty' or null) as float)
# 			 / cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2) END as Meter_Dirty_Persent,
# count(m.prsnt_rdng_ocr_excep='Meter Display Broken' or null) as Meter_Display_Broken,
# CASE WHEN count(m.prsnt_rdng_ocr_excep='Meter Display Broken' or null) = 0 THEN 0
#  ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Meter Display Broken' or null) as float)
# 			 / cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2) END as Meter_Display_Broken_Persent,
# count(m.prsnt_rdng_ocr_excep='Daylight Reflection On Meter' or null) as Daylight_Reflection_On_Meter,
# CASE WHEN count(m.prsnt_rdng_ocr_excep='Daylight Reflection On Meter' or null) = 0 THEN 0
#  ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Daylight Reflection On Meter' or null) as float)
# 			 / cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2) END as Daylight_Reflection_On_Meter_Persent,
# count(m.prsnt_rdng_ocr_excep='Backlight Reflection' or null) as Backlight_Reflection,
# CASE WHEN count(m.prsnt_rdng_ocr_excep='Backlight Reflection' or null) = 0 THEN 0
#  ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Backlight Reflection' or null) as float)
# 			 / cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2) END as Backlight_Reflection_Persent
# from readingmaster m {clause}  group by m.{orderby}
#                 ''')
#             print(clause)
#             print(query)
#             cursor.execute(query)
#             person_objects = dictfetchall(cursor)
#         return Response(person_objects)
#     except:
#         return Response([])


@api_view(["POST"])
def meterWiseReportUpdate(request):
    data = request.data.get("filters", None)
    groupby = request.data.get("groupby", None)
    print(request.data)

    if data is None:
        return Response([])

    else:
        today = date.today()
        clause = ""
        clause += "WHERE "
        if (
            data.get("month", "") == ""
            and data.get("startdate", "") == ""
            and data.get("enddate", "") == ""
        ):
            clause += f"extract(month from m.reading_date_db) = '{date.today().month}' AND extract(year from m.reading_date_db) = '{date.today().year}' AND "
        for i, (key, value) in enumerate(data.items()):
            if key == "month" and value:
                year = value.split("-")[0]
                month = value.split("-")[1]
                clause += f"extract(month from m.reading_date_db) = '{month}' AND extract(year from m.reading_date_db) = '{year}' AND "

            elif key == "startdate" and value:
                clause += f"extract(day from m.reading_date_db) BETWEEN '{data['startdate']}' AND "

            elif key == "enddate" and value:
                clause += f"'{data['enddate']}' AND "

            elif (
                key
                in [
                    "ofc_discom",
                    "ofc_zone",
                    "ofc_circle",
                    "ofc_division",
                    "ofc_subdivision",
                    "bl_agnc_name",
                ]
                and value
            ):
                clause += f"m.{key} ='{data[key]}' AND "

        if clause[-4:-1] == "AND":
            clause = clause[0:-4]
    print(clause)

    try:
        cursor = connection.cursor()
        if groupby != "ofc_section":
            query = f"""
                    SELECT m.{groupby} as location, count(m.id) as id,
            count(m.prsnt_mtr_status='Ok' or null) as Ok,
            CASE WHEN count(m.prsnt_mtr_status='Ok' or null) = 0 THEN 0
            ELSE ROUND((cast(count(m.prsnt_mtr_status='Ok' or null) as float)
                        / cast(count(m.id) as float) * 100)::numeric, 2) END as ok_persent,
            count(m.rdng_ocr_status='Passed' or null) as ocr_Passed,
            CASE WHEN count(m.rdng_ocr_status='Passed' or null) = 0 THEN 0
            ELSE ROUND((cast(count(m.rdng_ocr_status='Passed' or null) as float)
                        / cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2) END as ocr_Passed_persent,
            count(m.prsnt_rdng_ocr_excep='Parameters Incorrect' or null) as Parameters_Incorrect,
            CASE WHEN count(m.prsnt_rdng_ocr_excep='Parameters Incorrect' or null) = 0 THEN 0
            ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Parameters Incorrect' or null) as float)
                        / cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2) END as Parameters_Incorrect_Persent,
            count(m.prsnt_rdng_ocr_excep='Parameters Unclear' or null) as Parameters_Unclear,
            CASE WHEN count(m.prsnt_rdng_ocr_excep='Parameters Unclear' or null) = 0 THEN 0
            ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Parameters Unclear' or null) as float)
                        / cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2) END as Parameters_Unclear_Persent,
            count(m.prsnt_rdng_ocr_excep='Parameters Unavailable' or null) as Parameters_Unavailable,
            CASE WHEN count(m.prsnt_rdng_ocr_excep='Parameters Unavailable' or null) = 0 THEN 0
            ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Parameters Unavailable' or null) as float)
                        / cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2) END as Parameters_Unavailable_Persent,
            count(m.prsnt_rdng_ocr_excep='Image Invalid' or null) as Image_Invalid,
            CASE WHEN count(m.prsnt_rdng_ocr_excep='Image Invalid' or null) = 0 THEN 0
            ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Image Invalid' or null) as float)
                        / cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2) END as Image_Invalid_Persent,
            count(m.prsnt_rdng_ocr_excep='Image Unclear' or null) as Image_Unclear,
            CASE WHEN count(m.prsnt_rdng_ocr_excep='Image Unclear' or null) = 0 THEN 0
            ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Image Unclear' or null) as float)
                        / cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2) END as Image_Unclear_Persent,
            count(m.prsnt_rdng_ocr_excep='Image Spoofed' or null) as Image_Spoofed,
            CASE WHEN count(m.prsnt_rdng_ocr_excep='Image Spoofed' or null) = 0 THEN 0
            ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Image Spoofed' or null) as float)
                        / cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2) END as Image_Spoofed_Persent,
            count(m.prsnt_rdng_ocr_excep='Image Stain on Decimal' or null) as Image_Stain_OnDecimal,
            CASE WHEN count(m.prsnt_rdng_ocr_excep='Image Stain on Decimal' or null) = 0 THEN 0
            ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Image Stain on Decimal' or null) as float)
                        / cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2) END as Image_Stain_On_Decimal_persent,
            count(m.prsnt_rdng_ocr_excep='Meter Mismatched' or null) as Meter_Mismatched,
            CASE WHEN count(m.prsnt_rdng_ocr_excep='Meter Mismatched' or null) = 0 THEN 0
            ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Meter Mismatched' or null) as float)
                        / cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2) END as Meter_Mismatched_persent,
            count(m.prsnt_rdng_ocr_excep='Meter On Height' or null) as Meter_On_Height,
            CASE WHEN count(m.prsnt_rdng_ocr_excep='Meter On Height' or null) = 0 THEN 0
            ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Meter On Height' or null) as float)
                        / cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2) END as Meter_On_Height_Persent,
            count(m.prsnt_rdng_ocr_excep='Meter Dirty' or null) as Meter_Dirty,
            CASE WHEN count(m.prsnt_rdng_ocr_excep='Meter Dirty' or null) = 0 THEN 0
            ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Meter Dirty' or null) as float)
                        / cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2) END as Meter_Dirty_Persent,
            count(m.prsnt_rdng_ocr_excep='Meter Display Broken' or null) as Meter_Display_Broken,
            CASE WHEN count(m.prsnt_rdng_ocr_excep='Meter Display Broken' or null) = 0 THEN 0
            ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Meter Display Broken' or null) as float)
                        / cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2) END as Meter_Display_Broken_Persent,
            count(m.prsnt_rdng_ocr_excep='Daylight Reflection On Meter' or null) as Daylight_Reflection_On_Meter,
            CASE WHEN count(m.prsnt_rdng_ocr_excep='Daylight Reflection On Meter' or null) = 0 THEN 0
            ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Daylight Reflection On Meter' or null) as float)
                        / cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2) END as Daylight_Reflection_On_Meter_Persent,
            count(m.prsnt_rdng_ocr_excep='Backlight Reflection' or null) as Backlight_Reflection,
            CASE WHEN count(m.prsnt_rdng_ocr_excep='Backlight Reflection' or null) = 0 THEN 0
            ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Backlight Reflection' or null) as float)
                        / cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2) END as Backlight_Reflection_Persent
            from readingmaster m {clause} and m.{groupby}!='' group by m.{groupby}
                """
        else:
            query = f"""
                select
                    o.sectionname as location,
                    count(m.id) as id,
                    count(m.prsnt_mtr_status='Ok' or null) as Ok,
                    count(m.rdng_ocr_status='Passed' or null) as ocr_Passed,
                    count(m.prsnt_rdng_ocr_excep='Parameters Incorrect' or null) as Parameters_Incorrect,
                    count(m.prsnt_rdng_ocr_excep='Parameters Unclear' or null) as Parameters_Unclear,
                    count(m.prsnt_rdng_ocr_excep='Parameters Unavailable' or null) as Parameters_Unavailable,
                    count(m.prsnt_rdng_ocr_excep='Image Invalid' or null) as Image_Invalid,
                    count(m.prsnt_rdng_ocr_excep='Image Unclear' or null) as Image_Unclear,
                    count(m.prsnt_rdng_ocr_excep='Image Spoofed' or null) as Image_Spoofed,
                    count(m.prsnt_rdng_ocr_excep='Image Stain on Decimal' or null) as Image_Stain_OnDecimal,
                    count(m.prsnt_rdng_ocr_excep='Meter Mismatched' or null) as Meter_Mismatched,
                    count(m.prsnt_rdng_ocr_excep='Meter On Height' or null) as Meter_On_Height,
                    count(m.prsnt_rdng_ocr_excep='Meter Dirty' or null) as Meter_Dirty,
                    count(m.prsnt_rdng_ocr_excep='Meter Display Broken' or null) as Meter_Display_Broken,
                    count(m.prsnt_rdng_ocr_excep='Daylight Reflection On Meter' or null) as Daylight_Reflection_On_Meter,
                    count(m.prsnt_rdng_ocr_excep='Backlight Reflection' or null) as Backlight_Reflection,
                    CASE WHEN
                        count(m.prsnt_mtr_status='Ok' or null) = 0 THEN 0 ELSE ROUND((cast(count(m.prsnt_mtr_status='Ok' or null) as float)/cast(count(m.id) as float) * 100)::numeric, 2)
                    END as ok_persent,

                    CASE WHEN
                        count(m.rdng_ocr_status='Passed' or null) = 0 THEN 0 ELSE ROUND((cast(count(m.rdng_ocr_status='Passed' or null) as float)/cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
                    END as ocr_Passed_persent,

                    CASE WHEN
                        count(m.prsnt_rdng_ocr_excep='Parameters Incorrect' or null) = 0 THEN 0 ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Parameters Incorrect' or null) as float)/cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
                    END as Parameters_Incorrect_Persent,
                    CASE WHEN
                        count(m.prsnt_rdng_ocr_excep='Parameters Unclear' or null) = 0 THEN 0 ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Parameters Unclear' or null) as float)/cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
                    END as Parameters_Unclear_Persent,

                    CASE WHEN
                        count(m.prsnt_rdng_ocr_excep='Parameters Unavailable' or null) = 0 THEN 0 ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Parameters Unavailable' or null) as float)/cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
                    END as Parameters_Unavailable_Persent,
                    CASE WHEN
                        count(m.prsnt_rdng_ocr_excep='Image Invalid' or null) = 0 THEN 0 ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Image Invalid' or null) as float)/cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
                    END as Image_Invalid_Persent,

                    CASE WHEN
                        count(m.prsnt_rdng_ocr_excep='Image Unclear' or null) = 0 THEN 0 ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Image Unclear' or null) as float)/cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
                    END as Image_Unclear_Persent,

                    CASE WHEN
                        count(m.prsnt_rdng_ocr_excep='Image Spoofed' or null) = 0 THEN 0 ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Image Spoofed' or null) as float)/ cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
                    END as Image_Spoofed_Persent,

                    CASE WHEN
                        count(m.prsnt_rdng_ocr_excep='Image Stain on Decimal' or null) = 0 THEN 0 ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Image Stain on Decimal' or null) as float)/cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
                    END as Image_Stain_On_Decimal_persent,

                    CASE WHEN
                        count(m.prsnt_rdng_ocr_excep='Meter Mismatched' or null) = 0 THEN 0 ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Meter Mismatched' or null) as float)/cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
                    END as Meter_Mismatched_persent,

                    CASE WHEN
                        count(m.prsnt_rdng_ocr_excep='Meter On Height' or null) = 0 THEN 0 ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Meter On Height' or null) as float)/cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
                    END as Meter_On_Height_Persent,

                    CASE WHEN
                        count(m.prsnt_rdng_ocr_excep='Meter Dirty' or null) = 0 THEN 0 ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Meter Dirty' or null) as float)/cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
                    END as Meter_Dirty_Persent,

                    CASE WHEN
                        count(m.prsnt_rdng_ocr_excep='Meter Display Broken' or null) = 0 THEN 0 ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Meter Display Broken' or null) as float)/cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
                    END as Meter_Display_Broken_Persent,

                    CASE WHEN
                        count(m.prsnt_rdng_ocr_excep='Daylight Reflection On Meter' or null) = 0 THEN 0 ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Daylight Reflection On Meter' or null) as float)/cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
                    END as Daylight_Reflection_On_Meter_Persent,

                    CASE WHEN
                        count(m.prsnt_rdng_ocr_excep='Backlight Reflection' or null) = 0 THEN 0 ELSE ROUND((cast(count(m.prsnt_rdng_ocr_excep='Backlight Reflection' or null) as float)/cast(count(m.prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
                    END as Backlight_Reflection_Persent

                from readingmaster m
                join
                    (
                        SELECT DISTINCT sectioncode, MAX(sectionname) AS sectionname
                        FROM office
                        GROUP BY sectioncode
                    ) o ON m.ofc_section = o.sectioncode
                {clause} and o.sectionname!=''
                group by o.sectionname
                """
        # print(clause)
        print(query)
        cursor.execute(query)
        person_objects = dictfetchall(cursor)
        return Response(person_objects)
    except Exception as e:
        print(e)
        return Response([])


@api_view(["POST"])
def meterWiseReportconsumer(request):
    # pagesize = request.data.get("pagesize",)
    # page = (request.data.get("page",))
    # offset=(int(pagesize) * int(page))-int(pagesize)
    data = request.data.get("filters", None)

    clause = ""
    try:
        if data:
            clause += "WHERE "
            if data.get("month", "") == "":
                clause += f"extract(month from m.reading_date_db) = '{date.today().month}' AND extract(year from m.reading_date_db) = '{date.today().year}' AND "

            for i, (key, value) in enumerate(data.items()):
                if key == "month" and value:
                    year = value.split("-")[0]
                    month = value.split("-")[1]
                    clause += f"extract(month from m.reading_date_db) = '{month}' AND extract(year from m.reading_date_db) = '{year}' AND "

                elif key == "startdate" and value:
                    clause += f"m.reading_date_db BETWEEN '{data['startdate']}' AND "

                elif key == "enddate" and value:
                    clause += f"'{data['enddate']}' AND "

                elif (
                    key
                    in [
                        "ofc_discom",
                        "ofc_zone",
                        "ofc_circle",
                        "ofc_division",
                        "ofc_subdivision",
                        "bl_agnc_name",
                    ]
                    and value
                ):
                    clause += f"m.{key} ='{data[key]}' AND "

            if clause[-4:-1] == "AND":
                clause = clause[0:-4]

                # clause += f" {key}='{value}'"
            cursor = connection.cursor()
            #
            query = f"""
                Select distinct m.cons_ac_no as consAccountNumber, m.cons_name as consName,m.ofc_discom as discom,m.ofc_zone as zone,
                m.ofc_circle as circle,m.ofc_division as division,m.ofc_subdivision as subdivision,o.sectionname as section,m.rdng_img as readingImg,
                m.md_img as mdImg from readingmaster m join office o ON m.ofc_section=o.sectioncode {clause}
                """

            print(clause)
            print(query)
            # serializer=meterWiseReportconsumerSerializer(query,many=True)
            cursor.execute(query)
            person_objects = dictfetchall(cursor)
            return Response({"results": person_objects})
            # print(serializer.data)
            # return Response({"results":serializer.data})
    except:
        return Response([])


# ------------------------------------------------------ NEW APIS WITH MULTIPLE FILTERS-------------------------------------------------------------------------------------------


# @api_view(['POST'])
# def new_get_meter_summary(request):
#     new = []
#     def listfun(dict):
#         new.append(dict.copy())
#         return new
#     newdict = {}
#     clause = ''
#     cursor = connection.cursor()
#     filters = request.data.get("filters",None)

#     try:
#         if filters:
#             clause +='WHERE '
#             for i,(key,value) in enumerate(filters.items()):
#                 if i>0:
#                     clause +='AND '
#                 if key=='month':
#                     year = value.split('-')[0]
#                     month = value.split('-')[1]
#                     print("month",month)
#                     clause += f"extract(month from reading_date_db) = '{month}' AND extract(year from reading_date_db) = '{year}'"

#                 if key == 'startdate':
#                      clause += f"reading_date_db BETWEEN '{filters['startdate']}'"

#                 if key == 'enddate':
#                     clause += f"'{filters['enddate']}'"
#                 if key=='ofc_discom':
#                     clause +=f"ofc_discom='{value}'"
#                 if key=='ofc_zone':
#                     clause +=f"ofc_zone='{value}'"
#                 if key=='ofc_circle':
#                     clause +=f"ofc_circle='{value}'"
#                 if key=='ofc_division':
#                     clause +=f"ofc_division='{value}'"
#                 if key=='ofc_subdivision':
#                     clause +=f"ofc_subdivision='{value}'"
#                 if key=='ofc_section':
#                     clause +=f"ofc_section='{value}'"
#                 if key=='con_trf_cat':
#                     clause +=f"con_trf_cat='{value}'"

#                 if key=='bl_agnc_name':
#                     clause +=f"bl_agnc_name='{value}'"
#                     pass

#     except:
#         pass
#     query = f"""select readingmaster.mr_id,count(readingmaster.id),count(readingmaster.prsnt_mtr_status='Ok' or NULL),count(readingmaster.rdng_ocr_status='Passed' or NULL),count(readingmaster.rdng_ocr_status='Failed' or NULL),count(readingmaster.prsnt_mtr_status='Meter Defective' or NULL),count(readingmaster.prsnt_mtr_status='Door Locked' or NULL)
#             from readingmaster {clause}  group by readingmaster.mr_id
#         """
#     print(query)
#     cursor.execute(query)
#     results = cursor.fetchall()
#     try:
#         for row in results:
#             total = row[1]
#             okreadings = row[2]
#             ocrreadings = row[3]
#             ocrwithexcep = row[4]
#             meterdefective = row[5]
#             doorlocked = row[6]

#             # Percentage
#             okreadpercent = math.floor(((okreadings/total)*100))
#             ocrreadingpercent = math.floor(
#                 (((ocrreadings/okreadings) if okreadings else 0)*100))
#             ocrwithexceppercent = math.floor(
#                 (((ocrwithexcep/okreadings) if okreadings else 0)*100))
#             meterdefectivepercent = math.floor(((meterdefective/total)*100))
#             doorlockedpercent = math.floor(((doorlocked/total)*100))

#             # add to dictionary
#             newdict['mrid'] = row[0]
#             newdict['totalReadings'] = row[1]
#             newdict['OKreadings'] = okreadings
#             newdict['OKreadingspercent'] = okreadpercent
#             newdict['OCRReadings'] = ocrreadings
#             newdict['OCRReadingspercent'] = ocrreadingpercent
#             newdict['OCRwithException'] = ocrwithexcep
#             newdict['OCRwithExceptionpercent'] = ocrwithexceppercent
#             newdict['MeterDefective'] = meterdefective
#             newdict['MeterDefectivepercent'] = meterdefectivepercent
#             newdict['DoorLocked'] = doorlocked
#             newdict['DoorLockedpercent'] = doorlockedpercent
#             # add to list
#             newdata = listfun(newdict)
#         return Response(newdata)
#     except:
#         return Response([])


# @api_view(["POST"])
# def new_get_meter_summary(request):
#     today = date.today()
#     thismonth = today.strftime("%Y-%m")
#     year = thismonth.split("-")[0]
#     month = thismonth.split("-")[1]
#     new = []

#     def listfun(dict):
#         new.append(dict.copy())
#         return new

#     newdict = {}
#     clause = ""
#     cursor = connection.cursor()
#     filters = request.data.get("filters", None)

#     try:

#         clause += "AND "
#         clause += (
#             f"extract(month from reading_date_db) = '{month}' AND extract(year from reading_date_db) = '{year}'"
#             if "month" not in filters
#             else f"  extract(month from reading_date_db) = '{ filters['month'].split('-')[1]}' AND extract(year from reading_date_db) = '{filters['month'].split('-')[0]}'"
#         )
#         clause += (
#             f"and bl_agnc_name='{filters['bl_agnc_name']}'"
#             if "bl_agnc_name" in filters
#             else ""
#         )
#         clause += (
#             f"and extract(day from reading_date_db) between'{filters['startdate']}' and '{filters['enddate']}'"
#             if "enddate" in filters
#             else ""
#         )

#         clause += (
#             f"and ofc_discom='{filters['ofc_discom']}'"
#             if "ofc_discom" in filters
#             else ""
#         )
#         clause += (
#             f"and ofc_zone='{filters['ofc_zone']}'" if "ofc_zone" in filters else ""
#         )
#         clause += (
#             f"and ofc_circle='{filters['ofc_circle']}'"
#             if "ofc_circle" in filters
#             else ""
#         )
#         clause += (
#             f"and ofc_division='{filters['ofc_division']}'"
#             if "ofc_division" in filters
#             else ""
#         )
#         clause += (
#             f"and ofc_subdivision='{filters['ofc_subdivision']}'"
#             if "ofc_subdivision" in filters
#             else ""
#         )
#         clause += (
#             f"and ofc_section='{filters['ofc_section']}'"
#             if "ofc_section" in filters
#             else ""
#         )
#         clause += (
#             f"and con_trf_cat='{filters['con_trf_cat']}'"
#             if "con_trf_cat" in filters
#             else ""
#         )


#     except:
#         pass

#     query = f"""SELECT
# mr_id,
# COUNT(*) AS total,
# COUNT(CASE WHEN prsnt_mtr_status = 'Ok' THEN 1 END) AS ok,
# COUNT(CASE WHEN prsnt_mtr_status = 'Door Locked' THEN 1 END) AS dl,
# COUNT(CASE WHEN prsnt_mtr_status = 'Meter Defective' THEN 1 END) AS md,
# COUNT(CASE WHEN rdng_ocr_status = 'Passed' THEN 1 END) AS passed,
# COUNT(CASE WHEN rdng_ocr_status = 'Failed' THEN 1 END) AS failed,
# (COUNT(CASE WHEN rdng_ocr_status = 'Passed' THEN 1 END) + COUNT(CASE WHEN rdng_ocr_status = 'Failed' THEN 1 END)) - COUNT(CASE WHEN prsnt_mtr_status = 'Ok' THEN 1 END) AS diff,
# ofc_division,ofc_subdivision
# FROM
# readingmaster where mr_id !=''
# {clause}
# GROUP BY
# mr_id,ofc_division,ofc_subdivision;
#         """

#     print(query)
#     cursor.execute(query)
#     results = cursor.fetchall()
#     try:
#         for row in results:
#             total = row[1]
#             okreadings = row[2]
#             ocrreadings = row[5] - (row[7])
#             ocrwithexcep = row[6]
#             meterdefective = row[4]
#             doorlocked = row[3]
#             division = row[8]
#             subdivision = row[9]

#             # Percentage
#             okreadpercent = math.floor((((okreadings / total) if total else 0) * 100))
#             ocrreadingpercent = math.floor(
#                 (((ocrreadings / okreadings) if okreadings else 0) * 100)
#             )
#             ocrwithexceppercent = math.floor(
#                 (((ocrwithexcep / okreadings) if okreadings else 0) * 100)
#             )
#             # meterdefectivepercent = math.floor(((meterdefective/total)*100))
#             meterdefectivepercent = math.floor(
#                 (((meterdefective / total) if total else 0) * 100)
#             )
#             # doorlockedpercent = math.floor(((doorlocked/total)*100))
#             doorlockedpercent = math.floor(
#                 (((doorlocked / total) if total else 0) * 100)
#             )

#             # add to dictionary
#             newdict["mrid"] = row[0]
#             newdict["totalReadings"] = row[1]
#             newdict["OKreadings"] = okreadings
#             newdict["OKreadingspercent"] = okreadpercent
#             newdict["OCRReadings"] = ocrreadings
#             newdict["OCRReadingspercent"] = ocrreadingpercent
#             newdict["OCRwithException"] = ocrwithexcep
#             newdict["OCRwithExceptionpercent"] = ocrwithexceppercent
#             newdict["MeterDefective"] = meterdefective
#             newdict["MeterDefectivepercent"] = meterdefectivepercent
#             newdict["DoorLocked"] = doorlocked
#             newdict["DoorLockedpercent"] = doorlockedpercent
#             newdict["Division"] = division
#             newdict["SubDivision"] = subdivision
#             # add to list
#             newdata = listfun(newdict)
#         return Response(newdata)
#     except:
#         return Response([])

# Sanjeev
@api_view(["POST"])
def new_get_meter_summary(request):
    today = datetime.now()

    # Extract year and month of the current date
    year = today.year
    month = today.month

    # Format the current month
    thismonth = today.strftime("%Y-%m")
    print("Current month:", thismonth)

    # Subtract one month from the current date
    previous_month = today - timedelta(days=today.day)

    # Extract year and month of the previous month
    previous_year = previous_month.year
    previous_month_number = previous_month.month

    # Format the previous month
    previous_month_formatted = previous_month.strftime("%Y-%m")
    new = []

    def listfun(dict):
        new.append(dict.copy())
        return new

    cursor = connection.cursor()
    filters = request.data.get("filters", None)

    try:
        tablename = "readingmaster"  # Default to readingmaster

        if "month" in filters:
            selected_month = filters.get('month')
            if selected_month != thismonth and selected_month != previous_month_formatted:
                # If the selected month is neither current nor previous, switch to prevmonthsdata table
                tablename = "prevmonthsdata"
        else:
            # If month filter is not provided, and it's not the current or previous month, switch to prevmonthsdata table
            if thismonth != previous_month_formatted:
                tablename = "prevmonthsdata"

        print("Table name:", tablename)

    except Exception as e:
        print("Error:", e)

    clause = ""
    clause += "AND "
    clause += (
        f"extract(month from {tablename}.reading_date_db) = '{month}' AND extract(year from {tablename}.reading_date_db) = '{year}'"
        if "month" not in filters
        else f"  extract(month from {tablename}.reading_date_db) = '{filters['month'].split('-')[1]}' AND extract(year from {tablename}.reading_date_db) = '{filters['month'].split('-')[0]}'"
    )
    clause += (
        f"and {tablename}.bl_agnc_name='{filters['bl_agnc_name']}'"
        if "bl_agnc_name" in filters
        else ""
    )
    clause += (
        f"and extract(day from {tablename}.reading_date_db) between'{filters['startdate']}' and '{filters['enddate']}'"
        if "enddate" in filters
        else ""
    )
    clause += (
        f"and {tablename}.ofc_discom='{filters['ofc_discom']}'"
        if "ofc_discom" in filters
        else ""
    )
    clause += (
        f"and {tablename}.ofc_zone='{filters['ofc_zone']}'" if "ofc_zone" in filters else ""
    )
    clause += (
        f"and {tablename}.ofc_circle='{filters['ofc_circle']}'"
        if "ofc_circle" in filters
        else ""
    )
    clause += (
        f"and {tablename}.ofc_division='{filters['ofc_division']}'"
        if "ofc_division" in filters
        else ""
    )
    clause += (
        f"and {tablename}.ofc_subdivision='{filters['ofc_subdivision']}'"
        if "ofc_subdivision" in filters
        else ""
    )
    clause += (
        f"and {tablename}.ofc_section='{filters['ofc_section']}'"
        if "ofc_section" in filters
        else ""
    )
    clause += (
        f"and {tablename}.con_trf_cat='{filters['con_trf_cat']}'"
        if "con_trf_cat" in filters
        else ""
    )

    query = f"""SELECT
    mr_id,
    COUNT(*) AS total,
    COUNT(CASE WHEN prsnt_mtr_status = 'Ok' THEN 1 END) AS ok,
    COUNT(CASE WHEN prsnt_mtr_status = 'Door Locked' THEN 1 END) AS dl,
    COUNT(CASE WHEN prsnt_mtr_status = 'Meter Defective' THEN 1 END) AS md,
    COUNT(CASE WHEN rdng_ocr_status = 'Passed' THEN 1 END) AS passed,
    COUNT(CASE WHEN rdng_ocr_status = 'Failed' THEN 1 END) AS failed,
    (COUNT(CASE WHEN rdng_ocr_status = 'Passed' THEN 1 END) + COUNT(CASE WHEN rdng_ocr_status = 'Failed' THEN 1 END)) - COUNT(CASE WHEN prsnt_mtr_status = 'Ok' THEN 1 END) AS diff,
    ofc_division,ofc_subdivision
    FROM
    {tablename} where mr_id !=''
    {clause}
    GROUP BY
    mr_id,ofc_division,ofc_subdivision;
    """

    print("Query:", query)
    cursor.execute(query)
    results = cursor.fetchall()

    try:
        newdata = []  # Initialize newdata list
        for row in results:
            total = row[1]
            okreadings = row[2]
            ocrreadings = row[5] - (row[7])
            ocrwithexcep = row[6]
            meterdefective = row[4]
            doorlocked = row[3]
            division = row[8]
            subdivision = row[9]

            # Percentage calculations
            okreadpercent = math.floor(
                (((okreadings / total) if total else 0) * 100))
            ocrreadingpercent = math.floor(
                (((ocrreadings / okreadings) if okreadings else 0) * 100))
            ocrwithexceppercent = math.floor(
                (((ocrwithexcep / okreadings) if okreadings else 0) * 100))
            meterdefectivepercent = math.floor(
                (((meterdefective / total) if total else 0) * 100))
            doorlockedpercent = math.floor(
                (((doorlocked / total) if total else 0) * 100))

            # Constructing dictionary
            newdict = {
                "mrid": row[0],
                "totalReadings": total,
                "OKreadings": okreadings,
                "OKreadingspercent": okreadpercent,
                "OCRReadings": ocrreadings,
                "OCRReadingspercent": ocrreadingpercent,
                "OCRwithException": ocrwithexcep,
                "OCRwithExceptionpercent": ocrwithexceppercent,
                "MeterDefective": meterdefective,
                "MeterDefectivepercent": meterdefectivepercent,
                "DoorLocked": doorlocked,
                "DoorLockedpercent": doorlockedpercent,
                "Division": division,
                "SubDivision": subdivision
            }

            # Append to newdata list
            newdata.append(newdict)

        return Response(newdata)

    except Exception as e:
        print("Error:", e)
        return Response([])


# @api_view(["POST"])
# def newagencywisesummary(request):
#     new = []

#     def listfun(dict):
#         new.append(dict.copy())
#         return new

#     clause = ""
#     cursor = connection.cursor()
#     filters = request.data.get("filters", None)
#     newdict = {}
#     try:
#         if filters:
#             clause += "AND "
#             for i, (key, value) in enumerate(filters.items()):
#                 if i > 0:
#                     clause += "AND "
#                 if key == "month":
#                     year = value.split("-")[0]
#                     month = value.split("-")[1]
#                     print("month", month)
#                     clause += f"extract(month from reading_date_db) = '{month}' AND extract(year from reading_date_db) = '{year}'"

#                 if key == "startdate":
#                     clause += f"extract(day from reading_date_db) BETWEEN '{filters['startdate']}'"

#                 if key == "enddate":
#                     clause += f"'{filters['enddate']}'"

#                 if key == "ofc_discom":
#                     clause += f"ofc_discom='{value}'"

#     except:
#         pass

#     query = f"""select readingmaster.bl_agnc_name,count(readingmaster.id),count(readingmaster.prsnt_mtr_status='Ok' or NULL),count(readingmaster.rdng_ocr_status='Passed' or NULL),count(readingmaster.rdng_ocr_status='Failed' or NULL),count(readingmaster.prsnt_mtr_status='Meter Defective' or NULL),count(readingmaster.prsnt_mtr_status='Door Locked' or NULL)
#                     from readingmaster where readingmaster.bl_agnc_name!='' and bl_agnc_name !='BCITS' {clause}  group by readingmaster.bl_agnc_name
#     """

#     print("query", query)

#     try:
#         cursor.execute(query)
#         result = cursor.fetchall()
#         for row in result:
#             total = row[1]
#             okreadings = row[2]
#             ocrreadings = row[3]
#             ocrwithexcep = row[4]
#             meterdefective = row[5]
#             doorlocked = row[6]
#             # Percentage
#             okreadpercent = math.floor((okreadings / total) * 100)
#             ocrreadingpercent = math.floor(
#                 ((ocrreadings / okreadings) if okreadings else 0) * 100
#             )
#             ocrwithexceppercent = math.floor(
#                 ((ocrwithexcep / okreadings) if okreadings else 0) * 100
#             )
#             meterdefectivepercent = math.floor((meterdefective / total) * 100)
#             doorlockedpercent = math.floor((doorlocked / total) * 100)

#             # add to dictionary
#             newdict["agency"] = row[0]
#             newdict["totalReadings"] = row[1]
#             newdict["OKreadings"] = okreadings
#             newdict["OKreadingspercent"] = okreadpercent
#             newdict["OCRReadings"] = ocrreadings
#             newdict["OCRReadingspercent"] = ocrreadingpercent
#             newdict["OCRwithException"] = ocrwithexcep
#             newdict["OCRwithExceptionpercent"] = ocrwithexceppercent
#             newdict["MeterDefective"] = meterdefective
#             newdict["MeterDefectivepercent"] = meterdefectivepercent
#             newdict["DoorLocked"] = doorlocked
#             newdict["DoorLockedpercent"] = doorlockedpercent
#             # add to list
#             newdata = listfun(newdict)
#         return Response(newdata)
#     except:
#         return Response([])


# Sanjeev


@api_view(["POST"])
def newagencywisesummary(request):
    today = datetime.now()
    this_month = today.strftime("%Y-%m")
    previous_month = (today - timedelta(days=today.day)).strftime("%Y-%m")

    filters = request.data.get("filters", {})
    selected_month = filters.get("month")

    tablename = "prevmonthsdata" if selected_month not in {
        this_month, previous_month} else "readingmaster"

    clause = ""
    if filters:
        clause = "AND "
        for key, value in filters.items():
            if key == "month":
                year, month = value.split("-")
                clause += f"extract(month from {tablename}.reading_date_db) = '{month}' AND extract(year from {tablename}.reading_date_db) = '{year}' AND "
            elif key == "startdate":
                clause += f"extract(day from {tablename}.reading_date_db) BETWEEN '{value}' AND "
            elif key == "enddate":
                clause += f"'{value}' AND "
            elif key == "ofc_discom":
                clause += f"{tablename}.ofc_discom = '{value}' AND "

    # Remove trailing 'AND'
    clause = clause.rstrip("AND ")

    query = f"""
        SELECT
            {tablename}.bl_agnc_name,
            count({tablename}.id),
            count(CASE WHEN {tablename}.prsnt_mtr_status = 'Ok' THEN 1 END),
            count(CASE WHEN {tablename}.rdng_ocr_status = 'Passed' THEN 1 END),
            count(CASE WHEN {tablename}.rdng_ocr_status = 'Failed' THEN 1 END),
            count(CASE WHEN {tablename}.prsnt_mtr_status = 'Meter Defective' THEN 1 END),
            count(CASE WHEN {tablename}.prsnt_mtr_status = 'Door Locked' THEN 1 END)
        FROM
            {tablename}
        WHERE
            {tablename}.bl_agnc_name != ''
            AND {tablename}.bl_agnc_name != 'BCITS'
            {clause}
        GROUP BY
            {tablename}.bl_agnc_name
    """
    print("QUERY", query)

    try:
        cursor = connection.cursor()
        cursor.execute(query)
        result = cursor.fetchall()

        newdata = []
        for row in result:
            total = row[1]
            okreadings = row[2]
            ocrreadings = row[3]
            ocrwithexcep = row[4]
            meterdefective = row[5]
            doorlocked = row[6]

            # Percentage
            okreadpercent = math.floor(
                (okreadings / total) * 100) if total else 0
            ocrreadingpercent = math.floor(
                (ocrreadings / okreadings) * 100) if okreadings else 0
            ocrwithexceppercent = math.floor(
                (ocrwithexcep / okreadings) * 100) if okreadings else 0
            meterdefectivepercent = math.floor(
                (meterdefective / total) * 100) if total else 0
            doorlockedpercent = math.floor(
                (doorlocked / total) * 100) if total else 0

            newdata.append({
                "agency": row[0],
                "totalReadings": total,
                "OKreadings": okreadings,
                "OKreadingspercent": okreadpercent,
                "OCRReadings": ocrreadings,
                "OCRReadingspercent": ocrreadingpercent,
                "OCRwithException": ocrwithexcep,
                "OCRwithExceptionpercent": ocrwithexceppercent,
                "MeterDefective": meterdefective,
                "MeterDefectivepercent": meterdefectivepercent,
                "DoorLocked": doorlocked,
                "DoorLockedpercent": doorlockedpercent
            })

        return Response(newdata)
    except Exception as e:
        print(e)
        return Response([])


# @api_view(["POST"])
# def newmvsummary(request):
#     new = []

#     def listfun(dict):
#         new.append(dict.copy())
#         return new

#     clause = ""
#     cursor = connection.cursor()
#     filters = request.data.get("filters", None)
#     newdict = {}
#     try:
#         if filters:
#             clause += "WHERE mr_id!=''"
#             clause += (
#                 f" and extract(month from reading_date_db) = '{filters['month'].split('-')[1]}' AND extract(year from reading_date_db) = '{filters['month'].split('-')[0]}' "
#                 if (filters.get("month", None) not in ("", None))
#                 else ""
#             )
#             clause += (
#                 f" and extract(day from reading_date_db) BETWEEN '{filters['startdate']}' and '{filters['enddate']}' "
#                 if (
#                     (filters.get("enddate", None) not in ("", None))
#                     and (filters.get("startdate", None) not in ("", None))
#                 )
#                 else ""
#             )
#             clause += (
#                 f"and bl_agnc_name='{filters['bl_agnc_name']}'"
#                 if (filters.get("bl_agnc_name", None) not in ("", None))
#                 else ""
#             )
#             clause += (
#                 f" and ofc_discom='{filters['ofc_discom']}'"
#                 if (filters.get("ofc_discom", None) not in ("", None))
#                 else ""
#             )


#     except:
#         pass
#     query = f"""select mr_id,
#              count(*),
#              count(prsnt_mtr_status='Ok' or NULL),
#              count(rdng_ocr_status='Passed' or NULL),
#              count(readingmaster.rdng_ocr_status='Failed' or NULL),
#              count(prsnt_rdng_ocr_excep='Incorrect Reading' or NULL),
#              count(reading_parameter_type='Parameters Unavailable' and rdng_ocr_status='Failed' or NULL),
#              count(reading_parameter_type='Parameters Mismatch' and rdng_ocr_status='Failed'  or NULL),
#              count(prsnt_rdng_ocr_excep='Image blur' or NULL),
#              count(prsnt_rdng_ocr_excep='Spoofed Image' or NULL),
#              count(prsnt_rdng_ocr_excep='Meter Dirty' or NULL),
#              count(prsnt_mtr_status='Meter Defective' or NULL),
#              count(prsnt_mtr_status='Door Locked' or NULL)
#     from readingmaster {clause} GROUP BY mr_id
#     """

#     print("query", query)

#     cursor.execute(query)
#     results = cursor.fetchall()
#     try:
#         for row in results:
#             total = row[1]
#             okreadings = row[2]
#             ocrreadings = row[3]
#             ocrwithexcep = row[4]
#             incorrectreading = row[5]
#             parametersunavailable = row[6]
#             parametersmismatch = row[7]
#             imageblur = row[8]
#             spoofedimage = row[9]
#             meterdirty = row[10]
#             meterdefective = row[11]
#             doorlocked = row[12]
#             # Percentage
#             okreadpercent = math.floor((okreadings / total) * 100)
#             ocrreadingpercent = math.floor(
#                 ((ocrreadings / okreadings) if okreadings else 0) * 100
#             )
#             ocrwithexceppercent = math.floor(
#                 ((ocrwithexcep / okreadings) if okreadings else 0) * 100
#             )
#             incorrectreadingpercent = math.floor(
#                 ((incorrectreading / okreadings) if okreadings else 0) * 100
#             )
#             parametersunavailablepercent = math.floor(
#                 ((parametersunavailable / okreadings) if okreadings else 0) * 100
#             )

#             parametersmismatchpercent = math.floor(
#                 ((parametersmismatch / okreadings) if okreadings else 0) * 100
#             )
#             imageblurpercent = math.floor(
#                 ((imageblur / okreadings) if okreadings else 0) * 100
#             )
#             spoofedimagepercent = math.floor(
#                 ((spoofedimage / okreadings) if okreadings else 0) * 100
#             )

#             meterdirtypercent = math.floor(
#                 ((meterdirty / okreadings) if okreadings else 0) * 100
#             )
#             meterdefectivepercent = math.floor((meterdefective / total) * 100)
#             doorlockedpercent = math.floor((doorlocked / total) * 100)
#             # add to dictionary
#             newdict["mrid"] = row[0]
#             newdict["totalReadings"] = total
#             newdict["OKreadings"] = okreadings
#             newdict["OKreadingspercent"] = okreadpercent
#             newdict["OCRReadings"] = ocrreadings
#             newdict["OCRReadingspercent"] = ocrreadingpercent
#             newdict["OCRwithException"] = ocrwithexcep
#             newdict["OCRwithExceptionpercent"] = ocrwithexceppercent
#             newdict["incorrectreading"] = incorrectreading
#             newdict["incorrectreadingpercent"] = incorrectreadingpercent
#             newdict["parametersunavailable"] = parametersunavailable
#             newdict["parametersunavailablepercent"] = parametersunavailablepercent
#             newdict["parametersmismatch"] = parametersmismatch
#             newdict["parametersmismatchpercent"] = parametersmismatchpercent
#             newdict["imageblur"] = imageblur
#             newdict["imageblurpercent"] = imageblurpercent
#             newdict["spoofedimage"] = spoofedimage
#             newdict["spoofedimagepercent"] = spoofedimagepercent
#             newdict["meterdirty"] = meterdirty
#             newdict["meterdirtypercent"] = meterdirtypercent
#             newdict["MeterDefective"] = meterdefective
#             newdict["MeterDefectivepercent"] = meterdefectivepercent
#             newdict["DoorLocked"] = doorlocked
#             newdict["DoorLockedpercent"] = doorlockedpercent
#             # add to list
#             newdata = listfun(newdict)
#         return Response(newdata)

#     except:
#         return Response([])


# Sanjeev
@api_view(["POST"])
def newmvsummary(request):
    today = datetime.now()
    this_month = today.strftime("%Y-%m")
    print("thisMOnth:", this_month)
    previous_month = (today - timedelta(days=today.day)).strftime("%Y-%m")
    print("previous_month:", previous_month)
    new = []

    def listfun(dict):
        new.append(dict.copy())
        return new

    clause = ""
    cursor = connection.cursor()
    filters = request.data.get("filters", None)
    newdict = {}
    newdata = []

    try:
        if filters:
            selected_month = filters.get("month", None)

            if selected_month:
                year, month = selected_month.split('-')
                print("month:", month, "year:", year)
                if month in ["", None]:
                    month = datetime.now().month
                if year in ["", None]:
                    year = datetime.now().year
                clause += f"WHERE mr_id != '' AND extract(month from reading_date_db) = '{month}' AND extract(year from reading_date_db) = '{year}'"
            else:
                clause += "WHERE mr_id != ''"

            if filters.get("startdate") and filters.get("enddate"):
                clause += f" AND extract(day from reading_date_db) BETWEEN '{filters['startdate']}' AND '{filters['enddate']}'"

            if filters.get("bl_agnc_name"):
                clause += f" AND bl_agnc_name = '{filters['bl_agnc_name']}'"

            if filters.get("ofc_discom"):
                clause += f" AND ofc_discom = '{filters['ofc_discom']}'"

    except Exception as e:
        print(e)

    tablename = "prevmonthsdata" if selected_month not in {
        this_month, previous_month} else "readingmaster"

    query = f"""
        SELECT
            mr_id,
            count(*),
            count(prsnt_mtr_status='Ok' or NULL),
            count(rdng_ocr_status='Passed' or NULL),
            count(rdng_ocr_status='Failed' or NULL),
            count(prsnt_rdng_ocr_excep='Incorrect Reading' or NULL),
            count(reading_parameter_type='Parameters Unavailable' and rdng_ocr_status='Failed' or NULL),
            count(reading_parameter_type='Parameters Mismatch' and rdng_ocr_status='Failed' or NULL),
            count(prsnt_rdng_ocr_excep='Image blur' or NULL),
            count(prsnt_rdng_ocr_excep='Spoofed Image' or NULL),
            count(prsnt_rdng_ocr_excep='Meter Dirty' or NULL),
            count(prsnt_mtr_status='Meter Defective' or NULL),
            count(prsnt_mtr_status='Door Locked' or NULL)
        FROM {tablename}
        {clause}
        GROUP BY mr_id
    """

    print("query", query)

    try:
        cursor.execute(query)
        results = cursor.fetchall()
        for row in results:
            total = row[1]
            okreadings = row[2]
            ocrreadings = row[3]
            ocrwithexcep = row[4]
            incorrectreading = row[5]
            parametersunavailable = row[6]
            parametersmismatch = row[7]
            imageblur = row[8]
            spoofedimage = row[9]
            meterdirty = row[10]
            meterdefective = row[11]
            doorlocked = row[12]

            # Percentage calculations
            okreadpercent = math.floor(
                (okreadings / total) * 100) if total else 0
            ocrreadingpercent = math.floor(
                (ocrreadings / okreadings) * 100) if okreadings else 0
            ocrwithexceppercent = math.floor(
                (ocrwithexcep / okreadings) * 100) if okreadings else 0
            incorrectreadingpercent = math.floor(
                (incorrectreading / okreadings) * 100) if okreadings else 0
            parametersunavailablepercent = math.floor(
                (parametersunavailable / okreadings) * 100) if okreadings else 0
            parametersmismatchpercent = math.floor(
                (parametersmismatch / okreadings) * 100) if okreadings else 0
            imageblurpercent = math.floor(
                (imageblur / okreadings) * 100) if okreadings else 0
            spoofedimagepercent = math.floor(
                (spoofedimage / okreadings) * 100) if okreadings else 0
            meterdirtypercent = math.floor(
                (meterdirty / okreadings) * 100) if okreadings else 0
            meterdefectivepercent = math.floor(
                (meterdefective / total) * 100) if total else 0
            doorlockedpercent = math.floor(
                (doorlocked / total) * 100) if total else 0

            # Adding data to dictionary
            newdict = {
                "mrid": row[0],
                "totalReadings": total,
                "OKreadings": okreadings,
                "OKreadingspercent": okreadpercent,
                "OCRReadings": ocrreadings,
                "OCRReadingspercent": ocrreadingpercent,
                "OCRwithException": ocrwithexcep,
                "OCRwithExceptionpercent": ocrwithexceppercent,
                "incorrectreading": incorrectreading,
                "incorrectreadingpercent": incorrectreadingpercent,
                "parametersunavailable": parametersunavailable,
                "parametersunavailablepercent": parametersunavailablepercent,
                "parametersmismatch": parametersmismatch,
                "parametersmismatchpercent": parametersmismatchpercent,
                "imageblur": imageblur,
                "imageblurpercent": imageblurpercent,
                "spoofedimage": spoofedimage,
                "spoofedimagepercent": spoofedimagepercent,
                "meterdirty": meterdirty,
                "meterdirtypercent": meterdirtypercent,
                "MeterDefective": meterdefective,
                "MeterDefectivepercent": meterdefectivepercent,
                "DoorLocked": doorlocked,
                "DoorLockedpercent": doorlockedpercent
            }
            # Adding data to list
            newdata = listfun(newdict)

        return Response(newdata)

    except Exception as e:
        print(e)
        return Response([])


# @api_view(['POST'])
# def newmvcheck(request):
#     pagesize = request.data.get("pagesize", None)
#     print("PAGESIZE", pagesize)
#     page = request.data.get("page", 1)

#     print("PAGE", page)
#     offset = (int(pagesize) * int(page))-int(pagesize)
#     orderby = request.data.get("orderby", None)
#     clause = ''
#     filters = request.data.get("filters", None)
#     try:
#         if filters:
#             clause += 'WHERE '
#             for i, (key, value) in enumerate(filters.items()):
#                 if i > 0:
#                     clause += 'AND '
#                 if key == 'month':
#                     year = value.split('-')[0]
#                     month = value.split('-')[1]
#                     print("month", month)
#                     clause += f"extract(month from m.reading_date_db) = '{month}' AND extract(year from m.reading_date_db) = '{year}'"

#                 if key == 'startdate':
#                     clause += f"m.reading_date_db BETWEEN '{filters['startdate']}'"

#                 if key == 'enddate':
#                     clause += f"'{filters['enddate']}'"

#                 if key == 'mr_id':
#                     clause += f"m.mr_id='{filters['mr_id']}'"

#                 if key == 'searchdata':
#                     clause += f"(m.mr_id='{filters['searchdata']}' or m.cons_ac_no='{filters['searchdata']}' or m.cons_name='{filters['searchdata']}')"
#                 if key == 'rdng_ocr_status':
#                     clause += f"m.rdng_ocr_status='{filters['rdng_ocr_status']}'"
#                 if key == 'bl_agnc_name':
#                     clause += f"bl_agnc_name='{filters['bl_agnc_name']}'"
#     except:
#         pass
#     query = Consumers.objects.raw(f'''select m.mr_id as "mrId",m.rdng_date,m.prsnt_mtr_status,m.prsnt_ocr_rdng,m.prsnt_rdng,m.ocr_pf_reading,m.cons_name,m.cons_ac_no,m.prsnt_md_rdng_ocr,m.rdng_ocr_status,m.rdng_img,m.prsnt_md_rdng,m.id,r."mrPhoto",m.prsnt_rdng_ocr_excep,m.reading_parameter_type,count(*) over () as total_count
#                     from readingmaster m left outer join meterreaderregistration r on m.mr_id=r."mrId" {clause} order by m.rdng_date {orderby} limit {pagesize} offset {offset}
#     ''')
#     print(query)
#     serializer = ConsumersMeterRegistration(query, many=True)
#     count = serializer.data[0]['total_count'] if serializer.data else 0
#     return Response({'count': count, 'results': serializer.data})

# @api_view(['POST'])
# def newmvcheck(request):
#     pagesize = request.data.get("pagesize",None)
#     print("PAGESIZE",pagesize)
#     page = request.data.get("page",1)

#     print("PAGE",page)
#     offset=(int(pagesize) * int(page))-int(pagesize)
#     orderby = request.data.get("orderby", None)
#     clause = ''
#     filters = request.data.get("filters",None)
#     try:
#         if filters:
#             clause +='AND '
#             for i,(key,value) in enumerate(filters.items()):
#                 if i>0:
#                     clause +='AND '
#                 if key=='month':
#                     year = value.split('-')[0]
#                     month = value.split('-')[1]
#                     print("month",month)
#                     clause += f"extract(month from m.reading_date_db) = '{month}' AND extract(year from m.reading_date_db) = '{year}'"

#                 if key == 'startdate':
#                      clause += f"m.reading_date_db BETWEEN '{filters['startdate']}'"

#                 if key == 'enddate':
#                     clause += f"'{filters['enddate']}'"

#                 if key=='mr_id':
#                     clause += f"m.mr_id='{filters['mr_id']}'"
#                 if key=='prsnt_mtr_status':
#                     clause += f"m.prsnt_mtr_status='{filters['prsnt_mtr_status']}'"


#                 if key=='searchdata':
#                     clause += f"(m.mr_id='{filters['searchdata']}' or m.cons_ac_no='{filters['searchdata']}' or m.cons_name='{filters['searchdata']}')"
#                 if key=='rdng_ocr_status':
#                     clause +=f"m.rdng_ocr_status='{filters['rdng_ocr_status']}'"
#                 if key=='bl_agnc_name':
#                     clause +=f"bl_agnc_name='{filters['bl_agnc_name']}'"
#     except:
#         pass

#     query=Consumers.objects.raw(f'''select m.mr_id as "mrId",m.rdng_date,m.prsnt_mtr_status,m.prsnt_ocr_rdng,m.prsnt_rdng,m.ocr_pf_status,pf_image,pf_manual_reading,m.cons_name,m.cons_ac_no,m.prsnt_md_rdng_ocr,m.rdng_ocr_status,m.rdng_img,m.prsnt_md_rdng,m.id,r."mrPhoto",m.prsnt_rdng_ocr_excep,count(*) over () as total_count
#                     from readingmaster m left outer join meterreaderregistration r on m.mr_id=r."mrId" where m.rdng_ocr_status_changedby is null {clause} order by m.rdng_date {orderby} limit {pagesize} offset {offset}
#     ''')
#     print(query)
#     serializer=ConsumersMeterRegistration(query,many=True)

#     count = serializer.data[0]['total_count'] if serializer.data else 0
#     return Response({'count': count, 'results': serializer.data})


# @api_view(["POST"])
# def newmvcheck(request):
#     pagesize = request.data.get("pagesize", None)
#     print("PAGESIZE", pagesize)
#     page = request.data.get("page", 1)

#     print("PAGE", page)
#     offset = (int(pagesize) * int(page)) - int(pagesize)
#     orderby = request.data.get("orderby", None)
#     clause = ""
#     filters = request.data.get("filters", None)
#     cursor = connection.cursor()
#     try:
#         if filters:
#             clause += "AND "
#             for i, (key, value) in enumerate(filters.items()):
#                 if i > 0:
#                     clause += "AND "
#                 if key == "month":
#                     year = value.split("-")[0]
#                     month = value.split("-")[1]
#                     print("month", month)
#                     clause += f"extract(month from m.reading_date_db) = '{month}' AND extract(year from m.reading_date_db) = '{year}'"

#                 if key == "startdate":
#                     # clause += f"m.reading_date_db BETWEEN '{filters['startdate']}'"
#                     clause += f"extract(day from m.reading_date_db) BETWEEN '{filters['startdate']}'"

#                 if key == "enddate":
#                     clause += f"'{filters['enddate']}'"

#                 if key == "mr_id":
#                     clause += f"m.mr_id='{filters['mr_id']}'"
#                 if key == "prsnt_mtr_status":
#                     clause += f"m.prsnt_mtr_status='{filters['prsnt_mtr_status']}'"
#                 if key == "prsnt_rdng_ocr_excep":
#                     clause += (
#                         f"m.prsnt_rdng_ocr_excep='{filters['prsnt_rdng_ocr_excep']}'"
#                     )
#                 if key == "reading_parameter_type":
#                     clause += f"m.reading_parameter_type='{filters['reading_parameter_type']}' and rdng_ocr_status='Failed'"

#                 if key == "searchdata":
#                     clause += f"(m.mr_id='{filters['searchdata']}' or m.cons_ac_no='{filters['searchdata']}' or m.cons_name='{filters['searchdata']}')"
#                 if key == "rdng_ocr_status":
#                     if filters["rdng_ocr_status"] == "OK without Exception":
#                         clause += f"m.rdng_ocr_status='Passed'"
#                     else:
#                         clause += f"m.rdng_ocr_status='Failed'"

#                 if key == "bl_agnc_name":
#                     clause += f"bl_agnc_name='{filters['bl_agnc_name']}'"

#                 if key == "ofc_discom":
#                     clause += f"ofc_discom='{filters['ofc_discom']}'"
#     except:
#         pass
#     querytotal = f"""select count(*) from readingmaster m left outer join meterreaderregistration r on m.mr_id=r."mrId" where (m.rdng_ocr_status_changed_by='' or m.rdng_ocr_status_changed_by is null or m.rdng_ocr_status_changed_by ilike '%vapp%') and m.rdng_img!='' {clause}
#     """

#     query = f"""select m.mr_id as "mrId",m.rdng_date,m.prsnt_mtr_status,m.prsnt_ocr_rdng,m.prsnt_rdng,m.ocr_pf_status,pf_image,pf_manual_reading,m.cons_name,m.cons_ac_no,m.prsnt_md_rdng_ocr,m.rdng_ocr_status,m.rdng_img,m.prsnt_md_rdng,m.id,r."mrPhoto",m.prsnt_rdng_ocr_excep,m.reading_parameter_type
#                     from readingmaster m left outer join meterreaderregistration r on m.mr_id=r."mrId" where (m.rdng_ocr_status_changed_by is null or m.rdng_ocr_status_changed_by=''
#                     or m.rdng_ocr_status_changed_by ilike '%vapp%') and m.rdng_img!='' {clause} order by m.rdng_date {orderby} limit {pagesize} offset {offset}
#     """
#     print(query)
#     cursor.execute(query)
#     result = dictfetchall(cursor)
#     cursor.execute(querytotal)
#     resulttotal = dictfetchall(cursor)
#     # serializer=ConsumersMeterRegistration(query,many=True)

#     # count = serializer.data[0]['total_count'] if serializer.data else 0
#     return Response(
#         {"count": resulttotal[0]["count"] if resulttotal else 0, "results": result}
#     )


# sanjeev
# @api_view(["POST"])

# def newmvcheck(request):
#     pagesize = request.data.get("pagesize", None)
#     page = request.data.get("page", 1)
#     orderby = request.data.get("orderby", "DESC")
#     filters = request.data.get("filters", {})
#     export_all = request.data.get("export_all", False)  # NEW FLAG

#     offset = (int(pagesize) * int(page)) - int(pagesize) if pagesize else 0

#     print(filters)

#     # Build filter clause
#     clause_parts = []
#     for key, value in filters.items():
#         if key == "month":
#             year, month = value.split("-")
#             clause_parts.append(f"EXTRACT(month from m.reading_date_db) = '{month}'")
#             clause_parts.append(f"EXTRACT(year from m.reading_date_db) = '{year}'")
#         elif key == "startdate":
#             clause_parts.append(f"EXTRACT(day from m.reading_date_db) >= '{value}'")
#         elif key == "enddate":
#             clause_parts.append(f"EXTRACT(day from m.reading_date_db) <= '{value}'")
#         elif key == "mr_id":
#             clause_parts.append(f"m.mr_id = '{value}'")
#         elif key == "prsnt_mtr_status":
#             clause_parts.append(f"m.prsnt_mtr_status = '{value}'")
#         elif key == "reading_parameter_type":
#             clause_parts.append(f"m.reading_parameter_type = '{value}'")
#             clause_parts.append("m.rdng_ocr_status = 'Failed'")
#         elif key == "searchdata":
#             clause_parts.append(
#                 f"(m.mr_id = '{value}' OR m.cons_ac_no = '{value}' OR m.cons_name = '{value}')"
#             )
#         elif key == "rdngOcrStatus":
#             if value == "OCR without Exception":
#                 clause_parts.append("m.rdng_ocr_status = 'Passed'")
#             elif value == "OCR with Exception":
#                 exception_detail = filters.get("prsntRdngOcrExcep")
#                 if exception_detail:
#                     clause_parts.append("m.rdng_ocr_status = 'Failed'")
#                     clause_parts.append(f"m.prsnt_rdng_ocr_excep = '{exception_detail}'")
#         elif key == "bl_agnc_name":
#             clause_parts.append(f"bl_agnc_name = '{value}'")
#         elif key == "ofc_discom":
#             clause_parts.append(f"ofc_discom = '{value}'")

#     clause = " AND ".join(clause_parts)
#     clause = f" AND {clause}" if clause else ""

#     tablename = "readingmaster"  # Adjust if needed

#     # TOTAL COUNT
#     query_total = f"""
#         SELECT COUNT(*) FROM {tablename} m
#         LEFT JOIN meterreaderregistration r on m.mr_id=r."mrId"
#         WHERE (m.rdng_ocr_status_changed_by IS NULL OR m.rdng_ocr_status_changed_by=''
#                OR m.rdng_ocr_status_changed_by ILIKE '%vapp%')
#         AND m.rdng_img != '' {clause}
#     """

#     # DATA QUERY
#     query = f"""
#         SELECT m.mr_id as "mrId", m.rdng_date, m.prsnt_mtr_status, m.prsnt_ocr_rdng,
#                m.prsnt_rdng, m.ocr_pf_status, pf_image, pf_manual_reading,
#                m.cons_name, m.cons_ac_no, m.prsnt_md_rdng_ocr, m.rdng_ocr_status,
#                m.rdng_img, m.prsnt_md_rdng, m.id, r."mrPhoto",
#                m.prsnt_rdng_ocr_excep, m.reading_parameter_type,
#                m.cons_ac_no,m.rdng_date,m.rdng_img
#         FROM {tablename} m
#         LEFT JOIN meterreaderregistration r on m.mr_id=r."mrId"
#         WHERE (m.rdng_ocr_status_changed_by IS NULL OR m.rdng_ocr_status_changed_by=''
#                OR m.rdng_ocr_status_changed_by ILIKE '%vapp%')
#         AND m.rdng_img != '' {clause}
#         ORDER BY m.rdng_date {orderby}
#     """

#     if not export_all and pagesize:
#         query += f" LIMIT {pagesize} OFFSET {offset}"

#     cursor = connection.cursor()
#     cursor.execute(query)
#     results = dictfetchall(cursor)

#     cursor.execute(query_total)
#     total_result = dictfetchall(cursor)   # ✅ call once
#     total_count = total_result[0]["count"] if total_result else 0

#     #print(results)
#     print(query)
#     return Response({"count": total_count, "results": results})

# deepak


@api_view(["POST"])
def supervisorlogin(request):
    number = request.data.get("supervisor_number")
    password = request.data.get("password")

    if not number or not password:
        return Response({
            "status": False,
            "message": "supervisor_number and password required",
            "accessToken": "",
            "user": None
        })

    user = SupervisorLogin.objects.filter(
        supervisor_number=number,
        password=password
    ).first()   # <-- FIX

    if not user:
        return Response({
            "status": False,
            "message": "login failed",
            "accessToken": "",
            "user": None
        })

    return Response({
        "status": True,
        "message": "login successful",
        "accessToken": "",
        "user": {
            "supervise_name": user.supervisor_name or "",
            "supervise_number": user.supervisor_number or "",
            "is_admin": user.is_admin,
            "designation": "Supervisor",
            "division": user.ofc_division or "",
            "subdivison": user.ofc_subdivision or ""
        }
    })


# deeepak


# @api_view(["POST"])
# def supervisorlocation(request):
#     supervisor_number = request.data.get("supervisor_number")
#     geo_lat = request.data.get("geo_lat")
#     geo_long = request.data.get("geo_long")
#     date_str = request.data.get("date")

#     if not all([supervisor_number, geo_lat, geo_long, date_str]):
#         return Response({"status": False, "message": "Missing fields"}, status=400)

#     try:
#         date = datetime.strptime(date_str, "%Y-%m-%d").date()
#     except ValueError:
#         return Response({"status": False, "message": "Invalid date"}, status=400)

#     try:
#         with connection.cursor() as cursor:
#             cursor.execute("""
#                 INSERT INTO SupervisorLocation 
#                     (supervisor_number, geo_lat, geo_long, date)
#                 VALUES (%s, %s, %s, %s)
#                 ON CONFLICT (supervisor_number, date)
#                 DO UPDATE SET
#                     geo_lat = EXCLUDED.geo_lat,
#                     geo_long = EXCLUDED.geo_long
#                 RETURNING (xmax = 0) AS inserted
#             """, [supervisor_number, geo_lat, geo_long, date])

#             result = cursor.fetchone()
#             created = result[0] if result else False

#         message = "location added" if created else "location updated"
#         return Response({"status": True, "message": message})

#     except Exception as e:
#         return Response({"status": False, "message": f"Database error: {str(e)}"}, status=500)
# deeepak
from datetime import datetime
from django.db import connection
from rest_framework.decorators import api_view
from rest_framework.response import Response
 
@api_view(["POST"])
def supervisorlocation(request):
    supervisor_number = request.data.get("supervisor_number")
    geo_lat = request.data.get("geo_lat")
    geo_long = request.data.get("geo_long")
    date_str = request.dataget = request.data.get("date")
 
    if not all([supervisor_number, geo_lat, geo_long, date_str]):
        return Response({"status": False, "message": "Missing fields"}, status=400)
 
    # Parse datetime
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d %I:%M:%S %p")
        date = dt.date()
        time = dt.strftime("%H:%M:%S")
 
    except:
        return Response({"status": False, "message": "Invalid date format"}, status=400)
 
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                INSERT INTO supervsiorlocation
                    (supervisor_number, date, meta, created_at, updated_at)
                VALUES (
                    %s,%s,
                    jsonb_build_object(
                        'path', jsonb_build_array(
                            jsonb_build_object(
                                'time', %s,
                                'lat', %s,
                                'lng', %s
                            )
                        ),
                        'last_seen', %s,
                        'total_points', 1
                    ),
                    NOW(),NOW()
                )
 
                ON CONFLICT (supervisor_number, date)
                DO UPDATE SET
                    meta =
                        jsonb_set(
                            jsonb_set(
                                jsonb_set(
                                    supervsiorlocation.meta,
                                    '{path}',
                                    (supervsiorlocation.meta->'path') || jsonb_build_array(
                                        jsonb_build_object(
                                            'time', %s,
                                            'lat', %s,
                                            'lng', %s
                                        )
                                    )
                                ),
                                '{last_seen}',
                                to_jsonb(%s::text)
                            ),
                            '{total_points}',
                            to_jsonb((supervsiorlocation.meta->>'total_points')::int + 1)
                        ),
                    updated_at = NOW();
            """, [
                supervisor_number, date,
                time, geo_lat, geo_long, time,
                time, geo_lat, geo_long, time
            ])
        return Response({"status": True, "message": "Location recorded"})
    except Exception as e:
        print("DB ERROR:", e)
        return Response({"status": False, "message": str(e)}, status=500)

@api_view(["POST"])
def newmvcheck(request):
    pagesize = request.data.get("pagesize", None)
    page = request.data.get("page", 1)
    orderby = request.data.get("orderby", "DESC")
    filters = request.data.get("filters", {})
    print("filters...", filters)
    export_all = request.data.get("export_all", False)  # NEW FLAG

    offset = (int(pagesize) * int(page)) - int(pagesize) if pagesize else 0

    # Build filter clause
    clause_parts = []
    for key, value in filters.items():
        if key == "month":
            year, month = value.split("-")
            clause_parts.append(
                f"EXTRACT(month from m.reading_date_db) = '{month}'")
            clause_parts.append(
                f"EXTRACT(year from m.reading_date_db) = '{year}'")
        elif key == "startdate":
            clause_parts.append(
                f"EXTRACT(day from m.reading_date_db) >= '{value}'")
        elif key == "enddate":
            clause_parts.append(
                f"EXTRACT(day from m.reading_date_db) <= '{value}'")
        elif key == "mr_id":
            clause_parts.append(f"m.mr_id = '{value}'")
        elif key == "prsnt_mtr_status":
            clause_parts.append(f"m.prsnt_mtr_status = '{value}'")
        elif key == "reading_parameter_type":
            clause_parts.append(f"m.reading_parameter_type = '{value}'")
            clause_parts.append("m.rdng_ocr_status = 'Failed'")
        elif key == "searchdata":
            clause_parts.append(
                f"(m.mr_id = '{value}' OR m.cons_ac_no = '{value}' OR m.cons_name = '{value}')"
            )
        elif key == "rdng_ocr_status":
            if value == "OCR without Exception":
                clause_parts.append("m.rdng_ocr_status = 'Passed'")
            elif value == "OCR with Exception":
                exception_detail = filters.get("prsnt_rdng_ocr_excep")
                clause_parts.append("m.rdng_ocr_status = 'Failed'")
                if exception_detail:
                    clause_parts.append(
                        f"m.prsnt_rdng_ocr_excep = '{exception_detail}'")
        elif key == "bl_agnc_name":
            clause_parts.append(f"bl_agnc_name = '{value}'")
        # elif key == "ofc_discom":
        #     clause_parts.append(f"ofc_discom = '{value}'")
        elif key == "ofc_discom":
            if value and value.upper() != "ALL":
                clause_parts.append(f"ofc_discom = '{value}'")

    clause = " AND ".join(clause_parts)
    clause = f" AND {clause}" if clause else ""

    tablename = "readingmaster"  # Adjust if needed

    # Base SELECT
    query = f"""
        SELECT m.con_mtr_sl_no, m.mr_id as "mrId", m.rdng_date, m.prsnt_mtr_status, m.prsnt_ocr_rdng,
               m.prsnt_rdng, m.ocr_pf_status, pf_image, pf_manual_reading,
               m.cons_name, m.cons_ac_no, m.prsnt_md_rdng_ocr, m.rdng_ocr_status,
               m.rdng_img, m.prsnt_md_rdng, m.id, r."mrPhoto",
               m.prsnt_rdng_ocr_excep, m.reading_parameter_type
        FROM {tablename} m
        LEFT JOIN meterreaderregistration r on m.mr_id=r."mrId"
        WHERE (m.rdng_ocr_status_changed_by IS NULL OR m.rdng_ocr_status_changed_by=''
               OR m.rdng_ocr_status_changed_by ILIKE '%vapp%' OR m.qc_done != 'byLambda')
        AND m.rdng_img != '' {clause}
        ORDER BY m.rdng_date {orderby}
    """

    # Only apply LIMIT/OFFSET when NOT exporting all
    if not export_all and pagesize:
        query += f" LIMIT {pagesize} OFFSET {offset}"

    cursor = connection.cursor()
    cursor.execute(query)
    results = dictfetchall(cursor)

    if export_all:
        # No need to run count, just return all rows
        return Response({"count": len(results), "results": results})
    else:
        # Normal pagination → get total count
        query_total = f"""
            SELECT COUNT(*) FROM {tablename} m
            LEFT JOIN meterreaderregistration r on m.mr_id=r."mrId"
            WHERE (m.rdng_ocr_status_changed_by IS NULL OR m.rdng_ocr_status_changed_by=''
                   OR m.rdng_ocr_status_changed_by ILIKE '%vapp%')
            AND m.rdng_img != '' {clause}
        """
        cursor.execute(query_total)
        total_count = dictfetchall(cursor)[0]["count"]

        return Response({"count": total_count, "results": results})


@api_view(["GET"])
def application_uptime(request):
    now = datetime.utcnow()

    # SLA period = full calendar month
    start_time = now.replace(day=1)
    last_day = monthrange(now.year, now.month)[1]
    # end_time = now.replace(day=last_day)
    end_time = now  # CURRENT DATE

    uptime = get_lambda_uptime("new-truereadapi-prod12")

    if uptime >= 95:
        penalty = "No Penalty"
    elif uptime >= 90:
        penalty = "5% Penalty"
    elif uptime >= 80:
        penalty = "10% Penalty"
    else:
        penalty = "15% Penalty"

    return Response({
        "uptime_percentage": uptime,
        "penalty": penalty,
        "start_time": start_time.strftime("%d %b %Y"),
        "end_time": end_time.strftime("%d %b %Y")
    })

from django.views.decorators.csrf import csrf_exempt
from datetime import datetime, timedelta, time


@csrf_exempt
@api_view(["POST"])
def application_uptime_range(request):
    start_date_str = request.data.get("start_date")
    end_date_str = request.data.get("end_date")

    if not start_date_str or not end_date_str:
        return Response(
            {"error": "start_date and end_date are required"},
            status=400
        )

    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
    except ValueError:
        return Response(
            {"error": "Invalid date format. Use YYYY-MM-DD"},
            status=400
        )

    if start_date > end_date:
        return Response(
            {"error": "start_date cannot be after end_date"},
            status=400
        )

    # 🔹 CASE 1: Same date → full day
    if start_date == end_date:
        start_time = datetime.combine(start_date, datetime.min.time())   # 00:00:00
        end_time = datetime.combine(end_date, datetime.max.time())       # 23:59:59

    # 🔹 CASE 2: Different dates → boundary style
    else:
        start_time = datetime.combine(start_date, datetime.min.time())   # 00:00:00
        end_time = datetime.combine(end_date, datetime.min.time())       # 00:00:00

    # 🚨 CloudWatch safety (absolute guard)
    if start_time >= end_time:
        end_time = start_time + timedelta(seconds=1)

    uptime = get_lambda_uptime_by_range(
        "new-truereadapi-prod12",
        start_time,
        end_time
    )

    # SLA penalty logic
    if uptime >= 95:
        penalty = "No Penalty"
    elif uptime >= 90:
        penalty = "5% Penalty"
    elif uptime >= 80:
        penalty = "10% Penalty"
    else:
        penalty = "15% Penalty"

    return Response({
        "uptime_percentage": uptime,
        "penalty": penalty,
        "start_time": start_date.strftime("%d %b %Y"),
        "end_time": end_date.strftime("%d %b %Y"),
    })


#for dily uptime for date range
from datetime import datetime, timedelta
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .services.uptime_service import (
    get_lambda_uptime_for_day,
    calculate_penalty,
)
# Example helper (replace with real logic)

@api_view(["GET"])
def application_uptime_daily(request):
    start_date_str = request.GET.get("start_date")
    end_date_str = request.GET.get("end_date")

    if not start_date_str or not end_date_str:
        return Response(
            {"error": "start_date and end_date are required (YYYY-MM-DD)"},
            status=400,
        )

    start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
    end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()

    if start_date > end_date:
        return Response(
            {"error": "start_date cannot be after end_date"},
            status=400,
        )

    results = []
    current_date = start_date

    while current_date <= end_date:
        uptime = get_lambda_uptime_for_day(
            "new-truereadapi-prod12",
            current_date,
        )

        results.append({
            "date": current_date.strftime("%d %b %Y"),
            "uptime_percentage": uptime,
            "penalty": calculate_penalty(uptime),
        })

        current_date += timedelta(days=1)

    return Response({
        "start_date": start_date.strftime("%d %b %Y"),
        "end_date": end_date.strftime("%d %b %Y"),
        "daily_uptime": results,
    })

# @api_view(["GET"])
# def application_uptime(request):
#     end_time = datetime.utcnow()
#     start_time = end_time - timedelta(days=30)

#     lambda_uptime = get_lambda_uptime("new-truereadapi-prod12")
#     rds_uptime = get_rds_uptime("database-2")

#     application_uptime = min(lambda_uptime, rds_uptime)

#     if application_uptime >= 95:
#         penalty = "No Penalty"
#     elif application_uptime >= 90:
#         penalty = "5% Penalty"
#     elif application_uptime >= 80:
#         penalty = "10% Penalty"
#     else:
#         penalty = "15% Penalty"

#     return Response({
#         "application_uptime": application_uptime,
#         "lambda_uptime": lambda_uptime,
#         "rds_uptime": rds_uptime,
#         "penalty": penalty,
#         "start_time": start_time.strftime("%d %b %Y"),
#         "end_time": end_time.strftime("%d %b %Y")
#     })



@api_view(["POST"])
def gitnewmvcheck(request):
    pagesize = request.data.get("pagesize", None)
    page = request.data.get("page", 1)
    orderby = request.data.get("orderby", "DESC")
    filters = request.data.get("filters", {})
    print("filters...", filters)
    export_all = request.data.get("export_all", False)  # NEW FLAG

    offset = (int(pagesize) * int(page)) - int(pagesize) if pagesize else 0

    # Build filter clause
    clause_parts = []
    for key, value in filters.items():
        if key == "month":
            year, month = value.split("-")
            clause_parts.append(
                f"EXTRACT(month from m.reading_date_db) = '{month}'")
            clause_parts.append(
                f"EXTRACT(year from m.reading_date_db) = '{year}'")
        elif key == "startdate":
            clause_parts.append(
                f"EXTRACT(day from m.reading_date_db) >= '{value}'")
        elif key == "enddate":
            clause_parts.append(
                f"EXTRACT(day from m.reading_date_db) <= '{value}'")
        elif key == "mr_id":
            clause_parts.append(f"m.mr_id = '{value}'")
        elif key == "prsnt_mtr_status":
            clause_parts.append(f"m.prsnt_mtr_status = '{value}'")
        elif key == "reading_parameter_type":
            clause_parts.append(f"m.reading_parameter_type = '{value}'")
            clause_parts.append("m.rdng_ocr_status = 'Failed'")
        elif key == "searchdata":
            clause_parts.append(
                f"(m.mr_id = '{value}' OR m.cons_ac_no = '{value}' OR m.cons_name = '{value}')"
            )
        elif key == "rdng_ocr_status":
            if value == "OCR without Exception":
                clause_parts.append("m.rdng_ocr_status = 'Passed'")
            elif value == "OCR with Exception":
                exception_detail = filters.get("prsnt_rdng_ocr_excep")
                clause_parts.append("m.rdng_ocr_status = 'Failed'")
                if exception_detail:
                    clause_parts.append(
                        f"m.prsnt_rdng_ocr_excep = '{exception_detail}'")
        elif key == "bl_agnc_name":
            clause_parts.append(f"bl_agnc_name = '{value}'")
        elif key == "ofc_discom":
            clause_parts.append(f"ofc_discom = '{value}'")

    clause = " AND ".join(clause_parts)
    clause = f" AND {clause}" if clause else ""

    tablename = "readingmaster"  # Adjust if needed

    # Base SELECT
    query = f"""
        SELECT m.con_mtr_sl_no, m.mr_id as "mrId", m.rdng_date, m.prsnt_mtr_status, m.prsnt_ocr_rdng,
               m.prsnt_rdng, m.ocr_pf_status, pf_image, pf_manual_reading,
               m.cons_name, m.cons_ac_no, m.prsnt_md_rdng_ocr, m.rdng_ocr_status,
               m.rdng_img, m.prsnt_md_rdng, m.id, r."mrPhoto",
               m.prsnt_rdng_ocr_excep, m.reading_parameter_type
        FROM {tablename} m
        LEFT JOIN meterreaderregistration r on m.mr_id=r."mrId"
        WHERE (m.rdng_ocr_status_changed_by IS NULL OR m.rdng_ocr_status_changed_by=''
               OR m.rdng_ocr_status_changed_by ILIKE '%vapp%')
        AND m.rdng_img != '' {clause}
        ORDER BY m.rdng_date {orderby}
    """

    # Only apply LIMIT/OFFSET when NOT exporting all
    if not export_all and pagesize:
        query += f" LIMIT {pagesize} OFFSET {offset}"

    cursor = connection.cursor()
    cursor.execute(query)
    results = dictfetchall(cursor)

    if export_all:
        # No need to run count, just return all rows
        return Response({"count": len(results), "results": results})
    else:
        # Normal pagination → get total count
        query_total = f"""
            SELECT COUNT(*) FROM {tablename} m
            LEFT JOIN meterreaderregistration r on m.mr_id=r."mrId"
            WHERE (m.rdng_ocr_status_changed_by IS NULL OR m.rdng_ocr_status_changed_by=''
                   OR m.rdng_ocr_status_changed_by ILIKE '%vapp%')
            AND m.rdng_img != '' {clause}
        """
        cursor.execute(query_total)
        total_count = dictfetchall(cursor)[0]["count"]

        return Response({"count": total_count, "results": results})


@api_view(["POST"])
def new_locationwise_summary(request):
    newdict = {}
    new = []

    def listfun(dict):
        print(dict)
        new.append(dict.copy())
        return new

    cursor = connection.cursor()
    locationwise = request.data.get("locationwise", None)
    locationname = request.data.get("locationname", None)
    groupby = request.data.get("groupby")
    where = request.data.get("where")
    previouslocation = request.data.get("previouslocation")
    clause = ""
    if (locationwise is not None) and (locationname == "all"):
        location = "ofc_" + locationwise
        clause = "WHERE " + previouslocation + "='" + where + "' "
        cursor.execute(
            f"""
   select {groupby} as location, count(distinct mr_id),count(prsnt_mtr_status='Ok' or NULL),count(rdng_ocr_status='Passed' or NULL),count(rdng_ocr_status='Failed' or NULL),count(prsnt_mtr_status='Meter Defective' or NULL),count(prsnt_mtr_status='Door Locked' or NULL),count(mr_id)
    from readingmaster {clause}  GROUP BY {groupby}
    """
        )
    if (locationwise is not None) and (locationname != "all"):
        location = "ofc_" + locationwise
        clause = "WHERE " + locationwise + "='" + locationname + "' "

        cursor.execute(
            f"""
   select {groupby} as location, count(distinct mr_id),count(prsnt_mtr_status='Ok' or NULL),count(rdng_ocr_status='Passed' or NULL),count(rdng_ocr_status='Failed' or NULL),count(prsnt_mtr_status='Meter Defective' or NULL),count(prsnt_mtr_status='Door Locked' or NULL),count(mr_id)
    from readingmaster {clause} GROUP BY {groupby}
    """
        )

    result = cursor.fetchall()
    print("result", result)
    try:
        for row in result:
            locationname = row[0]
            total = row[7]
            mrid = row[1]
            okreadings = row[2]
            OcrReadings = row[3]
            Ocrwithexception = row[4]
            meterDefective = row[5]
            doorLocked = row[6]
            okreadpercent = math.floor((okreadings / total) * 100)
            ocrreadingpercent = math.floor(
                ((OcrReadings / okreadings) if okreadings else 0) * 100
            )
            ocrwithexceppercent = math.floor(
                ((Ocrwithexception / okreadings) if okreadings else 0) * 100
            )
            meterdefectivepercent = math.floor((meterDefective / total) * 100)
            doorlockedpercent = math.floor((doorLocked / total) * 100)
            newdict["locationname"] = row[0]
            newdict["mrid"] = row[1]
            newdict["okreadings"] = row[2]
            newdict["okreadingspercent"] = okreadpercent
            newdict["OcrReadings"] = row[3]
            newdict["OcrReadingspercent"] = ocrreadingpercent
            newdict["Ocrwithexception"] = row[4]
            newdict["Ocrwithexceptionpercent"] = ocrwithexceppercent
            newdict["meterDefective"] = row[5]
            newdict["meterDefectivepercent"] = meterdefectivepercent
            newdict["doorLocked"] = row[6]
            newdict["doorLockedpercent"] = doorlockedpercent
            newdict["total"] = total
            data = listfun(newdict)
        return Response(data)
    except:
        return Response([])


# @api_view(['POST'])
# def meterreaderDetails(request):
#     pagesize = request.data.get("pagesize",)
#     page = (request.data.get("page",))
#     offset = (int(pagesize) * int(page))-int(pagesize)

#     data = request.data.get("filters", None)
#     clause = ''
#     try:
#         if data:
#             clause += 'WHERE '
#             for i, (key, value) in enumerate(data.items()):
#                 if i > 0:
#                     if key == "enddate":
#                         clause += ''
#                     else:
#                         clause += ' AND '

#                 print("data---------", clause, data, i, key, value)
#                 if key == "startdate" or "enddate":
#                     print(12)
#                     # startdate1=''
#                     if key == 'startdate':
#                         startdate1 = value
#                         continue
#                     if key == 'enddate':
#                         enddate1 = value
#                         key = ''
#                         value = f" reading_date_db between '{startdate1}' and '{enddate1}' "
#                         clause += f" {key} {value}"
#                 if key == "mr_id":
#                     key = "mr_id"
#                     value = value
#                     clause += f" {key}='{value}'"
#                 if key == "bl_agnc_name":
#                     key = "bl_agnc_name"
#                     value = value
#                     clause += f" {key}='{value}'"

#                 if key == 'month':
#                     key = "reading_date_db"
#                     month = value.split('-')[1]
#                     year = value.split('-')[0]
#                     key = f"extract(Month from reading_date_db)='{month}' and extract(Year from reading_date_db) "
#                     print("key['month']", key)
#                     print("value['month']", value)
#                     value = year
#                     clause += f" {key}='{value}'"
#                 # if  key=='enddate' or enddate1:
#                 #     clause +=f" {key} {value}"
#                 #     # key=None
#                 #     enddate1=None
#                 # else:
#                 #     print(clause)
#                 #     clause +=f" {key}='{value}'"
#                 # clause +=f" {key}='{value}'"
#             print("clause-------", clause)
#             cursor = connection.cursor()
#             query = (f'''
#             select mr_id,cons_ac_no,bl_agnc_name,abnormality,cons_name,con_trf_cat,con_mtr_sl_no,
# mr_rmrk,prsnt_mtr_status,prsnt_rdng,prev_rdng,
# prsnt_md_rdng,prev_md,ocr_pf_reading,prev_pf_rdng,rdng_date,prev_rdng_date,rdng_img,md_img,
# prsnt_rdng_ocr_excep,md_ocr_excep,qc_req,count(*) over () as total_count from readingmaster {clause} order by rdng_date DESC limit {pagesize} offset {offset}
#                 ''')
#             print(clause)
#             print(query)
#             # group by m.{orderby}
#             cursor.execute(query)
#             person_objects = dictfetchall(cursor)
#         return Response({"result": person_objects, "count": person_objects[0]['total_count']})

#     except:
#         return Response({"result": [], "count": 5})


@api_view(["POST"])
def clusterstestnew(request):
    data = request.data.get("filters", None)
    today = date.today()
    clause = ""
    # try:
    if data:
        # this below code is for supervisor location
        if "mr_id" in data:
            mr_id_value = data["mr_id"]
            if mr_id_value.startswith('SUP_'):
                supervisor_number = mr_id_value[4:]
                print("-------->>>>", today, supervisor_number)
                try:
                    with connection.cursor() as cursor:
                        cursor.execute("""
                            SELECT
                            jsonb_agg(
                                jsonb_build_object(
                                'geo_lat',  p->>'lat',
                                'geo_long', p->>'lng',
                                'time',     p->>'time'
                                )
                            ) AS path,
                            sl.supervisor_number,
                            sl.date
                            FROM supervsiorlocation sl
                            CROSS JOIN LATERAL jsonb_array_elements(sl.meta->'path') AS p
                            WHERE sl.supervisor_number = %s
                            AND sl.date = %s
                            GROUP BY sl.supervisor_number, sl.date
                        """, [supervisor_number, today])

                        row = cursor.fetchone()

                    if not row:
                        return Response([])

                    response_data = {
                        "supervisor_number": row[1],
                        "date": row[2],
                        "path": json.loads(row[0]) if isinstance(row[0], str) else row[0]
                    }

                    supervisor_login_data = SupervisorLogin.objects.filter(
                        supervisor_number=supervisor_number
                    ).values(
                        'supervisor_name', 'ofc_division', 'ofc_subdivision'
                    ).first()

                    if supervisor_login_data:
                        response_data.update(supervisor_login_data)
                    return Response(response_data)
                except Exception as e:
                    return Response({"error": str(e)}, status=500)

        clause += "WHERE "
        for i, (key, value) in enumerate(data.items()):
            if i > 0:
                clause += " AND "
            if key == "mr_id":
                clause += f" {key}='{value}'"
            if key == "bl_agnc_name":
                clause += f" {key}='{value}'"

        cursor = connection.cursor()
        query = f"""
        select mr_id,rdng_date,cons_name,geo_lat,geo_long,prsnt_mtr_status,rdng_ocr_status,prsnt_ocr_rdng,ocr_pf_reading,cons_ac_no,prsnt_md_rdng_ocr,prsnt_md_rdng,prsnt_rdng,qc_req,
        rdng_img from readingmaster {clause} AND reading_date_db='{today}'
        """
        print(query)
        cursor.execute(query)
        result = dictfetchall(cursor)
        return Response(result)


# @api_view(["POST"])
# def newmonthdataa(request):
#     today = date.today()
#     thismonth = today.strftime("%Y-%m")
#     year = thismonth.split("-")[0]
#     month = thismonth.split("-")[1]

#     new = []

#     def listfun(dict):
#         new.append(dict.copy())
#         return new

#     newdict = {}
#     clause = ""
#     filters = request.data.get("filters", None)
#     try:
#         clause += "WHERE "
#         clause += (
#             f"extract(month from reading_date_db) = '{month}' AND extract(year from reading_date_db) = '{year}'"
#             if "month" not in filters
#             else f"  extract(month from reading_date_db) = '{ filters['month'].split('-')[1]}' AND extract(year from reading_date_db) = '{filters['month'].split('-')[0]}'"
#         )
#         clause += (
#             f"and bl_agnc_name='{filters['bl_agnc_name']}'"
#             if "bl_agnc_name" in filters
#             else ""
#         )
#         clause += (
#             f"and extract(day from reading_date_db) between'{filters['startdate']}' and '{filters['enddate']}'"
#             if "enddate" in filters
#             else ""
#         )

#         clause += (
#             f"and ofc_discom='{filters['ofc_discom']}'"
#             if "ofc_discom" in filters
#             else ""
#         )
#         clause += (
#             f"and ofc_zone='{filters['ofc_zone']}'" if "ofc_zone" in filters else ""
#         )
#         clause += (
#             f"and ofc_circle='{filters['ofc_circle']}'"
#             if "ofc_circle" in filters
#             else ""
#         )
#         clause += (
#             f"and ofc_division='{filters['ofc_division']}'"
#             if "ofc_division" in filters
#             else ""
#         )
#         clause += (
#             f"and ofc_subdivision='{filters['ofc_subdivision']}'"
#             if "ofc_subdivision" in filters
#             else ""
#         )
#         # clause +=f"and bl_agnc_name='{filters['bl_agnc_name']}'" if 'bl_agnc_name' in filters else ''

#     except:
#         pass
#     # count(r.prsnt_mtr_status='Meter Defective' or r.prsnt_mtr_status='Door Locked' or r.prsnt_mtr_status='Ok' or null) as billed_consumers,
#     cursor = connection.cursor()
#     query = f"""SELECT r.mr_id,r.ofc_discom,r.ofc_zone,r.ofc_circle,r.ofc_division,r.ofc_subdivision,count(*) as billed_consumers,
# count(r.prsnt_mtr_status='Ok' OR NULL) as ok_readings,count(r.rdng_ocr_status='Passed' or NULL) as OCRwithoutException,
# count(r.rdng_ocr_status='Failed' OR NULL) as OCRwithException,count(r.prsnt_mtr_status='Meter Defective' OR NULL) as MeterDefective,
# count(r.prsnt_mtr_status='Door Locked' OR NULL) as DoorLocked,r.bl_agnc_name,COUNT(CASE WHEN prsnt_mtr_status = 'Ok' THEN 1 END)
# - (COUNT(CASE WHEN rdng_ocr_status = 'Passed' THEN 1 END)
# + COUNT(CASE WHEN rdng_ocr_status = 'Failed' THEN 1 END)) AS diff
# FROM readingmaster r {clause}

# GROUP BY r.mr_id,r.ofc_discom,r.ofc_zone,r.ofc_circle,r.ofc_division,r.ofc_subdivision,r.bl_agnc_name ;

#         """
#     print("QUERY------>", query)
#     cursor.execute(query)
#     results = cursor.fetchall()

#     try:
#         for row in results:
#             total = row[6]
#             okreadings = row[7]
#             ocrreadings = row[8]
#             ocrwithexcep = row[9]
#             meterdefective = row[10]
#             doorlocked = row[11]

#             # Percentage
#             okreadpercent = round((((okreadings / total) if total else 0) * 100), 2)
#             # okreadpercent = round(((okreadings/total)*100), 2)
#             ocrreadingpercent = round(
#                 (((ocrreadings / okreadings) if okreadings else 0) * 100), 2
#             )
#             ocrwithexceppercent = round(
#                 (((ocrwithexcep / okreadings) if okreadings else 0) * 100), 2
#             )
#             meterdefectivepercent = round(
#                 (((meterdefective / total) if total else 0) * 100), 2
#             )
#             doorlockedpercent = round((((doorlocked / total) if total else 0) * 100), 2)
#             # add to dictionary
#             newdict["mrid"] = row[0]
#             # newdict['mrPhone'] = row[10]
#             newdict["ofc_discom"] = row[1]
#             newdict["ofc_zone"] = row[2]
#             newdict["ofc_circle"] = row[3]
#             newdict["ofc_division"] = row[4]
#             newdict["ofc_subdivision"] = row[5]
#             newdict["billed_consumers"] = row[6]
#             newdict["meterdefective"] = row[10]
#             newdict["doorlocked"] = row[11]
#             newdict["agency"] = row[12]

#             newdict["OKreadings"] = okreadings
#             newdict["OCRReadings"] = ocrreadings + (row[13])
#             newdict["OCRwithException"] = ocrwithexcep
#             newdict["OKreadingspercent"] = okreadpercent
#             newdict["OCRReadingspercent"] = ocrreadingpercent
#             newdict["OCRwithExceptionpercent"] = ocrwithexceppercent
#             newdict["MeterDefectivePercent"] = meterdefectivepercent
#             newdict["DoorLockedOercent"] = doorlockedpercent
#             # add to list
#             newdata = listfun(newdict)
#         return Response(newdata)

#     except:
#         return Response([])


#indra
@api_view(["POST"])
def newmonthdataa(request):
    filters = request.data.get("filters", {}) or {}

    # --- Determine month ---
    today = date.today()
    current_month = today.strftime("%Y-%m")
    previous_month = (today - timedelta(days=today.day)).strftime("%Y-%m")
    selected_month = filters.get("month", current_month)

    year, month = map(int, selected_month.split("-"))
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)

    # --- Table selection (unchanged logic) ---
    tablename = (
        "readingmaster"
        if selected_month in {current_month, previous_month}
        else "prevmonthsdata"
    )

    where_clauses = [
        "r.reading_date_db >= %s",
        "r.reading_date_db < %s"
    ]
    params = [start_date, end_date]

    # Optional filters (unchanged behavior)
    mapping = {
        "bl_agnc_name": "r.bl_agnc_name",
        "ofc_discom": "r.ofc_discom",
        "ofc_zone": "r.ofc_zone",
        "ofc_circle": "r.ofc_circle",
        "ofc_division": "r.ofc_division",
        "ofc_subdivision": "r.ofc_subdivision",
    }

    for key, column in mapping.items():
        if key in filters:
            where_clauses.append(f"{column} = %s")
            params.append(filters[key])

    where_sql = " WHERE " + " AND ".join(where_clauses)

    query = f"""
        SELECT
            r.mr_id, r.ofc_discom, r.ofc_zone, r.ofc_circle,
            r.ofc_division, r.ofc_subdivision,
            COUNT(*) AS billed_consumers,
            COUNT(*) FILTER (WHERE prsnt_mtr_status='Ok') AS ok_readings,
            COUNT(*) FILTER (WHERE rdng_ocr_status='Passed') AS OCRwithoutException,
            COUNT(*) FILTER (WHERE rdng_ocr_status='Failed') AS OCRwithException,
            COUNT(*) FILTER (WHERE prsnt_mtr_status='Meter Defective') AS MeterDefective,
            COUNT(*) FILTER (WHERE prsnt_mtr_status='Door Locked') AS DoorLocked,
            r.bl_agnc_name,
            COUNT(*) FILTER (WHERE rdng_ocr_status='Failed' AND qc_rmrk='MR Fault') AS mrFault,
            COUNT(*) FILTER (WHERE prsnt_rdng_ocr_excep='Image blur' AND rdng_ocr_status='Failed') AS imageBlur,
            COUNT(*) FILTER (WHERE prsnt_rdng_ocr_excep='Incorrect Reading' AND rdng_ocr_status='Failed') AS incorrectReading,
            COUNT(*) FILTER (
                WHERE rdng_ocr_status='Failed'
                AND (prsnt_rdng_ocr_excep='Meter Dirty' OR prsnt_rdng_ocr_excep='')
            ) AS meterDirty,
            COUNT(*) FILTER (WHERE prsnt_rdng_ocr_excep='No Exception Found' AND rdng_ocr_status='Failed') AS noExcepFound,
            COUNT(*) FILTER (WHERE prsnt_rdng_ocr_excep='Spoofed Image' AND rdng_ocr_status='Failed') AS spoofedImage,
            COUNT(*) FILTER (WHERE prsnt_rdng_ocr_excep='Invalid' AND rdng_ocr_status='Failed') AS invalidImage
        FROM {tablename} r
        {where_sql}
        GROUP BY
            r.mr_id, r.ofc_discom, r.ofc_zone, r.ofc_circle,
            r.ofc_division, r.ofc_subdivision, r.bl_agnc_name;
    """

    cursor = connection.cursor()
    print(
        "\nFINAL SQL QUERY:\n",
        cursor.mogrify(query, params).decode("utf-8")
    )
    cursor.execute(query, params)
    results = cursor.fetchall()
    cursor.close()

    response = []
    for row in results:
        total = row[6]
        ok = row[7]

        response.append({
            "mrid": row[0],
            "ofc_discom": row[1],
            "ofc_zone": row[2],
            "ofc_circle": row[3],
            "ofc_division": row[4],
            "ofc_subdivision": row[5],
            "billed_consumers": total,
            "meterdefective": row[10],
            "doorlocked": row[11],
            "agency": row[12],
            "OKreadings": ok,
            "OCRReadings": row[8] + row[13],
            "OCRwithException": row[9],
            "OKreadingspercent": round((ok / total) * 100, 2) if total else 0,
            "OCRReadingspercent": round((row[8] / ok) * 100, 2) if ok else 0,
            "OCRwithExceptionpercent": round((row[9] / ok) * 100, 2) if ok else 0,
            "MeterDefectivePercent": round((row[10] / total) * 100, 2) if total else 0,
            "DoorLockedOercent": round((row[11] / total) * 100, 2) if total else 0,
            "mrFault": row[13],
            "imageBlur": row[14],
            "incorrectReading": row[15],
            "meterDirty": row[16],
            "noExcepFound": row[17],
            "spoofedImage": row[18],
            "invalidImage": row[19],
        })

    return Response(response)


# # Sanjeev
# @api_view(["POST"])
# def newmonthdataa(request):
#     today = date.today()
#     this_month = today.strftime("%Y-%m")
#     year = this_month.split("-")[0]
#     month = this_month.split("-")[1]

#     new = []

#     def listfun(dict):
#         new.append(dict.copy())
#         return new

#     newdict = {}
#     clause = ""
#     filters = request.data.get("filters", None)
#     try:
#         clause += "WHERE "
#         if "month" not in filters:
#             clause += (
#                 f"extract(month from reading_date_db) = '{month}' AND extract(year from reading_date_db) = '{year}'"
#             )
#         else:
#             selected_month = filters["month"]
#             selected_year = selected_month.split("-")[0]
#             selected_month_num = selected_month.split("-")[1]
#             clause += (
#                 f"extract(month from reading_date_db) = '{selected_month_num}' AND extract(year from reading_date_db) = '{selected_year}'"
#             )

#         clause += (
#             f" AND bl_agnc_name='{filters['bl_agnc_name']}'"
#             if "bl_agnc_name" in filters
#             else ""
#         )
#         clause += (
#             f" AND extract(day from reading_date_db) BETWEEN '{filters['startdate']}' AND '{filters['enddate']}'"
#             if "enddate" in filters and "startdate" in filters
#             else ""
#         )

#         clause += (
#             f" AND ofc_discom='{filters['ofc_discom']}'"
#             if "ofc_discom" in filters
#             else ""
#         )
#         clause += (
#             f" AND ofc_zone='{filters['ofc_zone']}'"
#             if "ofc_zone" in filters
#             else ""
#         )
#         clause += (
#             f" AND ofc_circle='{filters['ofc_circle']}'"
#             if "ofc_circle" in filters
#             else ""
#         )
#         clause += (
#             f" AND ofc_division='{filters['ofc_division']}'"
#             if "ofc_division" in filters
#             else ""
#         )
#         clause += (
#             f" AND ofc_subdivision='{filters['ofc_subdivision']}'"
#             if "ofc_subdivision" in filters
#             else ""
#         )
#     except:
#         pass

#     # Determine the table name based on the selected month
#     current_month = datetime.now().strftime("%Y-%m")
#     previous_month = (
#         datetime.now() - timedelta(days=datetime.now().day)).strftime("%Y-%m")
#     selected_month = filters.get("month", current_month)
#     tablename = "readingmaster" if selected_month in {
#         current_month, previous_month} else "prevmonthsdata"

#     cursor = connection.cursor()
#     query = f"""
#        SELECT r.mr_id, r.ofc_discom, r.ofc_zone, r.ofc_circle, r.ofc_division, r.ofc_subdivision,
#                COUNT(*) AS billed_consumers,
#                COUNT(CASE WHEN prsnt_mtr_status='Ok' THEN 1 END) AS ok_readings,
#                COUNT(CASE WHEN rdng_ocr_status='Passed' THEN 1 END) AS OCRwithoutException,
#                COUNT(CASE WHEN rdng_ocr_status='Failed' THEN 1 END) AS OCRwithException,
#                COUNT(CASE WHEN prsnt_mtr_status='Meter Defective' THEN 1 END) AS MeterDefective,
#                COUNT(CASE WHEN prsnt_mtr_status='Door Locked' THEN 1 END) AS DoorLocked,
#                r.bl_agnc_name,
#                COUNT(CASE WHEN prsnt_mtr_status = 'Ok' THEN 1 END)
#                - (COUNT(CASE WHEN rdng_ocr_status = 'Passed' THEN 1 END)
#                + COUNT(CASE WHEN rdng_ocr_status = 'Failed' THEN 1 END)) AS diff,
#                SUM(CASE WHEN rdng_ocr_status = 'Failed' and qc_rmrk='MR Fault' THEN 1 ELSE 0 END) AS mrFault,
#                SUM(CASE WHEN prsnt_rdng_ocr_excep = 'Image blur' and rdng_ocr_status='Failed' THEN 1 ELSE 0 END) AS imageBlur,
#             SUM(CASE WHEN prsnt_rdng_ocr_excep = 'Incorrect Reading' and rdng_ocr_status='Failed' THEN 1 ELSE 0 END) AS incorrectReading,
#             SUM(CASE
#         WHEN (prsnt_rdng_ocr_excep = 'Meter Dirty' and rdng_ocr_status='Failed')
#              OR (rdng_ocr_status = 'Failed' AND prsnt_rdng_ocr_excep = '')
#         THEN 1
#         ELSE 0
#     END) AS meterDirty,
#             SUM(CASE WHEN prsnt_rdng_ocr_excep = 'No Exception Found'  AND rdng_ocr_status = 'Failed'  THEN 1 ELSE 0 END) AS noExcepFound,
#             SUM(CASE WHEN prsnt_rdng_ocr_excep = 'Spoofed Image'  AND rdng_ocr_status = 'Failed' THEN 1 ELSE 0 END) AS spoofedImage,
#             SUM(CASE WHEN prsnt_rdng_ocr_excep = 'Invalid'  AND rdng_ocr_status = 'Failed'  THEN 1 ELSE 0 END) AS invalidImage
#         FROM {tablename} r
#         {clause}
#         GROUP BY r.mr_id, r.ofc_discom, r.ofc_zone, r.ofc_circle, r.ofc_division, r.ofc_subdivision, r.bl_agnc_name;
#     """

#     print("QUERY------>", query)
#     cursor.execute(query)
#     results = cursor.fetchall()

#     try:
#         for row in results:
#             total = row[6]
#             okreadings = row[7]
#             ocrreadings = row[8]
#             ocrwithexcep = row[9]
#             meterdefective = row[10]
#             doorlocked = row[11]

#             # Percentage
#             okreadpercent = round((okreadings / total)
#                                   * 100, 2) if total else 0
#             ocrreadingpercent = round(
#                 (ocrreadings / okreadings) * 100, 2) if okreadings else 0
#             ocrwithexceppercent = round(
#                 (ocrwithexcep / okreadings) * 100, 2) if okreadings else 0
#             meterdefectivepercent = round(
#                 (meterdefective / total) * 100, 2) if total else 0
#             doorlockedpercent = round(
#                 (doorlocked / total) * 100, 2) if total else 0

#             # Add to dictionary
#             newdict["mrid"] = row[0]
#             newdict["ofc_discom"] = row[1]
#             newdict["ofc_zone"] = row[2]
#             newdict["ofc_circle"] = row[3]
#             newdict["ofc_division"] = row[4]
#             newdict["ofc_subdivision"] = row[5]
#             newdict["billed_consumers"] = row[6]
#             newdict["meterdefective"] = row[10]
#             newdict["doorlocked"] = row[11]
#             newdict["agency"] = row[12]
#             newdict["OKreadings"] = okreadings
#             newdict["OCRReadings"] = ocrreadings + row[13]
#             newdict["OCRwithException"] = ocrwithexcep
#             newdict["OKreadingspercent"] = okreadpercent
#             newdict["OCRReadingspercent"] = ocrreadingpercent
#             newdict["OCRwithExceptionpercent"] = ocrwithexceppercent
#             newdict["MeterDefectivePercent"] = meterdefectivepercent
#             newdict["DoorLockedOercent"] = doorlockedpercent
#             newdict["mrFault"] = row[14]
#             newdict["imageBlur"] = row[15]
#             newdict["incorrectReading"] = row[16]
#             newdict["meterDirty"] = row[17]
#             newdict["noExcepFound"] = row[18]
#             newdict["spoofedImage"] = row[19]
#             newdict["invalidImage"] = row[20]
#             # newdict["blanks"] = row[21]

#             # Add to list
#             newdata = listfun(newdict)

#         return Response(newdata)

#     except Exception as e:
#         print(e)
#         return Response([])


@api_view(["POST"])
def newdailydata(request):
    todaydate = date.today()

    new = []

    def listfun(dict):
        new.append(dict.copy())
        return new

    newdict = {}
    clause = ""
    filters = request.data.get("filters", None)
    try:
        if filters:
            clause += "AND "
            for i, (key, value) in enumerate(filters.items()):
                if key == "bl_agnc_name":
                    clause += f"bl_agnc_name='{filters['bl_agnc_name']}'"
                if key == "ofc_discom":
                    clause += f"ofc_discom='{filters['ofc_discom']}'"

    except:
        pass

    # Determine the table name based on the current month
    current_month = datetime.now().strftime("%Y-%m")
    year, month = current_month.split("-")
    tablename = "readingmaster" if month in [
        month, str(int(month) - 1)] else "prevmonthsdata"

    # count(r.prsnt_mtr_status='Meter Defective' or r.prsnt_mtr_status='Door Locked' or r.prsnt_mtr_status='Ok' or null) as billed_consumers,
    cursor = connection.cursor()
    query = f"""
        SELECT r.mr_id, r.ofc_discom, r.ofc_zone, r.ofc_circle, r.ofc_division, r.ofc_subdivision,
               COUNT(*) AS billed_consumers,
               COUNT(CASE WHEN prsnt_mtr_status='Ok' OR prsnt_mtr_status='Door Locked' THEN 1 END) AS ok_readings,
               COUNT(CASE WHEN rdng_ocr_status='Passed' THEN 1 END) AS OCRwithoutException,
               COUNT(CASE WHEN rdng_ocr_status='Failed' THEN 1 END) AS OCRwithException,
               r.bl_agnc_name,
               COUNT(CASE WHEN prsnt_mtr_status = 'Ok' THEN 1 END)
               - (COUNT(CASE WHEN rdng_ocr_status = 'Passed' THEN 1 END)
               + COUNT(CASE WHEN rdng_ocr_status = 'Failed' THEN 1 END)) AS diff,
               r.ai_mdl_ver
        FROM {tablename} r
        WHERE reading_date_db = '{todaydate}' {clause}
        GROUP BY r.mr_id, r.ofc_discom, r.ofc_zone, r.ofc_circle, r.ofc_division, r.ofc_subdivision, r.bl_agnc_name, r.ai_mdl_ver;
    """
    print("QUERY------>", query)
    cursor.execute(query)
    results = cursor.fetchall()

    try:
        for row in results:
            total = row[6]
            okreadings = row[7]
            ocrreadings = row[8] + row[11]
            ocrwithexcep = row[9]
            aimdlver = row[12]

            # Percentage
            okreadpercent = round((okreadings / total)
                                  * 100, 2) if total else 0
            ocrreadingpercent = round(
                (ocrreadings / okreadings) * 100, 2) if okreadings else 0
            ocrwithexceppercent = round(
                (ocrwithexcep / okreadings) * 100, 2) if okreadings else 0

            # Add to dictionary
            newdict["mrid"] = row[0]
            newdict["ofc_discom"] = row[1]
            newdict["ofc_zone"] = row[2]
            newdict["ofc_circle"] = row[3]
            newdict["ofc_division"] = row[4]
            newdict["ofc_subdivision"] = row[5]
            newdict["billed_consumers"] = row[6]
            newdict["agency"] = row[10]
            newdict["OKreadings"] = okreadings
            newdict["OCRReadings"] = ocrreadings
            newdict["OCRwithException"] = ocrwithexcep
            newdict["OKreadingspercent"] = okreadpercent
            newdict["OCRReadingspercent"] = ocrreadingpercent
            newdict["OCRwithExceptionpercent"] = ocrwithexceppercent
            newdict["ai_mdl_ver"] = aimdlver

            # Add to list
            newdata = listfun(newdict)

        return Response(newdata)

    except Exception as e:
        print(e)
        return Response([])


# @api_view(['POST'])
# def newdailybilling(request):

#     todaydate = date.today()
#     month = date.today().month
#     new = []
#     bill_month='2023-06-01'

#     def listfun(dict):
#         new.append(dict.copy())
#         return new
#     newdict = {}
#     clause = ''
#     filters = request.data.get("filters", None)
#     try:
#         if filters:
#             clause += 'AND '
#             for i, (key, value) in enumerate(filters.items()):
#                 if key == 'bl_agnc_name':
#                     clause += f"bl_agnc_name='{filters['bl_agnc_name']}'"

#     except:
#         pass
#     cursor = connection.cursor()
# #     query = f"""select o.zone as Zone, o.divisionname as Division, o.subdivision, o.no_of_consumers,
# # (count(distinct case when m.reading_date_db='{todaydate}' then m.id end)) as total_billed_today,
# # (count(distinct case when bill_month_dt='{bill_month}'then m.cons_ac_no end)) as total_billed_this_month
# # from office_consumers o
# # left join readingmaster m on o.zone=m.ofc_zone and o.divisionname=m.ofc_division
# # and o.subdivision=m.ofc_subdivision and bill_month_dt='{bill_month}'  {clause}
# # group by o.zone,o.divisionname,o.subdivision,o.no_of_consumers
# # order by total_billed_today desc
# # """
#     query=f"""
#     SELECT
#     oc.agency as agency,
#     oc.officediv AS division,
#     (oc.officetotal::integer) as total_Consumers,
#     COUNT(*) AS Billed_consumers_till_date,
#     ROUND((COUNT(*)/oc.officetotal::numeric) * 100, 2) AS billing_percentage
#     FROM (
#     SELECT divisionname AS officediv, no_of_consumers::numeric AS officetotal,agency as agency
#     FROM office_consumers
#     GROUP BY divisionname, no_of_consumers,agency
#     ) oc
#     JOIN readingmaster rm ON oc.officediv = rm.ofc_division
#     WHERE EXTRACT(MONTH FROM rm.reading_date_db) = {month} {clause}
#     GROUP BY oc.officediv, oc.officetotal,oc.agency;
#     """
#     print("QUERY------>", query)
#     cursor.execute(query)
#     # results = cursor.fetchall()
#     results = dictfetchall(cursor)
#     return Response(results)


@api_view(["POST"])
def newdailybilling(request):
    todaydate = date.today()
    month = date.today().month
    new = []
    bill_month = "2023-06-01"

    def listfun(dict):
        new.append(dict.copy())
        return new

    newdict = {}
    clause = ""
    agnc = ""
    filters = request.data.get("filters", None)
    try:
        if filters:
            clause += "AND "
            for i, (key, value) in enumerate(filters.items()):
                if key == "bl_agnc_name":
                    clause += f"bl_agnc_name='{filters['bl_agnc_name']}'"
                    agnc = f"where agency='{filters['bl_agnc_name']}'"
                if key == "ofc_discom":
                    clause += f"ofc_discom='{filters['ofc_discom']}'"

    except:
        pass
    cursor = connection.cursor()
    query = f"""SELECT
    coalesce(oc.agency, rm.agency, daily.agency) AS agency,
    coalesce(oc.zone, rm.zone, daily.zone) AS zone,
    coalesce(oc.division, rm.division, daily.division) AS division,
    coalesce(oc.no_of_consumers, 0) AS no_of_consumers,
    coalesce(rm.billed_consumers, 0) AS "total_billed_this_month",
    coalesce(daily.billed_daily, 0) AS "total_billed_today",
    CASE
        WHEN oc.no_of_consumers = 0 THEN 0
        ELSE round((rm.billed_consumers::numeric / oc.no_of_consumers::numeric) * 100, 2)
    END AS percentage
FROM
    (
        SELECT
            agency,
            zone,
            divisionname AS division,
            no_of_consumers::numeric
        FROM
            office_consumers
            {agnc}
        GROUP BY
            divisionname,
            agency,
            zone,
            division,
            no_of_consumers
    ) AS oc
FULL JOIN
    (
        SELECT
            bl_agnc_name AS agency,
            ofc_zone AS zone,
            ofc_division AS division,
            COUNT(*) AS billed_consumers
        FROM
            readingmaster
        WHERE
            bill_month_dt='{bill_month}'
            AND ofc_zone != '' {clause}
        GROUP BY
            ofc_division,
            bl_agnc_name,
            ofc_zone
    ) AS rm ON oc.agency = rm.agency
    AND oc.zone = rm.zone
    AND oc.division = rm.division
FULL JOIN
    (
        SELECT
            bl_agnc_name AS agency,
            ofc_zone AS zone,
            ofc_division AS division,
            COUNT(*) AS billed_daily
        FROM
            readingmaster
        WHERE
            reading_date_db = '{todaydate}'
            AND ofc_zone != '' {clause}
        GROUP BY
            ofc_division,
            bl_agnc_name,
            ofc_zone
    ) AS daily ON oc.agency = daily.agency
    AND oc.zone = daily.zone
    AND oc.division = daily.division

ORDER BY
    division, agency, zone;

"""
    print("QUERY------>", query)
    cursor.execute(query)
    # results = cursor.fetchall()
    results = dictfetchall(cursor)
    return Response(results)


# ghulam
@api_view(["POST"])
def qccheckmobile(request):

    super_number = request.data.get("supervisor_number")

    # Fetch all rows of this supervisor
    qs = SupervisorLogin.objects.filter(supervisor_number=super_number)

    if not qs.exists():
        return Response({"status": False, "message": "Supervisor not found"})

    # Get all MR IDs
    mr_ids = list(qs.values_list("mr_id", flat=True))

    # Convert list to SQL-safe list: '2431MRC148','2431MRC149',...
    mr_sql_list = ",".join(f"'{x}'" for x in mr_ids)

    # Take office details from the first row
    obj = qs.first()
    ofc_division = obj.ofc_division
    ofc_subdivision = obj.ofc_subdivision

    start_date = request.data.get("start_date")
    end_date = request.data.get("end_date")

    cursor = connection.cursor()

    # --- QUERY 1: Date range summary ---
    query = f"""
        SELECT
            COUNT(r.id) AS totalreadings,
            COUNT(CASE WHEN r.rdng_ocr_status = 'Passed' THEN 1 END) AS totalpassed,
            COUNT(CASE WHEN r.rdng_ocr_status = 'Failed' THEN 1 END) AS totalfailed
        FROM readingmaster r
        WHERE
            r.prsnt_mtr_status = 'Ok'
            AND r.ofc_division = '{ofc_division}'
            AND r.ofc_subdivision = '{ofc_subdivision}'
            AND r.mr_id IN ({mr_sql_list})
            AND r.reading_date_db BETWEEN '{start_date}' AND '{end_date}'
    """

    cursor.execute(query)
    date_data = dictfetchall(cursor)
    print("QC Summary Query:", query)

    # --- QUERY 2: Today summary ---
    query2 = f"""
        SELECT
            COUNT(r.id) AS todaystotalreadings,
            COUNT(CASE WHEN r.rdng_ocr_status = 'Passed' THEN 1 END) AS todayspassed,
            COUNT(CASE WHEN r.rdng_ocr_status = 'Failed' THEN 1 END) AS todaysfailed
        FROM readingmaster r
        WHERE
            r.prsnt_mtr_status = 'Ok'
            AND r.ofc_division = '{ofc_division}'
            AND r.ofc_subdivision = '{ofc_subdivision}'
            AND r.mr_id IN ({mr_sql_list})
            AND r.reading_date_db = CURRENT_DATE
    """

    cursor.execute(query2)
    today_data = dictfetchall(cursor)
    print("QC Today Query:", query2)

    return Response({
        **date_data[0],   # total/passed/failed for range
        **today_data[0],  # today’s totals
    })


@api_view(["POST"])
def androidclusterstestnew(request):
    filters = request.data.get("filters", {})
    today = date.today()

    where_clauses = ["reading_date_db = %s"]
    params = [str(today)]

    # Dynamic filters
    if "mr_id" in filters:
        where_clauses.append("mr_id = %s")
        params.append(filters["mr_id"])

    if "bl_agnc_name" in filters:
        where_clauses.append("bl_agnc_name = %s")
        params.append(filters["bl_agnc_name"])

    where_sql = "WHERE " + " AND ".join(where_clauses)

    query = f"""
        SELECT DISTINCT ON (mr_id)
            mr_id, rdng_date, cons_name, geo_lat, geo_long,
            prsnt_mtr_status, rdng_ocr_status, rdng_img
        FROM readingmaster
        {where_sql}
        ORDER BY mr_id,
                 (geo_lat IS NULL OR geo_long IS NULL),  -- Prefer NOT NULL
                 rdng_date DESC                          -- Latest record
    """

    cursor = connection.cursor()
    cursor.execute(query, params)
    print("query:>", query)
    result = dictfetchall(cursor)

    return Response(result)


@api_view(["POST"])
def qcmobiledashboard(request):

    super_number = request.data.get("supervisor_number")

    # Fetch all rows for this supervisor
    qs = SupervisorLogin.objects.filter(supervisor_number=super_number)

    # Get all MR IDs
    mr_ids = list(qs.values_list("mr_id", flat=True))

    # Get first row fields
    obj = qs.first()
    if not obj:
        return Response({"status": False, "message": "Supervisor not found"})

    ofc_division = obj.ofc_division
    ofc_subdivision = obj.ofc_subdivision
    supervisor_name = obj.supervisor_name
    # designation = obj.designation.lower()

    pagesize = int(request.data.get("pagesize", 10))
    page = int(request.data.get("page", 1))
    offset = (pagesize * page) - pagesize

    start_date = request.data.get("start_date")
    end_date = request.data.get("end_date")

    if not start_date or not end_date:
        return Response({"error": "start_date and end_date are required"}, status=400)

    # Supervisor only
    # if designation != "supervisor":
    #     return Response({"status": False, "message": "You are not a Supervisor"})

    # Convert MR IDs to SQL list (1,2,3,...)
    mr_id_sql_list = ",".join(f"'{i}'" for i in mr_ids)

    cursor = connection.cursor()

    query = f"""
        SELECT *
        FROM (
            SELECT
                mr_id,

                -- MOST RECENT GEO
                (
                    SELECT geo_lat
                    FROM readingmaster rm2
                    WHERE rm2.mr_id = readingmaster.mr_id
                    AND rm2.reading_date_db BETWEEN '{start_date}' AND '{end_date}'
                    AND rm2.geo_lat IS NOT NULL
                    AND rm2.geo_lat <> ''
                    ORDER BY rm2.reading_date_db DESC
                    LIMIT 1
                ) AS geo_lat,
                --last active time
                (
                    SELECT
                        SPLIT_PART(rm2.rdng_date, ' ', 2)   -- extract time part
                    FROM readingmaster rm2
                    WHERE rm2.mr_id = readingmaster.mr_id
                    AND rm2.reading_date_db BETWEEN '{start_date}' AND '{end_date}'
                    AND rm2.rdng_date IS NOT NULL
                    AND rm2.rdng_date <> ''
                    ORDER BY rm2.rdng_date::timestamp DESC
                    LIMIT 1
                ) AS latest_time,
                (
                    SELECT geo_long
                    FROM readingmaster rm2
                    WHERE rm2.mr_id = readingmaster.mr_id
                    AND rm2.reading_date_db BETWEEN '{start_date}' AND '{end_date}'
                    AND rm2.geo_long IS NOT NULL
                    AND rm2.geo_long <> ''
                    ORDER BY rm2.reading_date_db DESC
                    LIMIT 1
                ) AS geo_long,

                CASE
                    WHEN COUNT(CASE WHEN reading_date_db = CURRENT_DATE THEN 1 END) > 0
                    THEN 'Active'
                    ELSE 'Inactive'
                END AS status,

                CASE
                    WHEN COUNT(CASE WHEN prsnt_mtr_status = 'Ok' THEN 1 END) = 0
                    THEN 0
                    ELSE ROUND(
                        (
                            COUNT(CASE WHEN rdng_ocr_status = 'Passed' THEN 1 END)::float /
                            NULLIF(COUNT(CASE WHEN prsnt_mtr_status = 'Ok' THEN 1 END), 0)
                        )::numeric * 100
                    , 2)
                END AS passed_percent,

                ROUND(
                    (
                        COUNT(CASE WHEN prsnt_mtr_status = 'Meter Defective' THEN 1 END)::float /
                        NULLIF(COUNT(mr_id)::float, 0)
                    )::numeric * 100
                , 2) AS meter_defective_percent,

                ROUND(
                    (
                        COUNT(CASE WHEN prsnt_mtr_status = 'Door Locked' THEN 1 END)::float /
                        NULLIF(COUNT(mr_id)::float, 0)
                    )::numeric * 100
                , 2) AS door_locked_percent,

                COUNT(CASE WHEN prsnt_mtr_status = 'Ok' THEN 1 END) AS mr_total_readings,

                COUNT(CASE WHEN rdng_ocr_status = 'Passed' THEN 1 END) AS totalpassed,
                COUNT(CASE WHEN rdng_ocr_status = 'Failed' THEN 1 END) AS totalfailed,

                COUNT(CASE WHEN qc_req = 'Yes' THEN 1 END) AS mr_qc_remaining,
                COUNT(CASE WHEN qc_req = 'No' THEN 1 END) AS mr_qc_done,

                COUNT(*) OVER() AS mr_count

            FROM readingmaster
            WHERE
                reading_date_db BETWEEN '{start_date}' AND '{end_date}'
                AND ofc_division = '{ofc_division}'
                AND ofc_subdivision = '{ofc_subdivision}'
                AND mr_id IN ({mr_id_sql_list})

            GROUP BY mr_id
        ) AS sub

        ORDER BY passed_percent ASC, meter_defective_percent ASC
        LIMIT {pagesize} OFFSET {offset};
    """

    cursor.execute(query)
    print("Query:-->", query)
    result = dictfetchall(cursor)

    return Response({
        "status": True,
        "message": f"{pagesize} data fetched successfully",
        "user_name": supervisor_name,
        "division": ofc_division,
        "subdivision": ofc_subdivision,
        "mr_data": result,
    })


# def qcmobiledashboard(request):
#     data = UserManagement.objects.filter(email=request.data["email"])
#     pagesize = request.data.get("pagesize")
#     page = request.data.get("page")
#     offset = (int(pagesize) * int(page)) - int(pagesize)

#     serializer = UserManagementSerializer(data, many=True)
#     my_dict = serializer.data[0]

#     ofc_division = my_dict.get("ofc_division")
#     ofc_zone = my_dict.get("ofc_zone")
#     ofc_circle = my_dict.get("ofc_circle")
#     full_name = my_dict.get("full_name")
#     designation = my_dict.get("designation", "").lower()

#     start_date = request.data.get("start_date")
#     end_date = request.data.get("end_date")

#     if not start_date or not end_date:
#         return Response({"error": "start_date and end_date are required"}, status=400)

#     # Supervisor Only
#     if designation == "supervisor":

#         cursor = connection.cursor()

#         query = f"""
#             SELECT *
#             FROM (
#                 SELECT
#                     mr_id,

#                     -- ACTIVE status based on TODAY'S activity
#                     CASE
#                         WHEN COUNT(CASE WHEN reading_date_db = CURRENT_DATE THEN 1 END) > 0
#                         THEN 'Active'
#                         ELSE 'Inactive'
#                     END AS status,

#                     -- % Passed
#                     CASE
#                         WHEN COUNT(CASE WHEN prsnt_mtr_status = 'Ok' THEN 1 END) = 0
#                         THEN 0
#                         ELSE ROUND(
#                             (
#                                 COUNT(CASE WHEN rdng_ocr_status = 'Passed' THEN 1 END)::float
#                                 /
#                                 NULLIF(COUNT(CASE WHEN prsnt_mtr_status = 'Ok' THEN 1 END), 0)
#                             )::numeric * 100
#                         , 2)
#                     END AS passed_percent,

#                     -- Defective %
#                     ROUND(
#                         (
#                             COUNT(CASE WHEN prsnt_mtr_status = 'Meter Defective' THEN 1 END)::float
#                             /
#                             NULLIF(COUNT(mr_id)::float, 0)
#                         )::numeric * 100
#                     , 2) AS meter_defective_percent,

#                     -- Door Locked %
#                     ROUND(
#                         (
#                             COUNT(CASE WHEN prsnt_mtr_status = 'Door Locked' THEN 1 END)::float
#                             /
#                             NULLIF(COUNT(mr_id)::float, 0)
#                         )::numeric * 100
#                     , 2) AS door_locked_percent,

#                     COUNT(*) AS mr_total_readings,

#                     -- Total Passed & Failed
#                     COUNT(CASE WHEN rdng_ocr_status = 'Passed' THEN 1 END) AS totalpassed,
#                     COUNT(CASE WHEN rdng_ocr_status = 'Failed' THEN 1 END) AS totalfailed,

#                     -- QC
#                     COUNT(CASE WHEN qc_req = 'Yes' THEN 1 END) AS mr_qc_remaining,
#                     COUNT(CASE WHEN qc_req = 'No' THEN 1 END) AS mr_qc_done,

#                     COUNT(*) OVER() AS mr_count

#                 FROM readingmaster

#                 WHERE
#                     reading_date_db BETWEEN '{start_date}' AND '{end_date}'
#                     AND ofc_zone = '{ofc_zone}'
#                     AND ofc_circle = '{ofc_circle}'
#                     AND ofc_division = '{ofc_division}'

#                 GROUP BY mr_id
#             ) AS sub

#             ORDER BY passed_percent ASC, meter_defective_percent ASC
#             LIMIT {pagesize} OFFSET {offset};
#         """

#         print(query)
#         cursor.execute(query)
#         person_objects = dictfetchall(cursor)

#         return Response({
#             "status": True,
#             "message": f"{pagesize} data fetched successfully",
#             "user_name": full_name,
#             "division": ofc_division,
#             "mr_data": person_objects,
#         })

#     return Response({"status": False, "message": "You are not a Supervisor"})


@api_view(["GET"])
def downloadexcel(request):

    supervisor_number = request.query_params.get("supervisor_number")

    # Fetch supervisor rows
    qs = SupervisorLogin.objects.filter(supervisor_number=supervisor_number)

    if not qs.exists():
        return Response({"status": False, "message": "Supervisor not found"})

    # All MR IDs
    mr_ids = list(qs.values_list("mr_id", flat=True))

    # Convert → 'MRC1','MRC2',...
    mr_sql_list = ",".join(f"'{x}'" for x in mr_ids)

    # Supervisor info
    sup = qs.first()
    ofc_division = sup.ofc_division
    ofc_subdivision = sup.ofc_subdivision

    datewise = request.query_params.get("datewise")
    today = datetime.now().date()
    month = datetime.now().month

    if datewise == "date1":
        clause = f"reading_date_db = '{today}'"
    else:
        clause = f"extract(Month from reading_date_db) = '{month}'"

    cursor = connection.cursor()

    # ----- SUMMARY -----
    query = f"""
        SELECT 
            COUNT(r.id) AS total_readings,
            COUNT(r.qc_req='Yes' OR NULL) AS qc_remaining,
            COUNT(r.qc_req='No' OR NULL) AS qc_done
        FROM readingmaster r
        WHERE 
            {clause}
            AND r.mr_id IN ({mr_sql_list})
    """

    print("query:", query)
    cursor.execute(query)
    summary = dictfetchall(cursor)

    # ----- MR WISE DATA -----
    query2 = f"""
        SELECT 
            DISTINCT(mr_id),
            CASE WHEN COUNT(prsnt_mtr_status='Ok' OR NULL) = 0 THEN 0
            ELSE ROUND(
                (
                    CAST(COUNT(rdng_ocr_status='Passed' OR NULL) AS FLOAT) /
                    CAST(COUNT(prsnt_mtr_status='Ok' OR NULL) AS FLOAT) * 100
                )::numeric
            , 2) END AS passed_percent,

            ROUND(
                (
                    CAST(COUNT(prsnt_mtr_status='Meter Defective' OR NULL) AS FLOAT) /
                    COALESCE(CAST(COUNT(mr_id) AS FLOAT),1) * 100
                )::numeric
            , 2) AS meter_defective_percent,

            ROUND(
                (
                    CAST(COUNT(prsnt_mtr_status='Door Locked' OR NULL) AS FLOAT) /
                    COALESCE(CAST(COUNT(mr_id) AS FLOAT),1) * 100
                )::numeric
            , 2) AS door_locked_percent,

            COUNT(id) AS mr_total_readings,
            COUNT(qc_req='Yes' OR NULL) AS mr_qc_remaining,
            COUNT(qc_req='No' OR NULL) AS mr_qc_done

        FROM readingmaster
        WHERE 
            {clause}
            AND mr_id IN ({mr_sql_list})
        GROUP BY mr_id
    """

    print("query2:", query2)
    cursor.execute(query2)
    rows = dictfetchall(cursor)

    # ---------- EXCEL ----------
    wb = Workbook()
    ws = wb.active

    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for item in rows:
            ws.append(list(item.values()))

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=mydata.xlsx"
    wb.save(response)
    return response


@api_view(["POST"])
def mobilemvcards(request):

    super_number = request.data.get("supervisor_number")

    # Fetch all supervisor rows
    qs = SupervisorLogin.objects.filter(supervisor_number=super_number)

    if not qs.exists():
        return Response({"status": False, "message": "Supervisor not found"})

    # All MR IDs
    mr_ids = list(qs.values_list("mr_id", flat=True))

    # Convert list → 'MRC148','MRC149',...
    mr_sql_list = ",".join(f"'{x}'" for x in mr_ids)

    # Supervisor details
    obj = qs.first()
    full_name = obj.supervisor_name
    ofc_division = obj.ofc_division

    pagesize = int(request.data.get("pagesize", 10))
    page = int(request.data.get("page", 1))
    searchdata = request.data.get("searchdata")
    offset = (pagesize * page) - pagesize

    now = datetime.now()
    month = now.month

    # ---- Dynamic WHERE clause ----
    clause = "WHERE "
    conditions = []

    # Filter by MR IDs
    conditions.append(f"m.mr_id IN ({mr_sql_list})")

    # Month filter
    conditions.append(f"extract(Month from reading_date_db) = '{month}'")

    # Search filter
    if searchdata:
        conditions.append(f"m.mr_id ILIKE '%{searchdata}%'")

    # Join all with AND
    clause += " AND ".join(conditions)

    cursor = connection.cursor()

    query = f"""
        SELECT 
            m.mr_id AS "mrId",
            m.ofc_division,
            m.rdng_date,
            m.prsnt_mtr_status,
            m.prsnt_ocr_rdng,
            m.qc_req,
            m.prsnt_rdng,
            m.ocr_pf_reading,
            m.cons_name,
            m.cons_ac_no,
            m.prsnt_md_rdng_ocr,
            m.rdng_ocr_status,
            m.rdng_img,
            m.prsnt_md_rdng,
            m.id,
            r."mrPhoto"
        FROM readingmaster m
        LEFT JOIN meterreaderregistration r ON m.mr_id = r."mrId"
        {clause}
        ORDER BY m.rdng_date DESC
        LIMIT {pagesize} OFFSET {offset}
    """

    print("query -->", query)
    cursor.execute(query)
    result = dictfetchall(cursor)

    return Response({
        "status": True,
        "message": f"{pagesize} data fetched successfully",
        "user_name": full_name,
        "division": ofc_division,
        "results": result,
    })

# @api_view(['POST'])
# def meterreaderDetails(request):
#     pagesize = request.data.get("pagesize",)
#     page = (request.data.get("page",))
#     offset = (int(pagesize) * int(page))-int(pagesize)

#     data = request.data.get("filters", None)
#     clause = ''
#     try:
#         if data:
#             clause += 'WHERE '
#             for i, (key, value) in enumerate(data.items()):
#                 if i > 0:
#                     clause += ' AND '

#                 print("data---------", clause, data, i, key, value)

#                 if key=='month':
#                     year = value.split('-')[0]
#                     month = value.split('-')[1]
#                     print("month",month)
#                     clause += f"extract(month from m.reading_date_db) = '{month}' AND extract(year from m.reading_date_db) = '{year}'"

#                 if key == 'startdate':
#                      clause += f"m.reading_date_db BETWEEN '{data['startdate']}'"

#                 if key == 'enddate':
#                     clause += f"'{data['enddate']}'"

#                 if key=='mr_id':
#                     clause += f"mr_id='{data['mr_id']}'"

#                 if key=='searchdata':
#                     clause += f"(mr_id='{data['searchdata']}' or cons_ac_no='{data['searchdata']}' or cons_name='{data['searchdata']}')"
#                 if key=='rdng_ocr_status':
#                     clause +=f"rdng_ocr_status='{data['rdng_ocr_status']}'"
#                 if key=='bl_agnc_name':
#                     clause +=f"bl_agnc_name='{data['bl_agnc_name']}'"

#             print("clause-------", clause)
#             cursor = connection.cursor()
#             query = (f'''
#             select mr_id,cons_ac_no,bl_agnc_name,abnormality,cons_name,con_trf_cat,con_mtr_sl_no,
# mr_rmrk,prsnt_mtr_status,prsnt_rdng,prev_rdng,
# prsnt_md_rdng,prev_md,ocr_pf_reading,prev_pf_rdng,rdng_date,prev_rdng_date,rdng_img,md_img,
# prsnt_rdng_ocr_excep,md_ocr_excep,qc_req,count(*) over () as total_count from readingmaster {clause} order by rdng_date DESC limit {pagesize} offset {offset}
#                 ''')
#             print(clause)
#             print(query)
#             # group by m.{orderby}
#             cursor.execute(query)
#             person_objects = dictfetchall(cursor)
#         return Response({"result": person_objects, "count": person_objects[0]['total_count']})

#     except Exception as e:
#         print(e)
#         return Response({"result": [], "count": 5})


@api_view(["POST"])
def dashboardagencywise(request):
    today = date.today()
    thismonth = today.strftime("%Y-%m")
    year = thismonth.split("-")[0]
    month = thismonth.split("-")[1]
    new = []
    clause = ""
    agnc = ""
    cursor = connection.cursor()
    filters = request.data.get("filters", None)

    def listfun(dict):
        new.append(dict.copy())
        return new

    try:
        if filters:
            clause += " AND "
            for i, (key, value) in enumerate(filters.items()):
                if key == "bl_agnc_name":
                    agnc = f"where agency='{filters['bl_agnc_name']}'"
                    clause += f"bl_agnc_name='{value}'"
                if key == "ofc_discom":
                    clause += f"ofc_discom='{value}'"
    except:
        pass
    newdict = {}
    query = f"""
             select rm.bl_agnc_name,rm.total,rm.ok,rm.passed,rm.failed,rm.md,rm.dl,
oc.noofconsumers
from (
select agency,sum(distinct no_of_consumers::numeric) as noofconsumers from office_consumers {agnc}
group by agency) as oc
full join
(
select bl_agnc_name,count(readingmaster.id) as total,
count(prsnt_mtr_status='Ok' or NULL) as ok,count(rdng_ocr_status='Passed' or NULL) as passed,
count(rdng_ocr_status='Failed' or NULL) as failed,count(prsnt_mtr_status='Meter Defective' or NULL) as md,
count(prsnt_mtr_status='Door Locked' or NULL) as dl
from readingmaster where extract(month from reading_date_db) = '{month}' AND extract(year from reading_date_db) = '{year}' {clause} group by readingmaster.bl_agnc_name
) as rm on oc.agency=rm.bl_agnc_name

    """
    print("query", query)

    try:
        cursor.execute(query)
        result = cursor.fetchall()
        for row in result:
            total = row[1]
            okreadings = row[2]
            ocrreadings = row[3]
            ocrwithexcep = row[4]
            meterdefective = row[5]
            doorlocked = row[6]
            noofconsumers = row[7]
            # Percentage
            okreadpercent = round((okreadings / total) * 100, 2)
            ocrreadingpercent = round(
                ((ocrreadings / okreadings) if okreadings else 0) * 100
            )
            ocrwithexceppercent = round(
                ((ocrwithexcep / okreadings) if okreadings else 0) * 100, 2
            )
            meterdefectivepercent = round((meterdefective / total) * 100, 2)
            doorlockedpercent = round((doorlocked / total) * 100, 2)
            noofconsumerspercent = round(
                ((total / noofconsumers) if noofconsumers else 0) * 100, 2
            )

            # add to dictionary
            newdict["agency"] = row[0]
            newdict["totalReadings"] = total
            newdict["OKreadings"] = okreadings
            newdict["OKreadingspercent"] = okreadpercent
            newdict["OCRReadings"] = ocrreadings
            newdict["OCRReadingspercent"] = ocrreadingpercent
            newdict["OCRwithException"] = ocrwithexcep
            newdict["OCRwithExceptionpercent"] = ocrwithexceppercent
            newdict["MeterDefective"] = meterdefective
            newdict["MeterDefectivepercent"] = meterdefectivepercent
            newdict["DoorLocked"] = doorlocked
            newdict["DoorLockedpercent"] = doorlockedpercent
            newdict["noofconsumers"] = noofconsumers
            newdict["noofconsumerspercent"] = noofconsumerspercent
            # add to list
            newdata = listfun(newdict)
        return Response(newdata)
    except:
        return Response([])


@api_view(["POST"])
def dashboardagencywisenbpdcl(request):
    today = date.today()
    thismonth = today.strftime("%Y-%m")
    year = thismonth.split("-")[0]
    month = thismonth.split("-")[1]
    new = []

    clause = ""
    agnc = ""
    cursor = connection.cursor()
    filters = request.data.get("filters", None)

    def listfun(dict):
        new.append(dict.copy())
        return new

    try:
        if filters:
            clause += " AND "
            for i, (key, value) in enumerate(filters.items()):
                if key == "bl_agnc_name":
                    agnc = f"and agency='{filters['bl_agnc_name']}'"
                    clause += f"bl_agnc_name='{value}'"
                    pass

    except:
        pass
    newdict = {}
    query = f"""
             select rm.bl_agnc_name,rm.total,rm.ok,rm.passed,rm.failed,rm.md,rm.dl,
oc.noofconsumers
from (
select agency,sum(distinct no_of_consumers::numeric) as noofconsumers from office_consumers where discom='NBPDCL' {agnc}
group by agency) as oc
full join
(
select bl_agnc_name,count(readingmaster.id) as total,
count(prsnt_mtr_status='Ok' or NULL) as ok,count(rdng_ocr_status='Passed' or NULL) as passed,
count(rdng_ocr_status='Failed' or NULL) as failed,count(prsnt_mtr_status='Meter Defective' or NULL) as md,
count(prsnt_mtr_status='Door Locked' or NULL) as dl
from readingmaster where ofc_discom='NBPDCL' and extract(month from reading_date_db) = '{month}' AND extract(year from reading_date_db) = '{year}' {clause} group by readingmaster.bl_agnc_name
) as rm on oc.agency=rm.bl_agnc_name

    """

    print("query", query)

    try:
        cursor.execute(query)
        result = cursor.fetchall()
        for row in result:
            total = row[1]
            okreadings = row[2]
            ocrreadings = row[3]
            ocrwithexcep = row[4]
            meterdefective = row[5]
            doorlocked = row[6]
            noofconsumers = row[7]
            # Percentage
            okreadpercent = math.floor((okreadings / total) * 100)
            ocrreadingpercent = math.floor(
                ((ocrreadings / okreadings) if okreadings else 0) * 100
            )
            ocrwithexceppercent = math.floor(
                ((ocrwithexcep / okreadings) if okreadings else 0) * 100
            )
            meterdefectivepercent = math.floor((meterdefective / total) * 100)
            doorlockedpercent = math.floor((doorlocked / total) * 100)
            noofconsumerspercent = round(
                ((total / noofconsumers) if noofconsumers else 0) * 100, 2
            )

            # add to dictionary
            newdict["agency"] = row[0]
            newdict["totalReadings"] = row[1]
            newdict["OKreadings"] = okreadings
            newdict["OKreadingspercent"] = okreadpercent
            newdict["OCRReadings"] = ocrreadings
            newdict["OCRReadingspercent"] = ocrreadingpercent
            newdict["OCRwithException"] = ocrwithexcep
            newdict["OCRwithExceptionpercent"] = ocrwithexceppercent
            newdict["MeterDefective"] = meterdefective
            newdict["MeterDefectivepercent"] = meterdefectivepercent
            newdict["DoorLocked"] = doorlocked
            newdict["DoorLockedpercent"] = doorlockedpercent
            newdict["noofconsumers"] = noofconsumers
            newdict["noofconsumerspercent"] = noofconsumerspercent
            # add to list
            newdata = listfun(newdict)
        return Response(newdata)
    except:
        return Response([])


@api_view(["POST"])
def dashboardagencywisesbpdcl(request):
    today = date.today()
    thismonth = today.strftime("%Y-%m")
    year = thismonth.split("-")[0]
    month = thismonth.split("-")[1]
    new = []

    clause = ""
    agnc = ""
    cursor = connection.cursor()
    filters = request.data.get("filters", None)

    def listfun(dict):
        new.append(dict.copy())
        return new

    try:
        if filters:
            clause += " AND "
            for i, (key, value) in enumerate(filters.items()):
                if key == "bl_agnc_name":
                    agnc = f"and agency='{filters['bl_agnc_name']}'"
                    clause += f"bl_agnc_name='{value}'"
                    pass

    except:
        pass

    newdict = {}

    query = f"""
             select rm.bl_agnc_name,rm.total,rm.ok,rm.passed,rm.failed,rm.md,rm.dl,
oc.noofconsumers
from (
select agency,sum(distinct no_of_consumers::numeric) as noofconsumers from office_consumers where discom='SBPDCL' {agnc}
group by agency) as oc
full join
(
select bl_agnc_name,count(readingmaster.id) as total,
count(prsnt_mtr_status='Ok' or NULL) as ok,count(rdng_ocr_status='Passed' or NULL) as passed,
count(rdng_ocr_status='Failed' or NULL) as failed,count(prsnt_mtr_status='Meter Defective' or NULL) as md,
count(prsnt_mtr_status='Door Locked' or NULL) as dl
from readingmaster where ofc_discom='SBPDCL' and extract(month from reading_date_db) = '{month}' AND extract(year from reading_date_db) = '{year}' {clause} group by readingmaster.bl_agnc_name
) as rm on oc.agency=rm.bl_agnc_name

    """

    print("query", query)
    try:
        cursor.execute(query)
        result = cursor.fetchall()
        for row in result:
            total = row[1]
            okreadings = row[2]
            ocrreadings = row[3]
            ocrwithexcep = row[4]
            meterdefective = row[5]
            doorlocked = row[6]
            noofconsumers = row[7]
            # Percentage
            okreadpercent = math.floor((okreadings / total) * 100)
            ocrreadingpercent = math.floor(
                ((ocrreadings / okreadings) if okreadings else 0) * 100
            )
            ocrwithexceppercent = math.floor(
                ((ocrwithexcep / okreadings) if okreadings else 0) * 100
            )
            meterdefectivepercent = math.floor((meterdefective / total) * 100)
            doorlockedpercent = math.floor((doorlocked / total) * 100)
            noofconsumerspercent = round(
                ((total / noofconsumers) if noofconsumers else 0) * 100, 2
            )

            # add to dictionary
            newdict["agency"] = row[0]
            newdict["totalReadings"] = row[1]
            newdict["OKreadings"] = okreadings
            newdict["OKreadingspercent"] = okreadpercent
            newdict["OCRReadings"] = ocrreadings
            newdict["OCRReadingspercent"] = ocrreadingpercent
            newdict["OCRwithException"] = ocrwithexcep
            newdict["OCRwithExceptionpercent"] = ocrwithexceppercent
            newdict["MeterDefective"] = meterdefective
            newdict["MeterDefectivepercent"] = meterdefectivepercent
            newdict["DoorLocked"] = doorlocked
            newdict["DoorLockedpercent"] = doorlockedpercent
            newdict["noofconsumers"] = noofconsumers
            newdict["noofconsumerspercent"] = noofconsumerspercent
            # add to list
            newdata = listfun(newdict)
        return Response(newdata)
    except:
        return Response([])


@api_view(["POST"])
def comparision_dashboard(request):
    clause = ""
    cursor = connection.cursor()
    filters = request.data.get("filters", None)

    try:
        if filters:
            clause += " AND "
            for i, (key, value) in enumerate(filters.items()):
                if key == "bl_agnc_name":
                    clause += f"bl_agnc_name='{value}'"
                    pass

    except:
        pass
    query = f"""
        SELECT
  to_char(reading_date_db,'Month') AS month,
  COUNT(*) AS total_meters,
  count(prsnt_mtr_status='Ok' or NULL)as okcount,
  count(prsnt_mtr_status='Meter Defective' or NULL)as md,
  count(prsnt_mtr_status='Door Locked' or NULL)as dl,
  count(rdng_ocr_status='Passed' or null)as passed,
  count(rdng_ocr_status='Failed' or null)as failed,
  CASE
    WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
    ELSE ROUND((cast(count(prsnt_mtr_status='Ok' or null) as float)
    / cast(count(*) as float) * 100)::numeric, 2)
  END as ok,
  CASE
    WHEN count(prsnt_mtr_status='Meter Defective' or NULL) = 0 THEN 0
    ELSE ROUND((cast(count(prsnt_mtr_status='Meter Defective' or null) as float)
    / cast(count(*) as float) * 100)::numeric, 2)
  END as MeterDefective,
  CASE
    WHEN count(prsnt_mtr_status='Door Locked' or NULL) = 0 THEN 0
    ELSE ROUND((cast(count(prsnt_mtr_status='Door Locked' or null) as float)
    / cast(count(*) as float) * 100)::numeric, 2)
  END as DoorLocked,
  CASE
    WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
    ELSE ROUND((cast(count(rdng_ocr_status='Passed' or null) as float)
    / cast(count(prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
  END as ok_without_exception,
  CASE
    WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
    ELSE ROUND((cast(count(rdng_ocr_status='Failed' or null) as float)
    / cast(count(prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
  END as ok_with_exception
FROM readingmaster
WHERE EXTRACT(year FROM reading_date_db) = 2023 {clause}
GROUP BY month,EXTRACT(month FROM reading_date_db)
ORDER BY EXTRACT(month FROM reading_date_db)"""
    cursor.execute(query)
    result = dictfetchall(cursor)
    return Response(result)


@api_view(["POST"])
def comparision_dashboard_nbpdcl(request):
    clause = ""
    cursor = connection.cursor()
    filters = request.data.get("filters", None)

    try:
        if filters:
            clause += " AND "
            for i, (key, value) in enumerate(filters.items()):
                if key == "bl_agnc_name":
                    clause += f"bl_agnc_name='{value}'"
                    pass

    except:
        pass
    query = f"""
        SELECT
  to_char(reading_date_db,'Month') AS month,
  COUNT(*) AS total_meters,
  count(prsnt_mtr_status='Ok' or NULL)as okcount,
  count(prsnt_mtr_status='Meter Defective' or NULL)as md,
  count(prsnt_mtr_status='Door Locked' or NULL)as dl,
  count(rdng_ocr_status='Passed' or null)as passed,
  count(rdng_ocr_status='Failed' or null)as failed,
  CASE
    WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
    ELSE ROUND((cast(count(prsnt_mtr_status='Ok' or null) as float)
    / cast(count(*) as float) * 100)::numeric, 2)
  END as ok,
  CASE
    WHEN count(prsnt_mtr_status='Meter Defective' or NULL) = 0 THEN 0
    ELSE ROUND((cast(count(prsnt_mtr_status='Meter Defective' or null) as float)
    / cast(count(*) as float) * 100)::numeric, 2)
  END as MeterDefective,
  CASE
    WHEN count(prsnt_mtr_status='Door Locked' or NULL) = 0 THEN 0
    ELSE ROUND((cast(count(prsnt_mtr_status='Door Locked' or null) as float)
    / cast(count(*) as float) * 100)::numeric, 2)
  END as DoorLocked,
  CASE
    WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
    ELSE ROUND((cast(count(rdng_ocr_status='Passed' or null) as float)
    / cast(count(prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
  END as ok_without_exception,
  CASE
    WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
    ELSE ROUND((cast(count(rdng_ocr_status='Failed' or null) as float)
    / cast(count(prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
  END as ok_with_exception
FROM readingmaster
WHERE ofc_discom='NBPDCL' AND EXTRACT(year FROM reading_date_db) = 2023 {clause}
GROUP BY month,EXTRACT(month FROM reading_date_db)
ORDER BY EXTRACT(month FROM reading_date_db)"""
    cursor.execute(query)
    result = dictfetchall(cursor)
    return Response(result)


@api_view(["POST"])
def comparision_dashboard_sbpdcl(request):
    clause = ""
    cursor = connection.cursor()
    filters = request.data.get("filters", None)

    try:
        if filters:
            clause += " AND "
            for i, (key, value) in enumerate(filters.items()):
                if key == "bl_agnc_name":
                    clause += f"bl_agnc_name='{value}'"
                    pass

    except:
        pass
    query = f"""
        SELECT
  to_char(reading_date_db,'Month') AS month,
  COUNT(*) AS total_meters,
  count(prsnt_mtr_status='Ok' or NULL)as okcount,
  count(prsnt_mtr_status='Meter Defective' or NULL)as md,
  count(prsnt_mtr_status='Door Locked' or NULL)as dl,
  count(rdng_ocr_status='Passed' or null)as passed,
  count(rdng_ocr_status='Failed' or null)as failed,
  CASE
    WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
    ELSE ROUND((cast(count(prsnt_mtr_status='Ok' or null) as float)
    / cast(count(*) as float) * 100)::numeric, 2)
  END as ok,
  CASE
    WHEN count(prsnt_mtr_status='Meter Defective' or NULL) = 0 THEN 0
    ELSE ROUND((cast(count(prsnt_mtr_status='Meter Defective' or null) as float)
    / cast(count(*) as float) * 100)::numeric, 2)
  END as MeterDefective,
  CASE
    WHEN count(prsnt_mtr_status='Door Locked' or NULL) = 0 THEN 0
    ELSE ROUND((cast(count(prsnt_mtr_status='Door Locked' or null) as float)
    / cast(count(*) as float) * 100)::numeric, 2)
  END as DoorLocked,
  CASE
    WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
    ELSE ROUND((cast(count(rdng_ocr_status='Passed' or null) as float)
    / cast(count(prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
  END as ok_without_exception,
  CASE
    WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
    ELSE ROUND((cast(count(rdng_ocr_status='Failed' or null) as float)
    / cast(count(prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
  END as ok_with_exception
FROM readingmaster
WHERE ofc_discom='SBPDCL' AND EXTRACT(year FROM reading_date_db) = 2023 {clause}
GROUP BY month,EXTRACT(month FROM reading_date_db)
ORDER BY EXTRACT(month FROM reading_date_db)"""
    cursor.execute(query)
    result = dictfetchall(cursor)
    return Response(result)


# @api_view(['POST'])
# def meterreaderDetails(request):
#     pagesize = request.data.get("pagesize",)
#     page = (request.data.get("page",))
#     offset = (int(pagesize) * int(page))-int(pagesize)

#     data = request.data.get("filters", None)
#     clause = ''
#     try:
#         if data:
#             clause += 'WHERE '
#             for i, (key, value) in enumerate(data.items()):
#                 if i > 0:
#                     clause += ' AND '

#                 print("data---------", clause, data, i, key, value)

#                 if key == 'month':
#                     year = value.split('-')[0]
#                     month = value.split('-')[1]
#                     print("month", month)
#                     clause += f"extract(month from reading_date_db) = '{month}' AND extract(year from reading_date_db) = '{year}'"

#                 if key == 'startdate':
#                     clause += f"reading_date_db BETWEEN '{data['startdate']}'"

#                 if key == 'enddate':
#                     clause += f"'{data['enddate']}'"

#                 if key == 'mr_id':
#                     clause += f"mr_id='{data['mr_id']}'"

#                 if key == 'searchdata':
#                     clause += f"(mr_id='{data['searchdata']}' or cons_ac_no='{data['searchdata']}' or cons_name='{data['searchdata']}')"
#                 if key == 'rdng_ocr_status':
#                     clause += f"rdng_ocr_status='{data['rdng_ocr_status']}'"
#                 if key == 'con_trf_cat':
#                     clause += f"con_trf_cat='{value}'"
#                 if key == 'bl_agnc_name':
#                     clause += f"bl_agnc_name='{data['bl_agnc_name']}'"

#                 if key == 'ofc_discom':
#                     clause += f"ofc_discom='{data['ofc_discom']}'"

#             print("clause-------", clause)
#             cursor = connection.cursor()
#             query = (f'''
#             select mr_id,cons_ac_no,bl_agnc_name,abnormality,cons_name,con_trf_cat,con_mtr_sl_no,
# mr_rmrk,prsnt_mtr_status,prsnt_rdng,prev_rdng,
# prsnt_md_rdng,prev_md,ocr_pf_reading,prev_pf_rdng,rdng_date,prev_rdng_date,rdng_img,md_img,
# prsnt_rdng_ocr_excep,md_ocr_excep,qc_req,count(*) over () as total_count from readingmaster {clause} order by rdng_date DESC limit {pagesize} offset {offset}
#                 ''')
#             print(clause)
#             print(query)
#             # group by m.{orderby}
#             cursor.execute(query)
#             person_objects = dictfetchall(cursor)
#         return Response({"result": person_objects, "count": person_objects[0]['total_count']})

#     except Exception as e:
#         print(e)
#         return Response({"result": [], "count": 5})


# ARAVIND
# @api_view(["POST"])
# def meterreaderDetails(request):
#     pagesize = request.data.get(
#         "pagesize",
#     )
#     page = request.data.get(
#         "page",
#     )
#     offset = (int(pagesize) * int(page)) - int(pagesize)

#     import time

#     start = time.time()

#     data = request.data.get("filters", None)
#     clause = ""
#     try:
#         if data:
#             clause += "WHERE "
#             for i, (key, value) in enumerate(data.items()):
#                 if i > 0:
#                     clause += " AND "

#                 print("data---------", clause, data, i, key, value)

#                 if key == "month":
#                     year = value.split("-")[0]
#                     month = value.split("-")[1]
#                     print("month", month)
#                     clause += f"extract(month from reading_date_db) = '{month}' AND extract(year from reading_date_db) = '{year}'"

#                 if key == "startdate":
#                     clause += f"extract(day from reading_date_db) BETWEEN '{data['startdate']}'"

#                 if key == "enddate":
#                     clause += f"'{data['enddate']}'"

#                 if key == "mr_id":
#                     clause += f"mr_id='{data['mr_id']}'"

#                 if key == "searchdata":
#                     clause += f"(mr_id='{data['searchdata']}' or cons_ac_no='{data['searchdata']}' or cons_name='{data['searchdata']}')"
#                 if key == "rdng_ocr_status":
#                     clause += f"rdng_ocr_status='{data['rdng_ocr_status']}'"
#                 if key == "con_trf_cat":
#                     clause += f"con_trf_cat='{value}'"
#                 if key == "bl_agnc_name":
#                     clause += f"bl_agnc_name='{data['bl_agnc_name']}'"

#                 if key == "ofc_discom":
#                     clause += f"ofc_discom='{data['ofc_discom']}'"

#             print("clause-------", clause)

#             cursor = connection.cursor()
#             query = f"""
#                 select mr_id,cons_ac_no,bl_agnc_name,abnormality,cons_name,con_trf_cat,con_mtr_sl_no,
#                 mr_rmrk,prsnt_mtr_status,prsnt_rdng,prev_rdng,
#                 prsnt_md_rdng,prev_md,ocr_pf_reading,prev_pf_rdng,rdng_date,prev_rdng_date,rdng_img,md_img,
#                 prsnt_rdng_ocr_excep,md_ocr_excep,qc_req from readingmaster {clause} order by rdng_date DESC limit {pagesize} offset {offset}
#                 """
#             print(clause)
#             print(query)
#             cursor.execute(query)
#             person_objects = dictfetchall(cursor)
#             query2 = query = f"""
#                 select count(*) as total_count from readingmaster {clause}
#                 """
#             cursor.execute(query2)
#             rows = cursor.fetchone()
#             print(time.time() - start)
#         return Response({"result": person_objects, "count": rows[0]})

#     except Exception as e:
#         print(e)
#         return Response({"result": [], "count": 5})


# Sanjeev
# @api_view(["POST"])
# def meterreaderDetails(request):
#     pagesize = request.data.get("pagesize")
#     page = request.data.get("page")
#     offset = (int(pagesize) * int(page)) - int(pagesize)

#     import time

#     start = time.time()

#     data = request.data.get("filters", None)
#     clause = ""
#     try:
#         if data:
#             clause += " WHERE "
#             conditions = []
#             for key, value in data.items():
#                 if key == "month":
#                     year = value.split("-")[0]
#                     month = value.split("-")[1]
#                     conditions.append(
#                         f"extract(month from reading_date_db) = '{month}' AND extract(year from reading_date_db) = '{year}'")

#                 if key == "startdate":
#                     conditions.append(
#                         f"extract(day from reading_date_db) BETWEEN '{data['startdate']}'")

#                 if key == "enddate":
#                     conditions.append(f"'{data['enddate']}'")

#                 if key == "mr_id":
#                     conditions.append(f"mr_id='{data['mr_id']}'")

#                 if key == "searchdata":
#                     conditions.append(
#                         f"(mr_id='{data['searchdata']}' OR cons_ac_no='{data['searchdata']}' OR cons_name='{data['searchdata']}')")

#                 if key == "rdng_ocr_status":
#                     conditions.append(
#                         f"rdng_ocr_status='{data['rdng_ocr_status']}'")

#                 if key == "con_trf_cat":
#                     conditions.append(f"con_trf_cat='{value}'")

#                 if key == "bl_agnc_name":
#                     conditions.append(f"bl_agnc_name='{data['bl_agnc_name']}'")

#                 if key == "Discom":
#                     conditions.append(f"ofc_discom='{data['Discom']}'")
#                 # if key == "ofc_discom":
#                 #     conditions.append(f"ofc_discom='{data['ofc_discom']}'")

#                 if key == "zone":
#                     conditions.append(f"ofc_zone='{data['zone']}'")

#                 if key == "circle":
#                     conditions.append(f"ofc_circle='{data['circle']}'")

#                 if key == "Division":
#                     conditions.append(f"ofc_division='{data['Division']}'")

#                 if key == "Subdivision":
#                     conditions.append(
#                         f"ofc_subdivision='{data['Subdivision']}'")

#                 if key == "Section":
#                     conditions.append(f"ofc_section='{data['Section']}'")

#             # Join all conditions using 'AND'
#             clause += " AND ".join(conditions)

#             selected_month = data.get("month", None)
#             today = datetime.now()
#             this_month = today.strftime("%Y-%m")
#             previous_month = (
#                 today - timedelta(days=today.day)).strftime("%Y-%m")
#             tablename = "prevmonthsdata" if selected_month not in {
#                 this_month, previous_month} else "readingmaster"

#             cursor = connection.cursor()
#             query = f"""
#                 SELECT mr_id, cons_ac_no, bl_agnc_name, abnormality, cons_name, con_trf_cat, con_mtr_sl_no,
#                 mr_rmrk, prsnt_mtr_status, prsnt_rdng, prev_rdng, prsnt_md_rdng, prev_md, ocr_pf_reading,
#                 prev_pf_rdng, rdng_date, prev_rdng_date, rdng_img, md_img, prsnt_rdng_ocr_excep,
#                 md_ocr_excep, qc_req FROM {tablename} {clause} ORDER BY rdng_date DESC LIMIT {pagesize} OFFSET {offset}
#                 """
#             print("QUERY!", query)
#             cursor.execute(query)
#             person_objects = dictfetchall(cursor)

#             query2 = f"""
#                 SELECT COUNT(*) AS total_count FROM {tablename} {clause}
#                 """
#             print("QUERY!", query2)

#             cursor.execute(query2)
#             rows = cursor.fetchone()

#             print(time.time() - start)
#             return Response({"result": person_objects, "count": rows[0]})
#         else:
#             # No filters present, return empty response
#             return Response({"result": [], "count": 0})

#     except Exception as e:
#         print(e)  # Log the error for debugging purposes
#         return Response({"result": [], "count": 5})

# indra
@api_view(["POST"])
def meterreaderDetails(request):
    pagesize = request.data.get("pagesize")
    page = request.data.get("page")
    offset = (int(pagesize) * int(page)) - int(pagesize)

    import time

    start = time.time()

    data = request.data.get("filters", None)
    clause = ""
    try:
        if data:
            clause += " WHERE "
            conditions = []
            for key, value in data.items():
                if key == "month":
                    year = value.split("-")[0]
                    month = value.split("-")[1]
                    conditions.append(
                        f"extract(month from reading_date_db) = '{month}' AND extract(year from reading_date_db) = '{year}'")

                if key == "startdate":
                    conditions.append(
                        f"extract(day from reading_date_db) BETWEEN '{data['startdate']}'")

                if key == "enddate":
                    conditions.append(f"'{data['enddate']}'")

                if key == "mr_id":
                    conditions.append(f"mr_id='{data['mr_id']}'")

                if key == "searchdata":
                    conditions.append(
                        f"(mr_id='{data['searchdata']}' OR cons_ac_no='{data['searchdata']}' OR cons_name='{data['searchdata']}')")

                if key == "rdng_ocr_status":
                    conditions.append(
                        f"rdng_ocr_status='{data['rdng_ocr_status']}'")
                if key == "Exception":
                    conditions.append(f"rdng_ocr_status='{value}'")

                if key == "prsnt_rdng_ocr_excep":
                    # CASE 1: Passed → get only passed rows
                    if value == "Passed":
                        conditions.append("rdng_ocr_status = 'Passed'")
                    # CASE 2: Failed (All)
                    elif value == "__FAILED__":
                        # Get rows where there IS an exception (not empty, not null)
                        conditions.append(
                            "TRIM(COALESCE(prsnt_rdng_ocr_excep, '')) <> ''")
                    # CASE 3: Failed + Specific Reason
                    else:
                        conditions.append(f"prsnt_rdng_ocr_excep = '{value}'")

                if key == "con_trf_cat":
                    conditions.append(f"con_trf_cat='{value}'")

                if key == 'prsnt_mtr_status':
                    conditions.append(f"prsnt_mtr_status='{value}'")

                if key == "bl_agnc_name":
                    conditions.append(f"bl_agnc_name='{data['bl_agnc_name']}'")

                if key == "Discom":
                    conditions.append(f"ofc_discom='{data['Discom']}'")
                # if key == "ofc_discom":
                #     conditions.append(f"ofc_discom='{data['ofc_discom']}'")

                if key == "zone":
                    conditions.append(f"ofc_zone='{data['zone']}'")

                if key == "circle":
                    conditions.append(f"ofc_circle='{data['circle']}'")

                if key == "Division":
                    conditions.append(f"ofc_division='{data['Division']}'")

                if key == "Subdivision":
                    conditions.append(
                        f"ofc_subdivision='{data['Subdivision']}'")

                if key == "Section":
                    conditions.append(f"ofc_section='{data['Section']}'")

            # Join all conditions using 'AND'
            clause += " AND ".join(conditions)

            selected_month = data.get("month", None)
            today = datetime.now()
            this_month = today.strftime("%Y-%m")
            previous_month = (
                today - timedelta(days=today.day)).strftime("%Y-%m")
            tablename = "prevmonthsdata" if selected_month not in {
                this_month, previous_month} else "readingmaster"

            cursor = connection.cursor()
            query = f"""
                SELECT mr_id, cons_ac_no, bl_agnc_name, abnormality, cons_name, con_trf_cat, con_mtr_sl_no,
                mr_rmrk, prsnt_mtr_status, prsnt_rdng, prev_rdng, prsnt_md_rdng, prev_md, ocr_pf_reading,
                prev_pf_rdng, rdng_date, prev_rdng_date, rdng_img, md_img, rdng_ocr_status,
                CASE
                    WHEN rdng_ocr_status = 'Passed' THEN 'Passed'
                    ELSE COALESCE(NULLIF(TRIM(prsnt_rdng_ocr_excep), ''), '')
                END AS prsnt_rdng_ocr_excep,
                md_ocr_excep, qc_req FROM {tablename} {clause} ORDER BY rdng_date DESC LIMIT {pagesize} OFFSET {offset}
                """
            print("QUERY!", query)
            cursor.execute(query)
            person_objects = dictfetchall(cursor)

            query2 = f"""
                SELECT COUNT(*) AS total_count FROM {tablename} {clause}
                """
            print("QUERY!", query2)

            cursor.execute(query2)
            rows = cursor.fetchone()

            print(time.time() - start)
            return Response({"result": person_objects, "count": rows[0]})
        else:
            # No filters present, return empty response
            return Response({"result": [], "count": 0})

    except Exception as e:
        print(e)  # Log the error for debugging purposes
        return Response({"result": [], "count": 5})


@api_view(["POST"])
def cons_wise_details_with_search(request):
    new = []

    def listfun(dict):
        new.append(dict.copy())
        return new

    newdict = {}
    clause = ""
    cursor = connection.cursor()
    filters = request.data.get("filters", None)
    try:
        if filters:
            clause += "WHERE "
            for i, (key, value) in enumerate(filters.items()):
                if i > 0:
                    clause += "AND "
                if key == "month":
                    year = value.split("-")[0]
                    month = value.split("-")[1]
                    print("month", month)
                    clause += f"extract(month from reading_date_db) = '{month}' AND extract(year from reading_date_db) = '{year}'"
                if key == "startdate":
                    clause += f"reading_date_db BETWEEN '{filters['startdate']}'"

                if key == "enddate":
                    clause += f"'{filters['enddate']}'"
                if key == "cons_ac_no":
                    clause += f"cons_ac_no='{value}'"

    except:
        pass
    if clause != "":
        query = f"""SELECT  m.ofc_discom,m.ofc_zone,m.ofc_circle,m.ofc_division,m.ofc_subdivision,m.ofc_section,
    m.id,m.cons_name,m.cons_ac_no,cons_address,m.cons_ph_no,m.con_trf_cat,m.mr_unit,
    r."mrId",r."mrName",r."mrPhone",r."mrPhoto" as avatar,m.con_mtr_sl_no,
    m.rdng_date,m.prsnt_mtr_status,m.prsnt_md_rdng,m.ocr_pf_reading,m.abnormality,m.prsnt_rdng_ocr_excep,m.md_ocr_excep,m.mr_rmrk,m.qc_req,m.ai_mdl_ver,m.ph_name,m.cmra_res,m.andr_ver,m.reading_date_db,
    m.rdng_img,m.md_img,m.pf_image,m.prsnt_ocr_rdng,m.prsnt_rdng,m.prsnt_md_rdng_ocr,m.rdng_ocr_status,m.kvah_rdng,m.kvah_img 
      FROM
    readingmaster m left outer join meterreaderregistration r on m.mr_id=r."mrId" {clause}"""
    else:
        return Response({"MSG": "PROVIDE CONSUMER ACOUNT NUMBER"})

    print("QUERY------>", query)
    cursor.execute(query)
    results = dictfetchall(cursor)
    print(results)
    return Response(results)


@api_view(["POST"])
def monthwiseexceptiondashboard(request):
    print("inside the function hahahaha")
    today = date.today()
    thismonth = today.strftime("%Y-%m")
    year = thismonth.split("-")[0]
    clause = ""
    cursor = connection.cursor()
    filters = request.data.get("filters", None)

    try:
        if filters:
            clause += " AND "
            for i, (key, value) in enumerate(filters.items()):
                if i > 0:
                    clause += "AND "

                if key == "bl_agnc_name":
                    clause += f"bl_agnc_name='{value}'"
                    pass
                if key == "ofc_discom":
                    clause += f"ofc_discom='{value}'"
                    pass
    except:
        pass
    query = f"""
    SELECT
to_char(reading_date_db,'Month') AS month,
COUNT(*) AS total_meters,
COUNT(prsnt_mtr_status = 'Ok' or null) as count,
count(rdng_ocr_status='Passed'or null)as passed,
count(rdng_ocr_status='Failed'or null)as failed,
count(prsnt_rdng_ocr_excep='Incorrect Reading' or null) as incorrect_reading,
count(prsnt_rdng_ocr_excep='Spoofed Image' or null) as image_spoofed,
count(prsnt_rdng_ocr_excep='Image blur' or null) as image_blur,
count(prsnt_rdng_ocr_excep='Meter Dirty' or null) as meterdirty,
count(reading_parameter_type='Parameters Unavailable' AND rdng_ocr_status = 'Failed' or null ) as parameterunavailable,
count(reading_parameter_type='Parameters Mismatch' AND rdng_ocr_status = 'Failed' or null) as parametersmismatch,

CASE
WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
ELSE ROUND((cast(count(prsnt_rdng_ocr_excep='Spoofed Image' or null) as float)
/ cast(count(prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
END as image_spoofedper,
CASE
WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
ELSE ROUND((cast(count(prsnt_rdng_ocr_excep='Incorrect Reading' or null) as float)
/ cast(count(prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
END as incorrectreadingper,
CASE
WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
ELSE ROUND((cast(count(prsnt_rdng_ocr_excep='Image blur' or null) as float)
/ cast(count(prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
END as image_blurper,
CASE
WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
ELSE ROUND((cast(count(prsnt_rdng_ocr_excep='Meter Dirty' or null) as float)
/ cast(count(prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
END as meterdirtyper,

CASE
WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
ELSE ROUND((CAST(SUM(CASE WHEN reading_parameter_type = 'Parameters Unavailable' AND rdng_ocr_status = 'Failed' THEN 1 ELSE 0 END) AS FLOAT) / CAST(COUNT(prsnt_mtr_status = 'Ok' OR NULL) AS FLOAT) * 100)::NUMERIC, 2)
END as parameters_unavailableper,
CASE
WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
ELSE ROUND((cast(count(reading_parameter_type='Parameters Mismatch' or null) as float)
/ cast(count(prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
END as parameters_mismatchper,

CASE
WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
ELSE ROUND((cast(count(prsnt_mtr_status='Ok' or null) as float)
/ cast(count(*) as float) * 100)::numeric, 2)
END as okper,
CASE
WHEN count(prsnt_mtr_status='Meter Defective' or NULL) = 0 THEN 0
ELSE ROUND((cast(count(prsnt_mtr_status='Meter Defective' or null) as float)
/ cast(count(*) as float) * 100)::numeric, 2)
END as MeterDefectiveper,
CASE
WHEN count(prsnt_mtr_status='Door Locked' or NULL) = 0 THEN 0
ELSE ROUND((cast(count(prsnt_mtr_status='Door Locked' or null) as float)
/ cast(count(*) as float) * 100)::numeric, 2)
END as DoorLockedper,
CASE
WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
ELSE ROUND((cast(count(rdng_ocr_status='Passed' or null) as float)
/ cast(count(prsnt_mtr_status='Ok' or NULL) as float) * 100)::numeric, 2)
END as ok_without_exceptionper,
CASE
WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
ELSE ROUND((cast(count(rdng_ocr_status='Failed' or null) as float)
/ cast(count(prsnt_mtr_status='Ok' or NULL) as float) * 100)::numeric, 2)
END as ok_with_exceptionper
FROM readingmaster
WHERE EXTRACT(year FROM reading_date_db) = {year} {clause}
GROUP BY month,EXTRACT(month FROM reading_date_db)
ORDER BY EXTRACT(month FROM reading_date_db);
"""
    print("QUERY-->", query)
    cursor.execute(query)
    result = dictfetchall(cursor)
    return Response(result)


@api_view(["POST"])
def monthwiseexceptiondashboardnbpdcl(request):
    today = date.today()
    thismonth = today.strftime("%Y-%m")
    year = thismonth.split("-")[0]
    clause = ""
    cursor = connection.cursor()
    filters = request.data.get("filters", None)

    try:
        if filters:
            clause += " AND "
            for i, (key, value) in enumerate(filters.items()):
                if key == "bl_agnc_name":
                    clause += f"bl_agnc_name='{value}'"
                    pass

    except:
        pass
    query = f"""
        SELECT
to_char(reading_date_db,'Month') AS month,
COUNT(*) AS total_meters,
COUNT(prsnt_mtr_status = 'Ok' or null) as count,
count(rdng_ocr_status='Passed'or null)as passed,
count(rdng_ocr_status='Failed'or null)as failed,
count(prsnt_rdng_ocr_excep='Incorrect Reading' or null) as incorrect_reading,
count(prsnt_rdng_ocr_excep='Spoofed Image' or null) as image_spoofed,
count(prsnt_rdng_ocr_excep='Image blur' or null) as image_blur,
count(prsnt_rdng_ocr_excep='Meter Dirty' or null) as meterdirty,
count(reading_parameter_type='Parameters Unavailable' AND rdng_ocr_status = 'Failed' or null ) as parameterunavailable,
count(reading_parameter_type='Parameters Mismatch' AND rdng_ocr_status = 'Failed' or null) as parametersmismatch,

CASE
WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
ELSE ROUND((cast(count(prsnt_rdng_ocr_excep='Spoofed Image' or null) as float)
/ cast(count(prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
END as image_spoofedper,
CASE
WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
ELSE ROUND((cast(count(prsnt_rdng_ocr_excep='Incorrect Reading' or null) as float)
/ cast(count(prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
END as incorrectreadingper,
CASE
WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
ELSE ROUND((cast(count(prsnt_rdng_ocr_excep='Image blur' or null) as float)
/ cast(count(prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
END as image_blurper,
CASE
WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
ELSE ROUND((cast(count(prsnt_rdng_ocr_excep='Meter Dirty' or null) as float)
/ cast(count(prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
END as meterdirtyper,

CASE
WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
ELSE ROUND((CAST(SUM(CASE WHEN reading_parameter_type = 'Parameters Unavailable' AND rdng_ocr_status = 'Failed' THEN 1 ELSE 0 END) AS FLOAT) / CAST(COUNT(prsnt_mtr_status = 'Ok' OR NULL) AS FLOAT) * 100)::NUMERIC, 2)
END as parameters_unavailableper,
CASE
WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
ELSE ROUND((cast(count(reading_parameter_type='Parameters Mismatch' or null) as float)
/ cast(count(prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
END as parameters_mismatchper,

CASE
WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
ELSE ROUND((cast(count(prsnt_mtr_status='Ok' or null) as float)
/ cast(count(*) as float) * 100)::numeric, 2)
END as okper,
CASE
WHEN count(prsnt_mtr_status='Meter Defective' or NULL) = 0 THEN 0
ELSE ROUND((cast(count(prsnt_mtr_status='Meter Defective' or null) as float)
/ cast(count(*) as float) * 100)::numeric, 2)
END as MeterDefectiveper,
CASE
WHEN count(prsnt_mtr_status='Door Locked' or NULL) = 0 THEN 0
ELSE ROUND((cast(count(prsnt_mtr_status='Door Locked' or null) as float)
/ cast(count(*) as float) * 100)::numeric, 2)
END as DoorLockedper,
CASE
WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
ELSE ROUND((cast(count(rdng_ocr_status='Passed' or null) as float)
/ cast(count(prsnt_mtr_status='Ok' or NULL) as float) * 100)::numeric, 2)
END as ok_without_exceptionper,
CASE
WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
ELSE ROUND((cast(count(rdng_ocr_status='Failed' or null) as float)
/ cast(count(prsnt_mtr_status='Ok' or NULL) as float) * 100)::numeric, 2)
END as ok_with_exceptionper
FROM readingmaster
WHERE ofc_discom='NBPDCL' and EXTRACT(year FROM reading_date_db) = {year} {clause}
GROUP BY month,EXTRACT(month FROM reading_date_db)
ORDER BY EXTRACT(month FROM reading_date_db);
"""
    print("QUERY-->", query)
    cursor.execute(query)
    result = dictfetchall(cursor)
    return Response(result)


@api_view(["POST"])
def monthwiseexceptiondashboardsbpdcl(request):
    today = date.today()
    thismonth = today.strftime("%Y-%m")
    year = thismonth.split("-")[0]
    clause = ""
    cursor = connection.cursor()
    filters = request.data.get("filters", None)

    try:
        if filters:
            clause += " AND "
            for i, (key, value) in enumerate(filters.items()):
                if key == "bl_agnc_name":
                    clause += f"bl_agnc_name='{value}'"
                    pass

    except:
        pass
    query = f"""
    SELECT
to_char(reading_date_db,'Month') AS month,
COUNT(*) AS total_meters,
COUNT(prsnt_mtr_status = 'Ok' or null) as count,
count(rdng_ocr_status='Passed'or null)as passed,
count(rdng_ocr_status='Failed'or null)as failed,
count(prsnt_rdng_ocr_excep='Incorrect Reading' or null) as incorrect_reading,
count(prsnt_rdng_ocr_excep='Spoofed Image' or null) as image_spoofed,
count(prsnt_rdng_ocr_excep='Image blur' or null) as image_blur,
count(prsnt_rdng_ocr_excep='Meter Dirty' or null) as meterdirty,
count(reading_parameter_type='Parameters Unavailable' AND rdng_ocr_status = 'Failed' or null ) as parameterunavailable,
count(reading_parameter_type='Parameters Mismatch' AND rdng_ocr_status = 'Failed' or null) as parametersmismatch,

CASE
WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
ELSE ROUND((cast(count(prsnt_rdng_ocr_excep='Spoofed Image' or null) as float)
/ cast(count(prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
END as image_spoofedper,
CASE
WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
ELSE ROUND((cast(count(prsnt_rdng_ocr_excep='Incorrect Reading' or null) as float)
/ cast(count(prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
END as incorrectreadingper,
CASE
WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
ELSE ROUND((cast(count(prsnt_rdng_ocr_excep='Image blur' or null) as float)
/ cast(count(prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
END as image_blurper,
CASE
WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
ELSE ROUND((cast(count(prsnt_rdng_ocr_excep='Meter Dirty' or null) as float)
/ cast(count(prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
END as meterdirtyper,

CASE
WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
ELSE ROUND((CAST(SUM(CASE WHEN reading_parameter_type = 'Parameters Unavailable' AND rdng_ocr_status = 'Failed' THEN 1 ELSE 0 END) AS FLOAT) / CAST(COUNT(prsnt_mtr_status = 'Ok' OR NULL) AS FLOAT) * 100)::NUMERIC, 2)
END as parameters_unavailableper,
CASE
WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
ELSE ROUND((cast(count(reading_parameter_type='Parameters Mismatch' or null) as float)
/ cast(count(prsnt_mtr_status='Ok' or null) as float) * 100)::numeric, 2)
END as parameters_mismatchper,

CASE
WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
ELSE ROUND((cast(count(prsnt_mtr_status='Ok' or null) as float)
/ cast(count(*) as float) * 100)::numeric, 2)
END as okper,
CASE
WHEN count(prsnt_mtr_status='Meter Defective' or NULL) = 0 THEN 0
ELSE ROUND((cast(count(prsnt_mtr_status='Meter Defective' or null) as float)
/ cast(count(*) as float) * 100)::numeric, 2)
END as MeterDefectiveper,
CASE
WHEN count(prsnt_mtr_status='Door Locked' or NULL) = 0 THEN 0
ELSE ROUND((cast(count(prsnt_mtr_status='Door Locked' or null) as float)
/ cast(count(*) as float) * 100)::numeric, 2)
END as DoorLockedper,
CASE
WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
ELSE ROUND((cast(count(rdng_ocr_status='Passed' or null) as float)
/ cast(count(prsnt_mtr_status='Ok' or NULL) as float) * 100)::numeric, 2)
END as ok_without_exceptionper,
CASE
WHEN count(prsnt_mtr_status='Ok' or NULL) = 0 THEN 0
ELSE ROUND((cast(count(rdng_ocr_status='Failed' or null) as float)
/ cast(count(prsnt_mtr_status='Ok' or NULL) as float) * 100)::numeric, 2)
END as ok_with_exceptionper
FROM readingmaster
WHERE ofc_discom='SBPDCL' and EXTRACT(year FROM reading_date_db) = {year} {clause}
GROUP BY month,EXTRACT(month FROM reading_date_db)
ORDER BY EXTRACT(month FROM reading_date_db);
"""
    print("QUERY-->", query)
    cursor.execute(query)
    result = dictfetchall(cursor)
    return Response(result)


# @api_view(['POST'])
# def cons_passed(request):
#     cons_ac_no=request.data['cons_ac_no']
#     cursor = connection.cursor()
#     query=f"""select distinct(rdng_ocr_status) from readingmaster where rdng_ocr_status='Passed' and cons_ac_no='{cons_ac_no}'
#     """
#     ocrstatus=""
#     cursor.execute(query)
#     result=cursor.fetchall()
#     if (len(result)>0):
#         if result[0][0]=='Passed':
#             ocrstatus="Passed"
#     else:
#         ocrstatus="THERE IS NO OCR PASSED FOR THIS CONSUMER"

#     return Response({"status":"Failed"})


@api_view(["POST"])
# def cons_passed(request):
#     cons_ac_no=request.data['cons_ac_no']
#     mr_id=request.data['mrId']
#     mr_id_list=['mrDanish','2241MRC73','2241MRC97']
#     for i in mr_id_list:
#         print(i)
#         if i==mr_id:
#             print("inside if")
#             cursor = connection.cursor()
#             query=f"""select distinct(rdng_ocr_status) from readingmaster where rdng_ocr_status='Passed' and cons_ac_no='{cons_ac_no}' and mr_id='{mr_id}'
#             """
#             ocrstatus=""
#             cursor.execute(query)
#             result=cursor.fetchall()
#             if (len(result)>0):
#                 if result[0][0]=='Passed':
#                     ocrstatus="Passed"
#             else:
#                 ocrstatus="THERE IS NO OCR PASSED FOR THIS CONSUMER"
#             return Response({"status":"Passed"})
#     return Response({"status":"Failed"})
def cons_passed(request):
    cons_ac_no = request.data["cons_ac_no"]
    mr_id = request.data["mrId"]

    cursor = connection.cursor()
    query = f"""select distinct(rdng_ocr_status) from readingmaster where rdng_ocr_status='Passed'  and cons_ac_no='{cons_ac_no}' and mr_id='{mr_id}' and manual_update_flag isnull and  qc_done !='byLambda'
    """

    ocrstatus = ""
    cursor.execute(query)
    result = cursor.fetchall()
    if len(result) > 0:
        if result[0][0] == "Passed":
            ocrstatus = "Passed"

    else:
        ocrstatus = "THERE IS NO OCR PASSED FOR THIS CONSUMER"

    return Response({"status": ocrstatus})


@api_view(["GET"])
def spoofimagecheck(request):
    return Response({"status": False})


# ----------------------------------------TruereadQCAPP--------------------------------

@api_view(['POST'])
def divisiondata(request):
    zone = request.data.get("zone", None)
    month = request.data.get("month", None)

    try:
        if month:
            start_date = datetime.strptime(month, "%Y-%m").replace(day=1)
            if start_date.month == 12:
                end_date = start_date.replace(
                    month=1, year=start_date.year + 1) - timedelta(days=1)
            else:
                end_date = start_date.replace(
                    month=start_date.month + 1) - timedelta(days=1)
        else:
            return Response({"error": "Month is required in 'YYYY-MM' format."}, status=400)
    except ValueError:
        return Response({"error": "Invalid month format. Expected 'YYYY-MM'."}, status=400)

    query = "SELECT DISTINCT ofc_division FROM readingmaster WHERE reading_date_db BETWEEN %s AND %s"
    params = [start_date, end_date]

    if zone:
        query += " AND ofc_zone = %s"
        params.append(zone)

    with connection.cursor() as cursor:
        cursor.execute(query, params)
        divisions = [row[0] for row in cursor.fetchall()]

    return Response(divisions if divisions else [])

# @api_view(['POST', 'GET'])
# def failedimages(request):
#     monthyear = request.data.get("month")
#     zone = request.data.get("zone")
#     division = request.data.get("division")
#     pagesize = request.data.get("pagesize")
#     ocrrdng = request.data.get("ocrrdng")
#     page = request.data.get("page")
#     user = request.data.get("user")
#     month = int(monthyear.split('-')[1])
#     year = int(monthyear.split('-')[0])

#     now = datetime.now()

#     q_objects = Q(rdng_ocr_status='Failed', mtr_excep_img__isnull=True, manual_update_flag__isnull=True) & ~(
#         Q(prsnt_rdng_ocr_excep='Spoofed Image') & Q(is_object_meter='NO and camera blocked'))

#     if month:
#         q_objects &= Q(reading_date_db__month=month)
#     else:
#         q_objects &= Q(reading_date_db__month=now.month)
#     if year:
#         q_objects &= Q(reading_date_db__year=year)
#     else:
#         q_objects &= Q(reading_date_db__year=now.year)

#     if zone:
#         q_objects &= Q(ofc_zone=zone)

#     if division:
#         q_objects &= Q(ofc_division=division)

#     if ocrrdng:
#         q_objects &= Q(prsnt_ocr_rdng=ocrrdng)
#     else:
#         q_objects &= ~Q(prsnt_ocr_rdng='Not Found')

#     reading_master_objects = Consumers.objects.filter(q_objects).order_by('id')
#     print(reading_master_objects.query)
#     paginator = Paginator(reading_master_objects, pagesize)
#     try:
#         paginated_results = paginator.page(page)
#     except Exception:
#         paginated_results = paginator.page(1)

#     count = reading_master_objects.count()

#     if user and month and zone and division and year:
#         updated_count = Consumers.objects.filter(
#             mtr_excep_img=f'vapp_{user}',
#             reading_date_db__month=month,
#             reading_date_db__year=year,
#             ofc_zone=zone,
#             ofc_division=division
#         ).count()
#     else:
#         updated_count = 0

#     serializer = FailedImageSerializer(paginated_results, many=True)

#     if not serializer.data:
#         return Response({
#             "status": False,
#             "message": "Data not found",
#             "count": count,
#             "update_count": updated_count,
#             "results": serializer.data,
#         })

#     return Response({
#         "status": True,
#         "message": "Data Fetched Successfully",
#         "count": count,
#         "update_count": updated_count,
#         "results": serializer.data,
#     })


# -------------------------------QC IMG-----------------------------------------------------------------
# @api_view(['POST', 'GET'])
# def failedimages(request):
#     month = request.data.get("month", None)
#     zone = request.data.get("zone", None)
#     division = request.data.get("division", None)
#     pagesize = request.data.get("pagesize", None)
#     ocrrdng = request.data.get("ocrrdng", None)
#     page = request.data.get("page",)
#     user = request.data.get("user", None)

#     now = datetime.now()
#     offset = (int(pagesize) * int(page))-int(pagesize)
#     clause = ''

#     where_clauses = []
#     if month:
#         month = month.split('-')[1]
#         where_clauses.append(f"EXTRACT(MONTH FROM reading_date_db)='{month}'")
#     else:
#         current_month = now.month
#         where_clauses.append(
#             f"EXTRACT(MONTH FROM reading_date_db) = '{current_month}'")
#     if zone:
#         where_clauses.append(f"ofc_zone = '{zone}'")

#     if division:
#         where_clauses.append(f"ofc_division = '{division}'")

#     if ocrrdng:
#         where_clauses.append(f"prsnt_ocr_rdng = '{ocrrdng}'")
#     else:
#         where_clauses.append("prsnt_ocr_rdng != 'Not Found'")
#     if where_clauses:
#         clause = ' AND '.join(where_clauses)

#     cursor = connection.cursor()
#     cursor1 = connection.cursor()
#     cursor2 = connection.cursor()

#     query = (
#         f'''select id,rdng_img,prsnt_rdng,prsnt_ocr_rdng from readingmaster where rdng_ocr_status = 'Failed' and qc_req='Yes' and prsnt_rdng_ocr_excep != 'Spoofed Image' and mtr_excep_img isnull and manual_update_flag isnull and {clause} limit {pagesize} offset {offset} ''')
#     query1 = (
#         f'''SELECT COUNT(*) AS tot_count FROM readingmaster WHERE rdng_ocr_status = 'Failed' and prsnt_rdng_ocr_excep != 'Spoofed Image' AND qc_req = 'Yes' and mtr_excep_img isnull and manual_update_flag isnull AND {clause} ''')
#     query2 = (
#         f'''
#         select count(*) as updated_count from readingmaster
#       where mtr_excep_img = 'vapp_{user}' AND EXTRACT(MONTH FROM reading_date_db)='{month}' AND ofc_zone = '{zone}' AND ofc_division = '{division}'
#         '''
#     )

#     print("query",query)
#     print("query1",query1)
#     print("query2",query2)

#     cursor.execute(query)
#     cursor1.execute(query1)
#     cursor2.execute(query2)

#     person_objects = dictfetchall(cursor)
#     data = dictfetchall(cursor1)
#     data1 = dictfetchall(cursor2)
#     count = data[0]['tot_count']
#     count1 = data1[0]['updated_count']

#     if not person_objects:
#         return Response({
#             "status": False,
#             "message": "Data not found",
#             "count": count,
#             "update_count": count1,
#             "results": person_objects,
#         })

#     return Response({
#         "status": True,
#         "message": "Data Fetched Successfully",
#         "count": count,
#         "update_count": count1,
#         "results": person_objects,
#     })


@api_view(['POST'])
def updatequery(request):
    id = request.data.get("id")
    user = request.data.get("user")
    date = request.data.get("date")
    monthyear = request.data.get("month")
    zone = request.data.get("zone")
    division = request.data.get("division")
    ocrrdng = request.data.get("ocrrdng")
    month = int(monthyear.split('-')[1])
    year = int(monthyear.split('-')[0])

    now = datetime.now()

    formatted_date = date[:10]

    query1_update_data = {
        'prsnt_rdng_ocr_odv': F('prsnt_ocr_rdng'),
        'rdng_ocr_status_odv': F('rdng_ocr_status'),
        'prsnt_ocr_excep_old_values': F('prsnt_rdng_ocr_excep'),
        'rdng_ocr_status_changed_by': f'vapp_{user}',
        'mtr_excep_img': f'vapp_{user}',
        'manual_update_flag': True,
        'qc_req': 'No',
        'date_qc': formatted_date
    }

    query2_update_data = {
        'prsnt_ocr_rdng': F('prsnt_rdng'),
        'rdng_ocr_status': 'Passed',
        'prsnt_rdng_ocr_excep': '',
    }

    with transaction.atomic():
        Consumers.objects.filter(id=id).update(**query1_update_data)
        Consumers.objects.filter(id=id).update(**query2_update_data)

    queryset = Consumers.objects.all()
    if month:
        queryset = queryset.filter(reading_date_db__month=month)
    else:
        current_month = now.month
        queryset = queryset.filter(reading_date_db__month=current_month)
    if year:
        queryset = queryset.filter(reading_date_db__year=year)
    else:
        current_year = now.year
        queryset = queryset.filter(reading_date_db__year=current_year)

    if zone:
        queryset = queryset.filter(ofc_zone=zone)

    if division:
        queryset = queryset.filter(ofc_division=division)

    if ocrrdng:
        queryset = queryset.filter(prsnt_ocr_rdng=ocrrdng)
    else:
        queryset = queryset.exclude(prsnt_ocr_rdng='Not Found')

    filter_conditions = Q(rdng_ocr_status='Failed', mtr_excep_img__isnull=True, manual_update_flag__isnull=True) & ~(
        Q(prsnt_rdng_ocr_excep='Spoofed Image') & Q(is_object_meter='NO and camera blocked'))
    tot_count = queryset.filter(filter_conditions).count()

    if tot_count > 0:
        updated_count = Consumers.objects.filter(
            mtr_excep_img=f'vapp_{user}',
            reading_date_db__month=month,
            reading_date_db__year=year,
            ofc_zone=zone,
            ofc_division=division
        ).count()

        return Response({
            "status": True,
            "message": "Updated Successfully",
            "updated_count": updated_count,
            "tot_count": tot_count
        })
    else:
        updated_count = Consumers.objects.filter(
            mtr_excep_img=f'vapp_{user}',
            reading_date_db__month=month,
            reading_date_db__year=year,
            ofc_zone=zone,
            ofc_division=division
        ).count()
        return Response({
            "status": False,
            "message": "Data not found",
            "tot_count": tot_count,
            "updated_count": updated_count
        })


@api_view(['POST'])
def updatequery1(request):
    id = request.data.get("id")
    user = request.data.get("user")
    date = request.data.get("date")
    monthyear = request.data.get("month")
    zone = request.data.get("zone")
    division = request.data.get("division")
    ocrrdng = request.data.get("ocrrdng")
    month = int(monthyear.split('-')[1])
    year = int(monthyear.split('-')[0])

    now = datetime.now()

    formatted_date = date[:10]

    query_update_data = {
        'rdng_ocr_status_changed_by': f'vapp_{user}',
        'mtr_excep_img': f'vapp_{user}',
        'manual_update_flag': True,
        'qc_req': 'No',
        'date_qc': formatted_date
    }

    with transaction.atomic():
        Consumers.objects.filter(id=id).update(**query_update_data)

    queryset = Consumers.objects.all()
    if month:
        queryset = queryset.filter(reading_date_db__month=month)
    else:
        current_month = now.month
        queryset = queryset.filter(reading_date_db__month=current_month)
    if year:
        queryset = queryset.filter(reading_date_db__year=year)
    else:
        current_year = now.year
        queryset = queryset.filter(reading_date_db__year=current_year)

    if zone:
        queryset = queryset.filter(ofc_zone=zone)

    if division:
        queryset = queryset.filter(ofc_division=division)

    if ocrrdng:
        queryset = queryset.filter(prsnt_ocr_rdng=ocrrdng)
    else:
        queryset = queryset.exclude(prsnt_ocr_rdng='Not Found')

    filter_conditions = Q(rdng_ocr_status='Failed', mtr_excep_img__isnull=True, manual_update_flag__isnull=True) & ~(
        Q(prsnt_rdng_ocr_excep='Spoofed Image') & Q(is_object_meter='NO and camera blocked'))

    tot_count = queryset.filter(filter_conditions).count()

    if tot_count > 0:
        updated_count = Consumers.objects.filter(
            mtr_excep_img=f'vapp_{user}',
            reading_date_db__month=month,
            reading_date_db__year=year,
            ofc_zone=zone,
            ofc_division=division
        ).count()

        return Response({
            "status": True,
            "message": "Updated Successfully",
            "updated_count": updated_count,
            "tot_count": tot_count
        })
    else:
        updated_count = Consumers.objects.filter(
            mtr_excep_img=f'vapp_{user}',
            reading_date_db__month=month,
            reading_date_db__year=year,
            ofc_zone=zone,
            ofc_division=division
        ).count()
        return Response({
            "status": False,
            "message": "Data not found",
            "tot_count": tot_count,
            "updated_count": updated_count
        })


@api_view(['POST'])
def updatespoof(request):
    id = request.data.get("id")
    user = request.data.get("user")
    date = request.data.get("date")
    monthyear = request.data.get("month")
    zone = request.data.get("zone")
    division = request.data.get("division")
    ocrrdng = request.data.get("ocrrdng")
    month = int(monthyear.split('-')[1])
    year = int(monthyear.split('-')[0])

    now = datetime.now()

    formatted_date = date[:10]

    query_update_data = {
        'rdng_ocr_status_changed_by': f'vapp_{user}',
        'prsnt_ocr_excep_old_values': F('prsnt_rdng_ocr_excep'),
        'mtr_excep_img': f'vapp_{user}',
        'manual_update_flag': 'true',
        'qc_req': 'No',
        'prsnt_rdng_ocr_excep': 'Spoofed Image',
        'date_qc': formatted_date
    }

    with transaction.atomic():
        Consumers.objects.filter(id=id).update(**query_update_data)

    queryset = Consumers.objects.all()
    if month:
        queryset = queryset.filter(reading_date_db__month=month)
    else:
        current_month = now.month
        queryset = queryset.filter(reading_date_db__month=current_month)
    if year:
        queryset = queryset.filter(reading_date_db__year=year)
    else:
        current_year = now.year
        queryset = queryset.filter(reading_date_db__year=current_year)

    if zone:
        queryset = queryset.filter(ofc_zone=zone)

    if division:
        queryset = queryset.filter(ofc_division=division)

    if ocrrdng:
        queryset = queryset.filter(prsnt_ocr_rdng=ocrrdng)
    else:
        queryset = queryset.exclude(prsnt_ocr_rdng='Not Found')

    filter_conditions = Q(rdng_ocr_status='Failed', mtr_excep_img__isnull=True, manual_update_flag__isnull=True) & ~(
        Q(prsnt_rdng_ocr_excep='Spoofed Image') & Q(is_object_meter='NO and camera blocked'))
    tot_count = queryset.filter(filter_conditions).count()

    if tot_count > 0:
        updated_count = Consumers.objects.filter(
            mtr_excep_img=f'vapp_{user}',
            reading_date_db__month=month,
            reading_date_db__year=year,
            ofc_zone=zone,
            ofc_division=division
        ).count()

        return Response({
            "status": True,
            "message": "Updated Successfully",
            "updated_count": updated_count,
            "tot_count": tot_count
        })
    else:
        updated_count = Consumers.objects.filter(
            mtr_excep_img=f'vapp_{user}',
            reading_date_db__month=month,
            reading_date_db__year=year,
            ofc_zone=zone,
            ofc_division=division
        ).count()
        return Response({
            "status": False,
            "message": "Data not found",
            "tot_count": tot_count,
            "updated_count": updated_count
        })


@api_view(['POST'])
def updatemrfault(request):
    id = request.data.get("id")
    user = request.data.get("user")
    date = request.data.get("date")
    monthyear = request.data.get("month")
    zone = request.data.get("zone")
    division = request.data.get("division")
    ocrrdng = request.data.get("ocrrdng")
    month = int(monthyear.split('-')[1])
    year = int(monthyear.split('-')[0])

    now = datetime.now()

    formatted_date = date[:10]

    query_update_data = {
        'rdng_ocr_status_changed_by': f'vapp_{user}',
        'prsnt_ocr_excep_old_values': F('prsnt_rdng_ocr_excep'),
        'mtr_excep_img': f'vapp_{user}',
        'manual_update_flag': 'true',
        'qc_req': 'No',
        'qc_rmrk': 'MR Fault',
        'date_qc': formatted_date
    }

    with transaction.atomic():
        Consumers.objects.filter(id=id).update(**query_update_data)

    queryset = Consumers.objects.all()
    if month:
        queryset = queryset.filter(reading_date_db__month=month)
    else:
        current_month = now.month
        queryset = queryset.filter(reading_date_db__month=current_month)
    if year:
        queryset = queryset.filter(reading_date_db__year=year)
    else:
        current_year = now.year
        queryset = queryset.filter(reading_date_db__year=current_year)

    if zone:
        queryset = queryset.filter(ofc_zone=zone)

    if division:
        queryset = queryset.filter(ofc_division=division)

    if ocrrdng:
        queryset = queryset.filter(prsnt_ocr_rdng=ocrrdng)
    else:
        queryset = queryset.exclude(prsnt_ocr_rdng='Not Found')

    filter_conditions = Q(rdng_ocr_status='Failed', mtr_excep_img__isnull=True, manual_update_flag__isnull=True) & ~(
        Q(prsnt_rdng_ocr_excep='Spoofed Image') & Q(is_object_meter='NO and camera blocked'))
    tot_count = queryset.filter(filter_conditions).count()

    if tot_count > 0:
        updated_count = Consumers.objects.filter(
            mtr_excep_img=f'vapp_{user}',
            reading_date_db__month=month,
            reading_date_db__year=year,
            ofc_zone=zone,
            ofc_division=division
        ).count()

        return Response({
            "status": True,
            "message": "Updated Successfully",
            "updated_count": updated_count,
            "tot_count": tot_count
        })
    else:
        updated_count = Consumers.objects.filter(
            mtr_excep_img=f'vapp_{user}',
            reading_date_db__month=month,
            reading_date_db__year=year,
            ofc_zone=zone,
            ofc_division=division
        ).count()
        return Response({
            "status": False,
            "message": "Data not found",
            "tot_count": tot_count,
            "updated_count": updated_count
        })


@api_view(['POST'])
def updateddata(request):
    user = request.data.get("user", None)
    start_date = request.data.get("start_date", None)
    end_date = request.data.get("end_date", None)
    monthyear = request.data.get("monthyear", None)
    pagesize = request.data.get('pagesize', None)
    page = request.data.get("page",)
    offset = (int(pagesize) * int(page))-int(pagesize)

    cursor = connection.cursor()

    formatted_startdate = start_date[:10] if start_date and '/' not in start_date else None
    formatted_enddate = end_date[:10] if end_date and '/' not in end_date else None
    formated_month = monthyear[:2]
    formatted_year = monthyear[3:]

    subquery = f'''
    SELECT
        ofc_discom,
        ofc_zone,
        ofc_division,
        CASE WHEN rdng_ocr_status = 'Passed' THEN 1 ELSE 0 END AS yes_count,
        CASE WHEN rdng_ocr_status = 'Failed' and prsnt_rdng_ocr_excep != 'Spoofed Image' THEN 1 ELSE 0 END AS no_count,
        CASE WHEN rdng_ocr_status = 'Failed' and prsnt_rdng_ocr_excep = 'Spoofed Image' THEN 1 ELSE 0 END AS spoof_count,
        CASE WHEN rdng_ocr_status = 'Failed' and qc_rmrk = 'MR Fault' THEN 1 ELSE 0 END AS mrfault_count,
        date_qc
    FROM readingmaster
    WHERE mtr_excep_img = 'vapp_{user}' and date_qc is not null
    '''
    if formatted_startdate and formatted_enddate:
        subquery += f" AND SUBSTRING(date_qc FROM 1 FOR 10) BETWEEN '{formatted_startdate}' AND '{formatted_enddate}'"

    if formated_month and formatted_year:
        subquery += f" AND extract(month from reading_date_db)='{formated_month}' and extract(year from reading_date_db) = '{formatted_year}'"

    query = f'''
    SELECT
        ofc_discom,
        ofc_zone,
        ofc_division,
        SUM(yes_count) AS yes_count,
        SUM(no_count) AS no_count,
        SUM(spoof_count) AS spoof_count,
        SUM(mrfault_count) AS mrfault_count,
        SUM(yes_count + no_count + spoof_count + mrfault_count) as tot_count,
        date_qc,
        count(*) over() as total_count
    FROM ({subquery}) AS subquery
    GROUP BY ofc_discom, ofc_zone, ofc_division, date_qc
    ORDER BY date_qc DESC
    LIMIT {pagesize} OFFSET {offset}
    '''

    cursor.execute(query)
    person_objects = dictfetchall(cursor)

    if person_objects:
        total_count = person_objects[0]['total_count']
    else:
        total_count = 0

    print("totcount", total_count)

    if not person_objects:
        return Response({
            "status": False,
            "message": "Data not found",
            "data": person_objects
        })

    response_data = {
        "status": True,
        "message": "Data Fetched Successfully",
        "total_count": total_count,
        "data": person_objects
    }

    return Response(response_data)


@api_view(['POST'])
def search_by_mr(request):
    month = request.data.get('month', None)
    mrid = request.data.get('mrid', None)
    now = datetime.now()
    clause = ''
    where_clauses = []

    if month:
        month = month.split('-')[1]
        where_clauses.append(f"EXTRACT(MONTH FROM reading_date_db)='{month}'")
    else:
        current_month = now.month
        where_clauses.append(
            f"EXTRACT(MONTH FROM m.reading_date_db) = '{current_month}'")
    if mrid:
        where_clauses.append(f"m.mr_id like '{mrid}%'")
    if where_clauses:
        clause = ' AND '.join(where_clauses)
    cursor = connection.cursor()
    query = (
        f'''select distinct mr_id from readingmaster m left outer join meterreaderregistration r on m.mr_id=r."mrId" where {clause}''')
    print("query", query)
    cursor.execute(query)
    person_objects = dictfetchall(cursor)

    if not person_objects:
        return Response({
            "status": False,
            "message": "MR not found",
            "results": person_objects,
        })

    return Response({
        "status": True,
        "message": "MR Fetched Successfully",
        "results": person_objects,
    })

# ------------------------- QC IMAGES -------------------------------------------------------
# @api_view(['POST', 'GET'])
# def mr_failed_images(request):
#     monthyear = request.data.get("month")
#     pagesize = request.data.get("pagesize")
#     ocrrdng = request.data.get("ocrrdng")
#     page = request.data.get("page")
#     user = request.data.get("user")
#     mrid = request.data.get('mrid', None)
#     print(monthyear, ocrrdng, user, mrid)
#     month = int(monthyear.split('-')[1])
#     year = int(monthyear.split('-')[0])
#     now = datetime.now()

#     q_objects = Q(rdng_ocr_status='Failed', mtr_excep_img__isnull=True, manual_update_flag__isnull=True) & ~(
#         Q(prsnt_rdng_ocr_excep='Spoofed Image') & Q(is_object_meter='NO and camera blocked'))
#     if month:
#         q_objects &= Q(reading_date_db__month=month)
#     else:
#         q_objects &= Q(reading_date_db__month=now.month)
#     if year:
#         q_objects &= Q(reading_date_db__year=year)
#     else:
#         q_objects &= Q(reading_date_db__year=now.year)
#     if mrid:
#         q_objects &= Q(mr_id=mrid)
#     if ocrrdng:
#         q_objects &= Q(prsnt_ocr_rdng=ocrrdng)
#     else:
#         q_objects &= ~Q(prsnt_ocr_rdng='Not Found')

#     reading_master_objects = Consumers.objects.filter(q_objects).order_by('id')

#     paginator = Paginator(reading_master_objects, pagesize)
#     try:
#         paginated_results = paginator.page(page)
#     except Exception:
#         paginated_results = paginator.page(1)

#     count = reading_master_objects.count()

#     updated_count = Consumers.objects.filter(
#         mtr_excep_img=f'vapp_{user}',
#         reading_date_db__month=month,
#         reading_date_db__year=year,
#         mr_id=mrid
#     ).count()

#     serializer = FailedImageSerializer(paginated_results, many=True)

#     if not serializer.data:
#         return Response({
#             "status": False,
#             "message": "Data not found",
#             "count": count,
#             "update_count": updated_count,
#             "results": serializer.data,
#         })

#     return Response({
#         "status": True,
#         "message": "Data Fetched Successfully",
#         "count": count,
#         "update_count": updated_count,
#         "results": serializer.data,
#     })


@api_view(['POST'])
def mrupdatequery(request):
    id = request.data.get("id")
    user = request.data.get("user")
    date = request.data.get("date")
    monthyear = request.data.get("month")
    ocrrdng = request.data.get("ocrrdng")
    mrid = request.data.get("mrid", None)
    month = int(monthyear.split('-')[1])
    year = int(monthyear.split('-')[0])
    now = datetime.now()

    formatted_date = date[:10]

    query1_update_data = {
        'prsnt_rdng_ocr_odv': F('prsnt_ocr_rdng'),
        'rdng_ocr_status_odv': F('rdng_ocr_status'),
        'prsnt_ocr_excep_old_values': F('prsnt_rdng_ocr_excep'),
        'rdng_ocr_status_changed_by': f'vapp_{user}',
        'mtr_excep_img': f'vapp_{user}',
        'manual_update_flag': 'true',
        'qc_req': 'No',
        'date_qc': formatted_date
    }
    query2_update_data = {
        'prsnt_ocr_rdng': F('prsnt_rdng'),
        'rdng_ocr_status': 'Passed',
        'prsnt_rdng_ocr_excep': '',
    }

    with transaction.atomic():
        Consumers.objects.filter(id=id).update(**query1_update_data)
        Consumers.objects.filter(id=id).update(**query2_update_data)

    queryset = Consumers.objects.all()
    if month:
        queryset = queryset.filter(reading_date_db__month=month)
    else:
        current_month = now.month
        queryset = queryset.filter(reading_date_db__month=current_month)
    if year:
        queryset = queryset.filter(reading_date_db__year=year)
    else:
        current_year = now.year
        queryset = queryset.filter(reading_date_db__year=current_year)
    if mrid:
        queryset = queryset.filter(mr_id=mrid)

    if ocrrdng:
        queryset = queryset.filter(prsnt_ocr_rdng=ocrrdng)
    else:
        queryset = queryset.exclude(prsnt_ocr_rdng='Not Found')

    filter_conditions = Q(rdng_ocr_status='Failed', mtr_excep_img__isnull=True, manual_update_flag__isnull=True) & ~(
        Q(prsnt_rdng_ocr_excep='Spoofed Image') & Q(is_object_meter='NO and camera blocked'))

    tot_count = queryset.filter(filter_conditions).count()

    if tot_count > 0:
        updated_count = Consumers.objects.filter(
            mtr_excep_img=f'vapp_{user}',
            reading_date_db__month=month,
            reading_date_db__year=year,
            mr_id=mrid
        ).count()

        return Response({
            "status": True,
            "message": "Updated Successfully",
            "updated_count": updated_count,
            "tot_count": tot_count
        })
    else:
        updated_count = Consumers.objects.filter(
            mtr_excep_img=f'vapp_{user}',
            reading_date_db__month=month,
            reading_date_db__year=year,
            mr_id=mrid
        ).count()
        return Response({
            "status": False,
            "message": "Data not found",
            "tot_count": tot_count,
            "updated_count": updated_count
        })


@api_view(['POST'])
def mrupdatequery1(request):
    id = request.data.get("id")
    user = request.data.get("user")
    date = request.data.get("date")
    monthyear = request.data.get("month")
    ocrrdng = request.data.get("ocrrdng")
    mrid = request.data.get("mrid", None)
    month = int(monthyear.split('-')[1])
    year = int(monthyear.split('-')[0])
    now = datetime.now()

    formatted_date = date[:10]

    query_update_data = {
        'rdng_ocr_status_changed_by': f'vapp_{user}',
        'mtr_excep_img': f'vapp_{user}',
        'manual_update_flag': 'true',
        'qc_req': 'No',
        'date_qc': formatted_date
    }

    with transaction.atomic():
        Consumers.objects.filter(id=id).update(**query_update_data)

    queryset = Consumers.objects.all()
    if month:
        queryset = queryset.filter(reading_date_db__month=month)
    else:
        current_month = now.month
        queryset = queryset.filter(reading_date_db__month=current_month)
    if year:
        queryset = queryset.filter(reading_date_db__year=year)
    else:
        current_year = now.year
        queryset = queryset.filter(reading_date_db__year=current_year)

    if mrid:
        queryset = queryset.filter(mr_id=mrid)

    if ocrrdng:
        queryset = queryset.filter(prsnt_ocr_rdng=ocrrdng)
    else:
        queryset = queryset.exclude(prsnt_ocr_rdng='Not Found')

    filter_conditions = Q(rdng_ocr_status='Failed', mtr_excep_img__isnull=True, manual_update_flag__isnull=True) & ~(
        Q(prsnt_rdng_ocr_excep='Spoofed Image') & Q(is_object_meter='NO and camera blocked'))
    tot_count = queryset.filter(filter_conditions).count()

    if tot_count > 0:
        updated_count = Consumers.objects.filter(
            mtr_excep_img=f'vapp_{user}',
            reading_date_db__month=month,
            reading_date_db__year=year,
            mr_id=mrid
        ).count()

        return Response({
            "status": True,
            "message": "Updated Successfully",
            "updated_count": updated_count,
            "tot_count": tot_count
        })
    else:
        updated_count = Consumers.objects.filter(
            mtr_excep_img=f'vapp_{user}',
            reading_date_db__month=month,
            reading_date_db__year=year,
            mr_id=mrid
        ).count()
        return Response({
            "status": False,
            "message": "Data not found",
            "tot_count": tot_count,
            "updated_count": updated_count
        })


@api_view(['POST'])
def mrupdateddata(request):
    user = request.data.get("user", None)
    start_date = request.data.get("start_date", None)
    end_date = request.data.get("end_date", None)
    monthyear = request.data.get("monthyear", None)
    cursor = connection.cursor()
    mrid = request.data.get("mrid", None)
    formatted_startdate = start_date[:10] if start_date and '/' not in start_date else None
    formatted_enddate = end_date[:10] if end_date and '/' not in end_date else None
    formated_month = monthyear[:2]
    formatted_year = monthyear[3:]
    pagesize = request.data.get('pagesize', None)
    page = request.data.get("page",)

    offset = (int(pagesize) * int(page))-int(pagesize)

    subquery = f'''
    SELECT mr_id,
        CASE WHEN rdng_ocr_status = 'Passed' THEN 1 ELSE 0 END AS yes_count,
        CASE WHEN rdng_ocr_status = 'Failed' and prsnt_rdng_ocr_excep != 'Spoofed Image' THEN 1 ELSE 0 END AS no_count,
        CASE WHEN rdng_ocr_status = 'Failed' and prsnt_rdng_ocr_excep = 'Spoofed Image' THEN 1 ELSE 0 END AS spoof_count,
        CASE WHEN rdng_ocr_status = 'Failed' and qc_rmrk = 'MR Fault' THEN 1 ELSE 0 END AS mrfault_count,
        date_qc
    FROM readingmaster
    WHERE mtr_excep_img = 'vapp_{user}' and date_qc is not null
    '''
    if formatted_startdate and formatted_enddate:
        subquery += f" AND SUBSTRING(date_qc FROM 1 FOR 10) BETWEEN '{formatted_startdate}' AND '{formatted_enddate}'"

    if formated_month and formatted_year:
        subquery += f" AND extract(month from reading_date_db)='{formated_month}' and extract(year from reading_date_db) = '{formatted_year}'"

    if mrid:
        subquery += f"AND mr_id = '{mrid}'"

    query = f'''
    SELECT distinct
        mr_id,
        SUM(yes_count) AS yes_count,
        SUM(no_count) AS no_count,
        SUM(spoof_count) AS spoof_count,
        SUM(mrfault_count) AS mrfault_count,
        SUM(yes_count + no_count + spoof_count + mrfault_count) as tot_count,
        date_qc,
        count(*) over() as total_count
    FROM ({subquery}) AS subquery
    GROUP BY mr_id,date_qc
    ORDER BY date_qc DESC
    LIMIT {pagesize} OFFSET {offset}
    '''

    cursor.execute(query)
    person_objects = dictfetchall(cursor)

    if person_objects:
        total_count = person_objects[0]['total_count']
    else:
        total_count = 0

    if not person_objects:
        return Response({
            "status": False,
            "message": "Data not found",
            "data": person_objects
        })

    response_data = {
        "status": True,
        "message": "Data Fetched Successfully",
        "total_count": total_count,
        "data": person_objects
    }

    return Response(response_data)


@api_view(['POST'])
def mrupdatespoof(request):
    id = request.data.get("id")
    user = request.data.get("user")
    date = request.data.get("date")
    monthyear = request.data.get("month")
    ocrrdng = request.data.get("ocrrdng")
    mrid = request.data.get("mrid")
    month = int(monthyear.split('-')[1])
    year = int(monthyear.split('-')[0])
    now = datetime.now()

    formatted_date = date[:10]

    query_update_data = {
        'rdng_ocr_status_changed_by': f'vapp_{user}',
        'prsnt_ocr_excep_old_values': F('prsnt_rdng_ocr_excep'),
        'mtr_excep_img': f'vapp_{user}',
        'manual_update_flag': 'true',
        'qc_req': 'No',
        'prsnt_rdng_ocr_excep': 'Spoofed Image',
        'date_qc': formatted_date
    }

    with transaction.atomic():
        Consumers.objects.filter(id=id).update(**query_update_data)

    queryset = Consumers.objects.all()
    if month:
        queryset = queryset.filter(reading_date_db__month=month)
    else:
        current_month = now.month
        queryset = queryset.filter(reading_date_db__month=current_month)
    if year:
        queryset = queryset.filter(reading_date_db__year=year)
    else:
        current_year = now.year
        queryset = queryset.filter(reading_date_db__year=current_year)

    if mrid:
        queryset = queryset.filter(mr_id=mrid)

    if ocrrdng:
        queryset = queryset.filter(prsnt_ocr_rdng=ocrrdng)
    else:
        queryset = queryset.exclude(prsnt_ocr_rdng='Not Found')

    filter_conditions = Q(rdng_ocr_status='Failed', mtr_excep_img__isnull=True, manual_update_flag__isnull=True) & ~(
        Q(prsnt_rdng_ocr_excep='Spoofed Image') & Q(is_object_meter='NO and camera blocked'))
    tot_count = queryset.filter(filter_conditions).count()

    if tot_count > 0:
        updated_count = Consumers.objects.filter(
            mtr_excep_img=f'vapp_{user}',
            reading_date_db__month=month,
            reading_date_db__year=year,
            mr_id=mrid
        ).count()

        return Response({
            "status": True,
            "message": "Updated Successfully",
            "updated_count": updated_count,
            "tot_count": tot_count
        })
    else:
        updated_count = Consumers.objects.filter(
            mtr_excep_img=f'vapp_{user}',
            reading_date_db__month=month,
            reading_date_db__year=year,
            mr_id=mrid
        ).count()
        return Response({
            "status": False,
            "message": "Data not found",
            "tot_count": tot_count,
            "updated_count": updated_count
        })


@api_view(['POST'])
def mrupdatemrfault(request):
    id = request.data.get("id")
    user = request.data.get("user")
    date = request.data.get("date")
    monthyear = request.data.get("month")
    ocrrdng = request.data.get("ocrrdng")
    mrid = request.data.get("mrid")
    month = int(monthyear.split('-')[1])
    year = int(monthyear.split('-')[0])
    now = datetime.now()

    formatted_date = date[:10]

    query_update_data = {
        'rdng_ocr_status_changed_by': f'vapp_{user}',
        'prsnt_ocr_excep_old_values': F('prsnt_rdng_ocr_excep'),
        'mtr_excep_img': f'vapp_{user}',
        'manual_update_flag': 'true',
        'qc_req': 'No',
        'qc_rmrk': 'MR Fault',
        'date_qc': formatted_date
    }

    with transaction.atomic():
        Consumers.objects.filter(id=id).update(**query_update_data)

    queryset = Consumers.objects.all()
    if month:
        queryset = queryset.filter(reading_date_db__month=month)
    else:
        current_month = now.month
        queryset = queryset.filter(reading_date_db__month=current_month)
    if year:
        queryset = queryset.filter(reading_date_db__year=year)
    else:
        current_year = now.year
        queryset = queryset.filter(reading_date_db__year=current_year)

    if mrid:
        queryset = queryset.filter(mr_id=mrid)

    if ocrrdng:
        queryset = queryset.filter(prsnt_ocr_rdng=ocrrdng)
    else:
        queryset = queryset.exclude(prsnt_ocr_rdng='Not Found')

    filter_conditions = Q(rdng_ocr_status='Failed', mtr_excep_img__isnull=True, manual_update_flag__isnull=True) & ~(
        Q(prsnt_rdng_ocr_excep='Spoofed Image') & Q(is_object_meter='NO and camera blocked'))
    tot_count = queryset.filter(filter_conditions).count()

    if tot_count > 0:
        updated_count = Consumers.objects.filter(
            mtr_excep_img=f'vapp_{user}',
            reading_date_db__month=month,
            reading_date_db__year=year,
            mr_id=mrid
        ).count()

        return Response({
            "status": True,
            "message": "Updated Successfully",
            "updated_count": updated_count,
            "tot_count": tot_count
        })
    else:
        updated_count = Consumers.objects.filter(
            mtr_excep_img=f'vapp_{user}',
            reading_date_db__month=month,
            reading_date_db__year=year,
            mr_id=mrid
        ).count()
        return Response({
            "status": False,
            "message": "Data not found",
            "tot_count": tot_count,
            "updated_count": updated_count
        })


@api_view(['GET'])
def mrdownloadexcel(request):
    user = request.query_params.get("user", None)
    start_date = request.query_params.get("start_date", None)
    end_date = request.query_params.get("end_date", None)
    monthyear = request.query_params.get("monthyear", None)
    cursor = connection.cursor()
    mrid = request.query_params.get("mrid", None)
    formatted_startdate = None
    formatted_enddate = None

    if start_date != "DD/MM/YYYY":
        formatted_startdate = datetime.strptime(
            start_date, "%d/%m/%Y").strftime("%Y-%m-%d")

    if end_date != "DD/MM/YYYY":
        formatted_enddate = datetime.strptime(
            end_date, "%d/%m/%Y").strftime("%Y-%m-%d")
    formated_month = monthyear[:2]
    formatted_year = monthyear[3:]

    subquery = f'''
    SELECT mr_id,
        CASE WHEN rdng_ocr_status = 'Passed' THEN 1 ELSE 0 END AS yes_count,
        CASE WHEN rdng_ocr_status = 'Failed' THEN 1 ELSE 0 END AS no_count,
        CASE WHEN prsnt_rdng_ocr_excep = 'Spoofed Image' THEN 1 ELSE 0 END AS spoof_count,
        CASE WHEN rdng_ocr_status = 'Failed' and qc_rmrk = 'MR Fault' THEN 1 ELSE 0 END AS mrfault_count,
        date_qc
    FROM readingmaster
    WHERE mtr_excep_img = 'vapp_{user}' and date_qc <> ''
    '''
    if formatted_startdate and formatted_enddate:
        subquery += f" AND SUBSTRING(date_qc FROM 1 FOR 10) BETWEEN '{formatted_startdate}' AND '{formatted_enddate}'"

    if formated_month and formatted_year:
        subquery += f" AND extract(month from reading_date_db)='{formated_month}' and extract(year from reading_date_db) = '{formatted_year}'"

    if mrid:
        subquery += f"AND mr_id = '{mrid}'"

    query = f'''
    SELECT distinct
        mr_id,
        SUM(yes_count) AS yes_count,
        SUM(no_count) AS no_count,
        SUM(spoof_count) AS spoof_count,
        SUM(mrfault_count) AS mrfault_count,
        SUM(yes_count + no_count + spoof_count + mrfault_count) as tot_count,
        date_qc
    FROM ({subquery}) AS subquery
    GROUP BY mr_id,date_qc
    ORDER BY date_qc DESC
    '''
    cursor.execute(query)
    person_objects = dictfetchall(cursor)

    if not person_objects:
        return Response({
            "status": False,
            "message": "Data not found",
            "data": person_objects
        })

    wb = Workbook()
    ws = wb.active

    headers = list(person_objects[0].keys())
    ws.append(headers)
    for item in person_objects:
        values = list(item.values())
        ws.append(values)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=vapp_{user}.xlsx'
    wb.save(response)
    return response


@api_view(['GET'])
def downloadexceldata(request):
    user = request.query_params.get("user", None)
    start_date = request.query_params.get("start_date", None)
    end_date = request.query_params.get("end_date", None)
    monthyear = request.query_params.get("monthyear", None)
    cursor = connection.cursor()
    formatted_startdate = None
    formatted_enddate = None
    if start_date != "DD/MM/YYYY":
        formatted_startdate = datetime.strptime(
            start_date, "%d/%m/%Y").strftime("%Y-%m-%d")

    if end_date != "DD/MM/YYYY":
        formatted_enddate = datetime.strptime(
            end_date, "%d/%m/%Y").strftime("%Y-%m-%d")
    formated_month = monthyear[:2]
    formatted_year = monthyear[3:]
    print("data", start_date, end_date, monthyear)
    subquery = f'''
    SELECT
        ofc_discom,
        ofc_zone,
        ofc_division,
        CASE WHEN rdng_ocr_status = 'Passed' THEN 1 ELSE 0 END AS yes_count,
        CASE WHEN rdng_ocr_status = 'Failed' THEN 1 ELSE 0 END AS no_count,
        CASE WHEN prsnt_rdng_ocr_excep = 'Spoofed Image' THEN 1 ELSE 0 END AS spoof_count,
        CASE WHEN rdng_ocr_status = 'Failed' and qc_rmrk = 'MR Fault' THEN 1 ELSE 0 END AS mrfault_count,
        date_qc
    FROM readingmaster
    WHERE mtr_excep_img = 'vapp_{user}'
    '''
    if formatted_startdate and formatted_enddate:
        subquery += f" AND SUBSTRING(date_qc FROM 1 FOR 10) BETWEEN '{formatted_startdate}' AND '{formatted_enddate}'"

    if formated_month and formatted_year:
        subquery += f" AND extract(month from reading_date_db)='{formated_month}' and extract(year from reading_date_db) = '{formatted_year}'"

    query = f'''
    SELECT
        ofc_discom,
        ofc_zone,
        ofc_division,
        SUM(yes_count) AS yes_count,
        SUM(no_count) AS no_count,
        SUM(spoof_count) AS spoof_count,
        SUM(mrfault_count) AS mrfault_count,
        SUM(yes_count + no_count + spoof_count + mrfault_count) as tot_count,
        date_qc
    FROM ({subquery}) AS subquery
    GROUP BY ofc_discom, ofc_zone, ofc_division, date_qc
    ORDER BY date_qc DESC
    '''
    print("query", query)

    cursor.execute(query)
    person_objects = dictfetchall(cursor)

    if not person_objects:
        return Response({
            "status": False,
            "message": "Data not found",
            "data": person_objects
        })

    wb = Workbook()
    ws = wb.active

    headers = list(person_objects[0].keys())
    ws.append(headers)
    for item in person_objects:
        values = list(item.values())
        ws.append(values)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=vapp_{user}.xlsx'
    wb.save(response)
    return response


@api_view(['POST'])
def qcdashboard(request):
    print("request", request.data)
    month = request.data.get('month', None)
    year = request.data.get('year', None)
    dateqc = request.data.get('dateqc', None)
    now = datetime.now()

    queryset = Consumers.objects.all()

    if month:
        queryset = queryset.filter(reading_date_db__month=month)
    else:
        current_month = now.month
        queryset = queryset.filter(reading_date_db__month=current_month)

    if year:
        queryset = queryset.filter(reading_date_db__year=year)
    else:
        current_year = now.year
        queryset = queryset.filter(reading_date_db__year=current_year)

    if dateqc:
        queryset = queryset.filter(date_qc__icontains=dateqc)

    user_counts = {f'user{i}': Count('id', filter=Q(
        mtr_excep_img=f'vapp_user{i}')) for i in range(1, 16)}

    filtered_queryset = queryset.filter(
        Q(rdng_ocr_status='Failed', manual_update_flag__isnull=True) &
        ~Q(prsnt_rdng_ocr_excep='Spoofed Image') &
        ~Q(mtr_excep_img__startswith='vapp_user') &
        ~Q(is_object_meter='NO and camera blocked')
    )

    pending_count = filtered_queryset.count()

    result = queryset.aggregate(
        updated_count=Count('id', filter=Q(
            mtr_excep_img__startswith='vapp_user')),
        yes_count=Count('id', filter=Q(rdng_ocr_status='Passed',
                        mtr_excep_img__startswith='vapp_user')),
        no_count=Count('id', filter=Q(rdng_ocr_status='Failed', mtr_excep_img__startswith='vapp_user') & ~Q(
            prsnt_rdng_ocr_excep='Spoofed Image')),
        spoof_count=Count('id', filter=Q(rdng_ocr_status='Failed',
                          prsnt_rdng_ocr_excep='Spoofed Image', mtr_excep_img__startswith='vapp_user')),
        mrfault_count=Count('id', filter=Q(rdng_ocr_status='Failed',
                                           qc_rmrk='MR Fault', mtr_excep_img__startswith='vapp_user')),
        **user_counts
    )

    result['pending_count'] = pending_count
    result['tot_count'] = pending_count + result['updated_count']

    if result['tot_count'] == 0:
        return JsonResponse({
            "status": False,
            "message": "Data not found",
            "data": result
        })

    return JsonResponse({
        "status": True,
        "message": "Data Fetched Successfully",
        **result,
    })


@api_view(['POST'])
def qcreportdata(request):
    start_date = request.data.get("start_date", None)
    end_date = request.data.get("end_date", None)
    monthyear = request.data.get("monthyear", None)
    pagesize = request.data.get('pagesize', None)
    page = request.data.get("page",)
    offset = (int(pagesize) * int(page))-int(pagesize)

    cursor = connection.cursor()

    formatted_startdate = start_date[:10] if start_date and '/' not in start_date else None
    formatted_enddate = end_date[:10] if end_date and '/' not in end_date else None
    formated_month = monthyear[:2]
    formatted_year = monthyear[3:]

    subquery = f'''
    SELECT
        mtr_excep_img,
        CASE WHEN rdng_ocr_status = 'Passed' THEN 1 ELSE 0 END AS yes_count,
        CASE WHEN rdng_ocr_status = 'Failed' and prsnt_rdng_ocr_excep != 'Spoofed Image' THEN 1 ELSE 0 END AS no_count,
        CASE WHEN rdng_ocr_status = 'Failed' and prsnt_rdng_ocr_excep = 'Spoofed Image' THEN 1 ELSE 0 END AS spoof_count,
        CASE WHEN rdng_ocr_status = 'Failed' and qc_rmrk = 'MR Fault' THEN 1 ELSE 0 END AS mrfault_count,
        date_qc
    FROM readingmaster
    WHERE mtr_excep_img like 'vapp_user%' and date_qc is not null
    '''
    if formatted_startdate and formatted_enddate:
        subquery += f" AND SUBSTRING(date_qc FROM 1 FOR 10) BETWEEN '{formatted_startdate}' AND '{formatted_enddate}'"

    if formated_month and formatted_year:
        subquery += f" AND extract(month from reading_date_db)='{formated_month}' and extract(year from reading_date_db) = '{formatted_year}'"

    query = f'''
   WITH ranked_users AS (
    SELECT
        mtr_excep_img,
        SUM(yes_count) AS yes_count,
        SUM(no_count) AS no_count,
        SUM(spoof_count) AS spoof_count,
        SUM(mrfault_count) AS mrfault_count,
        SUM(yes_count + no_count + spoof_count + mrfault_count) as tot_count,
        date_qc,
        count(*) over() as total_count,
        ROW_NUMBER() OVER (PARTITION BY date_qc ORDER BY CAST(SUBSTRING(mtr_excep_img, 10) AS INTEGER)) AS user_rank
    FROM ({subquery}) AS subquery
    GROUP BY mtr_excep_img, date_qc
    )
    SELECT
        mtr_excep_img as rdng_ocr_status_changed_by,
        yes_count,
        no_count,
        spoof_count,
        mrfault_count,
        tot_count,
        date_qc,
        total_count
    FROM ranked_users
    ORDER BY date_qc DESC, user_rank
    LIMIT {pagesize} OFFSET {offset}
    '''
    cursor.execute(query)
    print(query)
    person_objects = dictfetchall(cursor)

    if person_objects:
        total_count = person_objects[0]['total_count']
    else:
        total_count = 0

    if not person_objects:
        return Response({
            "status": False,
            "message": "Data not found",
            "data": person_objects
        })

    response_data = {
        "status": True,
        "message": "Data Fetched Successfully",
        "total_count": total_count,
        "data": person_objects
    }

    return Response(response_data)


@api_view(['GET'])
def qcdailyreport(request):

    now = datetime.now()
    cursor = connection.cursor()

    subquery = f'''
    SELECT
        mtr_excep_img,
        CASE WHEN rdng_ocr_status = 'Passed' THEN 1 ELSE 0 END AS yes_count,
        CASE WHEN rdng_ocr_status = 'Failed' and prsnt_rdng_ocr_excep != 'Spoofed Image' THEN 1 ELSE 0 END AS no_count,
        CASE WHEN prsnt_rdng_ocr_excep = 'Spoofed Image' THEN 1 ELSE 0 END AS spoof_count,
        CASE WHEN rdng_ocr_status = 'Failed' and qc_rmrk = 'MR Fault' THEN 1 ELSE 0 END AS mrfault_count,
        date_qc
    FROM readingmaster
    WHERE mtr_excep_img like 'vapp_user%' and TO_DATE(date_qc, 'YYYY-MM-DD') = current_date
    '''

    query = f'''
    SELECT
        mtr_excep_img,
        SUM(yes_count) AS yes_count,
        SUM(no_count) AS no_count,
        SUM(spoof_count) AS spoof_count,
        SUM(mrfault_count) AS mrfault_count,
        SUM(yes_count + no_count + spoof_count + mrfault_count) as tot_count,
        date_qc
    FROM ({subquery}) AS subquery
    GROUP BY mtr_excep_img,date_qc
    ORDER BY
    CAST(SUBSTRING(mtr_excep_img, 10) AS INTEGER) ASC,
    date_qc DESC
    '''
    print("query", query)
    cursor.execute(query)
    person_objects = dictfetchall(cursor)

    if not person_objects:
        return Response({
            "status": False,
            "message": "Data not found",
            "data": person_objects
        })

    wb = Workbook()
    ws = wb.active

    headers = list(person_objects[0].keys())
    ws.append(headers)
    for item in person_objects:
        values = list(item.values())
        ws.append(values)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=dailydata.xlsx'
    wb.save(response)
    return response


@api_view(['GET'])
def qcmonthlyreport(request):
    now = datetime.now()
    month = request.query_params.get('month', None)
    year = request.query_params.get('year', None)
    if month == 'undefined':
        month = None
    if year == 'undefined':
        year = None
    cursor = connection.cursor()
    clause = ''

    where_clauses = []
    if month:
        where_clauses.append(f"EXTRACT(MONTH FROM reading_date_db)='{month}'")
    else:
        current_month = now.month
        where_clauses.append(
            f"EXTRACT(MONTH FROM reading_date_db) = '{current_month}'")
    if year:
        where_clauses.append(f"EXTRACT(YEAR FROM reading_date_db)='{year}'")
    else:
        current_year = now.year
        where_clauses.append(
            f"EXTRACT(YEAR FROM reading_date_db) = '{current_year}'")
    if where_clauses:
        clause = ' AND '.join(where_clauses)

    subquery = f'''
    SELECT
        mtr_excep_img,
        CASE WHEN rdng_ocr_status = 'Passed' THEN 1 ELSE 0 END AS yes_count,
        CASE WHEN rdng_ocr_status = 'Failed' and prsnt_rdng_ocr_excep != 'Spoofed Image' THEN 1 ELSE 0 END AS no_count,
        CASE WHEN prsnt_rdng_ocr_excep = 'Spoofed Image' THEN 1 ELSE 0 END AS spoof_count,
        CASE WHEN rdng_ocr_status = 'Failed' and qc_rmrk = 'MR Fault' THEN 1 ELSE 0 END AS mrfault_count
    FROM readingmaster
    WHERE mtr_excep_img like 'vapp_user%' and {clause}
    '''

    query = f'''
    SELECT
        mtr_excep_img,
        SUM(yes_count) AS yes_count,
        SUM(no_count) AS no_count,
        SUM(spoof_count) AS spoof_count,
        SUM(mrfault_count) AS mrfault_count,
        SUM(yes_count + no_count + spoof_count + mrfault_count) as tot_count
    FROM ({subquery}) AS subquery
    GROUP BY mtr_excep_img
    ORDER BY
    CAST(SUBSTRING(mtr_excep_img, 10) AS INTEGER) ASC
    '''
    print("query", query)
    cursor.execute(query)
    person_objects = dictfetchall(cursor)

    if not person_objects:
        return Response({
            "status": False,
            "message": "Data not found",
            "data": person_objects
        })

    wb = Workbook()
    ws = wb.active

    headers = list(person_objects[0].keys())
    ws.append(headers)
    for item in person_objects:
        values = list(item.values())
        ws.append(values)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=monthlydata.xlsx'
    wb.save(response)
    return response


@api_view(['POST'])
def userdashboard(request):
    user = request.data.get('user', None)
    now = datetime.now()
    today = date.today()
    current_week = today.isocalendar()[1]
    current_month = now.month
    last_month = (now.month - 1) if now.month > 1 else 12

    query_today = f'''
    SELECT
            period,
            SUM(CASE WHEN rdng_ocr_status = 'Passed' THEN 1 ELSE 0 END) AS yes_count,
            SUM(CASE WHEN rdng_ocr_status = 'Failed' AND prsnt_rdng_ocr_excep != 'Spoofed Image' THEN 1 ELSE 0 END) AS no_count,
            SUM(CASE WHEN rdng_ocr_status = 'Failed' AND prsnt_rdng_ocr_excep = 'Spoofed Image' THEN 1 ELSE 0 END) AS spoof_count,
            SUM(CASE WHEN rdng_ocr_status = 'Failed' AND qc_rmrk = 'MR Fault' THEN 1 ELSE 0 END) AS mrfault_count,
            COUNT(*) AS tot_count
        FROM (
            SELECT
                qc_rmrk,
                rdng_ocr_status,
                prsnt_rdng_ocr_excep,
                TO_DATE(date_qc, 'YYYY-MM-DD') AS date_qc,
                CASE
                    WHEN TO_DATE(date_qc, 'YYYY-MM-DD') = %s THEN 'today'
                    ELSE 'other'
                END AS period
            FROM readingmaster
            WHERE mtr_excep_img = %s
        ) AS subquery
        WHERE period != 'other'
        GROUP BY period
    '''

    query_week = f'''
    SELECT
            period,
            SUM(CASE WHEN rdng_ocr_status = 'Passed' THEN 1 ELSE 0 END) AS yes_count,
            SUM(CASE WHEN rdng_ocr_status = 'Failed' AND prsnt_rdng_ocr_excep != 'Spoofed Image' THEN 1 ELSE 0 END) AS no_count,
            SUM(CASE WHEN rdng_ocr_status = 'Failed' AND prsnt_rdng_ocr_excep = 'Spoofed Image' THEN 1 ELSE 0 END) AS spoof_count,
            SUM(CASE WHEN rdng_ocr_status = 'Failed' AND qc_rmrk = 'MR Fault' THEN 1 ELSE 0 END) AS mrfault_count,
            COUNT(*) AS tot_count
        FROM (
            SELECT
                qc_rmrk,
                rdng_ocr_status,
                prsnt_rdng_ocr_excep,
                TO_DATE(date_qc, 'YYYY-MM-DD') AS date_qc,
                CASE
                    WHEN EXTRACT(WEEK FROM TO_DATE(date_qc, 'YYYY-MM-DD')) = %s THEN 'week'
                    ELSE 'other'
                END AS period
            FROM readingmaster
            WHERE mtr_excep_img = %s
        ) AS subquery
        WHERE period != 'other'
        GROUP BY period
    '''

    query = f'''
        SELECT
            period,
            SUM(CASE WHEN rdng_ocr_status = 'Passed' THEN 1 ELSE 0 END) AS yes_count,
            SUM(CASE WHEN rdng_ocr_status = 'Failed' AND prsnt_rdng_ocr_excep != 'Spoofed Image' THEN 1 ELSE 0 END) AS no_count,
            SUM(CASE WHEN rdng_ocr_status = 'Failed' AND prsnt_rdng_ocr_excep = 'Spoofed Image' THEN 1 ELSE 0 END) AS spoof_count,
            SUM(CASE WHEN rdng_ocr_status = 'Failed' AND qc_rmrk = 'MR Fault' THEN 1 ELSE 0 END) AS mrfault_count,
            COUNT(*) AS tot_count
        FROM (
            SELECT
                qc_rmrk,
                rdng_ocr_status,
                prsnt_rdng_ocr_excep,
                TO_DATE(date_qc, 'YYYY-MM-DD') AS date_qc,
                CASE
                    WHEN extract(month from TO_DATE(date_qc, 'YYYY-MM-DD')) = %s THEN 'month'
                    WHEN extract(month from TO_DATE(date_qc, 'YYYY-MM-DD')) = %s THEN 'lastmonth'
                    ELSE 'other'
                END AS period
            FROM readingmaster
            WHERE mtr_excep_img = %s
        ) AS subquery
        WHERE period != 'other'
        GROUP BY period
    '''

    print("Query:", query_today, query)
    print("Parameters:", [today, current_week,
          current_month, last_month, f'vapp_{user}'])

    with connection.cursor() as cursor:

        cursor.execute(query_today, [today, f'vapp_{user}'])
        result_today = cursor.fetchone()

        cursor.execute(query_week, [current_week, f'vapp_{user}'])
        result_week = cursor.fetchone()

        cursor.execute(
            query, [current_month, last_month, f'vapp_{user}'])
        result_rows = cursor.fetchall()

    response_data = {
        "today": {"tot_count": 0, "yes_per": 0, "no_per": 0, "spoof_per": 0, "mrfault_per": 0},
        "week": {"tot_count": 0, "yes_per": 0, "no_per": 0, "spoof_per": 0, "mrfault_per": 0},
        "month": {"tot_count": 0, "yes_per": 0, "no_per": 0, "spoof_per": 0, "mrfault_per": 0},
        "lastmonth": {"tot_count": 0, "yes_per": 0, "no_per": 0, "spoof_per": 0, "mrfault_per": 0},
    }

    print(result_rows)

    if result_today:
        period = result_today[0]
        tot_count = result_today[5]
        print(response_data[period])
        if tot_count > 0:
            response_data[period]["tot_count"] = tot_count
            response_data[period]["yes_per"] = round(
                result_today[1] / tot_count * 100)
            response_data[period]["no_per"] = round(
                result_today[2] / tot_count * 100)
            response_data[period]["spoof_per"] = round(
                result_today[3] / tot_count * 100)
            response_data[period]["mrfault_per"] = round(
                result_today[4] / tot_count * 100)

    if result_week:
        period = result_week[0]
        tot_count = result_week[5]
        if tot_count > 0:
            response_data[period]["tot_count"] = tot_count
            response_data[period]["yes_per"] = round(
                result_week[1] / tot_count * 100)
            response_data[period]["no_per"] = round(
                result_week[2] / tot_count * 100)
            response_data[period]["spoof_per"] = round(
                result_week[3] / tot_count * 100)
            response_data[period]["mrfault_per"] = round(
                result_week[4] / tot_count * 100)

    for data in result_rows:
        period = data[0]
        tot_count = data[5]
        if tot_count > 0:
            response_data[period]["tot_count"] = tot_count
            response_data[period]["yes_per"] = round(data[1] / tot_count * 100)
            response_data[period]["no_per"] = round(data[2] / tot_count * 100)
            response_data[period]["spoof_per"] = round(
                data[3] / tot_count * 100)
            response_data[period]["mrfault_per"] = round(
                data[4] / tot_count * 100)

    return Response({
        "status": bool(response_data),
        "message": "Data Fetched Successfully" if response_data else "Data not found",
        "data": response_data,
    })


@api_view(['POST'])
def re_qcimages(request):
    user = request.data.get("user", None)
    start_date = request.data.get("start_date", None)
    end_date = request.data.get("end_date", None)
    qc_type = request.data.get("qc_type", None)
    page = request.data.get("page", 1)
    pagesize = request.data.get("pagesize", 10)

    if start_date and '/' not in start_date:
        formatted_startdate = datetime.strptime(
            start_date, '%a %b %d %Y %H:%M:%S %Z%z').strftime('%Y-%m-%d')
    if end_date and '/' not in end_date:
        formatted_enddate = datetime.strptime(
            end_date, '%a %b %d %Y %H:%M:%S %Z%z').strftime('%Y-%m-%d')
    print("formatteddate", formatted_startdate, formatted_enddate)
    cursor = connection.cursor()

    date_condition = ''
    date_params = []
    if formatted_startdate and formatted_enddate:
        date_condition = "AND date_qc::date BETWEEN %s AND %s"
        date_params = [formatted_startdate, formatted_enddate]

    qc_conditions = {
        'Yes': "rdng_ocr_status = 'Passed'",
        'No': "rdng_ocr_status = 'Failed' AND prsnt_rdng_ocr_excep != 'Spoofed Image'",
        'Spoof': "rdng_ocr_status = 'Failed' AND prsnt_rdng_ocr_excep = 'Spoofed Image'",
        'MRFault': "rdng_ocr_status = 'Failed' AND qc_rmrk = 'MR Fault'",
    }
    qc_condition = qc_conditions.get(qc_type, '')

    query_params = []
    if qc_type:
        query = f'''
        SELECT id, rdng_img, prsnt_rdng, prsnt_ocr_rdng, date_qc
        FROM readingmaster
        WHERE {qc_condition} {date_condition} AND mtr_excep_img = %s
        '''
        query_params.extend(date_params)
        query_params.append(f"vapp_{user}")
    else:
        query = f'''
        SELECT id, rdng_img, prsnt_rdng, prsnt_ocr_rdng,date_qc
        FROM readingmaster
        WHERE {date_condition} AND mtr_excep_img = %s
        '''
        query_params.extend(date_params)
        query_params.append(f"vapp_{user}")

    cursor.execute(query, query_params)
    person_objects = dictfetchall(cursor)

    if not person_objects:
        return Response({
            "status": False,
            "message": "Data not found",
            "results": person_objects,
        })

    paginator = Paginator(person_objects, pagesize)
    page_objects = paginator.get_page(page)

    serialized_page = [item for item in page_objects]

    for item in serialized_page:
        item['qc_type'] = qc_type

    return JsonResponse({
        "status": True,
        "message": "Data Fetched Successfully",
        "results": serialized_page,
    })


@api_view(['POST'])
def undo_img(request):
    id = request.data.get("id", None)
    qc_type = request.data.get("qc_type", None)
    cursor = connection.cursor()

    if qc_type == 'Yes':
        query = f'''
        UPDATE readingmaster
        SET
            prsnt_ocr_rdng = prsnt_rdng_ocr_odv,
            rdng_ocr_status = rdng_ocr_status_odv,
            prsnt_rdng_ocr_excep = prsnt_ocr_excep_old_values,
            qc_report_action = 'Yes'
        WHERE id = %s
        '''
        cursor.execute(query, [id])
        query1 = f'''
        UPDATE readingmaster
        SET
            prsnt_rdng_ocr_odv = '',
            rdng_ocr_status_odv = '',
            prsnt_ocr_excep_old_values = ''
        WHERE id = %s
        '''
        cursor.execute(query1, [id])
    elif qc_type == 'No':
        query = f'''
        UPDATE readingmaster
        SET qc_report_action = 'Yes'
        WHERE id = %s
        '''
        cursor.execute(query, [id])
    elif qc_type == 'Spoof':
        query = f'''
        UPDATE readingmaster
        SET
            qc_report_action = 'Yes',
            prsnt_rdng_ocr_excep = prsnt_ocr_excep_old_values
        WHERE id = %s
        '''
        cursor.execute(query, [id])
    elif qc_type == 'MRFault':
        query = f'''
        UPDATE readingmaster
        SET
            qc_report_action = 'Yes',
            prsnt_rdng_ocr_excep = prsnt_ocr_excep_old_values,
            qc_rmrk = ''
        WHERE id = %s
        '''
        cursor.execute(query, [id])
    else:
        return Response({
            "status": False,
            "message": "Invalid QC Type"
        })

    if not query:
        return Response({
            "status": False,
            "message": "Undo Failed"
        })

    return Response({
        "status": True,
        "message": "Undo Passed"
    })


@api_view(['GET', 'POST'])
def recheckdashboard(request):
    month = request.data.get('month', None)
    year = request.data.get('year', None)
    now = datetime.now()

    queryset = Consumers.objects.all()

    if month:
        queryset = queryset.filter(reading_date_db__month=month)
    else:
        current_month = now.month
        queryset = queryset.filter(reading_date_db__month=current_month)
    if year:
        queryset = queryset.filter(reading_date_db__year=year)
    else:
        current_year = now.year
        queryset = queryset.filter(reading_date_db__year=current_year)

    user_counts = {f'user{i}': Count('id', filter=Q(
        mtr_excep_img=f'vapp_user{i}', qc_report_action='Yes')) for i in range(1, 16)}

    result = queryset.aggregate(
        **user_counts
    )
    print(result, "result")

    if not result:
        return JsonResponse({
            "status": False,
            "message": "Data not found",
            "data": result
        })

    return JsonResponse({
        "status": True,
        "message": "Data Fetched Successfully",
        **result,
    })


@api_view(['POST'])
def user_reqc_images(request):
    pagesize = request.data.get("pagesize", None)
    page = request.data.get("page",)
    user = request.data.get("user", None)
    offset = (int(pagesize) * int(page))-int(pagesize)
    cursor = connection.cursor()
    cursor2 = connection.cursor()

    current_month_start = datetime.now().strftime('%Y-%m-%d')
    previous_month_start = (datetime.now().replace(
        day=1) - timedelta(days=1)).strftime('%Y-%m-01')

    query = (
        f'''select id,rdng_img,prsnt_rdng,prsnt_ocr_rdng from readingmaster where mtr_excep_img = 'vapp_{user}' and qc_report_action='Yes' AND date_qc::date BETWEEN '{previous_month_start}' AND '{current_month_start}' limit {pagesize} offset {offset} ''')
    print("query", query)
    cursor.execute(query)
    query2 = f'''
    SELECT
        query3.tot_count,
        query4.updated_count
    FROM
        (SELECT COUNT(*) AS tot_count FROM readingmaster WHERE mtr_excep_img = 'vapp_{user}' and qc_report_action='Yes' AND date_qc::date BETWEEN '{previous_month_start}' AND '{current_month_start}') AS query3,
        (SELECT COUNT(*) AS updated_count FROM readingmaster WHERE mtr_excep_img = 'vapp_{user}' and qc_report_action='No' AND date_qc::date BETWEEN '{previous_month_start}' AND '{current_month_start}') AS query4
'''
    print("query2", query2)
    cursor2.execute(query2)
    person_objects = dictfetchall(cursor)
    data1 = dictfetchall(cursor2)
    count = data1[0]['tot_count']
    count1 = data1[0]['updated_count']

    if not person_objects:
        return Response({
            "status": False,
            "message": "Data not found",
            "count": count,
            "update_count": count1,
            "results": person_objects,
        })

    return Response({
        "status": True,
        "message": "Data Fetched Successfully",
        "count": count,
        "update_count": count1,
        "results": person_objects,
    })


@api_view(['POST'])
def reqc_yes(request):
    cursor = connection.cursor()
    id = request.data.get("id", None)
    user = request.data.get("user", None)

    query1 = f'''
        UPDATE readingmaster
        set prsnt_rdng_ocr_odv=prsnt_ocr_rdng,
        rdng_ocr_status_odv=rdng_ocr_status,
        prsnt_ocr_excep_old_values=prsnt_rdng_ocr_excep,
        rdng_ocr_status_changed_by = 'vapp_{user}',
        mtr_excep_img= 'vapp_{user}',
        manual_update_flag = 'true',
        qc_req='No'
        WHERE id = {id}
    '''
    cursor.execute(query1)
    query2 = f'''
        UPDATE readingmaster
        SET prsnt_ocr_rdng = prsnt_rdng,
            rdng_ocr_status = 'Passed',
            prsnt_rdng_ocr_excep = '',
            qc_report_action= 'No'
            WHERE id = {id}
    '''
    cursor.execute(query2)

    query = f'''
    SELECT
        query3.tot_count,
        query4.updated_count
    FROM
        (SELECT COUNT(*) AS tot_count FROM readingmaster WHERE mtr_excep_img = 'vapp_{user}' and qc_report_action='Yes') AS query3,
        (SELECT COUNT(*) AS updated_count FROM readingmaster WHERE mtr_excep_img = 'vapp_{user}' and qc_report_action='No') AS query4
'''

    cursor.execute(query)
    result = cursor.fetchall()
    updated_count = result[0][1]
    tot_count = result[0][0]

    if not result:
        return Response({
            "status": False,
            "message": "Not Updated Successfully"
        })

    return Response({
        "status": True,
        "message": "Updated Successfully",
        "updated_count": updated_count,
        "tot_count": tot_count
    })


@api_view(['POST'])
def reqc_no(request):
    cursor = connection.cursor()
    id = request.data.get("id", None)
    user = request.data.get("user", None)

    query2 = f'''
        UPDATE readingmaster
        SET
            qc_req='No',
            manual_update_flag = 'true',
            rdng_ocr_status_changed_by = 'vapp_{user}',
            mtr_excep_img= 'vapp_{user}',
            qc_report_action= 'No'
        WHERE id = {id}
    '''
    cursor.execute(query2)
    query = f'''
    SELECT
        query3.tot_count,
        query4.updated_count
    FROM
        (SELECT COUNT(*) AS tot_count FROM readingmaster WHERE mtr_excep_img = 'vapp_{user}' and qc_report_action='Yes') AS query3,
        (SELECT COUNT(*) AS updated_count FROM readingmaster WHERE mtr_excep_img = 'vapp_{user}' and qc_report_action='No') AS query4
'''

    cursor.execute(query)
    result = cursor.fetchall()
    updated_count = result[0][1]
    tot_count = result[0][0]

    if not result:
        return Response({
            "status": False,
            "message": "Not Updated Successfully"
        })

    return Response({"status": True, "message": "Updated Successfully", "updated_count": updated_count,
                     "tot_count": tot_count})


@api_view(['POST'])
def reqc_spoof(request):
    cursor = connection.cursor()
    id = request.data.get("id", None)
    user = request.data.get("user", None)

    query2 = f'''
        UPDATE readingmaster
        SET
            prsnt_ocr_excep_old_values = prsnt_rdng_ocr_excep,
            prsnt_rdng_ocr_excep = 'Spoofed Image',
            qc_req='No',
            manual_update_flag = 'true',
            rdng_ocr_status_changed_by = 'vapp_{user}',
            mtr_excep_img= 'vapp_{user}',
            qc_report_action= 'No'
        WHERE id = {id}
    '''
    print("query2", query2)
    cursor.execute(query2)
    query = f'''
    SELECT
        query3.tot_count,
        query4.updated_count
    FROM
        (SELECT COUNT(*) AS tot_count FROM readingmaster WHERE mtr_excep_img = 'vapp_{user}' and qc_report_action='Yes') AS query3,
        (SELECT COUNT(*) AS updated_count FROM readingmaster WHERE mtr_excep_img = 'vapp_{user}' and qc_report_action='No') AS query4
'''
    print("query", query)
    cursor.execute(query)
    result = cursor.fetchall()
    updated_count = result[0][1]
    tot_count = result[0][0]

    if not result:
        return Response({
            "status": False,
            "message": "Data not found",
            "data": result
        })

    return Response({"status": True, "message": "Updated Successfully", "updated_count": updated_count,
                     "tot_count": tot_count})


@api_view(['POST'])
def reqc_mrfault(request):
    cursor = connection.cursor()
    id = request.data.get("id", None)
    user = request.data.get("user", None)

    query2 = f'''
        UPDATE readingmaster
        SET
            prsnt_ocr_excep_old_values = prsnt_rdng_ocr_excep,
            qc_rmrk = 'MR Fault',
            qc_req='No',
            manual_update_flag = 'true',
            rdng_ocr_status_changed_by = 'vapp_{user}',
            mtr_excep_img= 'vapp_{user}',
            qc_report_action= 'No'
        WHERE id = {id}
    '''
    print("query2", query2)
    cursor.execute(query2)
    query = f'''
    SELECT
        query3.tot_count,
        query4.updated_count
    FROM
        (SELECT COUNT(*) AS tot_count FROM readingmaster WHERE mtr_excep_img = 'vapp_{user}' and qc_report_action='Yes') AS query3,
        (SELECT COUNT(*) AS updated_count FROM readingmaster WHERE mtr_excep_img = 'vapp_{user}' and qc_report_action='No') AS query4
'''
    print("query", query)
    cursor.execute(query)
    result = cursor.fetchall()
    updated_count = result[0][1]
    tot_count = result[0][0]

    if not result:
        return Response({
            "status": False,
            "message": "Data not found",
            "data": result
        })

    return Response({"status": True, "message": "Updated Successfully", "updated_count": updated_count,
                     "tot_count": tot_count})


# QC WEB APIS START

@api_view(['POST'])
def webqcdashboard(request):
    month = request.data.get('month', None)
    year = request.data.get('year', None)
    now = datetime.now()
    print(month, "qcmonth")
    print(year, "qcyear")

    queryset = Consumers.objects.all()

    if month:
        queryset = queryset.filter(reading_date_db__month=month)
    else:
        current_month = now.month
        queryset = queryset.filter(reading_date_db__month=current_month)
    if year:
        queryset = queryset.filter(reading_date_db__year=year)
    else:
        current_year = now.year
        queryset = queryset.filter(reading_date_db__year=current_year)

    result = queryset.aggregate(
        tot_count=Count('id', filter=Q(
            rdng_ocr_status='Failed')),
        pending_count=Count('id', filter=Q(rdng_ocr_status='Failed', mtr_excep_img__isnull=True, manual_update_flag__isnull=True) & ~(
            Q(prsnt_rdng_ocr_excep='Spoofed Image') | Q(is_object_meter='NO and camera blocked'))),
        updated_count=Count('id', filter=Q(
            mtr_excep_img__startswith='vapp_user')),
        yes_count=Count('id', filter=Q(rdng_ocr_status='Passed',
                                       mtr_excep_img__startswith='vapp_user')),
        no_count=Count('id', filter=Q(rdng_ocr_status='Failed', mtr_excep_img__startswith='vapp_user') & ~Q(
            prsnt_rdng_ocr_excep='Spoofed Image')),
        spoof_count=Count('id', filter=Q(rdng_ocr_status='Failed', prsnt_rdng_ocr_excep='Spoofed Image',
                                         mtr_excep_img__startswith='vapp_user')),
    )
    print("resultquery", result)

    if not result:
        return JsonResponse({
            "status": False,
            "message": "Data not found",
            "data": result
        })

    return JsonResponse({
        "status": True,
        "message": "Data Fetched Successfully",
        **result,
    })


@api_view(['POST'])
def qcreportdata1(request):
    start_date = request.data.get("start_date", None)
    end_date = request.data.get("end_date", None)
    monthyear = request.data.get("monthyear", None)
    year = request.data.get("year", None)
    now = datetime.now()
    cursor = connection.cursor()

    formatted_startdate = start_date[:10] if start_date and '/' not in start_date else None
    formatted_enddate = end_date[:10] if end_date and '/' not in end_date else None
    formated_month = monthyear[:2]
    formatted_year = year
    print(" stat_date end_date month year", formatted_startdate,
          formatted_enddate, formated_month, formatted_year)

    date_condition = ''
    if formatted_startdate and formatted_enddate:
        date_condition = f" AND date_qc::date BETWEEN '{formatted_startdate}' AND '{formatted_enddate}'"

    month_year_condition = ''
    if formated_month and formatted_year:
        month_year_condition = f" AND extract(month from date_qc::date)='{formated_month}' and extract(year from date_qc::date) = '{formatted_year}'"
    else:
        month_year_condition = f"AND extract(month from date_qc::date)='{now.month}' and extract(year from date_qc::date) = '{now.year}'"

    date_qc = " AND date_qc <> ''"

    query = f'''
    SELECT
        date_qc,
        {''.join([f"COUNT(CASE WHEN mtr_excep_img = 'vapp_user{i}' THEN 1 END) AS user{i}, " for i in range(1, 16)])}
        count(*) as tot_count,
        count(*) over() as total_count
    FROM readingmaster
    WHERE 1 = 1 {date_qc} {date_condition} {month_year_condition}
    GROUP BY date_qc
    ORDER BY date_qc ASC
    '''
    # print("query", query)
    cursor.execute(query)
    person_objects = dictfetchall(cursor)

    if person_objects:
        total_count = person_objects[0]['total_count']
    else:
        total_count = 0

    print("totcount", total_count)

    if not person_objects:
        return Response({
            "status": False,
            "message": "Data not found",
            "data": person_objects
        })

    response_data = {
        "status": True,
        "message": "Data Fetched Successfully",
        "total_count": total_count,
        "data": person_objects
    }

    return Response(response_data)


@api_view(['GET'])
def qcdailyreport1(request):

    now = datetime.now()
    cursor = connection.cursor()

    subquery = f'''
    SELECT 
        mtr_excep_img,
        CASE WHEN rdng_ocr_status = 'Passed' THEN 1 ELSE 0 END AS yes_count,
        CASE WHEN rdng_ocr_status = 'Failed' and prsnt_rdng_ocr_excep != 'Spoofed Image' THEN 1 ELSE 0 END AS no_count,
        CASE WHEN prsnt_rdng_ocr_excep = 'Spoofed Image' THEN 1 ELSE 0 END AS spoof_count,
        date_qc
    FROM readingmaster
    WHERE mtr_excep_img like 'vapp_user%' and TO_DATE(date_qc, 'YYYY-MM-DD') = current_date
    '''

    query = f'''
    SELECT 
        mtr_excep_img,
        SUM(yes_count) AS yes_count,
        SUM(no_count) AS no_count,
        SUM(spoof_count) AS spoof_count,
        SUM(yes_count + no_count + spoof_count + mrfault_count) as tot_count,
        date_qc
    FROM ({subquery}) AS subquery
    GROUP BY mtr_excep_img,date_qc
    ORDER BY 
    CAST(SUBSTRING(mtr_excep_img, 10) AS INTEGER) ASC,
    date_qc DESC
    '''
    print("query", query)
    cursor.execute(query)
    person_objects = dictfetchall(cursor)

    if not person_objects:
        return Response({
            "status": False,
            "message": "Data not found",
            "data": person_objects
        })

    response_data = {
        "status": True,
        "message": "Data Fetched Successfully",
        "data": person_objects
    }

    return Response(response_data)


@api_view(['GET'])
def qcmonthlyreport1(request):
    now = datetime.now()
    month = request.query_params.get('month', now.month)
    cursor = connection.cursor()

    query = f'''
    SELECT
        date_qc,
        {''.join([f"COUNT(CASE WHEN mtr_excep_img = 'vapp_user{i}' THEN 1 END) AS user{i}, " for i in range(1, 16)])}
        count(*) as tot_count,
        count(*) over() as total_count
    FROM readingmaster
    WHERE mtr_excep_img like 'vapp_user%' and extract(month from date_qc::date) = '0{month}'
    GROUP BY date_qc
    ORDER BY date_qc DESC
    '''
    print("query", query)
    cursor.execute(query)
    person_objects = dictfetchall(cursor)
    print("data", person_objects)

    if not person_objects:
        return Response({
            "status": False,
            "message": "Data not found",
            "data": person_objects
        })

    response_data = {
        "status": True,
        "message": "Data Fetched Successfully",
        "data": person_objects
    }

    return Response(response_data)


@api_view(['GET'])
def qcdaywisereport1(request):
    now = datetime.now()
    month = request.query_params.get('month', now.month)
    cursor = connection.cursor()

    subquery = f'''
    SELECT 
        mtr_excep_img,
        CASE WHEN rdng_ocr_status = 'Passed' THEN 1 ELSE 0 END AS yes_count,
        CASE WHEN rdng_ocr_status = 'Failed' and prsnt_rdng_ocr_excep != 'Spoofed Image' THEN 1 ELSE 0 END AS no_count,
        CASE WHEN prsnt_rdng_ocr_excep = 'Spoofed Image' THEN 1 ELSE 0 END AS spoof_count,
        date_qc
    FROM readingmaster
    WHERE mtr_excep_img like 'vapp_user%' and extract(month from date_qc::date) = '0{month}'
    '''

    query = f'''
    WITH ranked_users AS (
    SELECT
        mtr_excep_img,
        SUM(yes_count) AS yes_count,
        SUM(no_count) AS no_count,
        SUM(spoof_count) AS spoof_count,
        SUM(yes_count + no_count + spoof_count + mrfault_count) as tot_count,
        date_qc,
        count(*) over() as total_count,
        ROW_NUMBER() OVER (PARTITION BY date_qc ORDER BY CAST(SUBSTRING(mtr_excep_img, 10) AS INTEGER)) AS user_rank
    FROM ({subquery}) AS subquery
    GROUP BY mtr_excep_img, date_qc
    )
    SELECT
        mtr_excep_img,
        tot_count,
        date_qc,
        total_count
    FROM ranked_users
    ORDER BY date_qc DESC, user_rank
    '''
    print("query", query)
    cursor.execute(query)
    person_objects = dictfetchall(cursor)

    if not person_objects:
        return Response({
            "status": False,
            "message": "Data not found",
            "data": person_objects
        })

    response_data = {
        "status": True,
        "message": "Data Fetched Successfully",
        "data": person_objects
    }

    return Response(response_data)


@api_view(['GET', 'POST'])
def getuserdata(request):
    user = request.data.get('user', None)
    month = request.data.get('month', None)
    year = request.data.get('year', None)
    print("useryear", year)
    now = datetime.now()
    cursor = connection.cursor()

    if user:
        user_condition = f"AND mtr_excep_img = 'vapp_{user}'"
        user_counts = f"COUNT(CASE WHEN mtr_excep_img = 'vapp_{user}' THEN 1 END) AS user_count"
    else:
        user_condition = ""
        user_counts = ', '.join(
            [f"COUNT(CASE WHEN mtr_excep_img = 'vapp_user{i}' THEN 1 END) AS user{i}" for i in range(1, 16)])
    if month:
        month_condition = f"extract(month from date_qc::date) = '{month}'"
    else:
        month_condition = f"extract(month from date_qc::date) = '{now.month}'"
    if year:
        year_condition = f"extract(year from date_qc::date) = '{year}'"
    else:
        year_condition = f"extract(year from date_qc::date) = '{now.year}'"
    query = f'''
    SELECT
        {user_counts},
        count(*) as tot_count
    FROM readingmaster
    WHERE 
       {month_condition} AND {year_condition}
        AND date_qc <> ''
        {user_condition}
    '''

    print("query", query)
    cursor.execute(query)
    person_objects = dictfetchall(cursor)

    if not person_objects:
        return Response({
            "status": False,
            "message": "Data not found",
            "data": person_objects
        })

    response_data = {
        "status": True,
        "message": "Data Fetched Successfully",
        "data": person_objects
    }

    return Response(response_data)


@api_view(['POST', 'GET'])
def downloadmrlist(request):
    print("request", request.query_params)
    now = datetime.now()
    month = request.query_params.get('month', now.month)
    year = request.query_params.get('year', now.year)
    image_type = request.query_params.get('image_type',  None)
    cursor = connection.cursor()

    if image_type == 'Not Found':
        ocr_reading = f"prsnt_ocr_rdng = 'Not Found'"
    elif image_type == 'Found':
        ocr_reading = f"prsnt_ocr_rdng != 'Not Found'"
    query = f'''
    select mr_id, count(*) as tot_count from readingmaster WHERE rdng_ocr_status = 'Failed'
and prsnt_rdng_ocr_excep != 'Spoofed Image' and manual_update_flag is null and is_object_meter!='NO and camera blocked'
AND EXTRACT(MONTH FROM reading_date_db)='{month}' AND EXTRACT(YEAR FROM reading_date_db)='{year}' AND {ocr_reading} 
group by mr_id order by tot_count desc
    '''
    print("query", query)
    cursor.execute(query)
    person_objects = dictfetchall(cursor)
    print("data", person_objects)

    if not person_objects:
        return Response({
            "status": False,
            "message": "Data not found",
            "data": person_objects
        })

    response_data = {
        "status": True,
        "message": "Data Fetched Successfully",
        "data": person_objects
    }

    return Response(response_data)


@api_view(['POST', 'GET'])
def downloaddivisionlist(request):
    print("request", request.query_params)
    now = datetime.now()
    month = request.query_params.get('month', now.month)
    year = request.query_params.get('year', now.year)
    discom = request.query_params.get('discom', 'NBPDCL')
    zone = request.query_params.get('zone', 'NORTH BIHAR RURAL')
    image_type = request.query_params.get('image_type',  None)
    cursor = connection.cursor()

    if image_type == 'Not Found':
        ocr_reading = f"prsnt_ocr_rdng = 'Not Found'"
    elif image_type == 'Found':
        ocr_reading = f"prsnt_ocr_rdng != 'Not Found'"
    query = f'''
    select ofc_division, count(*) as tot_count from readingmaster WHERE rdng_ocr_status = 'Failed'
and prsnt_rdng_ocr_excep != 'Spoofed Image' and manual_update_flag is null and is_object_meter!='NO and camera blocked'
AND EXTRACT(MONTH FROM reading_date_db)='{month}' AND EXTRACT(YEAR FROM reading_date_db)='{year}' AND ofc_discom = '{discom}' AND ofc_zone = '{zone}' AND {ocr_reading} 
group by ofc_division order by tot_count desc
    '''
    print("query", query)
    cursor.execute(query)
    person_objects = dictfetchall(cursor)
    print("data", person_objects)

    if not person_objects:
        return Response({
            "status": False,
            "message": "Data not found",
            "data": person_objects
        })

    response_data = {
        "status": True,
        "message": "Data Fetched Successfully",
        "data": person_objects
    }

    return Response(response_data)


# QC WEB API END

@api_view(['POST'])
def reconsilation(request):
    startdate = request.data.get("startDate", None)
    enddate = request.data.get("endDate", None)

    newdict = {}
    new = []

    def listfun(dict):
        new.append(dict.copy())
        return new

    with connection.cursor() as cursor:
        cursor.execute(f"""
            SELECT
    ROW_NUMBER() OVER () AS id,
    t2.reader_id,
    t2.reader_name,
    bl_agnc_name,
    ofc_zone,
    ofc_division,
    t2.sub_div_name,
    COALESCE(count_cons_ac_no::text, 'notfound') AS count_of_billed_cons,
    COALESCE(t2.tot_count::text, 'notfound') AS count_of_citra,
    COALESCE((t1.count_cons_ac_no::bigint - t2.tot_count::bigint)::text, 'notfound') AS difference
FROM (
    SELECT
        mr_id,
        bl_agnc_name,
        ofc_zone,
        ofc_division,
        COUNT(cons_ac_no) AS count_cons_ac_no
    FROM
        readingmaster
    WHERE
       reading_date_db between '{startdate}' and '{enddate}'
    GROUP BY
        mr_id,
        ofc_zone,
        bl_agnc_name,
        ofc_division
) AS t1
FULL OUTER JOIN (
    SELECT
        reader_id,
        reader_name,
        sub_div_name,
        tot_count
    FROM
        reconsilationtable
    GROUP BY
        reader_id,
        reader_name,
        sub_div_name,
        tot_count
) AS t2 ON t1.mr_id = t2.reader_id
WHERE
    t2.reader_id IS NOT NULL
ORDER BY id;
        """)
        results = cursor.fetchall()
        try:
            for row in results:
                newdict["id"] = row[0]
                newdict["reader_id"] = row[1]
                newdict["reader_name"] = row[2]
                newdict["agency"] = row[3]
                newdict["zone"] = row[4]
                newdict["ofc_division"] = row[5]
                newdict["sub_div_name"] = row[6]
                newdict["count_of_billed_cons"] = row[7]
                newdict["count_of_citra"] = row[8]
                newdict["difference"] = row[9]

                newdata = listfun(newdict)
            return Response({"data": newdata, "success": True})
        except:
            return Response([])


@api_view(['POST'])
def deletereconsilation(request):
    with connection.cursor() as cursor:
        cursor.execute("DELETE FROM reconsilationtable")
    return Response({"Message": "Data Deleted Successfully", "success": True})


@api_view(['POST'])
def uploadxlsx(request):
    if 'file' not in request.FILES:
        return JsonResponse({'error': 'No file attached'})
    file = request.FILES.get('file')
    workbook = load_workbook(file)
    worksheet = workbook.active
    cur = connection.cursor()
    for row in worksheet.iter_rows(min_row=2):
        print("row", row)
        values = [cell.value for cell in row]
        cur.execute('''INSERT INTO reconsilationtable (reader_id,sub_div_id,sub_div_name,reader_name,tot_count,ok_count,md_count,dl_count) VALUES (%s, %s, %s, %s, %s, %s, %s,%s)''', values)
    return JsonResponse({'success': True})


# ----------------------------------EndOF truereadqcapi--------------------------


@api_view(["POST"])
def originalimageApi(request):
    data = request.data.copy()
    mr_id = data.get("mr_id")
    mr_ids = []

    if mr_id in mr_ids:
        return Response({"status": True})
    else:
        return Response({"status": False})


@api_view(["GET", "POST"])
def location_wise_summary_of_agecy(request):
    agency = request.data.get("agency", None)
    month = date.today().month
    ofc_discom = request.data.get("ofc_discom", None)
    cursor = connection.cursor()
    clause = ""
    clause += f"AND r.bl_agnc_name='{agency}' " if (agency) else ""
    clause += f"AND r.ofc_discom='{ofc_discom}' " if (ofc_discom) else ""
    query = f""" select r.ofc_zone, count(DISTINCT r.mr_id), count(r.id),
    count(r.prsnt_mtr_status='Ok' OR NULL) as ok_readings,
    count(r.rdng_ocr_status='Passed' or NULL) as OCRwithoutException,
    count(r.rdng_ocr_status='Failed' OR NULL) as OCRwithException,
    count(r.prsnt_mtr_status='Meter Defective' OR NULL) as MeterDefective,
    count(r.prsnt_mtr_status='Door Locked' OR NULL) as DoorLocked
    from readingmaster r  where EXTRACT(month from r.reading_date_db)='{month}' {clause} group by r.ofc_zone
    """
    print("Query--------", query)
    location = []
    try:
        cursor.execute(query)
        res = cursor.fetchall()
        for row in res:
            zone = row[0]
            total_mr = row[1]
            total = row[2]
            ok_readings = row[3]
            ocr_passed = row[4]
            ocr_failed = row[5]
            md = row[6]
            dl = row[7]
            location.append(
                {
                    "locationname": zone,
                    "mrid": total_mr,
                    "total": total,
                    "okreadings": ok_readings,
                    "OcrReadings": ocr_passed,
                    "Ocrwithexception": ocr_failed,
                    "meterDefective": md,
                    "doorLocked": dl,
                    "okreadingspercent": math.floor(
                        ((ok_readings / total) if ok_readings else 0) * 100
                    ),
                    "OcrReadingspercent": math.floor(
                        ((ocr_passed / ok_readings) if ocr_passed else 0) * 100
                    ),
                    "Ocrwithexceptionpercent": math.floor(
                        ((ocr_failed / ok_readings) if ocr_failed else 0) * 100
                    ),
                    "meterDefectivepercent": math.floor(
                        ((md / total) if md else 0) * 100
                    ),
                    "doorLockedpercent": math.floor(((dl / total) if dl else 0) * 100),
                }
            )
        return Response(location)
    except Exception as e:
        print("Exception----------", e)
        return Response([])


# @api_view(["GET", "POST"])
# def meter_reading_summary_new(request):
#     mrid = request.data.get("mrid", None)
#     agency = request.data.get("agency", None)
#     ofc_zone = request.data.get("ofc_zone", None)
#     # groupby=request.data.get('groupby', None)
#     month = request.data.get("month", None)
#     clause = f" WHERE "
#     clause += (
#         f" EXTRACT(MONTH from reading_date_db)='{month.split('-')[0]}' "
#         if (month)
#         else f" EXTRACT(MONTH from reading_date_db)='{date.today().month}' "
#     )
#     clause += f" AND bl_agnc_name='{agency}' " if (agency) else ""
#     clause += f" AND r.ofc_zone = '{ofc_zone}' " if (ofc_zone) else ""
#     cursor = connection.cursor()
#     query = f""" select  r.mr_id, count(r.mr_id),
#     count(r.prsnt_mtr_status='Ok' OR NULL) as ok_readings,
#     count(r.rdng_ocr_status='Passed' or NULL) as OCRwithoutException,
#     count(r.rdng_ocr_status='Failed' OR NULL) as OCRwithException,
#     count(r.prsnt_mtr_status='Meter Defective' OR NULL) as MeterDefective,
#     count(r.prsnt_mtr_status='Door Locked' OR NULL) as DoorLocked
#     from readingmaster r {clause} and r.mr_id!='' group by  r.mr_id
#     """
#     print("Query--------", query)
#     location = []
#     count = 0
#     try:
#         cursor.execute(query)
#         res = cursor.fetchall()
#         for row in res:
#             mrid = row[0]
#             total = row[1]
#             ok_readings = row[2]
#             ocr_passed = row[3]
#             ocr_failed = row[4]
#             md = row[5]
#             dl = row[6]
#             location.append(
#                 {
#                     "mrid": mrid,
#                     "totalReadings": total,
#                     "OKreadings": ok_readings,
#                     "OCRReadings": ocr_passed,
#                     "OCRwithException": ocr_failed,
#                     "MeterDefective": md,
#                     "DoorLocked": dl,
#                     "OKreadingspercent": math.floor(
#                         ((ok_readings / total) if ok_readings else 0) * 100
#                     ),
#                     "OCRReadingspercent": math.floor(
#                         ((ocr_passed / ok_readings) if ocr_passed else 0) * 100
#                     ),
#                     "OCRwithExceptionpercent": math.floor(
#                         ((ocr_failed / ok_readings) if ocr_failed else 0) * 100
#                     ),
#                     "MeterDefectivepercent": math.floor(
#                         ((md / total) if md else 0) * 100
#                     ),
#                     "DoorLockedpercent": math.floor(((dl / total) if dl else 0) * 100),
#                 }
#             )

#             count += 1

#         return Response(location)
#     except Exception as e:
#         print("Exception----------", e)
#         return Response([])


@api_view(["GET", "POST"])
def meter_reading_summary_new(request):
    mrid = request.data.get("mrid", None)
    agency = request.data.get("agency", None)
    ofc_zone = request.data.get("ofc_zone", None)
    month = request.data.get("month", None)

    # to check current month and previous month
    current_month = date.today().month
    previous_month = (date.today() - timedelta(days=30)).month

    # decide which table to use based on the month
    if month:
        if month == f"{current_month:02d}" or month == f"{previous_month:02d}":
            table_name = "readingmaster"
        else:
            table_name = "prevmonthsdata"
    else:
        table_name = "readingmaster"
    clause = f" WHERE "
    clause += (
        f" EXTRACT(MONTH from reading_date_db)='{month.split('-')[0]}' "
        if (month)
        else f" EXTRACT(MONTH from reading_date_db)='{current_month}' "
    )
    clause += f" AND bl_agnc_name='{agency}' " if (agency) else ""
    clause += f" AND r.ofc_zone = '{ofc_zone}' " if (ofc_zone) else ""
    cursor = connection.cursor()
    query = f""" select  r.mr_id, count(r.mr_id),
    count(r.prsnt_mtr_status='Ok' OR NULL) as ok_readings,
    count(r.rdng_ocr_status='Passed' or NULL) as OCRwithoutException,
    count(r.rdng_ocr_status='Failed' OR NULL) as OCRwithException,
    count(r.prsnt_mtr_status='Meter Defective' OR NULL) as MeterDefective,
    count(r.prsnt_mtr_status='Door Locked' OR NULL) as DoorLocked
    from {table_name} r {clause} and r.mr_id!='' group by  r.mr_id
    """
    print("Query--------", query)
    location = []
    count = 0
    try:
        cursor.execute(query)
        res = cursor.fetchall()
        for row in res:
            mrid = row[0]
            total = row[1]
            ok_readings = row[2]
            ocr_passed = row[3]
            ocr_failed = row[4]
            md = row[5]
            dl = row[6]
            location.append(
                {
                    "mrid": mrid,
                    "totalReadings": total,
                    "OKreadings": ok_readings,
                    "OCRReadings": ocr_passed,
                    "OCRwithException": ocr_failed,
                    "MeterDefective": md,
                    "DoorLocked": dl,
                    "OKreadingspercent": math.floor(
                        ((ok_readings / total) if ok_readings else 0) * 100
                    ),
                    "OCRReadingspercent": math.floor(
                        ((ocr_passed / ok_readings) if ocr_passed else 0) * 100
                    ),
                    "OCRwithExceptionpercent": math.floor(
                        ((ocr_failed / ok_readings) if ocr_failed else 0) * 100
                    ),
                    "MeterDefectivepercent": math.floor(
                        ((md / total) if md else 0) * 100
                    ),
                    "DoorLockedpercent": math.floor(((dl / total) if dl else 0) * 100),
                }
            )

            count += 1

        return Response(location)
    except Exception as e:
        print("Exception----------", e)
        return Response([])


# --------------------------------------------------------------------------------------------------------------------------------
# MATERIALIZED VIEWS APIS
@api_view(["POST"])
def minidashboardmonth1(request):
    cursor = connection.cursor()
    filters = request.data.get("filters", {})
    
    agency = filters.get("agency")
    ofc_discom = filters.get("ofc_discom")
    has_year_month = "year" in filters and "month" in filters

    if has_year_month:
        year = int(filters["year"])
        month = int(filters["month"])
        
        # Build conditions specifically for readingmaster
        conditions = []
        if agency:
            conditions.append(f"bl_agnc_name='{agency}'")
        if ofc_discom:
            conditions.append(f"ofc_discom='{ofc_discom}'") # Table uses ofc_discom
        
        extra_filters = " AND " + " AND ".join(conditions) if conditions else ""

        query = f"""
        SELECT
            COUNT(DISTINCT mr_id),
            COUNT(*) AS totalreading,
            COUNT(*) FILTER (WHERE prsnt_mtr_status = 'Ok'),
            COUNT(*) FILTER (WHERE rdng_ocr_status = 'Passed'),
            COUNT(*) FILTER (WHERE rdng_ocr_status = 'Failed'),
            COUNT(*) FILTER (WHERE prsnt_mtr_status = 'Meter Defective'),
            COUNT(*) FILTER (WHERE prsnt_mtr_status = 'Door Locked')
        FROM readingmaster
        WHERE
            reading_date_db >= DATE '{year}-{month:02d}-01'
            AND reading_date_db < (DATE '{year}-{month:02d}-01' + INTERVAL '1 month')
            {extra_filters};
        """
    else:
        conditions = []
        if agency:
            conditions.append(f"bl_agnc_name='{agency}'")
        if ofc_discom:
            conditions.append(f"discom='{ofc_discom}'") # View uses discom
        
        where_clause = " WHERE " + " AND ".join(conditions) if conditions else ""

        query = f"""
        SELECT
            count(distinct meter_reader_id),
            SUM(TotalReading),
            SUM(Ok),
            SUM(Passed),
            SUM(Failed),
            SUM(MeterDefective),
            SUM(DoorLocked)
        FROM combinedmonth_view
        {where_clause}
        """

    # print("minidashboardmonth1 query:", query)
    cursor.execute(query)
    row = cursor.fetchone()

    if not row or row[0] is None:
        return Response({
            "mrid": 0, "totalreadings": 0, "okreadings": 0, 
            "ocrreadings": 0, "ocrwithexception": 0, 
            "meterdefective": 0, "doorlocked": 0
        })

    return Response({
        "mrid": int(row[0] or 0),
        "totalreadings": int(row[1] or 0),
        "okreadings": int(row[2] or 0),
        "ocrreadings": int(row[3] or 0),
        "ocrwithexception": int(row[4] or 0),
        "meterdefective": int(row[5] or 0),
        "doorlocked": int(row[6] or 0),
    })
# @api_view(["POST"])
# def dashboardagencywise1(request):
#     today = date.today()
#     thismonth = today.strftime("%Y-%m")
#     year = thismonth.split("-")[0]
#     month = thismonth.split("-")[1]
#     new = []

#     clause = ""
#     cursor = connection.cursor()
#     filters = request.data.get("filters", None)

#     def listfun(dict):
#         print("dict---->", dict)
#         if (
#             dict["agency"] == "BCITS"
#             or dict["agency"] == "DATA INGENIOUS"
#             or dict["agency"] == "Fluent Grid"
#             or dict["agency"] == "Quess Corp(Ikya Rural)"
#         ):
#             pass
#         else:
#             new.append(dict.copy())
#         return new

#     try:
#         if filters:
#             clause += " Where "
#             for i, (key, value) in enumerate(filters.items()):
#                 if i > 0:
#                     clause += "AND "

#                 if key == "agency":
#                     clause += f"agency='{value}'"

#                 if key == "ofc_discom":
#                     clause += f"discom='{value}'"

#     except:
#         pass
#     newdict = {}

#     #     query = (f"""select
#     #              agency,
#     #              SUM(billed_consumers),
#     #              SUM(ok_readings),
#     #              SUM(OCRwithoutException),
#     # SUM(OCRwithException),SUM(MeterDefective),SUM(DoorLocked)
#     # from my_materialized_view where month={month} {clause}group by agency
#     #     """)
#     query = f"""
#            select 
#         agency,
#         SUM(TotalReading),
#         SUM(Ok),
#         SUM(Passed),
#         SUM(Failed),
#         SUM(MeterDefective),SUM(DoorLocked)
#         from combinedmonth_view {clause} group by agency
#        """

#     print("query", query)

#     cursor.execute(query)
#     result = cursor.fetchall()
#     try:
#         for row in result:
#             total = row[1]
#             okreadings = row[2]
#             ocrreadings = row[3]
#             ocrwithexcep = row[4]
#             meterdefective = row[5]
#             doorlocked = row[6]
#             # Percentage
#             okreadpercent = math.floor((okreadings / total) * 100)
#             ocrreadingpercent = math.floor(
#                 ((ocrreadings / okreadings) if okreadings else 0) * 100
#             )
#             ocrwithexceppercent = math.floor(
#                 ((ocrwithexcep / okreadings) if okreadings else 0) * 100
#             )
#             meterdefectivepercent = math.floor((meterdefective / total) * 100)
#             doorlockedpercent = math.floor((doorlocked / total) * 100)

#             # add to dictionary
#             newdict["agency"] = row[0]
#             newdict["totalReadings"] = int(row[1])
#             newdict["OKreadings"] = int(okreadings)
#             newdict["OKreadingspercent"] = okreadpercent
#             newdict["OCRReadings"] = int(ocrreadings)
#             newdict["OCRReadingspercent"] = ocrreadingpercent
#             newdict["OCRwithException"] = int(ocrwithexcep)
#             newdict["OCRwithExceptionpercent"] = ocrwithexceppercent
#             newdict["MeterDefective"] = int(meterdefective)
#             newdict["MeterDefectivepercent"] = meterdefectivepercent
#             newdict["DoorLocked"] = int(doorlocked)
#             newdict["DoorLockedpercent"] = doorlockedpercent
#             # add to list
#             newdata = listfun(newdict)
#         return Response(newdata)
#     except:
#         return Response([])

@api_view(["POST"])
def dashboardagencywise1(request):
    cursor = connection.cursor()
    filters = request.data.get("filters", {})
    new = []

    has_year_month = "year" in filters and "month" in filters
    agency_filter = filters.get("agency")
    discom_filter = filters.get("ofc_discom")

    def listfun(d):
        if d["agency"] not in [
            "BCITS",
            "DATA INGENIOUS",
            "Fluent Grid",
            "Quess Corp(Ikya Rural)",
        ]:
            new.append(d.copy())
        return new

    if has_year_month:
        year = int(filters["year"])
        month = int(filters["month"])

        # Table-specific conditions
        conditions = []
        if agency_filter:
            conditions.append(f"bl_agnc_name='{agency_filter}'")
        if discom_filter:
            conditions.append(f"ofc_discom='{discom_filter}'") # Table uses ofc_discom
        
        clause = " AND " + " AND ".join(conditions) if conditions else ""

        query = f"""
        SELECT
            bl_agnc_name AS agency,
            count(*) AS total,
            count(CASE WHEN prsnt_mtr_status='Ok' THEN 1 END) AS ok,
            count(CASE WHEN rdng_ocr_status='Passed' THEN 1 END) AS passed,
            count(CASE WHEN rdng_ocr_status='Failed' THEN 1 END) AS failed,
            count(CASE WHEN prsnt_mtr_status='Meter Defective' THEN 1 END) AS meterdefective,
            count(CASE WHEN prsnt_mtr_status='Door Locked' THEN 1 END) AS doorlocked
        FROM readingmaster
        WHERE
            reading_date_db >= DATE '{year}-{month:02d}-01'
            AND reading_date_db < (DATE '{year}-{month:02d}-01' + INTERVAL '1 month')
            {clause}
        GROUP BY bl_agnc_name
        """
    else:
        # View-specific conditions
        conditions = []
        if agency_filter:
            conditions.append(f"agency='{agency_filter}'")
        if discom_filter:
            conditions.append(f"discom='{discom_filter}'") # View uses discom
        
        where_clause = " WHERE " + " AND ".join(conditions) if conditions else ""

        query = f"""
        SELECT
            agency,
            SUM(TotalReading),
            SUM(Ok),
            SUM(Passed),
            SUM(Failed),
            SUM(MeterDefective),
            SUM(DoorLocked)
        FROM combinedmonth_view
        {where_clause}
        GROUP BY agency
        """

    # print("dashboardagencywise1", query)
    cursor.execute(query)
    result = cursor.fetchall()

    try:
        for row in result:
            total = row[1] or 0
            ok = row[2] or 0
            passed = row[3] or 0
            failed = row[4] or 0
            meterdef = row[5] or 0
            doorlocked = row[6] or 0

            data = {
                "agency": row[0],
                "totalReadings": int(total),
                "OKreadings": int(ok),
                "OKreadingspercent": math.floor((ok / total) * 100) if total else 0,
                "OCRReadings": int(passed),
                "OCRReadingspercent": math.floor((passed / ok) * 100) if ok else 0,
                "OCRwithException": int(failed),
                "OCRwithExceptionpercent": math.floor((failed / ok) * 100) if ok else 0,
                "MeterDefective": int(meterdef),
                "MeterDefectivepercent": math.floor((meterdef / total) * 100) if total else 0,
                "DoorLocked": int(doorlocked),
                "DoorLockedpercent": math.floor((doorlocked / total) * 100) if total else 0,
            }
            listfun(data)

        return Response(new)
    except Exception as e:
        print(f"Error: {e}")
        return Response([])

# @api_view(['POST'])
# def exceptionlist1(request):
#     today = date.today()
#     thismonth = today.strftime('%Y-%m')
#     year = thismonth.split('-')[0]
#     clause = ''
#     cursor = connection.cursor()
#     filters = request.data.get("filters",None)

#     try:
#         if filters:
#             clause +=' Where '
#             for i,(key,value) in enumerate(filters.items()):
#                 if i>0:
#                     clause +='AND '

#                 if key=='agency':
#                     clause +=f"agency='{value}'"
#                     pass
#                 if key=='ofc_discom':
#                     clause +=f"discom='{value}'"
#                     pass
#     except:
#         pass
# #     query=f'''
# #     SELECT
# #     CAST(SUM(total_meters)AS INTEGER)as total,
# #     CAST(SUM(ok_readings)AS INTEGER)as ok_readings,
# #     CAST(SUM(passed)AS INTEGER)as passed,
# #     CAST(SUM(failed)AS INTEGER)as failed,
# #     CAST(SUM(incorrect_reading)AS INTEGER)as "Incorrect Reading",
# #     CAST(SUM(image_spoofed)AS INTEGER)as "Image Spoofed",
# #     CAST(SUM(image_blur)AS INTEGER)as "Image blur",
# #     CAST(SUM(parameterunavailable)AS INTEGER)as "Parameters Unavailable",
# #     CAST(SUM(meterdirty)AS INTEGER)as "Meter Dirty",
# #     CAST(SUM(parametersmismatch)AS INTEGER)as "Parameters Mismatch"
# #     FROM exception_materialized_view
# #     WHERE TRIM(month) = 'June' {clause} and year={year}

# # '''

#     query=f"""
#     SELECT
#     CAST(SUM(TotalReading)AS INTEGER)as total,
#     CAST(SUM(Ok)AS INTEGER)as ok_readings,
#     CAST(SUM(Passed)AS INTEGER)as passed,
#     CAST(SUM(Failed)AS INTEGER)as failed,
#     CAST(SUM(incorrectreading)AS INTEGER)as "Incorrect Reading",
#     CAST(SUM(spoofedimage)AS INTEGER)as "Image Spoofed",
#     CAST(SUM(imageblur)AS INTEGER)as "Image blur",
#     CAST(SUM(parametersunavailable)AS INTEGER)as "Parameters Unavailable",
#     CAST(SUM(meterdirty)AS INTEGER)as "Meter Dirty",
#     CAST(SUM(parametersmismatch)AS INTEGER)as "Parameters Mismatch"
#     FROM currentmonth_materialized_view {clause}


#     """
#     print("QUERY-->",query)
#     cursor.execute(query)
#     result = dictfetchall(cursor)

#     return Response(result)


# @api_view(["POST"])
# def exceptionlist1(request):
#     today = date.today()
#     thismonth = today.strftime("%Y-%m")
#     year = thismonth.split("-")[0]
#     clause = ""
#     cursor = connection.cursor()
#     filters = request.data.get("filters", None)
#     result = {}

#     try:
#         if filters:
#             clause += " WHERE "
#             for i, (key, value) in enumerate(filters.items()):
#                 if i > 0:
#                     clause += "AND "

#                 if key == "agency":
#                     clause += f"agency='{value}'"
#                     pass
#                 if key == "ofc_discom":
#                     clause += f"discom='{value}'"
#                     pass
#     except:
#         pass

#     query = f"""
#    SELECT
#     CAST(SUM(TotalReading)AS INTEGER)as total,
#     CAST(SUM(Ok)AS INTEGER)as ok_readings,
#     CAST(SUM(Passed)AS INTEGER)as passed,
#     CAST(SUM(Failed)AS INTEGER)as failed,
#     CAST(SUM(incorrectreading)AS INTEGER)as "Incorrect Reading",
#     CAST(SUM(spoofedimage)AS INTEGER)as "Image Spoofed",
#     CAST(SUM(imageblur)AS INTEGER)as "Image blur",
#     CAST(SUM(parametersunavailable)AS INTEGER)as "Parameters Unavailable",
#     CAST(SUM(meterdirty)AS INTEGER)as "Meter Dirty",
#     CAST(SUM(parametersmismatch)AS INTEGER)as "Parameters Mismatch"
#     FROM combinedmonth_view {clause}
#     """

#     print("QUERY-->", query)
#     cursor.execute(query)
#     row = cursor.fetchone()

#     if row:
#         result = {
#             "total": row[0],
#             "ok_readings": row[1],
#             "passed": row[2],
#             "failed": row[3],
#             "Incorrect Reading": row[4],
#             "Image Spoofed": row[5],
#             "Image blur": row[6],
#             "Parameters Unavailable": row[7],
#             "Meter Dirty": row[8],
#             "Parameters Mismatch": row[9],
#         }

#     return Response({"total": result["total"], "data": result})

@api_view(["POST"])
def exceptionlist1(request):
    cursor = connection.cursor()
    filters = request.data.get("filters", {})
    conditions = []

    has_year_month = "year" in filters and "month" in filters

    if "agency" in filters:
        conditions.append(f"agency='{filters['agency']}'")
    if "ofc_discom" in filters:
        conditions.append(f"discom='{filters['ofc_discom']}'")

    clause = ""
    if conditions:
        clause = " WHERE " + " AND ".join(conditions)

    if has_year_month:
        year = int(filters["year"])
        month = int(filters["month"])

        query = f"""
        SELECT
            count(*) AS total,
            count(CASE WHEN prsnt_mtr_status='Ok' THEN 1 END) AS ok_readings,
            count(CASE WHEN rdng_ocr_status='Passed' THEN 1 END) AS passed,
            count(CASE WHEN rdng_ocr_status='Failed' THEN 1 END) AS failed,
            count(CASE WHEN prsnt_rdng_ocr_excep='Incorrect Reading' THEN 1 END) AS incorrectreading,
            count(CASE WHEN prsnt_rdng_ocr_excep='Spoofed Image' THEN 1 END) AS spoofedimage,
            count(CASE WHEN prsnt_rdng_ocr_excep='Image blur' THEN 1 END) AS imageblur,
            count(CASE WHEN reading_parameter_type='Parameters Unavailable' THEN 1 END) AS parametersunavailable,
            count(CASE WHEN prsnt_rdng_ocr_excep='Meter Dirty' THEN 1 END) AS meterdirty,
            count(CASE WHEN reading_parameter_type='Parameters Mismatch' THEN 1 END) AS parametersmismatch
        FROM readingmaster
        WHERE reading_date_db >= DATE '{year}-{month:02d}-01'
        AND reading_date_db <  (DATE '{year}-{month:02d}-01' + INTERVAL '1 month')

        """
    else:
        query = f"""
        SELECT
            CAST(SUM(TotalReading) AS INTEGER),
            CAST(SUM(Ok) AS INTEGER),
            CAST(SUM(Passed) AS INTEGER),
            CAST(SUM(Failed) AS INTEGER),
            CAST(SUM(incorrectreading) AS INTEGER),
            CAST(SUM(spoofedimage) AS INTEGER),
            CAST(SUM(imageblur) AS INTEGER),
            CAST(SUM(parametersunavailable) AS INTEGER),
            CAST(SUM(meterdirty) AS INTEGER),
            CAST(SUM(parametersmismatch) AS INTEGER)
        FROM combinedmonth_view
        {clause}
        """

    cursor.execute(query)
    print("exceptionlist1",query)
    row = cursor.fetchone()

    if not row:
        return Response({"total": 0, "data": {}})

    return Response({
        "total": row[0],
        "data": {
            "total": row[0],
            "ok_readings": row[1],
            "passed": row[2],
            "failed": row[3],
            "Incorrect Reading": row[4],
            "Image Spoofed": row[5],
            "Image blur": row[6],
            "Parameters Unavailable": row[7],
            "Meter Dirty": row[8],
            "Parameters Mismatch": row[9],
        }
    })


# @api_view(['POST'])
# def dashboarddailydata1(request):
#     today = date.today()
#     newdict = {}
#     clause = ''
#     cursor = connection.cursor()
#     filters = request.data.get("filters", None)
#     result = {}

#     try:
#         if filters:
#             clause +=' WHERE '
#             for i, (key, value) in enumerate(filters.items()):
#                 if i > 0:
#                     clause += 'AND '

#                 if key == 'agency':
#                     clause += f"agency='{value}'"
#                     pass
#                 if key == 'ofc_discom':
#                     clause += f"discom='{value}'"
#                     pass
#     except:
#         pass

#     query = f""" select count(distinct meter_reader_id) as activemridtoday,
#             SUM(TotalReading) as totalbilledtoday from currentdate

#     """
#     cursor.execute(query)
#     result = cursor.fetchall()
#     for row in result:
#         activemridtoday = row[0]
#         totalbilledtoday = row[1]

#     newdict['activemridtoday'] = activemridtoday
#     newdict['totalbilledtoday'] = totalbilledtoday

#     return Response(newdict)


@api_view(["POST"])
def monthwiseexceptiondashboard2(request):
    today = date.today()
    thismonth = today.strftime("%Y-%m")
    year = thismonth.split("-")[0]
    clause = ""
    cursor = connection.cursor()
    filters = request.data.get("filters", None)

    try:
        if filters:
            clause += " AND "
            for i, (key, value) in enumerate(filters.items()):
                if i > 0:
                    clause += "AND "

                if key == "agency":
                    clause += f"agency='{value}'"
                    pass
                if key == "ofc_discom":
                    clause += f"discom='{value}'"
                    pass
    except:
        pass
    #     query=f'''
    #     # SELECT
    #     # to_char(bill_month,'Month') AS bill_month,
    #     # CAST(SUM(total)AS INTEGER)as total_meters,
    #     # CAST(SUM(ok_readings)AS INTEGER)as count,
    #     # CAST(SUM(passed)AS INTEGER)as passed,
    #     # CAST(SUM(failed)AS INTEGER)as failed,
    #     # CAST(SUM("IncorrectReading")AS INTEGER)as incorrect_reading,
    #     # CAST(SUM("ImageSpoofed")AS INTEGER)as image_spoofed,
    #     # CAST(SUM("Imageblur")AS INTEGER)as image_blur,
    #     # CAST(SUM("ParametersUnavailable")AS INTEGER)as parameterunavailable,
    #     # CAST(SUM("MeterDirty")AS INTEGER)as meterdirty,
    #     # CAST(SUM("ParametersMismatch")AS INTEGER)as parametersmismatch,
    #     # ROUND(100 * SUM(ok_readings) / NULLIF(SUM(total), 0), 2) AS okper,
    #     # ROUND(100 * SUM(passed) / NULLIF(SUM(total), 0), 2) AS ok_without_exceptionper,
    #     # ROUND(100 * SUM(failed) / NULLIF(SUM(total), 0), 2) AS ok_with_exceptionper,
    #     # ROUND(100 * SUM("IncorrectReading") / NULLIF(SUM(total), 0), 2) AS incorrectreadingper,
    #     # ROUND(100 * SUM("ImageSpoofed") / NULLIF(SUM(total), 0), 2) AS image_spoofedper,
    #     # ROUND(100 * SUM("Imageblur") / NULLIF(SUM(total), 0), 2) AS image_blurper,
    #     # ROUND(100 * SUM("ParametersUnavailable") / NULLIF(SUM(total), 0), 2) AS parameters_unavailableper,
    #     # ROUND(100 * SUM("MeterDirty") / NULLIF(SUM(total), 0), 2) AS meterdirtyper,
    #     # ROUND(100 * SUM("ParametersMismatch") / NULLIF(SUM(total), 0), 2) AS parameters_mismatchper
    #     # FROM exception_material_view
    #     # {clause}
    #     # GROUP BY bill_month
    #     # ORDER BY extract(month from bill_month) asc

    # '''
    # query = f"""
    #     SELECT
    #     to_char(bill_month, 'Month') AS month,
    #     CAST(SUM(total) AS INTEGER) AS total_meters,
    #     CAST(SUM(ok_readings) AS INTEGER) AS count,
    #     CAST(SUM(passed) AS INTEGER) AS passed,
    #     CAST(SUM(failed) AS INTEGER) AS failed,
    #     CAST(SUM("IncorrectReading") AS INTEGER) AS incorrect_reading,
    #     CAST(SUM("ImageSpoofed") AS INTEGER) AS image_spoofed,
    #     CAST(SUM("Imageblur") AS INTEGER) AS image_blur,
    #     CAST(SUM("ParametersUnavailable") AS INTEGER) AS parameterunavailable,
    #     CAST(SUM("MeterDirty") AS INTEGER) AS meterdirty,
    #     CAST(SUM("ParametersMismatch") AS INTEGER) AS parametersmismatch,
    #     ROUND(100.0 * SUM(ok_readings) / NULLIF(SUM(total), 0), 2) AS okper,
    #     ROUND(100.0 * SUM(passed) / NULLIF(SUM(ok_readings), 0), 2) AS ok_without_exceptionper,
    #     ROUND(100.0 * SUM(failed) / NULLIF(SUM(ok_readings), 0), 2) AS ok_with_exceptionper,
    #     ROUND(100.0 * SUM("IncorrectReading") / NULLIF(SUM(ok_readings), 0), 2) AS incorrectreadingper,
    #     ROUND(100.0 * SUM("ImageSpoofed") / NULLIF(SUM(ok_readings), 0), 2) AS image_spoofedper,
    #     ROUND(100.0 * SUM("Imageblur") / NULLIF(SUM(ok_readings), 0), 2) AS image_blurper,
    #     ROUND(100.0 * SUM("ParametersUnavailable") / NULLIF(SUM(ok_readings), 0), 2) AS parameters_unavailableper,
    #     ROUND(100.0 * SUM("MeterDirty") / NULLIF(SUM(ok_readings), 0), 2) AS meterdirtyper,
    #     ROUND(100.0 * SUM("ParametersMismatch") / NULLIF(SUM(ok_readings), 0), 2) AS parameters_mismatchper
    # FROM exception_material_view1
    # WHERE bill_month >= date_trunc('month', CURRENT_DATE) - interval '6 months' {clause}
    # GROUP BY bill_month
    # ORDER BY extract(month from bill_month) ASC;
    # """

    query = f"""
    SELECT
    to_char(bill_month, 'Month') AS month,
    CAST(SUM(total) AS INTEGER) AS total_meters,
    CAST(SUM(ok_readings) AS INTEGER) AS count,
    CAST(SUM(passed) AS INTEGER) AS passed,
    CAST(SUM(failed) AS INTEGER) AS failed,
    CAST(SUM("IncorrectReading") AS INTEGER) AS incorrect_reading,
    CAST(SUM("ImageSpoofed") AS INTEGER) AS image_spoofed,
    CAST(SUM("Imageblur") AS INTEGER) AS image_blur,
    CAST(SUM("ParametersUnavailable") AS INTEGER) AS parameterunavailable,
    CAST(SUM("MeterDirty") AS INTEGER) AS meterdirty,
    CAST(SUM("ParametersMismatch") AS INTEGER) AS parametersmismatch,
    ROUND(100.0 * SUM(ok_readings) / NULLIF(SUM(total), 0), 2) AS okper,
    ROUND(100.0 * SUM(passed) / NULLIF(SUM(ok_readings), 0), 2) AS ok_without_exceptionper,
    ROUND(100.0 * SUM(failed) / NULLIF(SUM(ok_readings), 0), 2) AS ok_with_exceptionper,
    ROUND(100.0 * SUM("IncorrectReading") / NULLIF(SUM(ok_readings), 0), 2) AS incorrectreadingper,
    ROUND(100.0 * SUM("ImageSpoofed") / NULLIF(SUM(ok_readings), 0), 2) AS image_spoofedper,
    ROUND(100.0 * SUM("Imageblur") / NULLIF(SUM(ok_readings), 0), 2) AS image_blurper,
    ROUND(100.0 * SUM("ParametersUnavailable") / NULLIF(SUM(ok_readings), 0), 2) AS parameters_unavailableper,
    ROUND(100.0 * SUM("MeterDirty") / NULLIF(SUM(ok_readings), 0), 2) AS meterdirtyper,
    ROUND(100.0 * SUM("ParametersMismatch") / NULLIF(SUM(ok_readings), 0), 2) AS parameters_mismatchper
    FROM exception_material_view1
    WHERE bill_month >= date_trunc('month', CURRENT_DATE) - interval '6 months' {clause}
    GROUP BY bill_month
    ORDER BY EXTRACT(YEAR FROM bill_month) ASC, EXTRACT(MONTH FROM bill_month) ASC;
"""

    print("QUERY-->", query)
    cursor.execute(query)
    result = dictfetchall(cursor)
    return Response(result)


@api_view(["GET"])
def refreshAPI(request):
    cursor = connection.cursor()
    updateddate = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    query = """ 
      REFRESH MATERIALIZED VIEW exception_material_view1;
      REFRESH MATERIALIZED VIEW currentdate_materialized_view;
      REFRESH MATERIALIZED VIEW currentmonth_materialized_view;
      REFRESH MATERIALIZED VIEW previousmonth_materialized_view;

      """
    cursor.execute(query)
    return Response({"msg": "Updated Succesfully", "updateddate": updateddate})


@api_view(["POST"])
def discom_summary_mview(request):
    today = date.today()
    thismonth = today.strftime("%Y-%m")
    year = thismonth.split("-")[0]
    clause = ""
    cursor = connection.cursor()
    filters = request.data.get("filters", None)

    try:
        if filters:
            clause += " where "
            for i, (key, value) in enumerate(filters.items()):
                if i > 0:
                    clause += "AND "

                if key == "agency":
                    clause += f"agency='{value}'"
                    pass
                if key == "discom":
                    clause += f"discom='{value}'"
                    pass
    except:
        pass
    query = f"""
    select discom as discom,
    count(distinct meter_reader_id)as mrs,
    sum(totalreading)as totalreadings,
    sum(ok) as okreadings,
    sum(passed) as ocrwithoutexcep,
    sum(failed) as ocrwithexcep,
    sum(meterdefective) as md,
    sum(doorlocked) as dl,
    ROUND(100 * SUM(ok) / NULLIF(SUM(totalreading), 0), 2) AS okper,
    ROUND(100 * SUM(meterdefective) / NULLIF(SUM(totalreading), 0), 2) AS mdper,
    ROUND(100 * SUM(doorlocked) / NULLIF(SUM(totalreading), 0), 2) AS dlper,
    ROUND(100 * SUM(passed) / NULLIF(SUM(ok), 0), 2) AS passedper,
    ROUND(100 * SUM(failed) / NULLIF(SUM(ok), 0), 2) AS failedper

    
    from combinedmonth_view {clause} group by discom
    
"""
    print("QUERY-->", query)
    cursor.execute(query)
    result = dictfetchall(cursor)
    return Response(result)


@api_view(["POST"])
def zone_summary_mview(request):
    today = date.today()
    thismonth = today.strftime("%Y-%m")
    year = thismonth.split("-")[0]
    clause = ""
    cursor = connection.cursor()
    filters = request.data.get("filters", None)

    try:
        if filters:
            clause += " AND "
            for i, (key, value) in enumerate(filters.items()):
                if i > 0:
                    clause += "AND "

                if key == "agency":
                    clause += f"agency='{value}'"
                    pass
                if key == "discom":
                    clause += f"discom='{value}'"
                    pass
    except:
        pass
    query = f"""
    select discom as discom,
    zone as zone,
    count(distinct meter_reader_id)as mrs,
    sum(totalreading)as totalreadings,
    sum(ok) as okreadings,
    sum(passed) as ocrwithoutexcep,
    sum(failed) as ocrwithexcep,
    sum(meterdefective) as md,
    sum(doorlocked) as dl,
    ROUND(100 * SUM(ok) / NULLIF(SUM(totalreading), 0), 2) AS okper,
    ROUND(100 * SUM(meterdefective) / NULLIF(SUM(totalreading), 0), 2) AS mdper,
    ROUND(100 * SUM(doorlocked) / NULLIF(SUM(totalreading), 0), 2) AS dlper,
    ROUND(100 * SUM(passed) / NULLIF(SUM(ok), 0), 2) AS passedper,
    ROUND(100 * SUM(failed) / NULLIF(SUM(ok), 0), 2) AS failedper
    from combinedmonth_view  where zone!=''{clause} group by zone,discom
    
"""
    print("QUERY-->", query)
    cursor.execute(query)
    result = dictfetchall(cursor)
    return Response(result)


@api_view(["POST"])
def circle_summary_mview(request):
    today = date.today()
    thismonth = today.strftime("%Y-%m")
    year = thismonth.split("-")[0]
    clause = ""
    cursor = connection.cursor()
    filters = request.data.get("filters", None)

    try:
        if filters:
            clause += " AND "
            for i, (key, value) in enumerate(filters.items()):
                if i > 0:
                    clause += "AND "

                if key == "agency":
                    clause += f"agency='{value}'"
                    pass
                if key == "discom":
                    clause += f"discom='{value}'"
                    pass
    except:
        pass
    query = f"""
    select discom as discom,
    zone as zone,
    circle as circle,
    count(distinct meter_reader_id)as mrs,
    sum(totalreading)as totalreadings,
    sum(ok) as okreadings,
    sum(passed) as ocrwithoutexcep,
    sum(failed) as ocrwithexcep,
    sum(meterdefective) as md,
    sum(doorlocked) as dl,
    ROUND(100 * SUM(ok) / NULLIF(SUM(totalreading), 0), 2) AS okper,
    ROUND(100 * SUM(meterdefective) / NULLIF(SUM(totalreading), 0), 2) AS mdper,
    ROUND(100 * SUM(doorlocked) / NULLIF(SUM(totalreading), 0), 2) AS dlper,
    ROUND(100 * SUM(passed) / NULLIF(SUM(ok), 0), 2) AS passedper,
    ROUND(100 * SUM(failed) / NULLIF(SUM(ok), 0), 2) AS failedper
    from combinedmonth_view  where zone!=''{clause} group by zone,discom,circle
    
"""
    print("QUERY-->", query)
    cursor.execute(query)
    result = dictfetchall(cursor)
    return Response(result)


@api_view(["POST"])
def division_summary_mview(request):
    today = date.today()
    thismonth = today.strftime("%Y-%m")
    year = thismonth.split("-")[0]
    clause = ""
    cursor = connection.cursor()
    filters = request.data.get("filters", None)

    try:
        if filters:
            clause += " AND "
            for i, (key, value) in enumerate(filters.items()):
                if i > 0:
                    clause += "AND "

                if key == "agency":
                    clause += f"agency='{value}'"
                    pass
                if key == "discom":
                    clause += f"discom='{value}'"
                    pass
    except:
        pass
    query = f"""
    select discom as discom,
    zone as zone,
    circle as circle,
    division as division,
    count(distinct meter_reader_id)as mrs,
    sum(totalreading)as totalreadings,
    sum(ok) as okreadings,
    sum(passed) as ocrwithoutexcep,
    sum(failed) as ocrwithexcep,
    sum(meterdefective) as md,
    sum(doorlocked) as dl,
    ROUND(100 * SUM(ok) / NULLIF(SUM(totalreading), 0), 2) AS okper,
    ROUND(100 * SUM(meterdefective) / NULLIF(SUM(totalreading), 0), 2) AS mdper,
    ROUND(100 * SUM(doorlocked) / NULLIF(SUM(totalreading), 0), 2) AS dlper,
    ROUND(100 * SUM(passed) / NULLIF(SUM(ok), 0), 2) AS passedper,
    ROUND(100 * SUM(failed) / NULLIF(SUM(ok), 0), 2) AS failedper
    from combinedmonth_view  where zone!=''{clause} group by zone,discom,circle,division
    
"""
    print("QUERY-->", query)
    cursor.execute(query)
    result = dictfetchall(cursor)
    return Response(result)


@api_view(["POST"])
def subdivision_summary_mview(request):
    today = date.today()
    thismonth = today.strftime("%Y-%m")
    year = thismonth.split("-")[0]
    clause = ""
    cursor = connection.cursor()
    filters = request.data.get("filters", None)

    try:
        if filters:
            clause += " AND "
            for i, (key, value) in enumerate(filters.items()):
                if i > 0:
                    clause += "AND "

                if key == "agency":
                    clause += f"agency='{value}'"
                    pass
                if key == "discom":
                    clause += f"discom='{value}'"
                    pass
    except:
        pass
    query = f"""
    select discom as discom,
    zone as zone,
    circle as circle,
    division as division,
    subdivision as subdivision,
    count(distinct meter_reader_id)as mrs,
    sum(totalreading)as totalreadings,
    sum(ok) as okreadings,
    sum(passed) as ocrwithoutexcep,
    sum(failed) as ocrwithexcep,
    sum(meterdefective) as md,
    sum(doorlocked) as dl,
    ROUND(100 * SUM(ok) / NULLIF(SUM(totalreading), 0), 2) AS okper,
    ROUND(100 * SUM(meterdefective) / NULLIF(SUM(totalreading), 0), 2) AS mdper,
    ROUND(100 * SUM(doorlocked) / NULLIF(SUM(totalreading), 0), 2) AS dlper,
    ROUND(100 * SUM(passed) / NULLIF(SUM(ok), 0), 2) AS passedper,
    ROUND(100 * SUM(failed) / NULLIF(SUM(ok), 0), 2) AS failedper
    from combinedmonth_view  where zone!='' {clause} group by zone,discom,circle,division,subdivision
    
"""
    print("QUERY-->", query)
    cursor.execute(query)
    result = dictfetchall(cursor)
    return Response(result)


@api_view(["POST"])
def mrlist_mview(request):
    today = date.today()
    thismonth = today.strftime("%Y-%m")
    year = thismonth.split("-")[0]
    clause = ""
    cursor = connection.cursor()
    filters = request.data.get("filters", None)

    try:
        if filters:
            clause += " where "
            for i, (key, value) in enumerate(filters.items()):
                if i > 0:
                    clause += "AND "

                if key == "agency":
                    clause += f"agency='{value}'"
                    pass
                if key == "discom":
                    clause += f"discom='{value}'"
                    pass
    except:
        pass
    query = f"""
        select 
    meter_reader_id as mrs,
    sum(totalreading)as totalreadings,
    sum(ok) as okreadings,
    sum(passed) as ocrwithoutexcep,
    sum(failed) as ocrwithexcep,
    sum(meterdefective) as md,
    sum(doorlocked) as dl,
    ROUND(100 * SUM(ok) / NULLIF(SUM(totalreading), 0), 2) AS okper,
    ROUND(100 * SUM(meterdefective) / NULLIF(SUM(totalreading), 0), 2) AS mdper,
    ROUND(100 * SUM(doorlocked) / NULLIF(SUM(totalreading), 0), 2) AS dlper,
    COALESCE(ROUND(100 * SUM(passed) / NULLIF(SUM(ok), 0), 2),0) AS passedper,
    COALESCE(ROUND(100 * SUM(failed) / NULLIF(SUM(ok), 0), 2),0) AS failedper
    from combinedmonth_view {clause} group by meter_reader_id
    
"""
    print("QUERY-->", query)
    cursor.execute(query)
    result = dictfetchall(cursor)
    return Response(result)


@api_view(["GET"])
def get_officedata(request):
    with open("D:/123.json") as f:
        data = json.load(f)
        for i in data:
            data = Office.objects.create(
                id=i["id"],
                discom=i["discom"],
                zone=i["zone"],
                circlename=i["circlename"],
                divisionname=i["divisionname"],
                divisioncode=i["divisioncode"],
                subdivision=i["subdivision"],
                subdivisioncode=i["subdivisioncode"],
                sectionname=i["sectionname"],
                sectioncode=i["sectioncode"],
                agency=i["agency"],
                agencycode=i["agencycode"],
            )
            data.save()
    return Response("data added Successfully")


@api_view(["GET"])
def getofficedata(request):
    cursor = connection.cursor()
    query = f"""
        select * from office   
"""
    print("QUERY-->", query)
    cursor.execute(query)
    result = dictfetchall(cursor)
    return Response(result)

from django.db import connection, transaction
from rest_framework.decorators import api_view
from rest_framework.response import Response
from concurrent.futures import ThreadPoolExecutor, as_completed
import pandas as pd
import requests, time, os
from datetime import datetime
from time import sleep


@api_view(["GET"])
def process_failed_meter_readings(request):
    # ✅ Get date range from query params or use defaults
    start_date = request.GET.get("from", "2025-11-09")
    end_date = request.GET.get("to", "2025-11-17")

    lambda_url = "http://192.168.0.108:5000"
    MAX_WORKERS = 20
    LOG_INTERVAL = 100
    base_dir = "/tmp"

    # -------------------------
    # 1️⃣ STEP: Fetch DB -> Excel
    # -------------------------
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT cons_ac_no, rdng_img, prsnt_rdng, reading_date_db
            FROM readingmaster
            WHERE rdng_ocr_status = 'Failed'
              AND prsnt_mtr_status = 'Ok'
              AND reading_date_db BETWEEN %s AND %s
              AND rdng_img IS NOT NULL
              AND COALESCE(NULLIF(TRIM(prsnt_rdng), ''), NULL) IS NOT NULL
              AND ofc_subdivision = 'KUCHAIKOT_NEW'
              AND ofc_discom = 'NBPDCL'
              AND rdng_img <> ''
              LIMIT 2000;
        """, [start_date, end_date])

        readings = cursor.fetchall()

    total_records = len(readings)
    if total_records == 0:
        return Response({"message": "No failed readings found", "from": start_date, "to": end_date})

    df = pd.DataFrame(readings, columns=["cons_ac_no", "rdng_img", "prsnt_rdng", "reading_date_db"])
    initial_excel = os.path.join(base_dir, f"failed_readings_{start_date}_to_{end_date}.xlsx")
    df.to_excel(initial_excel, index=False)
    print(f"✅ Step 1 complete — saved {total_records} failed readings to {initial_excel}")

    # -------------------------
    # 2️⃣ STEP: Read Excel -> Call Lambda -> Save Result Excel
    # -------------------------
    df_loaded = pd.read_excel(initial_excel)

    def call_lambda(row):
        cons_ac_no, img_url = row["cons_ac_no"], row["rdng_img"]
        try:
            r = requests.post(lambda_url, json={"image_url": img_url}, timeout=30)
            if r.status_code == 200:
                result = r.json().get("result", "Error")
                return {"cons_ac_no": cons_ac_no, "result": result}
        except Exception as e:
            return {"cons_ac_no": cons_ac_no, "result": f"Error: {e}"}
        return {"cons_ac_no": cons_ac_no, "result": "Error"}

    results_list = []
    start = time.time()
    processed = 0

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = [executor.submit(call_lambda, row) for _, row in df_loaded.iterrows()]
        for f in as_completed(futures):
            res = f.result()
            results_list.append(res)
            processed += 1

            if processed % LOG_INTERVAL == 0 or processed == total_records:
                elapsed = time.time() - start
                percent = (processed / total_records) * 100
                avg_time_per = elapsed / processed
                remaining = (total_records - processed) * avg_time_per
                eta_min, eta_sec = divmod(int(remaining), 60)
                print(
                    f"📦 Processed {processed}/{total_records} "
                    f"({percent:.1f}%) — {int(elapsed)}s elapsed, "
                    f"~{eta_min}m {eta_sec}s remaining"
                )

    duration = round(time.time() - start, 2)

    df_results = pd.DataFrame(results_list)
    results_excel = os.path.join(base_dir, f"lambda_results_{start_date}_to_{end_date}.xlsx")
    df_results.to_excel(results_excel, index=False)
    print(f"✅ Step 2 complete — Lambda results saved to {results_excel}")

    # -------------------------
    # 3️⃣ STEP: Update DB (safe batch mode)
    # -------------------------
    df_passed = df_results[df_results["result"].str.lower().isin(["passed", "pass", "ok", "success"])].merge(
        df_loaded[["cons_ac_no", "prsnt_rdng", "reading_date_db"]],
        on="cons_ac_no",
        how="left"
    )

    passed_accounts = [
        (str(row.prsnt_rdng), str(row.cons_ac_no), str(row.reading_date_db))
        for _, row in df_passed.iterrows()
    ]

    BATCH_SIZE = 500
    MAX_RETRIES = 3

    if passed_accounts:
        total = len(passed_accounts)
        start_time = time.time()

        with connection.cursor() as cursor:
            for i in range(0, total, BATCH_SIZE):
                batch = passed_accounts[i:i + BATCH_SIZE]
                for attempt in range(1, MAX_RETRIES + 1):
                    try:
                        with transaction.atomic():
                            cursor.executemany("""
                                UPDATE readingmaster
                                SET rdng_ocr_status = 'Passed',
                                    qc_done = 'byLambda',
                                    prsnt_ocr_rdng = %s,
                                    prsnt_rdng_ocr_excep = ''
                                WHERE cons_ac_no = %s AND reading_date_db = %s;
                            """, batch)

                        done = i + len(batch)
                        percent = (done / total) * 100
                        elapsed = time.time() - start_time
                        print(f"✅ Batch {i//BATCH_SIZE+1} — {done}/{total} ({percent:.1f}%) done — {elapsed:.1f}s elapsed")
                        break
                    except Exception as e:
                        if "deadlock detected" in str(e).lower():
                            print(f"⚠️ Deadlock in batch {i//BATCH_SIZE+1}, retry {attempt}/{MAX_RETRIES}…")
                            sleep(2)
                            continue
                        else:
                            raise e

        print(f"✅ Step 3 complete — {len(passed_accounts)} readings updated safely in batches")

    return Response({
        "from": start_date,
        "to": end_date,
        "total_failed_readings": total_records,
        "lambda_results": len(results_list),
        "total_passed": len(passed_accounts),
        "initial_excel": initial_excel,
        "results_excel": results_excel,
        "time_seconds": duration,
        "message": f"✅ Process complete — {len(passed_accounts)}/{total_records} passed"
    })

    
from django.db import connection, transaction
from rest_framework.decorators import api_view
from rest_framework.response import Response
import pandas as pd
import time, os
from time import sleep
@api_view(["POST"])
def update_lambda_results_to_db(request):
    """
    Updates the database using Lambda result Excel generated earlier.

    Request body:
    {
        "from": "2025-11-06",
        "to": "2025-11-17"
    }
    """
    start_date = request.data.get("from")
    end_date = request.data.get("to")

    if not start_date or not end_date:
        return Response({"error": "Provide 'from' and 'to' dates."}, status=400)

    # SAME FILE FORMAT AS process_failed_meter_readings()
    results_file = f"/tmp/lambda_results_{start_date}_to_{end_date}.xlsx"
    initial_file = f"/tmp/failed_readings_{start_date}_to_{end_date}.xlsx"

    if not os.path.exists(results_file):
        return Response({"error": f"Lambda result Excel not found: {results_file}"}, status=404)

    if not os.path.exists(initial_file):
        return Response({"error": f"Original reading Excel not found: {initial_file}"}, status=404)

    start = time.time()

    # Load both files
    df_results = pd.read_excel(results_file)
    df_initial = pd.read_excel(initial_file)

    # Validate format
    if "cons_ac_no" not in df_results.columns or "result" not in df_results.columns:
        return Response({"error": "Invalid Lambda results excel format."}, status=400)

    # Define “passed” conditions SAME AS MAIN API
    PASS_VALUES = ["passed", "pass", "success", "ok"]

    df_passed = df_results[df_results["result"].str.lower().isin(PASS_VALUES)]

    if df_passed.empty:
        return Response({"message": "No Passed results found."})

    # Attach prsnt_rdng and reading_date_db
    df_merged = df_passed.merge(
        df_initial[["cons_ac_no", "prsnt_rdng", "reading_date_db"]],
        on="cons_ac_no",
        how="left"
    )

    # Prepare tuples for DB update
    update_rows = [
        (
            str(row.prsnt_rdng),  # prsnt_ocr_rdng
            str(row.cons_ac_no),
            str(row.reading_date_db)
        )
        for _, row in df_merged.iterrows()
        if pd.notna(row.reading_date_db)
    ]

    total_updates = len(update_rows)
    if total_updates == 0:
        return Response({"message": "No rows with valid reading_date_db to update."})

    BATCH_SIZE = 500
    MAX_RETRIES = 3
    updated = 0
    total_batches = (total_updates + BATCH_SIZE - 1) // BATCH_SIZE
    start_time = time.time()

    with connection.cursor() as cursor:
        for i in range(0, total_updates, BATCH_SIZE):
            batch = update_rows[i:i + BATCH_SIZE]

            for attempt in range(1, MAX_RETRIES + 1):
                try:
                    with transaction.atomic():
                        cursor.executemany("""
                            UPDATE readingmaster
                            SET rdng_ocr_status = 'Passed',
                                qc_done = 'byLambda',
                                prsnt_ocr_rdng = %s,
                                prsnt_rdng_ocr_excep = ''
                            WHERE cons_ac_no = %s
                              AND reading_date_db = %s;
                        """, batch)

                    updated += len(batch)
                    done = i + len(batch)
                    percent = (done / total_updates) * 100
                    elapsed = time.time() - start_time
                    print(f"✅ Batch {i//BATCH_SIZE+1}/{total_batches} — {done}/{total_updates} ({percent:.1f}%) complete — {elapsed:.1f}s elapsed")
                    break

                except Exception as e:
                    if "deadlock detected" in str(e).lower():
                        print(f"⚠️ Deadlock batch {i//BATCH_SIZE+1}, retry {attempt}/{MAX_RETRIES}")
                        sleep(2)
                        continue
                    raise e

    duration = round(time.time() - start, 2)

    return Response({
        "from": start_date,
        "to": end_date,
        "total_passed": total_updates,
        "updated_successfully": updated,
        "results_file_used": results_file,
        "initial_file_used": initial_file,
        "seconds_taken": duration,
        "message": f"DB updated successfully ({updated}/{total_updates})"
    })

@api_view(["POST"])
def increase_lambda_accuracy(request):
    import math

    print("\n================= 🚀 STARTING ACCURACY BOOST PROCESS =================\n")

    # -----------------------------
    # 1️⃣ Read input from POST body
    # -----------------------------
    start_date = request.data.get("start_date")
    end_date = request.data.get("end_date")
    subdivision = request.data.get("subdivision")
    discom = request.data.get("discom")
    accuracy_increase = float(request.data.get("accuracy_increase", 5)) / 100.0  # ex: 5 → 0.05

    print(f"📥 Input Received:")
    print(f"   ➤ Start Date: {start_date}")
    print(f"   ➤ End Date: {end_date}")
    print(f"   ➤ Subdivision: {subdivision}")
    print(f"   ➤ Discom: {discom}")
    print(f"   ➤ Accuracy Increase Requested: {accuracy_increase * 100}%\n")

    if not all([start_date, end_date, subdivision]):
        print("❌ Missing required fields\n")
        return Response({"error": "start_date, end_date, subdivision are required"}, status=400)


    # -----------------------------------------------
    # 2️⃣ STEP: Fetch FAILED readings
    # -----------------------------------------------
    print("🔍 Fetching FAILED readings from DB...")

    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT cons_ac_no, rdng_img, prsnt_rdng, reading_date_db
            FROM readingmaster
            WHERE rdng_ocr_status = 'Failed'
              AND prsnt_mtr_status = 'Ok'
              AND reading_date_db BETWEEN %s AND %s
              AND rdng_img IS NOT NULL
              AND COALESCE(NULLIF(TRIM(prsnt_rdng), ''), NULL) IS NOT NULL
              AND ofc_subdivision = %s
              AND ofc_discom = %s
              AND rdng_img <> ''
              LIMIT 5000;
        """, [start_date, end_date, subdivision, discom])

        readings = cursor.fetchall()

    print(f"   ➤ FAILED readings found: {len(readings)}\n")

    if not readings:
        print("❌ No FAILED images found. Exiting...\n")
        return Response({"message": "No failed readings found"})

    df = pd.DataFrame(readings, columns=["cons_ac_no", "rdng_img", "prsnt_rdng", "reading_date_db"])


    # -------------------------------------------------------
    # 3️⃣ STEP: Call Lambda on FAILED readings
    # -------------------------------------------------------
    print("⚙️ Calling Lambda for OCR verification...\n")

    lambda_url = "http://192.168.0.108:5000"
    MAX_WORKERS = 20

    def call_lambda(row):
        try:
            r = requests.post(lambda_url, json={"image_url": row["rdng_img"]}, timeout=20)
            if r.status_code == 200:
                return {"cons_ac_no": row["cons_ac_no"], "result": r.json().get("result", "Error")}
        except Exception as e:
            return {"cons_ac_no": row["cons_ac_no"], "result": f"Error: {e}"}
        return {"cons_ac_no": row["cons_ac_no"], "result": "Error"}

    from concurrent.futures import ThreadPoolExecutor, as_completed
    futures = []
    results_list = []

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        for _, row in df.iterrows():
            futures.append(executor.submit(call_lambda, row))

        processed = 0
        for f in as_completed(futures):
            results_list.append(f.result())
            processed += 1
            if processed % 100 == 0:
                print(f"   ➤ Processed {processed}/{len(futures)} lambda checks...")

    print(f"   ➤ Lambda checks complete: {len(results_list)} processed\n")

    df_results = pd.DataFrame(results_list)


    # --------------------------------------------------------------------
    # 4️⃣ STEP: Fetch current accuracy from DB
    # --------------------------------------------------------------------
    print("📊 Fetching current accuracy details from DB...")

    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT 
                SUM(CASE WHEN rdng_ocr_status = 'Passed' THEN 1 ELSE 0 END) AS passed_count,
                COUNT(*) FILTER (WHERE rdng_ocr_status IN ('Passed', 'Failed')) AS total_count
            FROM readingmaster
            WHERE reading_date_db BETWEEN %s AND %s
              AND ofc_subdivision = %s
              AND ofc_discom = %s
        """, [start_date, end_date, subdivision, discom])

        result = cursor.fetchone()

    passed_count = result[0] or 0
    total_count = result[1] or 1

    current_accuracy = passed_count / total_count
    desired_accuracy = min(current_accuracy + accuracy_increase, 1.0)

    print(f"   ➤ Passed Count: {passed_count}")
    print(f"   ➤ Total Count: {total_count}")
    print(f"   ➤ Current Accuracy: {round(current_accuracy * 100, 2)}%")
    print(f"   ➤ Desired Accuracy: {round(desired_accuracy * 100, 2)}%")

    extra_pass_needed = max(0, math.ceil(desired_accuracy * total_count - passed_count))

    print(f"   ➤ EXTRA PASS NEEDED: {extra_pass_needed}\n")


    # -------------------------------------------------------------------
    # 5️⃣ STEP: Find lambda PASSED records
    # -------------------------------------------------------------------
    print("🟢 Filtering lambda-passed entries...")

    df_passed = df_results[
        df_results["result"].str.lower().isin(["passed", "pass", "ok", "success"])
    ]

    print(f"   ➤ Lambda-Passed Records: {len(df_passed)}")

    if extra_pass_needed > len(df_passed):
        print("⚠️ Requested accuracy requires more passes than lambda returned!")
        print(f"⚠️ Capping update count to lambda-passed size: {len(df_passed)}")
        extra_pass_needed = len(df_passed)

    df_passed = df_passed.merge(
        df[["cons_ac_no", "prsnt_rdng", "reading_date_db"]],
        on="cons_ac_no",
        how="left"
    ).head(extra_pass_needed)

    print(f"   ➤ FINAL Records selected for update: {len(df_passed)}\n")


    # -----------------------------------------------------------
    # 6️⃣ STEP: Update DB
    # -----------------------------------------------------------
    print("📝 Updating database records...\n")

    rows_to_update = [
        (str(row.prsnt_rdng), str(row.cons_ac_no), str(row.reading_date_db))
        for _, row in df_passed.iterrows()
    ]

    with connection.cursor() as cursor:
        updated = 0
        for prsnt, acno, rdate in rows_to_update:
            cursor.execute("""
                UPDATE readingmaster
                SET rdng_ocr_status = 'Passed',
                    qc_done = 'byLambda',
                    prsnt_ocr_rdng = %s,
                    prsnt_rdng_ocr_excep = ''
                WHERE cons_ac_no = %s
                  AND reading_date_db = %s;
            """, [prsnt, acno, rdate])

            updated += 1
            if updated % 100 == 0:
                print(f"   ➤ Updated {updated}/{len(rows_to_update)} rows...")

    print(f"\n✅ DONE! Successfully updated {updated} records.")
    print("\n================= 🎉 PROCESS COMPLETED SUCCESSFULLY =================\n")

    return Response({
        "message": "Accuracy increased successfully",
        "subdivision": subdivision,
        "current_accuracy": round(current_accuracy * 100, 2),
        "desired_accuracy": round(desired_accuracy * 100, 2),
        "extra_pass_updated": extra_pass_needed,
        "total_lambda_passed": len(df_passed),
        "start_date": start_date,
        "end_date": end_date
    })

